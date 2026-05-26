"""Export do DRE em CSV ou XLSX (com folhas separadas).

Estrutura do XLSX:
  - DRE: KPIs principais + linhas por natureza (Directo/Variável/Fixo/Financeiro)
  - Centro de Custo: breakdown Bills + PFs + total por centro
  - Evolução 6 Meses: matriz centro × mês com totais

CSV: tudo num único ficheiro, secções separadas por linhas em branco.
"""
import csv
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from .views import (
    _compute_cc_evolution,
    _compute_dre_metrics,
    _resolve_dre_period,
)


def _dec(v):
    if isinstance(v, Decimal):
        return float(v)
    return v or 0


@login_required
def dre_export(request):
    """Export DRE em CSV (default) ou XLSX (?format=xlsx)."""
    date_from, date_to = _resolve_dre_period(request)
    cur = _compute_dre_metrics(date_from, date_to)
    evol = _compute_cc_evolution(date_to, n_months=6)
    fmt = (request.GET.get("format") or "csv").lower()
    fname_base = f"dre_{date_from:%Y%m%d}_{date_to:%Y%m%d}"

    if fmt == "xlsx":
        return _xlsx_response(cur, evol, fname_base)
    return _csv_response(cur, evol, fname_base, date_from, date_to)


def _csv_response(cur, evol, fname, date_from, date_to):
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{fname}.csv"'
    resp.write("﻿")  # BOM para Excel abrir UTF-8
    w = csv.writer(resp, delimiter=";")

    # Cabeçalho
    w.writerow(["DRE — Demonstração de Resultados"])
    w.writerow([f"Período: {date_from:%d/%m/%Y} a {date_to:%d/%m/%Y}"])
    w.writerow([])

    # 1. KPIs principais
    w.writerow(["KPI", "Valor (€)"])
    rows = [
        ("Receita total", cur["total_revenue"]),
        ("Custos directos (motoristas)", cur["total_driver_cost"]),
        ("Custos directos (returns)", cur["total_returns_cost"]),
        ("Custos directos extra (Bills)", cur["total_direct_extra"]),
        ("Total custos directos", cur["total_direct_op"]),
        ("Margem bruta", cur["margem_bruta"]),
        ("Custos variáveis", cur["total_variavel"]),
        ("Margem contribuição", cur["margem_contribuicao"]),
        ("Custos fixos", cur["total_fixo"]),
        ("EBITDA", cur["ebitda"]),
        ("Custos financeiros", cur["total_financeiro"]),
        ("Resultado líquido", cur["resultado_liquido"]),
    ]
    for label, val in rows:
        w.writerow([label, f"{_dec(val):.2f}"])
    w.writerow([])

    # 2. Receita por hub
    w.writerow(["Receita por HUB", "Entregas", "Receita (€)"])
    for h in cur["revenues_by_hub"]:
        w.writerow([h["hub_name"], h["deliveries"], f"{_dec(h['revenue']):.2f}"])
    w.writerow([])

    # 3. Linhas por natureza
    for nature_key, label in [
        ("direct_extra_lines", "Custos Directos Extra"),
        ("variavel_lines", "Custos Variáveis"),
        ("fixo_lines", "Custos Fixos"),
        ("financeiro_lines", "Custos Financeiros"),
    ]:
        lines = cur.get(nature_key, [])
        if not lines:
            continue
        w.writerow([label, "Nº", "Total (€)"])
        for it in lines:
            w.writerow([it["name"], it["n"], f"{_dec(it['total']):.2f}"])
        w.writerow([])

    # 4. Por centro de custo
    w.writerow([
        "Centro de Custo", "Tipo", "Bills (€)", "PFs Motorista (€)", "Total (€)",
    ])
    for cc_name, info in cur["by_cost_center"].items():
        w.writerow([
            cc_name,
            info.get("type") or "—",
            f"{_dec(info.get('from_bills')):.2f}",
            f"{_dec(info.get('from_drivers')):.2f}",
            f"{_dec(info.get('total')):.2f}",
        ])
    w.writerow([])

    # 5. Evolução 6 meses
    if evol["rows"]:
        header = ["Centro"] + [m["label"] for m in evol["months"]] + ["Total", "Δ%"]
        w.writerow(header)
        for row in evol["rows"]:
            cells = [row["cc"]]
            cells += [f"{_dec(v):.2f}" for v in row["values"]]
            cells.append(f"{_dec(row['total']):.2f}")
            cells.append(
                "novo" if row["delta_pct"] is None
                else f"{row['delta_pct']:.1f}%"
            )
            w.writerow(cells)
        w.writerow(
            ["Total mensal"]
            + [f"{_dec(t):.2f}" for t in evol["totals_by_month"]]
            + [f"{_dec(evol['grand_total']):.2f}", ""]
        )

    return resp


def _xlsx_response(cur, evol, fname):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(row_cells):
        for c in row_cells:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

    # ── Sheet 1: DRE ──
    ws = wb.active
    ws.title = "DRE"
    ws.append(["KPI", "Valor (€)"])
    style_header(ws[1])
    rows = [
        ("Receita total", cur["total_revenue"]),
        ("Custos directos (motoristas)", cur["total_driver_cost"]),
        ("Custos directos (returns)", cur["total_returns_cost"]),
        ("Custos directos extra (Bills)", cur["total_direct_extra"]),
        ("Total custos directos", cur["total_direct_op"]),
        ("Margem bruta", cur["margem_bruta"]),
        ("Custos variáveis", cur["total_variavel"]),
        ("Margem contribuição", cur["margem_contribuicao"]),
        ("Custos fixos", cur["total_fixo"]),
        ("EBITDA", cur["ebitda"]),
        ("Custos financeiros", cur["total_financeiro"]),
        ("Resultado líquido", cur["resultado_liquido"]),
    ]
    for label, val in rows:
        r = ws.append([label, _dec(val)])
        ws.cell(row=ws.max_row, column=2).number_format = "#,##0.00"
        if label in (
            "Receita total", "Margem bruta", "Margem contribuição",
            "EBITDA", "Resultado líquido",
        ):
            ws.cell(row=ws.max_row, column=1).font = bold
            ws.cell(row=ws.max_row, column=2).font = bold
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # ── Sheet 2: Receita por Hub ──
    ws2 = wb.create_sheet("Receita por HUB")
    ws2.append(["HUB", "Entregas", "Receita (€)"])
    style_header(ws2[1])
    for h in cur["revenues_by_hub"]:
        ws2.append([h["hub_name"], h["deliveries"], _dec(h["revenue"])])
        ws2.cell(row=ws2.max_row, column=3).number_format = "#,##0.00"
    ws2.column_dimensions["A"].width = 30

    # ── Sheet 3: Detalhe por natureza ──
    ws3 = wb.create_sheet("Detalhe por Natureza")
    ws3.append(["Natureza", "Categoria", "Nº", "Total (€)"])
    style_header(ws3[1])
    for nature_key, label in [
        ("direct_extra_lines", "Directo (extra)"),
        ("variavel_lines", "Variável"),
        ("fixo_lines", "Fixo"),
        ("financeiro_lines", "Financeiro"),
    ]:
        for it in cur.get(nature_key, []):
            ws3.append([label, it["name"], it["n"], _dec(it["total"])])
            ws3.cell(row=ws3.max_row, column=4).number_format = "#,##0.00"
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 35

    # ── Sheet 4: Centro de custo ──
    ws4 = wb.create_sheet("Centro de Custo")
    ws4.append(["Centro", "Tipo", "Bills (€)", "PFs Motorista (€)", "Total (€)"])
    style_header(ws4[1])
    for cc_name, info in cur["by_cost_center"].items():
        ws4.append([
            cc_name,
            info.get("type") or "—",
            _dec(info.get("from_bills")),
            _dec(info.get("from_drivers")),
            _dec(info.get("total")),
        ])
        for col in (3, 4, 5):
            ws4.cell(row=ws4.max_row, column=col).number_format = "#,##0.00"
    ws4.column_dimensions["A"].width = 28

    # ── Sheet 5: Evolução 6 meses ──
    if evol["rows"]:
        ws5 = wb.create_sheet("Evolução 6 Meses")
        header = (
            ["Centro"]
            + [m["label"] for m in evol["months"]]
            + ["Total", "Δ%"]
        )
        ws5.append(header)
        style_header(ws5[1])
        for row in evol["rows"]:
            cells = [row["cc"]]
            cells += [_dec(v) for v in row["values"]]
            cells.append(_dec(row["total"]))
            cells.append(
                "novo" if row["delta_pct"] is None
                else round(row["delta_pct"], 1)
            )
            ws5.append(cells)
            for col in range(2, 2 + len(evol["months"]) + 1):
                ws5.cell(row=ws5.max_row, column=col).number_format = "#,##0.00"
        # Linha total mensal
        totals_row = (
            ["Total mensal"]
            + [_dec(t) for t in evol["totals_by_month"]]
            + [_dec(evol["grand_total"]), ""]
        )
        ws5.append(totals_row)
        for col in range(1, len(totals_row) + 1):
            ws5.cell(row=ws5.max_row, column=col).font = bold
        for col in range(2, 2 + len(evol["months"]) + 1):
            ws5.cell(row=ws5.max_row, column=col).number_format = "#,##0.00"
        ws5.column_dimensions["A"].width = 28

    # Serialize
    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}.xlsx"'
    return resp
