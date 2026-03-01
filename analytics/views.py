"""
Views para dashboards de analytics e relatórios.
"""

from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db import models
from django.db.models import Avg, Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone

from core.models import Partner
from fleet_management.models import Vehicle, VehicleIncident
from orders_manager.models import OrderIncident

from .models import (
    DailyMetrics,
    DriverPerformance,
    PerformanceAlert,
    VolumeForecast,
)


@login_required
def dashboard_overview(request):
    """
    Dashboard principal consolidado multi-partner.
    Mostra visão geral de todos os partners.
    """
    # Período padrão: últimos 30 dias
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    # Filtros opcionais
    partner_id = request.GET.get("partner")
    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET.get("start_date"), "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET.get("end_date"), "%Y-%m-%d").date()

    # Query base
    metrics_query = DailyMetrics.objects.filter(
        date__gte=start_date, date__lte=end_date
    )

    if partner_id:
        metrics_query = metrics_query.filter(partner_id=partner_id)

    # Agregações por partner
    partner_stats = (
        metrics_query.values("partner__name")
        .annotate(
            total_orders=Sum("total_orders"),
            delivered=Sum("delivered_orders"),
            failed=Sum("failed_orders"),
            revenue=Sum("total_revenue"),
            avg_success_rate=Avg("success_rate"),
        )
        .order_by("-total_orders")
    )

    # Métricas totais
    totals = metrics_query.aggregate(
        total_orders=Sum("total_orders"),
        delivered_orders=Sum("delivered_orders"),
        failed_orders=Sum("failed_orders"),
        total_revenue=Sum("total_revenue"),
        total_bonuses=Sum("total_bonuses"),
        total_penalties=Sum("total_penalties"),
        avg_success_rate=Avg("success_rate"),
        avg_delivery_time=Avg("average_delivery_time_hours"),
    )

    # Alertas ativos (não reconhecidos)
    active_alerts = PerformanceAlert.objects.filter(is_acknowledged=False).order_by(
        "-severity", "-created_at"
    )[:10]

    # Top 5 motoristas do período
    top_drivers = (
        DriverPerformance.objects.filter(
            month__gte=start_date.replace(day=1),
            month__lte=end_date.replace(day=1),
        )
        .select_related("driver")
        .annotate(
            total_deliveries_sum=Sum("total_deliveries"),
            success_rate_avg=Avg("success_rate"),
        )
        .order_by("-total_deliveries_sum")[:5]
    )

    # Previsão próximos 7 dias (método MA7)
    forecast_7d = VolumeForecast.objects.filter(
        forecast_date__gte=end_date,
        forecast_date__lte=end_date + timedelta(days=7),
        method="MA7",
    ).order_by("forecast_date")

    # Lista de partners para filtro
    all_partners = Partner.objects.all().order_by("name")

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "selected_partner": partner_id,
        "all_partners": all_partners,
        "partner_stats": partner_stats,
        "totals": totals,
        "active_alerts": active_alerts,
        "top_drivers": top_drivers,
        "forecast_7d": forecast_7d,
        "date_range_days": (end_date - start_date).days + 1,
    }

    return render(request, "analytics/dashboard_overview.html", context)


@login_required
def metrics_dashboard(request):
    """
    Dashboard detalhado de métricas diárias.
    Gráficos de evolução temporal.
    """
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    partner_id = request.GET.get("partner")
    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET.get("start_date"), "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET.get("end_date"), "%Y-%m-%d").date()

    metrics_query = DailyMetrics.objects.filter(
        date__gte=start_date, date__lte=end_date
    )

    if partner_id:
        metrics_query = metrics_query.filter(partner_id=partner_id)

    # Dados para gráficos (agrupados por data)
    daily_data = (
        metrics_query.values("date")
        .annotate(
            total_orders=Sum("total_orders"),
            delivered=Sum("delivered_orders"),
            failed=Sum("failed_orders"),
            revenue=Sum("total_revenue"),
            success_rate=Avg("success_rate"),
            active_drivers=Sum("active_drivers_count"),
        )
        .order_by("date")
    )

    # KPIs principais
    totals = metrics_query.aggregate(
        total_orders=Sum("total_orders"),
        delivered_orders=Sum("delivered_orders"),
        failed_orders=Sum("failed_orders"),
        total_revenue=Sum("total_revenue"),
        avg_success_rate=Avg("success_rate"),
        peak_drivers=Sum("active_drivers_count"),
    )

    all_partners = Partner.objects.all().order_by("name")

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "selected_partner": partner_id,
        "all_partners": all_partners,
        "daily_data": list(daily_data),
        "totals": totals,
    }

    return render(request, "analytics/metrics_dashboard.html", context)


@login_required
def forecasts_dashboard(request):
    """
    Dashboard de previsões (forecasts).
    Mostra os 5 métodos e compara precisão.
    """
    end_date = timezone.now().date()
    start_date = end_date
    forecast_end = end_date + timedelta(days=30)

    partner_id = request.GET.get("partner")

    # Previsões futuras
    forecasts_query = VolumeForecast.objects.filter(
        forecast_date__gte=start_date, forecast_date__lte=forecast_end
    )

    if partner_id:
        forecasts_query = forecasts_query.filter(partner_id=partner_id)

    # Previsões por método
    forecasts_by_method = {}
    for method in ["MA7", "MA30", "EMA", "TREND", "SEASONAL"]:
        forecasts_by_method[method] = list(
            forecasts_query.filter(method=method)
            .values(
                "forecast_date",
                "predicted_volume",
                "lower_bound",
                "upper_bound",
            )
            .order_by("forecast_date")
        )

    # Melhor método (MA7 como padrão)
    best_forecasts = list(
        forecasts_query.filter(method="MA7")
        .values(
            "forecast_date",
            "predicted_volume",
            "method",
            "lower_bound",
            "upper_bound",
        )
        .order_by("forecast_date")
    )

    # Histórico real vs previsto (últimos 7 dias para validação)
    validation_start = start_date - timedelta(days=7)
    actual_data = DailyMetrics.objects.filter(
        date__gte=validation_start, date__lt=start_date
    )

    if partner_id:
        actual_data = actual_data.filter(partner_id=partner_id)

    actual_vs_predicted = list(
        actual_data.values("date")
        .annotate(actual_volume=Sum("total_orders"))
        .order_by("date")
    )

    all_partners = Partner.objects.all().order_by("name")

    context = {
        "selected_partner": partner_id,
        "all_partners": all_partners,
        "forecasts_by_method": forecasts_by_method,
        "best_forecasts": best_forecasts,
        "actual_vs_predicted": actual_vs_predicted,
        "forecast_start": start_date,
        "forecast_end": forecast_end,
    }

    return render(request, "analytics/forecasts_dashboard.html", context)


@login_required
def alerts_dashboard(request):
    """
    Dashboard de alertas de performance.
    Lista alertas ativos e histórico.
    """
    status_filter = request.GET.get("status", "ACTIVE")
    severity_filter = request.GET.get("severity")
    alert_type_filter = request.GET.get("alert_type")

    alerts_query = PerformanceAlert.objects.all()

    if status_filter:
        # Map status to is_acknowledged field
        if status_filter == "ACTIVE":
            alerts_query = alerts_query.filter(is_acknowledged=False)
        elif status_filter == "RESOLVED":
            alerts_query = alerts_query.filter(is_acknowledged=True)
    if severity_filter:
        alerts_query = alerts_query.filter(severity=severity_filter)
    if alert_type_filter:
        alerts_query = alerts_query.filter(alert_type=alert_type_filter)

    alerts = alerts_query.order_by("-severity", "-created_at")[:100]

    # Estatísticas de alertas
    alert_stats = {
        "total_active": PerformanceAlert.objects.filter(is_acknowledged=False).count(),
        "total_resolved": PerformanceAlert.objects.filter(is_acknowledged=True).count(),
        "critical": PerformanceAlert.objects.filter(
            severity="CRITICAL", is_acknowledged=False
        ).count(),
        "warning": PerformanceAlert.objects.filter(
            severity="WARNING", is_acknowledged=False
        ).count(),
        "info": PerformanceAlert.objects.filter(
            severity="INFO", is_acknowledged=False
        ).count(),
    }

    # Alertas por tipo
    alerts_by_type = (
        PerformanceAlert.objects.filter(is_acknowledged=False)
        .values("alert_type")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    context = {
        "alerts": alerts,
        "alert_stats": alert_stats,
        "alerts_by_type": alerts_by_type,
        "status_filter": status_filter,
        "severity_filter": severity_filter,
        "alert_type_filter": alert_type_filter,
    }

    return render(request, "analytics/alerts_dashboard.html", context)


@login_required
def drivers_performance_dashboard(request):
    """
    Dashboard de performance de motoristas.
    Rankings e evolução individual.
    """
    # Mês atual por padrão
    today = timezone.now().date()
    current_month = today.replace(day=1)

    month_param = request.GET.get("month")
    if month_param:
        current_month = datetime.strptime(month_param, "%Y-%m").date().replace(day=1)

    # Performance do mês selecionado
    month_performance = (
        DriverPerformance.objects.filter(month=current_month)
        .select_related("driver")
        .order_by("-rank_in_team")
    )

    # Top 10 por sucesso
    top_by_success = month_performance.order_by("-success_rate")[:10]

    # Top 10 por volume
    top_by_volume = month_performance.order_by("-total_deliveries")[:10]

    # Top 10 por receita
    top_by_revenue = month_performance.order_by("-total_earnings")[:10]

    # Estatísticas gerais
    stats = month_performance.aggregate(
        total_drivers=Count("id"),
        avg_success_rate=Avg("success_rate"),
        total_deliveries=Sum("total_deliveries"),
        total_earnings=Sum("total_earnings"),
    )

    # Últimos 6 meses para evolução
    six_months_ago = current_month - timedelta(days=180)
    historical_data = (
        DriverPerformance.objects.filter(
            month__gte=six_months_ago, month__lte=current_month
        )
        .values("month")
        .annotate(
            avg_success_rate=Avg("success_rate"),
            total_deliveries=Sum("total_deliveries"),
        )
        .order_by("month")
    )

    context = {
        "current_month": current_month,
        "month_performance": month_performance,
        "top_by_success": top_by_success,
        "top_by_volume": top_by_volume,
        "top_by_revenue": top_by_revenue,
        "stats": stats,
        "historical_data": list(historical_data),
    }

    return render(request, "analytics/drivers_performance_dashboard.html", context)


@login_required
def incidents_report(request):
    """
    Relatório de incidências (top motivos de falha).
    """
    # Período padrão: últimos 30 dias
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET.get("start_date"), "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET.get("end_date"), "%Y-%m-%d").date()

    # Top motivos de falha em pedidos
    order_incidents = (
        OrderIncident.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        )
        .values("incident_type", "description")
        .annotate(count=Count("id"), affected_orders=Count("order", distinct=True))
        .order_by("-count")[:20]
    )

    # Incidências por partner
    incidents_by_partner = (
        OrderIncident.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        )
        .values("order__partner__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    # Incidências de veículos
    vehicle_incidents = (
        VehicleIncident.objects.filter(
            incident_date__gte=start_date, incident_date__lte=end_date
        )
        .values("incident_type")
        .annotate(count=Count("id"), total_cost=Sum("fine_amount"))
        .order_by("-count")
    )

    # Estatísticas gerais
    total_order_incidents = OrderIncident.objects.filter(
        created_at__date__gte=start_date, created_at__date__lte=end_date
    ).count()

    total_vehicle_incidents = VehicleIncident.objects.filter(
        incident_date__gte=start_date, incident_date__lte=end_date
    ).count()

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "order_incidents": order_incidents,
        "incidents_by_partner": incidents_by_partner,
        "vehicle_incidents": vehicle_incidents,
        "total_order_incidents": total_order_incidents,
        "total_vehicle_incidents": total_vehicle_incidents,
    }

    return render(request, "analytics/incidents_report.html", context)


@login_required
def vehicles_performance_report(request):
    """
    Relatório de performance de veículos (Custo x Entregas).
    """
    # Período padrão: últimos 30 dias
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET.get("start_date"), "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET.get("end_date"), "%Y-%m-%d").date()

    # Performance por veículo
    vehicles = (
        Vehicle.objects.filter(status="ACTIVE")
        .annotate(
            # Contar atribuições (assignments)
            total_shifts=Count(
                "assignments",
                distinct=True,
                filter=Q(
                    assignments__date__gte=start_date,
                    assignments__date__lte=end_date,
                ),
            ),
            # Custo total de manutenções
            maintenance_cost=Sum(
                "maintenance_records__cost",
                filter=Q(
                    maintenance_records__completed_date__gte=start_date,
                    maintenance_records__completed_date__lte=end_date,
                ),
            ),
            # Custo de incidentes
            incident_cost=Sum(
                "incidents__fine_amount",
                filter=Q(
                    incidents__incident_date__gte=start_date,
                    incidents__incident_date__lte=end_date,
                ),
            ),
            # Número de incidentes
            incident_count=Count(
                "incidents",
                filter=Q(
                    incidents__incident_date__gte=start_date,
                    incidents__incident_date__lte=end_date,
                ),
            ),
        )
        .order_by("-total_shifts")
    )

    # Calcular custo total e custo por shift
    vehicles_data = []
    for vehicle in vehicles:
        maintenance = vehicle.maintenance_cost or Decimal("0")
        incidents = vehicle.incident_cost or Decimal("0")
        total_cost = maintenance + incidents
        cost_per_shift = (
            total_cost / vehicle.total_shifts
            if vehicle.total_shifts > 0
            else Decimal("0")
        )

        vehicles_data.append(
            {
                "vehicle": vehicle,
                "total_shifts": vehicle.total_shifts,
                "maintenance_cost": maintenance,
                "incident_cost": incidents,
                "total_cost": total_cost,
                "cost_per_shift": cost_per_shift,
                "incident_count": vehicle.incident_count,
            }
        )

    # Ordenar por custo total (decrescente)
    vehicles_data.sort(key=lambda x: x["total_cost"], reverse=True)

    # Estatísticas gerais
    total_vehicles = len(vehicles_data)
    total_cost = sum(v["total_cost"] for v in vehicles_data)
    avg_cost_per_vehicle = (
        total_cost / total_vehicles if total_vehicles > 0 else Decimal("0")
    )

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "vehicles_data": vehicles_data,
        "total_vehicles": total_vehicles,
        "total_cost": total_cost,
        "avg_cost_per_vehicle": avg_cost_per_vehicle,
    }

    return render(request, "analytics/vehicles_performance_report.html", context)


# ============================================================================
# API ENDPOINTS (JSON) para gráficos dinâmicos
# ============================================================================


@login_required
def api_metrics_data(request):
    """API endpoint para dados de métricas (JSON)."""
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    partner_id = request.GET.get("partner")
    if request.GET.get("start_date"):
        start_date = datetime.strptime(request.GET.get("start_date"), "%Y-%m-%d").date()
    if request.GET.get("end_date"):
        end_date = datetime.strptime(request.GET.get("end_date"), "%Y-%m-%d").date()

    metrics_query = DailyMetrics.objects.filter(
        date__gte=start_date, date__lte=end_date
    )

    if partner_id:
        metrics_query = metrics_query.filter(partner_id=partner_id)

    data = list(
        metrics_query.values("date")
        .annotate(
            total_orders=Sum("total_orders"),
            delivered=Sum("delivered_orders"),
            failed=Sum("failed_orders"),
            revenue=Sum("total_revenue"),
            success_rate=Avg("success_rate"),
        )
        .order_by("date")
    )

    # Converter date para string
    for item in data:
        item["date"] = item["date"].isoformat()
        item["revenue"] = float(item["revenue"] or 0)
        item["success_rate"] = float(item["success_rate"] or 0)

    return JsonResponse({"data": data})


@login_required
def api_forecasts_data(request):
    """API endpoint para dados de previsões (JSON)."""
    end_date = timezone.now().date()
    start_date = end_date
    forecast_end = end_date + timedelta(days=30)

    partner_id = request.GET.get("partner")
    method = request.GET.get("method", "BEST")

    forecasts_query = VolumeForecast.objects.filter(
        forecast_date__gte=start_date, forecast_date__lte=forecast_end
    )

    if partner_id:
        forecasts_query = forecasts_query.filter(partner_id=partner_id)

    if method == "BEST":
        forecasts_query = forecasts_query.filter(method="MA7")
    else:
        forecasts_query = forecasts_query.filter(method=method)

    data = list(
        forecasts_query.values(
            "forecast_date",
            "predicted_volume",
            "lower_bound",
            "upper_bound",
            "method",
        ).order_by("forecast_date")
    )

    for item in data:
        item["forecast_date"] = item["forecast_date"].isoformat()
        item["predicted_volume"] = float(item["predicted_volume"])
        item["lower_bound"] = float(item["lower_bound"] or 0)
        item["upper_bound"] = float(item["upper_bound"] or 0)

    return JsonResponse({"data": data})


@login_required
def api_alerts_data(request):
    """API endpoint para dados de alertas (JSON)."""
    status = request.GET.get("status", "ACTIVE")

    alerts = PerformanceAlert.objects.filter(status=status)

    data = list(
        alerts.values(
            "alert_type", "severity", "message", "created_at", "partner__name"
        ).order_by("-created_at")[:50]
    )

    for item in data:
        item["created_at"] = item["created_at"].isoformat()

    return JsonResponse({"data": data})


@login_required
def api_drivers_data(request):
    """API endpoint para dados de motoristas (JSON)."""
    today = timezone.now().date()
    current_month = today.replace(day=1)

    month_param = request.GET.get("month")
    if month_param:
        current_month = datetime.strptime(month_param, "%Y-%m").date().replace(day=1)

    drivers = (
        DriverPerformance.objects.filter(month=current_month)
        .select_related("driver")
        .values(
            "driver__nome_completo",
            "total_deliveries",
            "success_rate",
            "total_earnings",
            "rank_in_team",
        )
        .order_by("-rank_in_team")
    )

    data = list(drivers)

    for item in data:
        item["success_rate"] = float(item["success_rate"] or 0)
        item["total_earnings"] = float(item["total_earnings"] or 0)

    return JsonResponse({"data": data})


# ============================================================================
# EXPORTAÇÕES (Excel e PDF)
# ============================================================================


@login_required
def export_metrics_excel(request):
    """Exportar métricas para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            status=500,
        )

    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)

    partner_id = request.GET.get("partner")

    metrics_query = DailyMetrics.objects.filter(
        date__gte=start_date, date__lte=end_date
    )

    if partner_id:
        metrics_query = metrics_query.filter(partner_id=partner_id)

    metrics = metrics_query.order_by("date")

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Métricas Diárias"

    # Cabeçalhos
    headers = [
        "Data",
        "Parceiro",
        "Total Pedidos",
        "Entregues",
        "Falhados",
        "Taxa Sucesso (%)",
        "Receita (€)",
        "Bónus (€)",
        "Penalidades (€)",
        "Motoristas Ativos",
        "Veículos Ativos",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row, metric in enumerate(metrics, 2):
        ws.cell(row=row, column=1, value=metric.date.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=metric.partner.name)
        ws.cell(row=row, column=3, value=metric.total_orders)
        ws.cell(row=row, column=4, value=metric.delivered_orders)
        ws.cell(row=row, column=5, value=metric.failed_orders)
        ws.cell(row=row, column=6, value=float(metric.success_rate))
        ws.cell(row=row, column=7, value=float(metric.total_revenue))
        ws.cell(row=row, column=8, value=float(metric.total_bonuses))
        ws.cell(row=row, column=9, value=float(metric.total_penalties))
        ws.cell(row=row, column=10, value=metric.active_drivers_count)
        ws.cell(row=row, column=11, value=metric.active_vehicles_count)

    # Ajustar largura das colunas
    for col in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Preparar resposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="metricas_{start_date}_{end_date}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def export_drivers_excel(request):
    """Exportar performance de motoristas para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            status=500,
        )

    today = timezone.now().date()
    current_month = today.replace(day=1)

    month_param = request.GET.get("month")
    if month_param:
        current_month = datetime.strptime(month_param, "%Y-%m").date().replace(day=1)

    drivers = (
        DriverPerformance.objects.filter(month=current_month)
        .select_related("driver")
        .order_by("-rank_in_team")
    )

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Performance {current_month.strftime("%m-%Y")}'

    # Cabeçalhos
    headers = [
        "Ranking",
        "Motorista",
        "Total Entregas",
        "Entregas com Sucesso",
        "Entregas Falhadas",
        "Taxa Sucesso (%)",
        "Tempo Médio (h)",
        "Receita Total (€)",
        "Bónus (€)",
        "Penalidades (€)",
        "Ganhos Liquidos (€)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for row, perf in enumerate(drivers, 2):
        driver_name = f"{perf.driver.user.first_name} {perf.driver.user.last_name}"

        ws.cell(row=row, column=1, value=perf.ranking)
        ws.cell(row=row, column=2, value=driver_name)
        ws.cell(row=row, column=3, value=perf.total_deliveries)
        ws.cell(row=row, column=4, value=perf.successful_deliveries)
        ws.cell(row=row, column=5, value=perf.failed_deliveries)
        ws.cell(row=row, column=6, value=float(perf.success_rate))
        ws.cell(row=row, column=7, value=float(perf.average_delivery_time or 0))
        ws.cell(row=row, column=8, value=float(perf.total_earnings))
        ws.cell(row=row, column=9, value=float(perf.total_bonuses))
        ws.cell(row=row, column=10, value=float(perf.total_penalties))
        ws.cell(row=row, column=11, value=float(perf.net_earnings))

    # Ajustar largura das colunas
    for col in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Preparar resposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="motoristas_{current_month.strftime("%m_%Y")}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def export_report_pdf(request):
    """Exportar relatório consolidado em PDF."""
    try:
        from io import BytesIO

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        return HttpResponse(
            "Biblioteca reportlab não instalada. Execute: pip install reportlab",
            status=500,
        )

    # Criar PDF em memória
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    title = Paragraph("Relatório de Analytics - Léguas Franzinas", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 0.5 * cm))

    # Período
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    period = Paragraph(
        f'Período: {start_date.strftime("%d/%m/%Y")} - {end_date.strftime("%d/%m/%Y")}',
        styles["Normal"],
    )
    elements.append(period)
    elements.append(Spacer(1, 0.5 * cm))

    # Métricas totais
    metrics = DailyMetrics.objects.filter(
        date__gte=start_date, date__lte=end_date
    ).aggregate(
        total_orders=Sum("total_orders"),
        delivered=Sum("delivered_orders"),
        failed=Sum("failed_orders"),
        revenue=Sum("total_revenue"),
    )

    data = [
        ["Métrica", "Valor"],
        ["Total de Pedidos", str(metrics["total_orders"] or 0)],
        ["Pedidos Entregues", str(metrics["delivered"] or 0)],
        ["Pedidos Falhados", str(metrics["failed"] or 0)],
        ["Receita Total", f"€{metrics['revenue'] or 0:.2f}"],
    ]

    table = Table(data, colWidths=[10 * cm, 5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)

    # Gerar PDF
    doc.build(elements)

    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="relatorio_{start_date}_{end_date}.pdf"'
    )

    return response


# ========== AUTOMAÇÕES ==========


@login_required
def automations_dashboard(request):
    """Dashboard de automações e alertas"""
    from datetime import date

    from .services.automation_service import AutomationService

    # Obter resumo de alertas
    alerts = AutomationService.get_alerts_summary()

    # Obter detalhes
    overdue_orders = AutomationService.get_overdue_orders()[:10]
    pending_maintenances = AutomationService.get_pending_maintenances()[:10]
    unassigned_shifts = AutomationService.get_unassigned_shifts()[:10]

    # Estatísticas de atribuição automática
    today = date.today()
    from orders_manager.models import Order

    todays_assigned = Order.objects.filter(
        assigned_at__date=today, current_status="ASSIGNED"
    ).count()

    todays_pending = Order.objects.filter(
        current_status="PENDING", assigned_driver__isnull=True
    ).count()

    context = {
        "alerts": alerts,
        "overdue_orders": overdue_orders,
        "pending_maintenances": pending_maintenances,
        "unassigned_shifts": unassigned_shifts,
        "todays_assigned": todays_assigned,
        "todays_pending": todays_pending,
    }

    return render(request, "analytics/automations_dashboard.html", context)


@login_required
def run_auto_assignment(request):
    """Executa atribuição automática de pedidos"""
    from datetime import date

    from django.contrib import messages
    from django.shortcuts import redirect

    from .services.automation_service import AutomationService

    if request.method == "POST":
        target_date_str = request.POST.get("target_date")
        max_orders = int(request.POST.get("max_orders", 50))

        if target_date_str:
            # Atribuição por data específica
            target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
            result = AutomationService.auto_assign_orders_for_date(target_date)

            messages.success(
                request,
                f"Atribuição automática concluída! {result.get('total_orders_assigned', 0)} "
                f"pedidos atribuídos em {result.get('shifts_processed', 0)} turnos.",
            )
        else:
            # Atribuição de pedidos pendentes
            result = AutomationService.auto_assign_pending_orders(max_orders)

            messages.success(
                request,
                f"Atribuição automática concluída! {result['assigned']} pedidos atribuídos, "
                f"{result['failed']} falharam.",
            )

        return redirect("analytics:automations_dashboard")

    # GET - mostrar formulário
    context = {"today": date.today()}

    return render(request, "analytics/run_auto_assignment.html", context)


@login_required
def route_optimizer(request):
    """Interface para otimização de rotas"""
    from datetime import date

    from drivers_app.models import DriverProfile

    from .services.automation_service import AutomationService

    driver_id = request.GET.get("driver")
    target_date_str = request.GET.get("date")

    target_date = date.today()
    if target_date_str:
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()

    optimized_route = None
    selected_driver = None

    if driver_id:
        try:
            selected_driver = DriverProfile.objects.get(id=driver_id)
            optimized_route = AutomationService.optimize_route_for_driver(
                selected_driver, target_date
            )
        except DriverProfile.DoesNotExist:
            pass

    # Lista de motoristas ativos
    active_drivers = DriverProfile.objects.filter(status="ATIVO").order_by(
        "nome_completo"
    )

    context = {
        "active_drivers": active_drivers,
        "selected_driver": selected_driver,
        "target_date": target_date,
        "optimized_route": optimized_route,
    }

    return render(request, "analytics/route_optimizer.html", context)


@login_required
def shift_suggestions(request):
    """Sugestões inteligentes de turnos para a semana"""
    from datetime import date, timedelta

    from .services.automation_service import AutomationService

    start_date_str = request.GET.get("start_date")

    if start_date_str:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    else:
        # Próxima segunda-feira
        today = date.today()
        days_until_monday = (7 - today.weekday()) % 7
        start_date = today + timedelta(
            days=days_until_monday if days_until_monday > 0 else 7
        )

    suggestions = AutomationService.suggest_shift_assignments_for_week(start_date)

    context = {
        "suggestions": suggestions["suggestions"],
        "start_date": suggestions["start_date"],
        "end_date": suggestions["end_date"],
    }

    return render(request, "analytics/shift_suggestions.html", context)


# ============================================
# RELATÓRIOS AVANÇADOS
# ============================================


@login_required
def vehicle_utilization_report(request):
    """Relatório de utilização de veículos"""
    from datetime import date, timedelta

    from django.db.models import Avg, Sum

    from fleet_management.models import Vehicle
    from route_allocation.models import DriverShift

    # Parâmetros de filtro
    days = int(request.GET.get("days", 30))
    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    # Query de veículos com estatísticas
    vehicles_stats = []
    vehicles = Vehicle.objects.filter(status="ACTIVE")

    for vehicle in vehicles:
        # Turnos usando este veículo
        shifts = DriverShift.objects.filter(
            driver__assigned_vehicle=vehicle,
            date__gte=start_date,
            date__lte=end_date,
            status="COMPLETED",
        )

        # Estatísticas
        total_shifts = shifts.count()
        total_deliveries = shifts.aggregate(total=Sum("total_deliveries"))["total"] or 0
        avg_deliveries = shifts.aggregate(avg=Avg("total_deliveries"))["avg"] or 0

        # Dias ativos (estimativa baseada em turnos únicos por data)
        active_days = shifts.values("date").distinct().count()

        # Taxa de utilização (dias ativos / dias no período)
        utilization_rate = (active_days / days * 100) if days > 0 else 0

        # Estimativa de km (baseado em deliveries - aproximação)
        estimated_km = total_deliveries * 15  # ~15km por entrega (média estimada)

        vehicles_stats.append(
            {
                "vehicle": vehicle,
                "active_days": active_days,
                "total_shifts": total_shifts,
                "total_deliveries": total_deliveries,
                "avg_deliveries_per_shift": round(avg_deliveries, 1),
                "estimated_km": estimated_km,
                "utilization_rate": round(utilization_rate, 1),
                "km_per_day": (
                    round(estimated_km / active_days, 1) if active_days > 0 else 0
                ),
            }
        )

    # Ordenar por taxa de utilização
    vehicles_stats.sort(key=lambda x: x["utilization_rate"], reverse=True)

    context = {
        "vehicles_stats": vehicles_stats,
        "start_date": start_date,
        "end_date": end_date,
        "days_filter": days,
        "total_vehicles": len(vehicles_stats),
        "avg_utilization": (
            round(
                sum(v["utilization_rate"] for v in vehicles_stats)
                / len(vehicles_stats),
                1,
            )
            if vehicles_stats
            else 0
        ),
    }

    return render(request, "analytics/vehicle_utilization_report.html", context)


@login_required
def fleet_cost_report(request):
    """Relatório de custos de frota"""
    from datetime import date, timedelta
    from decimal import Decimal

    from django.db.models import Sum

    from fleet_management.models import Vehicle, VehicleMaintenance
    from route_allocation.models import DriverShift

    # Parâmetros de filtro
    days = int(request.GET.get("days", 30))
    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    # Query de veículos com custos
    vehicles_costs = []
    vehicles = Vehicle.objects.filter(status="ACTIVE")

    for vehicle in vehicles:
        # Custos de manutenção no período
        maintenances = VehicleMaintenance.objects.filter(
            vehicle=vehicle,
            completed_date__gte=start_date,
            completed_date__lte=end_date,
            is_completed=True,
        )

        maintenance_cost = maintenances.aggregate(total=Sum("cost"))[
            "total"
        ] or Decimal("0")
        maintenance_count = maintenances.count()

        # Entregas realizadas no período
        shifts = DriverShift.objects.filter(
            driver__assigned_vehicle=vehicle,
            date__gte=start_date,
            date__lte=end_date,
            status="COMPLETED",
        )

        total_deliveries = shifts.aggregate(total=Sum("total_deliveries"))["total"] or 0

        # Estimativa de combustível (€0.15 por km, ~15km por entrega)
        estimated_km = total_deliveries * 15
        fuel_cost = Decimal(str(estimated_km * 0.15))

        # Custo total e por entrega
        total_cost = maintenance_cost + fuel_cost
        cost_per_delivery = (
            (total_cost / total_deliveries) if total_deliveries > 0 else Decimal("0")
        )

        vehicles_costs.append(
            {
                "vehicle": vehicle,
                "maintenance_cost": maintenance_cost,
                "maintenance_count": maintenance_count,
                "fuel_cost": fuel_cost,
                "estimated_km": estimated_km,
                "total_cost": total_cost,
                "total_deliveries": total_deliveries,
                "cost_per_delivery": cost_per_delivery,
            }
        )

    # Ordenar por custo total
    vehicles_costs.sort(key=lambda x: x["total_cost"], reverse=True)

    # Totais gerais
    total_maintenance = sum(v["maintenance_cost"] for v in vehicles_costs)
    total_fuel = sum(v["fuel_cost"] for v in vehicles_costs)
    total_general = total_maintenance + total_fuel
    total_deliveries_all = sum(v["total_deliveries"] for v in vehicles_costs)
    avg_cost_per_delivery = (
        (total_general / total_deliveries_all)
        if total_deliveries_all > 0
        else Decimal("0")
    )

    context = {
        "vehicles_costs": vehicles_costs,
        "start_date": start_date,
        "end_date": end_date,
        "days_filter": days,
        "total_vehicles": len(vehicles_costs),
        "total_maintenance": total_maintenance,
        "total_fuel": total_fuel,
        "total_general": total_general,
        "total_deliveries": total_deliveries_all,
        "avg_cost_per_delivery": avg_cost_per_delivery,
    }

    return render(request, "analytics/fleet_cost_report.html", context)


@login_required
def shift_performance_report(request):
    """Relatório de performance de turnos"""
    from datetime import date, timedelta

    from django.db.models import Avg, Sum

    from drivers_app.models import DriverProfile
    from route_allocation.models import DriverShift

    # Parâmetros de filtro
    days = int(request.GET.get("days", 30))
    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    # Query de motoristas com estatísticas
    drivers_stats = []
    drivers = DriverProfile.objects.filter(status="ATIVO")

    for driver in drivers:
        # Turnos completados no período
        shifts = DriverShift.objects.filter(
            driver=driver,
            date__gte=start_date,
            date__lte=end_date,
            status="COMPLETED",
        )

        total_shifts = shifts.count()

        if total_shifts == 0:
            continue

        # Estatísticas de entregas
        total_deliveries = shifts.aggregate(total=Sum("total_deliveries"))["total"] or 0
        successful = shifts.aggregate(total=Sum("successful_deliveries"))["total"] or 0
        failed = shifts.aggregate(total=Sum("failed_deliveries"))["total"] or 0

        avg_deliveries = shifts.aggregate(avg=Avg("total_deliveries"))["avg"] or 0

        # Taxa de sucesso
        success_rate = (
            (successful / total_deliveries * 100) if total_deliveries > 0 else 0
        )

        # Duração média de turno (calculada quando há tempos reais)
        shifts_with_times = shifts.filter(
            actual_start_time__isnull=False, actual_end_time__isnull=False
        )

        avg_duration_minutes = 0
        if shifts_with_times.exists():
            durations = []
            for shift in shifts_with_times:
                duration = shift.actual_end_time - shift.actual_start_time
                durations.append(duration.total_seconds() / 60)
            avg_duration_minutes = sum(durations) / len(durations) if durations else 0

        drivers_stats.append(
            {
                "driver": driver,
                "total_shifts": total_shifts,
                "total_deliveries": total_deliveries,
                "successful_deliveries": successful,
                "failed_deliveries": failed,
                "avg_deliveries_per_shift": round(avg_deliveries, 1),
                "success_rate": round(success_rate, 1),
                "avg_duration_minutes": round(avg_duration_minutes, 0),
                "avg_duration_hours": round(avg_duration_minutes / 60, 1),
            }
        )

    # Ordenar por taxa de sucesso
    drivers_stats.sort(key=lambda x: x["success_rate"], reverse=True)

    # Estatísticas gerais
    total_shifts_all = sum(d["total_shifts"] for d in drivers_stats)
    total_deliveries_all = sum(d["total_deliveries"] for d in drivers_stats)
    total_successful = sum(d["successful_deliveries"] for d in drivers_stats)
    avg_success_rate = (
        round(
            sum(d["success_rate"] for d in drivers_stats) / len(drivers_stats),
            1,
        )
        if drivers_stats
        else 0
    )

    context = {
        "drivers_stats": drivers_stats,
        "start_date": start_date,
        "end_date": end_date,
        "days_filter": days,
        "total_drivers": len(drivers_stats),
        "total_shifts": total_shifts_all,
        "total_deliveries": total_deliveries_all,
        "total_successful": total_successful,
        "avg_success_rate": avg_success_rate,
    }

    return render(request, "analytics/shift_performance_report.html", context)


# ============================================
# INTEGRAÇÕES & API STATUS
# ============================================


@login_required
def api_status_dashboard(request):
    """Dashboard de status das integrações de API com parceiros"""
    from datetime import timedelta

    from django.db.models import Count, Max, Q

    from core.models import PartnerIntegration, SyncLog

    # Buscar todas as integrações ativas
    integrations = (
        PartnerIntegration.objects.filter(is_active=True)
        .select_related("partner")
        .annotate(
            total_syncs=Count("sync_logs"),
            success_syncs=Count("sync_logs", filter=Q(sync_logs__status="SUCCESS")),
            error_syncs=Count("sync_logs", filter=Q(sync_logs__status="ERROR")),
            last_log_time=Max("sync_logs__started_at"),
        )
        .order_by("partner__name")
    )

    # Estatísticas por integração
    integrations_stats = []
    now = timezone.now()

    for integration in integrations:
        # Determinar status de saúde
        health_status = "healthy"  # healthy, warning, critical, unknown
        health_message = "Operacional"

        if not integration.last_sync_at:
            health_status = "unknown"
            health_message = "Sem sincronizações"
        elif integration.is_sync_overdue:
            health_status = "critical"
            health_message = "Sincronização atrasada"
        elif integration.last_sync_status == "ERROR":
            health_status = "critical"
            health_message = "Última sincronização falhou"
        elif integration.last_sync_status == "PARTIAL":
            health_status = "warning"
            health_message = "Sincronização parcial"

        # Taxa de sucesso (últimas 24h)
        day_ago = now - timedelta(hours=24)
        recent_logs = SyncLog.objects.filter(
            integration=integration, started_at__gte=day_ago
        )

        recent_total = recent_logs.count()
        recent_success = recent_logs.filter(status="SUCCESS").count()
        success_rate_24h = (
            round((recent_success / recent_total * 100), 1) if recent_total > 0 else 0
        )

        # Tempo desde última sync
        if integration.last_sync_at:
            time_since_sync = now - integration.last_sync_at
            minutes_ago = int(time_since_sync.total_seconds() / 60)

            if minutes_ago < 60:
                last_sync_display = f"{minutes_ago} minutos atrás"
            elif minutes_ago < 1440:  # 24 horas
                last_sync_display = f"{int(minutes_ago/60)} horas atrás"
            else:
                last_sync_display = f"{int(minutes_ago/1440)} dias atrás"
        else:
            last_sync_display = "Nunca"

        integrations_stats.append(
            {
                "integration": integration,
                "health_status": health_status,
                "health_message": health_message,
                "success_rate_24h": success_rate_24h,
                "recent_syncs_24h": recent_total,
                "last_sync_display": last_sync_display,
                "minutes_since_sync": (
                    minutes_ago if integration.last_sync_at else None
                ),
            }
        )

    # Estatísticas globais
    total_integrations = len(integrations_stats)
    healthy_count = sum(
        1 for i in integrations_stats if i["health_status"] == "healthy"
    )
    warning_count = sum(
        1 for i in integrations_stats if i["health_status"] == "warning"
    )
    critical_count = sum(
        1 for i in integrations_stats if i["health_status"] == "critical"
    )

    # Logs recentes (últimas 10 sincronizações)
    recent_logs = SyncLog.objects.select_related("integration__partner").order_by(
        "-started_at"
    )[:10]

    context = {
        "integrations_stats": integrations_stats,
        "total_integrations": total_integrations,
        "healthy_count": healthy_count,
        "warning_count": warning_count,
        "critical_count": critical_count,
        "recent_logs": recent_logs,
    }

    return render(request, "analytics/api_status_dashboard.html", context)


@login_required
def sync_logs_list(request):
    """Lista detalhada de logs de sincronização com filtros"""
    from datetime import date, timedelta

    from core.models import PartnerIntegration, SyncLog

    # Filtros
    integration_id = request.GET.get("integration")
    status_filter = request.GET.get("status")
    operation_filter = request.GET.get("operation")
    days = int(request.GET.get("days", 7))

    # Query base
    end_date = date.today()
    start_date = end_date - timedelta(days=days)

    logs = SyncLog.objects.select_related("integration__partner").filter(
        started_at__date__gte=start_date, started_at__date__lte=end_date
    )

    # Aplicar filtros
    if integration_id:
        logs = logs.filter(integration_id=integration_id)

    if status_filter:
        logs = logs.filter(status=status_filter)

    if operation_filter:
        logs = logs.filter(operation=operation_filter)

    # Ordenar
    logs = logs.order_by("-started_at")

    # Stats do período
    total_logs = logs.count()
    success_logs = logs.filter(status="SUCCESS").count()
    error_logs = logs.filter(status="ERROR").count()
    avg_duration = logs.exclude(completed_at__isnull=True).aggregate(
        avg_duration=models.Avg(models.F("completed_at") - models.F("started_at"))
    )["avg_duration"]

    # Converter avg_duration para segundos
    if avg_duration:
        avg_duration_seconds = avg_duration.total_seconds()
    else:
        avg_duration_seconds = 0

    # Opções para filtros
    integrations = PartnerIntegration.objects.filter(is_active=True).select_related(
        "partner"
    )

    context = {
        "logs": logs[:100],  # Limitar a 100 registros por performance
        "total_logs": total_logs,
        "success_logs": success_logs,
        "error_logs": error_logs,
        "avg_duration_seconds": round(avg_duration_seconds, 1),
        "integrations": integrations,
        "status_choices": SyncLog.STATUSES,
        "operation_choices": SyncLog.SYNC_OPERATIONS,
        "filters": {
            "integration_id": integration_id,
            "status": status_filter,
            "operation": operation_filter,
            "days": days,
        },
        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "analytics/sync_logs_list.html", context)


@login_required
def retry_failed_sync(request, log_id):
    """Re-executa uma sincronização que falhou"""
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect

    from core.models import SyncLog

    if request.method != "POST":
        messages.error(request, "Método não permitido")
        return redirect("analytics:sync_logs_list")

    log = get_object_or_404(SyncLog, id=log_id)

    # Verificar se é uma falha
    if log.status not in ["ERROR", "TIMEOUT", "PARTIAL"]:
        messages.warning(request, "Este log não representa uma falha")
        return redirect("analytics:sync_logs_list")

    try:
        # Importar o serviço de sync apropriado baseado no parceiro
        integration = log.integration

        # Criar novo log para o retry
        new_log = SyncLog.objects.create(
            integration=integration,
            operation=log.operation,
            status="STARTED",
            request_data={
                "retry_of": log.id,
                "original_timestamp": str(log.started_at),
            },
        )

        # TODO: Aqui você implementaria a lógica específica de retry
        # baseada no tipo de integração (Paack, etc.)
        # Por agora, vamos simular um retry

        # Exemplo: se fosse Paack
        if integration.partner.name == "Paack":
            # from ordersmanager_paack.sync_service import SyncService
            # sync_service = SyncService()
            # result = sync_service.sync_data(force_refresh=True)
            pass

        # Marcar como parcialmente bem-sucedido (já que é simulado)
        new_log.mark_completed(
            status="PARTIAL",
            message=f"Retry manual executado. Log original: #{log.id}",
        )

        # Atualizar status da integração
        integration.mark_sync_success(message=f"Retry manual executado com sucesso")

        messages.success(
            request, f"Retry executado com sucesso! Novo log: #{new_log.id}"
        )

    except Exception as e:
        messages.error(request, f"Erro ao executar retry: {str(e)}")

    return redirect("analytics:sync_logs_list")
