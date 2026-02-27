import io
import json
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .ai_detector import IntelligentDataDetector
from .models import ProcessingHistory, LearningPattern

@login_required
def converter_view(request):
    """
    View principal do conversor de listas com IA.
    Gerencia o fluxo de processamento, aprendizado e geração de XLSX.
    """
    # Se não é POST, mostra formulário inicial com estatísticas da IA
    if request.method != "POST":
        # Carrega estatísticas da IA para mostrar na página inicial
        detector = IntelligentDataDetector()
        ai_stats = {
            'patterns': {},
            'total_history': ProcessingHistory.objects.count(),
            'accuracy': detector.calculate_accuracy(),
            'last_processed': None
        }
        
        # Contar padrões de aprendizado
        for field_type in ['endereco', 'codigo_id', 'hora', 'data', 'horario', 'litros', 'quantidade']:
            ai_stats['patterns'][field_type] = LearningPattern.objects.filter(pattern_type=field_type).count()
        
        # Data do último processamento
        last_process = ProcessingHistory.objects.order_by('-created_at').first()
        if last_process:
            ai_stats['last_processed'] = last_process.created_at
            
        return render(request, "converter.html", {'ai_stats': ai_stats})
    
    # Processa o POST
    action = request.POST.get("action", "parse")
    
    if action == "parse":
        # Primeira etapa: parsear a lista usando IA
        raw_text = request.POST.get("lista", "")
        
        if not raw_text.strip():
            return render(request, "converter.html", {
                'error': 'Por favor, insira dados para processar.',
            })
        
        try:
            # Inicializar detector IA
            detector = IntelligentDataDetector()
            
            # Parse inteligente dos dados
            dados_estruturados = detector.parse_intelligent_blocks(raw_text)
            
            if not dados_estruturados:
                return render(request, "converter.html", {
                    'error': 'Não foi possível detectar dados válidos. Verifique o formato da sua lista.',
                    'raw_text': raw_text
                })
            
            # Gerar relatório de confiança
            confidence_report = detector.get_confidence_report(dados_estruturados)
            
            # Carregar estatísticas da IA usando o sistema de ranking
            stats = {
                'patterns': {},
                'ranking': detector.ranking_data.get('stats', {}),
                'accuracy': detector.calculate_accuracy(),
                'total_history': ProcessingHistory.objects.count(),
                'last_update': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            
            # Contar padrões de aprendizado
            for field_type in ['endereco', 'codigo_id', 'hora', 'data', 'horario', 'litros', 'quantidade']:
                stats['patterns'][field_type] = LearningPattern.objects.filter(pattern_type=field_type).count()
            
            # Padrões aprendidos com ranking
            stats['ranking_patterns'] = {}
            for field_type, patterns in detector.ranking_data.get('patterns', {}).items():
                stats['ranking_patterns'][field_type] = len(patterns)
                
            # Confiança por campo
            stats['field_confidence'] = confidence_report.get('field_averages', {})

            return render(request, "converter_table.html", {
                'dados': dados_estruturados,
                'raw_text': raw_text,
                'confidence_report': confidence_report,
                'ai_stats': stats
            })
            
        except Exception as e:
            import traceback
            traceback_str = traceback.format_exc()
            return render(request, "converter.html", {
                'error': f'Erro ao processar dados: {str(e)}',
                'debug_info': traceback_str,
                'raw_text': raw_text
            })
    
    elif action == "generate_xlsx":
        # Segunda etapa: gerar XLSX com dados editados
        dados_json = request.POST.get("dados_editados", "[]")
        raw_text = request.POST.get("raw_text", "")
        
        try:
            dados = json.loads(dados_json)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Dados inválidos'}, status=400)

        # Aprender com as correções do usuário
        if raw_text and dados:
            detector = IntelligentDataDetector()
            detector.learn_from_corrections(raw_text, dados)

        # Criar XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista Convertida"

        # Cabeçalhos
        headers = ["Endereço", "ID", "Número", "Hora", "Data", "Horário", "Litros", "Quantidade"]
        ws.append(headers)
        
        # Aplicar estilo aos cabeçalhos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        centered = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = thin_border

        # Adicionar dados
        for row_idx, item in enumerate(dados, start=2):
            ws.append([
                item.get('endereco', ''),
                item.get('codigo_id', ''),
                item.get('numero', ''),
                item.get('hora', ''),
                item.get('data', ''),
                item.get('horario', ''),
                item.get('litros', ''),
                item.get('quantidade', ''),
            ])
            
            # Aplicar estilos nas células de dados
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Alinhar números à direita
                if col_idx in [3, 7, 8]:  # Número, Litros, Quantidade
                    cell.alignment = Alignment(horizontal='right')

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # +3 para dar um pouco mais de espaço
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Adicionar rodapé com informações
        ws.append([])  # linha vazia
        footer_row = len(dados) + 3
        ws.cell(row=footer_row, column=1).value = f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws.cell(row=footer_row, column=1).font = Font(italic=True, size=9)
        
        ws.cell(row=footer_row+1, column=1).value = f"Total de itens: {len(dados)}"
        ws.cell(row=footer_row+1, column=1).font = Font(bold=True)

        # Salvar em memória
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="lista_convertida_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        with io.BytesIO() as buffer:
            wb.save(buffer)
            response.write(buffer.getvalue())
        return response

    return render(request, "converter.html")
