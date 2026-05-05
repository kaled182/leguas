import json
import re
from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Max, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl.utils import get_column_letter

from customauth.models import DriverAccess
from ordersmanager_paack.models import Driver

from .models import DriverDocument, DriverProfile, EmpresaParceira, Vehicle, VehicleDocument

# ===== HELPER FUNCTIONS =====


def serialize_driver_data(driver):
    """Serializa os dados do motorista para JSON"""
    return {
        "id": driver.id,
        "nome_completo": driver.nome_completo,
        "nif": driver.nif,
        "niss": driver.niss or "",
        "data_nascimento": (
            driver.data_nascimento.strftime("%d/%m/%Y")
            if driver.data_nascimento
            else ""
        ),
        "data_nascimento_iso": (
            driver.data_nascimento.strftime("%Y-%m-%d")
            if driver.data_nascimento
            else ""
        ),
        "nacionalidade": driver.nacionalidade or "",
        "email": driver.email,
        "telefone": driver.telefone or "",
        "endereco_completo": (
            f"{driver.endereco_residencia or ''}, "
            f"{driver.codigo_postal or ''} {driver.cidade or ''}"
        ).strip(),
        "tipo_vinculo": driver.tipo_vinculo,
        "nome_frota": driver.nome_frota or "",
        "empresa_parceira_id": driver.empresa_parceira_id,
        "empresa_parceira_nome": driver.empresa_parceira.nome if driver.empresa_parceira else "",
        "price_per_package": (
            str(driver.price_per_package) if driver.price_per_package else ""
        ),
        "status": driver.status,
        "approved_at": (
            driver.approved_at.strftime("%d/%m/%Y %H:%M") if driver.approved_at else ""
        ),
        "approved_by": str(driver.approved_by) if driver.approved_by else "",
        "documents": [
            {
                "id": doc.id,
                "tipo_display": doc.get_tipo_documento_display(),
                "validade": (
                    doc.data_validade.strftime("%d/%m/%Y") if doc.data_validade else ""
                ),
                "url": doc.arquivo.url,
                "file_extension": doc.file_extension,
                "is_expired": doc.is_expired,
                "expiring_soon": (
                    doc.days_until_expiration <= 30 and doc.days_until_expiration > 0
                    if doc.days_until_expiration
                    else False
                ),
            }
            for doc in driver.documents.all()
        ],
        "vehicles": [
            {
                "id": vehicle.id,
                "matricula": vehicle.matricula,
                "marca": vehicle.marca,
                "modelo": vehicle.modelo,
                "tipo_display": vehicle.get_tipo_veiculo_display(),
                "documents": [
                    {
                        "id": vdoc.id,
                        "tipo_display": vdoc.get_tipo_documento_display(),
                        "validade": (
                            vdoc.data_validade.strftime("%d/%m/%Y")
                            if vdoc.data_validade
                            else ""
                        ),
                        "url": vdoc.arquivo.url,
                        "file_extension": vdoc.file_extension,
                    }
                    for vdoc in vehicle.vehicle_documents.all()
                ],
            }
            for vehicle in driver.vehicles.all()
        ],
    }


# ===== VIEWS =====


def driver_dashboard_view(request):
    """
    Dashboard principal para motoristas autenticados.

    Mostra os pedidos do motorista com filtros por status e data.
    """
    # Verificar se é motorista autenticado
    if not request.session.get("is_driver_authenticated"):
        messages.error(request, "Acesso negado. Faça login como motorista.")
        return redirect("customauth:login")

    try:
        driver_access_id = request.session.get("driver_access_id")
        driver_access = DriverAccess.objects.get(id=driver_access_id)
    except DriverAccess.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
        return redirect("customauth:login")

    # Obter parâmetros de filtro
    status = request.GET.get("status", "")
    date_str = request.GET.get("date", "")
    today = timezone.now().date()

    # Usar data fornecida ou hoje
    if date_str:
        try:
            date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            date = today
    else:
        date = today

    # Obter rota do motorista
    route = driver_access.get_route()
    if not route:
        messages.warning(request, "Nenhum motorista vinculado ao seu acesso.")
        orders_qs = []
        stats = {"to_attempt": 0, "delivered": 0, "failed": 0, "total": 0}
    else:
        # Obter pedidos filtrados
        orders_qs = route.get_orders(date=date)
        if status:
            orders_qs = orders_qs.filter(simplified_order_status=status)

        # Calcular estatísticas
        stats = {
            "to_attempt": route.get_pending_count(date=date),
            "delivered": route.get_delivered_count(date=date),
            "failed": route.get_failed_count(date=date),
            "total": route.get_orders_count(date=date),
        }

    # Paginação
    paginator = Paginator(orders_qs, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Formatar dados dos pedidos
    orders = [
        {
            "order_id": order.order_id,
            "retailer": order.retailer,
            "client_address": order.client_address,
            "intended_delivery_date": order.intended_delivery_date,
            "status": order.simplified_order_status,
            "is_delivered": order.is_delivered,
            "is_failed": order.is_failed,
        }
        for order in page_obj.object_list
    ]

    context = {
        "welcome_name": driver_access.full_name,
        "orders": orders,
        "page_obj": page_obj,
        "stats": stats,
        "today": today,
        "selected_date": date,
        "selected_status": status,
        "driver_access": driver_access,
    }

    return render(request, "drivers_app/driver_dashboard.html", context)


def driver_export_xlsx(request):
    """
    Exporta os pedidos do motorista para XLSX.
    """
    # Verificar autenticação
    if not request.session.get("is_driver_authenticated"):
        return JsonResponse({"error": "Não autorizado"}, status=401)

    try:
        driver_access_id = request.session.get("driver_access_id")
        driver_access = DriverAccess.objects.get(id=driver_access_id)
        route = driver_access.get_route()

        if not route:
            return JsonResponse({"error": "Motorista não vinculado"}, status=400)

        # Obter data do filtro
        date_str = request.GET.get("date", "")
        if date_str:
            try:
                date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                date = timezone.now().date()
        else:
            date = timezone.now().date()

        # Criar workbook
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
            )
        )
        filename = f"pedidos_{driver_access.first_name}_{date}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Pedidos {date}"

        # Headers
        headers = [
            "ID Pedido",
            "Varejista",
            "Endereço",
            "Data Prevista",
            "Status",
            "Entregue",
            "Falhado",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Dados
        orders = route.get_orders(date=date)
        for row_num, order in enumerate(orders, 2):
            worksheet.cell(row=row_num, column=1, value=order.order_id)
            worksheet.cell(row=row_num, column=2, value=order.retailer)
            worksheet.cell(row=row_num, column=3, value=order.client_address)
            worksheet.cell(row=row_num, column=4, value=order.intended_delivery_date)
            worksheet.cell(row=row_num, column=5, value=order.simplified_order_status)
            worksheet.cell(
                row=row_num,
                column=6,
                value="Sim" if order.is_delivered else "Não",
            )
            worksheet.cell(
                row=row_num,
                column=7,
                value="Sim" if order.is_failed else "Não",
            )

        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(response)
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def driver_logout_view(request):
    """
    Logout específico para motoristas.
    """
    # Limpar sessão do motorista
    request.session.pop("driver_access_id", None)
    request.session.pop("driver_name", None)
    request.session.pop("is_driver_authenticated", None)

    messages.success(request, "Você foi desconectado com sucesso.")
    return redirect("customauth:login")


@csrf_exempt
@require_http_methods(["POST"])
def register_driver_typebot(request):
    """
    Endpoint API para registro de motoristas via Typebot.

    Recebe dados do Typebot e cria um registro de motorista pendente.

    Body JSON:
    {
        "nif": "123456789",
        "nome": "João Silva",
        "telefone": "+351911111111",
        "email": "joao@example.com"
    }

    Retorna JSON:
    - Sucesso: {"success": true, "driver_id": "ID"}
    - Erro: {"success": false, "error": "mensagem"}
    """
    try:
        # Parse JSON body
        data = json.loads(request.body)

        # Validar campos obrigatórios
        required_fields = ["nif", "nome", "telefone", "email"]
        missing_fields = [field for field in required_fields if not data.get(field)]

        if missing_fields:
            return JsonResponse(
                {
                    "success": False,
                    "error": f'Campos obrigatórios faltando: {", ".join(missing_fields)}',
                },
                status=400,
            )

        nif = data["nif"].strip()
        nome = data["nome"].strip()
        data["telefone"].strip()
        email = data["email"].strip().lower()

        # Validar NIF (9 dígitos)
        if not re.match(r"^\d{9}$", nif):
            return JsonResponse(
                {
                    "success": False,
                    "error": "NIF inválido. Deve conter exatamente 9 dígitos.",
                },
                status=400,
            )

        # Validar email
        if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
            return JsonResponse(
                {"success": False, "error": "Email inválido."}, status=400
            )

        # Verificar se NIF já existe
        if Driver.objects.filter(driver_id=nif).exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": "Este NIF já está registrado no sistema.",
                },
                status=409,
            )

        # Criar Driver com status pendente
        driver = Driver.objects.create(
            driver_id=nif,
            name=nome,
            vehicle="PENDENTE",
            vehicle_norm="PENDENTE",
            is_active=False,  # Inativo até aprovação manual
        )

        return JsonResponse(
            {
                "success": True,
                "driver_id": driver.driver_id,
                "message": "Cadastro recebido com sucesso! Aguarde aprovação da equipe.",
            },
            status=201,
        )

    except json.JSONDecodeError:
        return JsonResponse(
            {
                "success": False,
                "error": "JSON inválido no corpo da requisição.",
            },
            status=400,
        )

    except Exception as e:
        # Log error (consider using proper logging in production)
        return JsonResponse(
            {
                "success": False,
                "error": f"Erro ao processar cadastro: {str(e)}",
            },
            status=500,
        )


@require_http_methods(["GET", "POST"])
def public_driver_register(request):
    """
    Pagina publica de cadastro de motoristas (formulario web).

    GET: Renderiza formulario HTML
    POST: Processa cadastro e cria Driver pendente
    """
    if request.method == "GET":
        # Renderizar formulario
        return render(
            request,
            "drivers_app/driver_public_register.html",
            {"current_year": datetime.now().year},
        )

    # POST - Processar formulario
    try:
        # Obter dados do formulario
        nif = request.POST.get("nif", "").strip()
        nome = request.POST.get("nome", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        email = request.POST.get("email", "").strip().lower()
        terms_accepted = request.POST.get("terms") == "on"

        # Validar campos obrigatorios
        if not all([nif, nome, telefone, email]):
            messages.error(request, "Todos os campos sao obrigatorios.")
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Validar termos
        if not terms_accepted:
            messages.error(request, "Voce deve concordar com os termos para continuar.")
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Validar NIF (9 digitos)
        if not re.match(r"^\d{9}$", nif):
            messages.error(
                request,
                "NIF invalido. Deve conter exatamente 9 digitos numericos.",
            )
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Validar email
        if not re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email):
            messages.error(request, "Email invalido. Verifique o formato.")
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Validar telefone
        if not re.match(r"^\+?[0-9]{9,15}$", telefone):
            messages.error(
                request,
                "Telefone invalido. Use formato internacional (ex: +351911111111).",
            )
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Verificar se NIF ja existe
        if DriverProfile.objects.filter(nif=nif).exists():
            messages.error(
                request,
                "Este NIF ja esta registrado no sistema. Entre em contato conosco se houver algum problema.",
            )
            return render(
                request,
                "drivers_app/driver_public_register.html",
                {"current_year": datetime.now().year},
            )

        # Criar DriverProfile com status pendente
        driver_profile = DriverProfile.objects.create(
            nif=nif,
            nome_completo=nome,
            telefone=telefone,
            email=email,
            status="PENDENTE",
            is_active=False,
        )

        # Mensagem de sucesso
        messages.success(
            request,
            f"Cadastro enviado com sucesso, {nome.split()[0]}! "
            f"Voce recebera um email em {email} com os proximos passos em ate 48 horas uteis. "
            f"Seu codigo de referencia e: {nif}. "
            f"Status atual: PENDENTE (aguardando analise da equipe).",
        )

        # Redirecionar para evitar reenvio ao atualizar
        return redirect("drivers_app:public_register")

    except Exception as e:
        messages.error(
            request,
            f"Erro ao processar cadastro: {str(e)}. Por favor, tente novamente ou entre em contato.",
        )
        return render(
            request,
            "drivers_app/driver_public_register.html",
            {"current_year": datetime.now().year},
        )


@require_http_methods(["GET", "POST"])
def public_driver_register_full(request):
    """
    Formulario completo de cadastro de motoristas com todos os dados e documentos.
    Multiplos steps, uploads de arquivos e validacao completa.
    """
    if request.method == "GET":
        return render(
            request,
            "drivers_app/driver_public_register_full.html",
            {"current_year": datetime.now().year},
        )

    # POST - Processar formulario completo
    try:
        # === VALIDAR TERMOS ===
        if not request.POST.get("terms"):
            messages.error(request, "Voce deve concordar com os termos para continuar.")
            return redirect("drivers_app:public_register_full")

        # === A. DADOS PESSOAIS ===
        nif = request.POST.get("nif", "").strip()
        nome_completo = request.POST.get("nome_completo", "").strip()
        niss = request.POST.get("niss", "").strip()
        data_nascimento = request.POST.get("data_nascimento")
        nacionalidade = request.POST.get("nacionalidade", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        email = request.POST.get("email", "").strip().lower()
        endereco_residencia = request.POST.get("endereco_residencia", "").strip()
        codigo_postal = request.POST.get("codigo_postal", "").strip()
        cidade = request.POST.get("cidade", "").strip()
        tipo_vinculo = request.POST.get("tipo_vinculo")
        nome_frota = request.POST.get("nome_frota", "").strip()

        # Validar campos obrigatorios
        if not all(
            [
                nif,
                nome_completo,
                niss,
                data_nascimento,
                nacionalidade,
                telefone,
                email,
                endereco_residencia,
                codigo_postal,
                cidade,
                tipo_vinculo,
            ]
        ):
            messages.error(
                request, "Todos os campos obrigatorios devem ser preenchidos."
            )
            return redirect("drivers_app:public_register_full")

        # Validar NIF
        if not re.match(r"^\d{9}$", nif):
            messages.error(request, "NIF invalido. Deve conter exatamente 9 digitos.")
            return redirect("drivers_app:public_register_full")

        # Validar NISS
        if not re.match(r"^\d{11}$", niss):
            messages.error(request, "NISS invalido. Deve conter exatamente 11 digitos.")
            return redirect("drivers_app:public_register_full")

        # Verificar se NIF ja existe
        if DriverProfile.objects.filter(nif=nif).exists():
            messages.error(request, "Este NIF ja esta registrado no sistema.")
            return redirect("drivers_app:public_register_full")

        # === CRIAR PERFIL DO MOTORISTA ===
        driver_profile = DriverProfile.objects.create(
            nif=nif,
            nome_completo=nome_completo,
            niss=niss,
            data_nascimento=data_nascimento,
            nacionalidade=nacionalidade,
            telefone=telefone,
            email=email,
            endereco_residencia=endereco_residencia,
            codigo_postal=codigo_postal,
            cidade=cidade,
            tipo_vinculo=tipo_vinculo,
            nome_frota=nome_frota if tipo_vinculo == "PARCEIRO" else "",
            status="PENDENTE",
            is_active=False,
        )

        # === B. PROCESSAR DOCUMENTOS DO MOTORISTA ===

        # Documento de Identificacao
        if "doc_identificacao" in request.FILES:
            tipo_doc_id = request.POST.get("tipo_doc_id")
            validade_doc_id = request.POST.get("validade_doc_id") or None
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento=tipo_doc_id,
                arquivo=request.FILES["doc_identificacao"],
                data_validade=validade_doc_id,
            )

        # Carta de Conducao - Frente
        if "cnh_frente" in request.FILES:
            categoria_cnh = request.POST.get("categoria_cnh", "").strip()
            validade_cnh = request.POST.get("validade_cnh")
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento="CNH_FRENTE",
                arquivo=request.FILES["cnh_frente"],
                data_validade=validade_cnh,
                categoria_cnh=categoria_cnh,
            )

        # Carta de Conducao - Verso
        if "cnh_verso" in request.FILES:
            categoria_cnh = request.POST.get("categoria_cnh", "").strip()
            validade_cnh = request.POST.get("validade_cnh")
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento="CNH_VERSO",
                arquivo=request.FILES["cnh_verso"],
                data_validade=validade_cnh,
                categoria_cnh=categoria_cnh,
            )

        # Registo Criminal
        if "registo_criminal" in request.FILES:
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento="RC",
                arquivo=request.FILES["registo_criminal"],
            )

        # Certificado ADR (Opcional)
        if "certificado_adr" in request.FILES:
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento="ADR",
                arquivo=request.FILES["certificado_adr"],
            )

        # Declaracao de Atividade (apenas para DIRETO)
        if tipo_vinculo == "DIRETO" and "declaracao_atividade" in request.FILES:
            DriverDocument.objects.create(
                motorista=driver_profile,
                tipo_documento="DECLARACAO_ATIVIDADE",
                arquivo=request.FILES["declaracao_atividade"],
            )

        # === C. DADOS DO VEICULO ===
        matricula = request.POST.get("matricula", "").strip().upper()
        marca = request.POST.get("marca", "").strip()
        modelo = request.POST.get("modelo", "").strip()
        tipo_veiculo = request.POST.get("tipo_veiculo")
        ano = request.POST.get("ano") or None

        if all([matricula, marca, modelo, tipo_veiculo]):
            # Criar veiculo
            vehicle = Vehicle.objects.create(
                motorista=driver_profile,
                matricula=matricula,
                marca=marca,
                modelo=modelo,
                tipo_veiculo=tipo_veiculo,
                ano=ano,
                is_active=True,
            )

            # DUA
            if "dua" in request.FILES:
                VehicleDocument.objects.create(
                    veiculo=vehicle,
                    tipo_documento="DUA",
                    arquivo=request.FILES["dua"],
                )

            # IPO
            if "ipo" in request.FILES:
                validade_ipo = request.POST.get("validade_ipo") or None
                VehicleDocument.objects.create(
                    veiculo=vehicle,
                    tipo_documento="IPO",
                    arquivo=request.FILES["ipo"],
                    data_validade=validade_ipo,
                )

            # Seguro
            if "seguro" in request.FILES:
                validade_seguro = request.POST.get("validade_seguro")
                VehicleDocument.objects.create(
                    veiculo=vehicle,
                    tipo_documento="SEGURO",
                    arquivo=request.FILES["seguro"],
                    data_validade=validade_seguro,
                )

        # === MENSAGEM DE SUCESSO ===
        messages.success(
            request,
            f"Cadastro completo enviado com sucesso, {nome_completo.split()[0]}! "
            f"Voce recebera um email em {email} com os proximos passos em ate 48 horas uteis. "
            f"Seu codigo de referencia e: {nif}. "
            f"Status atual: PENDENTE (aguardando analise da equipe).",
        )

        # Atualizar status para EM_ANALISE se todos os documentos foram enviados
        if driver_profile.documents.count() >= 4:  # Minimo de docs esperados
            driver_profile.status = "EM_ANALISE"
            driver_profile.save()

        return redirect("drivers_app:public_register_full")

    except Exception as e:
        messages.error(
            request,
            f"Erro ao processar cadastro: {str(e)}. Por favor, tente novamente.",
        )
        return redirect("drivers_app:public_register_full")


# === VIEWS ADMINISTRATIVAS ===


@login_required
def admin_create_driver(request):
    """
    Pagina administrativa para criar motorista ou frota diretamente (sem formulario publico).
    Acesso apenas para equipe interna autenticada.
    """
    if request.method == "POST":
        try:
            # Criar perfil
            nif = request.POST.get("nif", "").strip()
            nome_completo = request.POST.get("nome_completo", "").strip()
            email = request.POST.get("email", "").strip().lower()
            telefone = request.POST.get("telefone", "").strip()
            tipo_vinculo = request.POST.get("tipo_vinculo", "DIRETO")
            nome_frota = request.POST.get("nome_frota", "").strip()

            # Validar NIF unico
            if DriverProfile.objects.filter(nif=nif).exists():
                messages.error(request, f"NIF {nif} ja cadastrado no sistema.")
                return redirect("drivers_app:admin_create_driver")

            # Criar perfil com status ATIVO (criacao interna ja e aprovada)
            driver_profile = DriverProfile.objects.create(
                nif=nif,
                nome_completo=nome_completo,
                email=email,
                telefone=telefone,
                tipo_vinculo=tipo_vinculo,
                nome_frota=nome_frota if tipo_vinculo == "PARCEIRO" else "",
                status="ATIVO",
                is_active=True,
                approved_at=timezone.now(),
                approved_by=request.user.username,
            )

            messages.success(
                request,
                f"Motorista {nome_completo} criado com sucesso! Status: ATIVO",
            )
            return redirect("drivers_app:admin_active_drivers")

        except Exception as e:
            messages.error(request, f"Erro ao criar motorista: {str(e)}")
            return redirect("drivers_app:admin_create_driver")

    return render(
        request,
        "drivers_app/admin_create_driver.html",
        {"current_year": datetime.now().year},
    )


@login_required
def admin_approve_drivers(request):
    """
    Area para aprovacao de motoristas pendentes e em analise.
    Permite visualizar documentos, aprovar ou rejeitar cadastros.
    """
    # Filtrar motoristas pendentes e em analise
    status_filter = request.GET.get("status", "")
    search_query = request.GET.get("q", "")

    drivers = (
        DriverProfile.objects.filter(status__in=["PENDENTE", "EM_ANALISE"])
        .select_related()
        .prefetch_related("documents", "vehicles")
    )

    if status_filter:
        drivers = drivers.filter(status=status_filter)

    if search_query:
        drivers = drivers.filter(
            Q(nif__icontains=search_query)
            | Q(nome_completo__icontains=search_query)
            | Q(apelido__icontains=search_query)
            | Q(courier_id_cainiao__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    drivers = drivers.order_by("-created_at")

    # Paginacao
    paginator = Paginator(drivers, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Contadores
    total_pendentes = DriverProfile.objects.filter(status="PENDENTE").count()
    total_em_analise = DriverProfile.objects.filter(status="EM_ANALISE").count()

    return render(
        request,
        "drivers_app/admin_approve_drivers.html",
        {
            "page_obj": page_obj,
            "total_pendentes": total_pendentes,
            "total_em_analise": total_em_analise,
            "status_filter": status_filter,
            "search_query": search_query,
        },
    )


@login_required
def admin_approve_driver_action(request, driver_id):
    """
    Acao para aprovar ou rejeitar um motorista.
    POST com action=approve ou action=reject
    """
    if request.method != "POST":
        return redirect("drivers_app:admin_approve_drivers")

    try:
        driver = DriverProfile.objects.get(id=driver_id)
        action = request.POST.get("action")

        if action == "approve":
            driver.status = "ATIVO"
            driver.is_active = True
            driver.approved_at = timezone.now()
            driver.approved_by = request.user.username
            driver.save()
            messages.success(
                request,
                f"Motorista {driver.nome_completo} aprovado com sucesso!",
            )

        elif action == "reject":
            observacao = request.POST.get(
                "observacao", "Cadastro rejeitado pela equipe"
            )
            driver.status = "IRREGULAR"
            driver.is_active = False
            driver.observacoes = observacao
            driver.save()
            messages.warning(request, f"Motorista {driver.nome_completo} rejeitado.")

        return redirect("drivers_app:admin_approve_drivers")

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista nao encontrado.")
        return redirect("drivers_app:admin_approve_drivers")


@login_required
def admin_active_drivers(request):
    """
    Lista de motoristas aprovados e ativos.
    Permite buscar, filtrar e gerenciar motoristas ativos.
    """
    search_query = request.GET.get("q", "")
    tipo_vinculo = request.GET.get("tipo_vinculo", "")
    status_filter = request.GET.get("status_filter", "")
    hub_filter = request.GET.get("hub_id", "")
    sort_by = request.GET.get("sort", "name")  # name | recent
    has_open_complaint = request.GET.get("complaint", "")  # "1" ativo
    inactivity_filter = request.GET.get("inactive", "")  # "30" / "60"

    # Filtro base - mostrar motoristas aprovados (ATIVO, BLOQUEADO, IRREGULAR)
    # Excluir apenas PENDENTE e EM_ANALISE
    drivers = (
        DriverProfile.objects.exclude(status__in=["PENDENTE", "EM_ANALISE"])
        .select_related()
        .prefetch_related("documents", "vehicles")
    )

    # Filtro de status específico
    if status_filter:
        drivers = drivers.filter(status=status_filter)

    if search_query:
        drivers = drivers.filter(
            Q(nif__icontains=search_query)
            | Q(nome_completo__icontains=search_query)
            | Q(apelido__icontains=search_query)
            | Q(courier_id_cainiao__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(telefone__icontains=search_query)
        )

    if tipo_vinculo:
        drivers = drivers.filter(tipo_vinculo=tipo_vinculo)

    # Filtro por HUB: drivers com operação em CP4s do HUB nos últimos 90 dias.
    if hub_filter:
        try:
            from settlements.models import (
                CainiaoHub, CainiaoHubCP4, CainiaoOperationTask,
            )
            from datetime import timedelta
            from django.utils import timezone as _tz
            hub = CainiaoHub.objects.filter(id=int(hub_filter)).first()
            if hub:
                cp4s = list(
                    CainiaoHubCP4.objects.filter(hub=hub)
                    .values_list("cp4", flat=True)
                )
                if cp4s:
                    cutoff = _tz.now().date() - timedelta(days=90)
                    zip_q = Q()
                    for cp4 in cp4s:
                        zip_q |= Q(zip_code__startswith=cp4)
                    courier_ids = set(
                        CainiaoOperationTask.objects.filter(
                            zip_q, task_date__gte=cutoff,
                        )
                        .exclude(courier_id_cainiao="")
                        .values_list("courier_id_cainiao", flat=True)
                        .distinct()
                    )
                    drivers = drivers.filter(
                        courier_id_cainiao__in=courier_ids,
                    )
        except (ValueError, TypeError, ImportError):
            pass

    # Filtro: drivers com reclamação aberta
    if has_open_complaint == "1":
        from .models import CustomerComplaint
        open_status = ["ABERTO", "NOTIFICADO", "RESPONDIDO"]
        ids_with_open = set(
            CustomerComplaint.objects.filter(
                status__in=open_status,
            ).values_list("driver_id", flat=True)
        )
        drivers = drivers.filter(id__in=ids_with_open)

    # Pré-calcula última operação por courier_id_cainiao (para filtro
    # de inactividade e enriquecimento da página)
    from datetime import timedelta
    from settlements.models import CainiaoOperationTask
    today = timezone.now().date()

    last_op_by_courier = dict(
        CainiaoOperationTask.objects
        .exclude(courier_id_cainiao="")
        .values_list("courier_id_cainiao")
        .annotate(last=Max("task_date"))
        .values_list("courier_id_cainiao", "last")
    )

    if inactivity_filter:
        try:
            n_days = int(inactivity_filter)
            cutoff = today - timedelta(days=n_days)
            inactive_couriers = {
                cid for cid, dt in last_op_by_courier.items()
                if dt and dt < cutoff
            }
            # Drivers sem courier_id também são "inactivos"
            drivers = drivers.filter(
                Q(courier_id_cainiao__in=inactive_couriers)
                | Q(courier_id_cainiao="")
            )
        except (ValueError, TypeError):
            pass

    # Ordenação: alfabética por defeito (apelido → nome completo)
    if sort_by == "recent":
        drivers = drivers.order_by("-approved_at")
    elif sort_by == "last_op":
        # Ordenação em Python depois (precisa de last_op_date)
        drivers = drivers.order_by("apelido", "nome_completo")
    else:
        drivers = drivers.order_by("apelido", "nome_completo")

    # Paginacao
    paginator = Paginator(drivers, 30)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ─── Enriquecer cada driver da página com dados auxiliares ───────
    # Evita N+1: query batch de complaints abertas e docs a expirar
    page_drivers = list(page_obj.object_list)
    page_obj.object_list = page_drivers
    page_driver_ids = [d.id for d in page_drivers]

    # Complaints abertas por driver
    from .models import CustomerComplaint
    open_status = ["ABERTO", "NOTIFICADO", "RESPONDIDO"]
    open_complaints_by_driver = dict(
        CustomerComplaint.objects.filter(
            driver_id__in=page_driver_ids,
            status__in=open_status,
        ).values_list("driver_id")
        .annotate(n=Count("id"))
        .values_list("driver_id", "n")
    )

    # Documentos a expirar nos próximos 30 dias por driver
    cutoff_30 = today + timedelta(days=30)
    expiring_docs_by_driver = {}
    expired_docs_by_driver = {}
    for doc in DriverDocument.objects.filter(
        motorista_id__in=page_driver_ids,
    ).exclude(data_validade__isnull=True):
        d_id = doc.motorista_id
        if doc.data_validade < today:
            expired_docs_by_driver.setdefault(d_id, 0)
            expired_docs_by_driver[d_id] += 1
        elif doc.data_validade <= cutoff_30:
            expiring_docs_by_driver.setdefault(d_id, 0)
            expiring_docs_by_driver[d_id] += 1

    # Anexa atributos transientes a cada driver
    for d in page_drivers:
        last_op = (
            last_op_by_courier.get(d.courier_id_cainiao)
            if d.courier_id_cainiao else None
        )
        d.last_op_date = last_op
        d.days_since_op = (today - last_op).days if last_op else None
        d.has_open_complaint = (
            open_complaints_by_driver.get(d.id, 0) > 0
        )
        d.n_open_complaints = open_complaints_by_driver.get(d.id, 0)
        d.n_expiring_docs = expiring_docs_by_driver.get(d.id, 0)
        d.n_expired_docs = expired_docs_by_driver.get(d.id, 0)

        # Health status: red > yellow > green
        if (
            d.status in ("BLOQUEADO", "IRREGULAR")
            or d.n_expired_docs > 0
            or (d.days_since_op is not None and d.days_since_op > 30)
        ):
            d.health = "red"
            if d.status == "BLOQUEADO":
                d.health_label = "Bloqueado"
            elif d.status == "IRREGULAR":
                d.health_label = "Irregular"
            elif d.n_expired_docs > 0:
                d.health_label = (
                    f"{d.n_expired_docs} doc(s) expirado(s)"
                )
            else:
                d.health_label = (
                    f"Sem operação há {d.days_since_op}d"
                )
        elif (
            d.has_open_complaint
            or d.n_expiring_docs > 0
            or (d.days_since_op is not None and d.days_since_op > 7)
        ):
            d.health = "yellow"
            parts = []
            if d.has_open_complaint:
                parts.append(f"{d.n_open_complaints} reclamação(ões)")
            if d.n_expiring_docs > 0:
                parts.append(f"{d.n_expiring_docs} doc(s) a expirar")
            if (
                d.days_since_op is not None
                and d.days_since_op > 7
            ):
                parts.append(f"Sem op. há {d.days_since_op}d")
            d.health_label = " · ".join(parts)
        else:
            d.health = "green"
            d.health_label = "Saudável"

    # Re-ordena se for sort=last_op (None vai para o fim)
    if sort_by == "last_op":
        page_drivers.sort(
            key=lambda x: (
                x.last_op_date is None,
                -(x.last_op_date.toordinal() if x.last_op_date else 0),
            )
        )
        page_obj.object_list = page_drivers

    # Estatisticas (para os KPI cards)
    total_ativos = DriverProfile.objects.filter(
        status="ATIVO", is_active=True,
    ).count()
    total_diretos = DriverProfile.objects.filter(
        status="ATIVO", tipo_vinculo="DIRETO",
    ).count()
    total_parceiros = DriverProfile.objects.filter(
        status="ATIVO", tipo_vinculo="PARCEIRO",
    ).count()
    total_bloqueados = DriverProfile.objects.filter(
        status="BLOQUEADO",
    ).count()

    # Mini-cards extra
    # Drivers com algum doc a expirar em <=30d
    n_drivers_expiring = (
        DriverDocument.objects
        .exclude(data_validade__isnull=True)
        .filter(data_validade__gte=today, data_validade__lte=cutoff_30)
        .values("motorista_id").distinct().count()
    )
    # Drivers sem operação há >30d (entre os ATIVO+is_active)
    cutoff_30d_back = today - timedelta(days=30)
    inactive_couriers_30 = {
        cid for cid, dt in last_op_by_courier.items()
        if dt and dt < cutoff_30d_back
    }
    n_drivers_inactive_30 = DriverProfile.objects.filter(
        status="ATIVO", is_active=True,
    ).filter(
        Q(courier_id_cainiao__in=inactive_couriers_30)
        | Q(courier_id_cainiao="")
    ).count()
    # Drivers com reclamação aberta
    n_drivers_open_complaints = (
        CustomerComplaint.objects.filter(status__in=open_status)
        .values("driver_id").distinct().count()
    )

    # Alertas de documentos proximos a vencer (mantém compat)
    drivers_with_expiring_docs = []
    for driver in page_drivers[:10]:
        expiring = driver.get_documents_expiring_soon(days=30)
        if expiring.exists():
            drivers_with_expiring_docs.append(
                {"driver": driver, "docs": expiring}
            )

    empresas_parceiras = EmpresaParceira.objects.filter(
        ativo=True,
    ).order_by("nome")

    # HUBs disponíveis para o filtro
    try:
        from settlements.models import CainiaoHub
        hubs = CainiaoHub.objects.all().order_by("name")
    except ImportError:
        hubs = []

    return render(
        request,
        "drivers_app/admin_active_drivers.html",
        {
            "page_obj": page_obj,
            "total_ativos": total_ativos,
            "total_diretos": total_diretos,
            "total_parceiros": total_parceiros,
            "total_bloqueados": total_bloqueados,
            "n_drivers_expiring": n_drivers_expiring,
            "n_drivers_inactive_30": n_drivers_inactive_30,
            "n_drivers_open_complaints": n_drivers_open_complaints,
            "drivers_with_expiring_docs": drivers_with_expiring_docs,
            "search_query": search_query,
            "tipo_vinculo": tipo_vinculo,
            "status_filter": status_filter,
            "hub_filter": hub_filter,
            "sort_by": sort_by,
            "has_open_complaint": has_open_complaint,
            "inactivity_filter": inactivity_filter,
            "hubs": hubs,
            "empresas_parceiras": empresas_parceiras,
        },
    )


# === VIEWS DE GESTÃO DE MOTORISTAS ATIVOS ===


@login_required
@require_http_methods(["POST"])
def admin_edit_driver_personal(request, driver_id):
    """Editar dados pessoais de um motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        driver.nif = request.POST.get("nif", driver.nif)
        driver.niss = request.POST.get("niss", driver.niss)
        driver.nome_completo = request.POST.get("nome_completo", driver.nome_completo)
        driver.data_nascimento = (
            request.POST.get("data_nascimento") or driver.data_nascimento
        )
        driver.nacionalidade = request.POST.get("nacionalidade", driver.nacionalidade)
        driver.email = request.POST.get("email", driver.email)
        driver.telefone = request.POST.get("telefone", driver.telefone)

        # Parse endereço completo (simplificado)
        endereco = request.POST.get("endereco", "")
        if endereco:
            driver.endereco_residencia = endereco

        driver.save()
        messages.success(
            request,
            f"Dados pessoais de {driver.nome_completo} atualizados com sucesso!",
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao atualizar dados: {str(e)}")

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Dados atualizados com sucesso!",
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )
        return JsonResponse(
            {"success": False, "message": "Erro ao atualizar"}, status=400
        )

    return redirect("drivers_app:admin_active_drivers")


@login_required
def admin_get_driver_data(request, driver_id):
    """API para obter dados atualizados do motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)
        return JsonResponse(serialize_driver_data(driver))
    except DriverProfile.DoesNotExist:
        return JsonResponse({"error": "Motorista não encontrado"}, status=404)


@login_required
@require_http_methods(["POST"])
def admin_edit_driver_professional(request, driver_id):
    """Editar dados profissionais de um motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        driver.tipo_vinculo = request.POST.get("tipo_vinculo", driver.tipo_vinculo)
        driver.nome_frota = request.POST.get("nome_frota", driver.nome_frota)

        empresa_id = request.POST.get("empresa_parceira_id")
        if empresa_id:
            driver.empresa_parceira_id = int(empresa_id) if empresa_id != "0" else None
        elif request.POST.get("tipo_vinculo") == "DIRETO":
            driver.empresa_parceira = None

        # Override de valor por pacote (se toggle activo + valor > 0, grava;
        # caso contrário, limpa para usar o preço do parceiro)
        from decimal import Decimal, InvalidOperation
        override_on = request.POST.get("price_override_enabled") in (
            "on", "true", "1", "yes",
        )
        raw_price = (request.POST.get("price_per_package") or "").strip()
        if override_on and raw_price:
            try:
                v = Decimal(raw_price.replace(",", "."))
                driver.price_per_package = v if v > 0 else None
            except InvalidOperation:
                pass
        else:
            driver.price_per_package = None

        driver.save()
        messages.success(
            request,
            f"Dados profissionais de {driver.nome_completo} atualizados!",
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao atualizar: {str(e)}")

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Dados profissionais atualizados!",
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )
        return JsonResponse({"success": False, "message": "Erro"}, status=400)

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_attach_document(request, driver_id):
    """Anexar novo documento a um motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        tipo_documento = request.POST.get("tipo_documento")
        data_validade = request.POST.get("data_validade") or None
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            messages.error(request, "Nenhum arquivo foi enviado.")
            return redirect("drivers_app:admin_active_drivers")

        DriverDocument.objects.create(
            motorista=driver,
            tipo_documento=tipo_documento,
            arquivo=arquivo,
            data_validade=data_validade,
        )

        messages.success(
            request,
            f'Documento "{DriverDocument.TIPO_DOCUMENTO_CHOICES[0][1]}" anexado com sucesso!',
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao anexar documento: {str(e)}")

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Documento anexado com sucesso!",
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )
        return JsonResponse({"success": False, "message": "Erro ao anexar"}, status=400)

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_delete_document(request, document_id):
    """Deletar documento de motorista"""
    try:
        documento = DriverDocument.objects.get(id=document_id)
        tipo = documento.get_tipo_documento_display()

        # Deletar arquivo físico
        if documento.arquivo:
            documento.arquivo.delete()

        documento.delete()
        messages.success(request, f'Documento "{tipo}" deletado com sucesso!')
        driver_id = documento.motorista.id

    except DriverDocument.DoesNotExist:
        messages.error(request, "Documento não encontrado.")
        driver_id = None
    except Exception as e:
        messages.error(request, f"Erro ao deletar documento: {str(e)}")
        driver_id = None

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" and driver_id:
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": f'Documento "{tipo}" deletado!',
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_add_vehicle(request, driver_id):
    """Adicionar novo veículo a um motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        matricula = request.POST.get("matricula", "").strip().upper()
        marca = request.POST.get("marca", "").strip()
        modelo = request.POST.get("modelo", "").strip()
        tipo_veiculo = request.POST.get("tipo_veiculo")
        ano = request.POST.get("ano") or None

        # Verificar se matrícula já existe
        if Vehicle.objects.filter(matricula=matricula).exists():
            messages.error(request, f"A matrícula {matricula} já está cadastrada.")
            return redirect("drivers_app:admin_active_drivers")

        Vehicle.objects.create(
            motorista=driver,
            matricula=matricula,
            marca=marca,
            modelo=modelo,
            tipo_veiculo=tipo_veiculo,
            ano=ano,
            is_active=True,
        )

        messages.success(
            request,
            f"Veículo {matricula} ({marca} {modelo}) adicionado com sucesso!",
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao adicionar veículo: {str(e)}")

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Veículo adicionado com sucesso!",
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )
        return JsonResponse({"success": False, "message": "Erro"}, status=400)

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_delete_vehicle(request, vehicle_id):
    """Deletar veículo de motorista"""
    try:
        veiculo = Vehicle.objects.get(id=vehicle_id)
        matricula = veiculo.matricula

        # Deletar documentos do veículo
        for doc in veiculo.vehicle_documents.all():
            if doc.arquivo:
                doc.arquivo.delete()

        veiculo.delete()
        messages.success(request, f"Veículo {matricula} deletado com sucesso!")
        driver_id = veiculo.motorista.id

    except Vehicle.DoesNotExist:
        messages.error(request, "Veículo não encontrado.")
        driver_id = None
    except Exception as e:
        messages.error(request, f"Erro ao deletar veículo: {str(e)}")
        driver_id = None

    # Se for AJAX, retornar JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" and driver_id:
        driver = DriverProfile.objects.filter(id=driver_id).first()
        if driver:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Veículo {matricula} deletado!",
                    "driver_id": driver.id,
                    "driver_data": serialize_driver_data(driver),
                }
            )

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_delete_vehicle_document(request, document_id):
    """Deletar documento de veículo"""
    try:
        documento = VehicleDocument.objects.get(id=document_id)
        tipo = documento.get_tipo_documento_display()

        if documento.arquivo:
            documento.arquivo.delete()

        documento.delete()
        messages.success(request, f'Documento "{tipo}" do veículo deletado!')

    except VehicleDocument.DoesNotExist:
        messages.error(request, "Documento não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao deletar: {str(e)}")

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_deactivate_driver(request, driver_id):
    """Desativar motorista"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        driver.status = "BLOQUEADO"
        driver.is_active = False
        driver.observacoes = f'Desativado por {request.user.username} em {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        driver.save()

        messages.warning(request, f"Motorista {driver.nome_completo} desativado.")

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao desativar: {str(e)}")

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_activate_driver(request, driver_id):
    """Ativar motorista (reverter bloqueio)"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)

        driver.status = "ATIVO"
        driver.is_active = True
        driver.observacoes = f'Reativado por {request.user.username} em {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        driver.save()

        messages.success(
            request, f"Motorista {driver.nome_completo} reativado com sucesso!"
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao ativar: {str(e)}")

    return redirect("drivers_app:admin_active_drivers")


@login_required
@require_http_methods(["POST"])
def admin_delete_driver(request, driver_id):
    """Excluir permanentemente um motorista e todos os seus dados"""
    try:
        driver = DriverProfile.objects.get(id=driver_id)
        nome = driver.nome_completo

        # Deletar todos os arquivos de documentos
        for doc in driver.documents.all():
            if doc.arquivo:
                doc.arquivo.delete()

        # Deletar todos os arquivos de documentos de veículos
        for vehicle in driver.vehicles.all():
            for vdoc in vehicle.vehicle_documents.all():
                if vdoc.arquivo:
                    vdoc.arquivo.delete()

        # Django vai deletar em cascata: documents, vehicles, vehicle_documents
        driver.delete()

        messages.success(
            request,
            f"Motorista {nome} e todos os seus dados foram excluídos permanentemente.",
        )

    except DriverProfile.DoesNotExist:
        messages.error(request, "Motorista não encontrado.")
    except Exception as e:
        messages.error(request, f"Erro ao excluir: {str(e)}")

    return redirect("drivers_app:admin_active_drivers")


# ============================================================================
# INDICAÇÕES / REFERRALS
# ============================================================================

@login_required
def referral_search_drivers(request):
    """Lista todos os motoristas disponíveis para indicação."""
    from .models import DriverReferral
    exclude_id = request.GET.get("exclude", "")
    qs = DriverProfile.objects.all().order_by("nome_completo")
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    referral_map = {
        r.referred_id: r.referrer.nome_completo
        for r in DriverReferral.objects.select_related("referrer").all()
    }
    results = [
        {
            "id": d.id,
            "nome": d.nome_completo,
            "nif": d.nif or "",
            "ativo": d.is_active,
            "ja_indicado": d.id in referral_map,
            "indicado_por": referral_map.get(d.id, ""),
        }
        for d in qs[:500]
    ]
    return JsonResponse({"results": results})


@login_required
@require_POST
def referral_add(request, driver_id):
    """Adiciona uma indicação: driver_id é o REFERRER (quem indicou)."""
    import json
    from decimal import Decimal, InvalidOperation
    from .models import DriverReferral

    referrer = get_object_or_404(DriverProfile, id=driver_id)
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    referred_id = body.get("referred_id")
    comissao = body.get("comissao_por_pacote", "0.05")
    notas = body.get("notas", "")

    if not referred_id:
        return JsonResponse({"success": False, "error": "Motorista indicado não especificado."}, status=400)

    referred = get_object_or_404(DriverProfile, id=referred_id)

    if referred.id == referrer.id:
        return JsonResponse({"success": False, "error": "Um motorista não pode indicar-se a si próprio."}, status=400)

    if hasattr(referred, "referral_received"):
        return JsonResponse({"success": False, "error": f"{referred.nome_completo} já tem uma indicação registada."}, status=400)

    try:
        comissao_dec = Decimal(str(comissao).replace(",", "."))
    except (InvalidOperation, TypeError):
        comissao_dec = Decimal("0.05")

    ref = DriverReferral.objects.create(
        referrer=referrer,
        referred=referred,
        comissao_por_pacote=comissao_dec,
        notas=notas,
    )
    return JsonResponse({
        "success": True,
        "referral": {
            "id": ref.id,
            "referred_id": referred.id,
            "referred_nome": referred.nome_completo,
            "comissao_por_pacote": str(ref.comissao_por_pacote),
            "ativo": ref.ativo,
            "notas": ref.notas,
        },
    })


@login_required
@require_POST
def referral_update(request, referral_id):
    """Actualiza comissão/notas/ativo de uma indicação."""
    import json
    from decimal import Decimal, InvalidOperation
    from .models import DriverReferral

    ref = get_object_or_404(DriverReferral, id=referral_id)
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    if "comissao_por_pacote" in body:
        try:
            ref.comissao_por_pacote = Decimal(str(body["comissao_por_pacote"]).replace(",", "."))
        except (InvalidOperation, TypeError):
            pass
    if "ativo" in body:
        ref.ativo = bool(body["ativo"])
    if "notas" in body:
        ref.notas = body["notas"]
    ref.save()
    return JsonResponse({"success": True})


@login_required
@require_POST
def referral_delete(request, referral_id):
    """Remove uma indicação."""
    from .models import DriverReferral
    ref = get_object_or_404(DriverReferral, id=referral_id)
    ref.delete()
    return JsonResponse({"success": True})


@login_required
def referral_list(request, driver_id):
    """Devolve as indicações de um motorista (dadas e recebida)."""
    from .models import DriverReferral
    driver = get_object_or_404(DriverProfile, id=driver_id)

    given = DriverReferral.objects.filter(referrer=driver).select_related("referred")
    given_data = [
        {
            "id": r.id,
            "referred_id": r.referred.id,
            "referred_nome": r.referred.nome_completo,
            "comissao_por_pacote": str(r.comissao_por_pacote),
            "ativo": r.ativo,
            "notas": r.notas,
        }
        for r in given
    ]

    received = None
    if hasattr(driver, "referral_received"):
        r = driver.referral_received
        received = {
            "id": r.id,
            "referrer_id": r.referrer.id,
            "referrer_nome": r.referrer.nome_completo,
            "comissao_por_pacote": str(r.comissao_por_pacote),
            "ativo": r.ativo,
        }

    return JsonResponse({"given": given_data, "received": received})


# ============================================================================
# EMPRESAS PARCEIRAS
# ============================================================================

@login_required
def empresas_parceiras_list(request):
    """Lista e gestão de empresas parceiras."""
    import json as _json
    empresas = EmpresaParceira.objects.all().order_by("nome")
    empresas_json = _json.dumps([_empresa_to_dict(e) for e in empresas])
    return render(request, "drivers_app/empresas_parceiras.html", {
        "empresas": empresas,
        "empresas_json": empresas_json,
    })


@login_required
@require_POST
def empresa_parceira_create(request):
    """Cria uma nova empresa parceira."""
    import json as _json
    try:
        body = _json.loads(request.body)
    except Exception:
        body = request.POST.dict()

    from decimal import Decimal as _D, InvalidOperation
    nome = body.get("nome", "").strip()
    if not nome:
        return JsonResponse({"success": False, "error": "Nome é obrigatório."})

    try:
        taxa = _D(str(body.get("taxa_iva", "23.00")).replace(",", "."))
    except InvalidOperation:
        taxa = _D("23.00")

    # Default price per package (opcional)
    raw_dpp = (body.get("driver_default_price_per_package") or "").strip() if isinstance(body.get("driver_default_price_per_package"), str) else body.get("driver_default_price_per_package")
    dpp = None
    if raw_dpp not in (None, ""):
        try:
            v = _D(str(raw_dpp).replace(",", "."))
            dpp = v if v > 0 else None
        except InvalidOperation:
            pass

    empresa = EmpresaParceira.objects.create(
        nome=nome,
        nif=body.get("nif", "").strip(),
        morada=body.get("morada", "").strip(),
        codigo_postal=body.get("codigo_postal", "").strip(),
        cidade=body.get("cidade", "").strip(),
        email=body.get("email", "").strip(),
        telefone=body.get("telefone", "").strip(),
        contacto_nome=body.get("contacto_nome", "").strip(),
        iban=body.get("iban", "").strip(),
        taxa_iva=taxa,
        driver_default_price_per_package=dpp,
        notas=body.get("notas", "").strip(),
        ativo=bool(body.get("ativo", True)),
    )
    return JsonResponse({"success": True, "empresa": _empresa_to_dict(empresa)})


@login_required
@require_POST
def empresa_parceira_update(request, empresa_id):
    """Atualiza uma empresa parceira."""
    import json as _json
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    try:
        body = _json.loads(request.body)
    except Exception:
        body = request.POST.dict()

    from decimal import Decimal as _D, InvalidOperation
    for field in ["nome", "nif", "morada", "codigo_postal", "cidade",
                  "email", "telefone", "contacto_nome", "iban", "notas"]:
        if field in body:
            setattr(empresa, field, str(body[field]).strip())
    if "taxa_iva" in body:
        try:
            empresa.taxa_iva = _D(str(body["taxa_iva"]).replace(",", "."))
        except InvalidOperation:
            pass
    if "driver_default_price_per_package" in body:
        raw = body["driver_default_price_per_package"]
        if raw in (None, "", "0", 0):
            empresa.driver_default_price_per_package = None
        else:
            try:
                v = _D(str(raw).replace(",", "."))
                empresa.driver_default_price_per_package = v if v > 0 else None
            except InvalidOperation:
                pass
    if "ativo" in body:
        empresa.ativo = bool(body["ativo"])
    empresa.save()
    return JsonResponse({"success": True, "empresa": _empresa_to_dict(empresa)})


@login_required
@require_POST
def empresa_parceira_delete(request, empresa_id):
    """Elimina uma empresa parceira (apenas se não tiver motoristas associados)."""
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    if empresa.motoristas.exists():
        count = empresa.motoristas.count()
        return JsonResponse({
            "success": False,
            "error": f"Não é possível eliminar: {count} motorista(s) associado(s). Reatribua-os primeiro."
        })
    empresa.delete()
    return JsonResponse({"success": True})


@login_required
def empresa_parceira_motoristas(request, empresa_id):
    """Devolve motoristas de uma empresa parceira."""
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    drivers = empresa.motoristas.all().order_by("nome_completo")
    return JsonResponse({
        "empresa": _empresa_to_dict(empresa),
        "motoristas": [
            {
                "id": d.id,
                "nome": d.nome_completo,
                "nif": d.nif,
                "status": d.status,
                "status_display": d.get_status_display(),
                "is_active": d.is_active,
            }
            for d in drivers
        ],
    })


@login_required
def empresa_parceira_assign_driver(request, empresa_id, driver_id):
    """Associa um motorista a uma empresa parceira."""
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    driver = get_object_or_404(DriverProfile, id=driver_id)
    driver.empresa_parceira = empresa
    driver.tipo_vinculo = "PARCEIRO"
    driver.save(update_fields=["empresa_parceira", "tipo_vinculo"])
    return JsonResponse({
        "success": True,
        "driver": {
            "id": driver.id,
            "nome": driver.nome_completo,
            "nif": driver.nif,
            "status": driver.status,
            "status_display": driver.get_status_display(),
            "is_active": driver.is_active,
        },
    })


@login_required
@require_POST
def empresa_parceira_remove_driver(request, empresa_id, driver_id):
    """Remove um motorista de uma empresa parceira."""
    driver = get_object_or_404(DriverProfile, id=driver_id, empresa_parceira_id=empresa_id)
    driver.empresa_parceira = None
    driver.save(update_fields=["empresa_parceira"])
    return JsonResponse({"success": True})


@login_required
def empresa_parceira_search_drivers(request, empresa_id):
    """Pesquisa motoristas para associar à empresa."""
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    q = request.GET.get("q", "").strip()
    drivers = DriverProfile.objects.filter(is_active=True)
    if q:
        drivers = drivers.filter(
            Q(nome_completo__icontains=q) | Q(nif__icontains=q)
        )
    results = []
    for d in drivers.select_related("empresa_parceira")[:40]:
        results.append({
            "id": d.id,
            "nome": d.nome_completo,
            "nif": d.nif,
            "empresa_atual": d.empresa_parceira.nome if d.empresa_parceira else None,
            "empresa_atual_id": d.empresa_parceira_id,
            "nesta_empresa": d.empresa_parceira_id == empresa_id,
        })
    return JsonResponse({"drivers": results})


@login_required
def empresa_parceira_prefaturas(request, empresa_id):
    """Pré-faturas agregadas de todos os motoristas da empresa num período."""
    from decimal import Decimal as D
    from settlements.models import DriverPreInvoice
    from .models import EmpresaParceiraLancamento
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    try:
        mes = int(request.GET.get("mes", timezone.now().month))
        ano = int(request.GET.get("ano", timezone.now().year))
    except (ValueError, TypeError):
        mes = timezone.now().month
        ano = timezone.now().year

    driver_ids = list(empresa.motoristas.values_list("id", flat=True))
    pre_invoices = (
        DriverPreInvoice.objects
        .filter(driver_id__in=driver_ids, periodo_inicio__month=mes, periodo_inicio__year=ano)
        .select_related("driver")
        .order_by("driver__nome_completo")
    )

    rows = []
    total_base = D("0")
    total_bonus = D("0")
    total_liquido = D("0")
    for pf in pre_invoices:
        rows.append({
            "id": pf.id,
            "tipo": "prefatura",
            "numero": pf.numero,
            "driver_id": pf.driver_id,
            "driver_nome": pf.driver.nome_completo,
            "driver_nif": pf.driver.nif,
            "periodo": f"{pf.periodo_inicio.strftime('%d/%m')} – {pf.periodo_fim.strftime('%d/%m/%Y')}",
            "base_entregas": str(pf.base_entregas),
            "total_bonus": str(pf.total_bonus),
            "total_a_receber": str(pf.total_a_receber),
            "status": pf.status,
            "status_display": pf.get_status_display(),
        })
        total_base += pf.base_entregas
        total_bonus += pf.total_bonus
        total_liquido += pf.total_a_receber

    # Lançamentos manuais do mesmo período
    lancamentos = EmpresaParceiraLancamento.objects.filter(
        empresa=empresa,
        periodo_inicio__month=mes,
        periodo_inicio__year=ano,
    ).order_by("periodo_inicio")

    lanc_rows = []
    total_lanc_base = D("0")
    total_lanc_bonus = D("0")
    total_lanc_total = D("0")
    for lc in lancamentos:
        lanc_rows.append({
            "id": lc.id,
            "descricao": lc.descricao,
            "qtd_entregas": lc.qtd_entregas,
            "valor_por_entrega": str(lc.valor_por_entrega),
            "valor_base": str(lc.valor_base),
            "valor_bonus": str(lc.valor_bonus),
            "pacotes_perdidos": str(lc.pacotes_perdidos),
            "adiantamentos": str(lc.adiantamentos),
            "total_a_receber": str(lc.total_a_receber),
            "periodo": f"{lc.periodo_inicio.strftime('%d/%m')} – {lc.periodo_fim.strftime('%d/%m/%Y')}",
            "periodo_inicio_iso": lc.periodo_inicio.isoformat(),
            "periodo_fim_iso": lc.periodo_fim.isoformat(),
            "status": lc.status,
            "status_display": lc.get_status_display(),
            "notas": lc.notas,
        })
        total_lanc_base += lc.valor_base
        total_lanc_bonus += lc.valor_bonus
        total_lanc_total += lc.total_a_receber

    grand_total = total_liquido + total_lanc_total
    iva_rate = empresa.taxa_iva / 100
    total_iva = grand_total * iva_rate
    total_com_iva = grand_total + total_iva

    return JsonResponse({
        "empresa_nome": empresa.nome,
        "taxa_iva": str(empresa.taxa_iva),
        "mes": mes,
        "ano": ano,
        "pre_invoices": rows,
        "lancamentos": lanc_rows,
        "totais": {
            "base_entregas": str(total_base),
            "total_bonus": str(total_bonus),
            "total_liquido": str(total_liquido),
            "total_lancamentos": str(total_lanc_total),
            "grand_total": str(grand_total),
            "total_iva": str(total_iva),
            "total_com_iva": str(total_com_iva),
        },
    })


@login_required
def empresa_auto_emit_config(request, empresa_id):
    """GET ou POST da config FleetAutoEmitConfig de uma empresa.

    GET → devolve config actual (ou defaults se não existe)
    POST → cria/actualiza
    """
    from .models import FleetAutoEmitConfig
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)

    if request.method == "GET":
        cfg = FleetAutoEmitConfig.objects.filter(empresa=empresa).first()
        if not cfg:
            return JsonResponse({"success": True, "config": {
                "enabled": False, "period_type": "monthly",
                "day_of_month": 1, "weekday": 0,
                "auto_send_whatsapp": False,
                "last_emitted_at": None,
                "last_emitted_period_from": None,
                "last_emitted_period_to": None,
                "last_summary": {},
            }})
        return JsonResponse({"success": True, "config": {
            "enabled": cfg.enabled,
            "period_type": cfg.period_type,
            "day_of_month": cfg.day_of_month,
            "weekday": cfg.weekday,
            "auto_send_whatsapp": cfg.auto_send_whatsapp,
            "last_emitted_at": (
                cfg.last_emitted_at.strftime("%Y-%m-%d %H:%M")
                if cfg.last_emitted_at else None
            ),
            "last_emitted_period_from": (
                cfg.last_emitted_period_from.strftime("%Y-%m-%d")
                if cfg.last_emitted_period_from else None
            ),
            "last_emitted_period_to": (
                cfg.last_emitted_period_to.strftime("%Y-%m-%d")
                if cfg.last_emitted_period_to else None
            ),
            "last_summary": cfg.last_summary or {},
        }})

    # POST — guardar
    import json as _json
    try:
        body = _json.loads(request.body)
    except Exception:
        body = request.POST.dict()

    day_of_month = max(1, min(28, int(body.get("day_of_month") or 1)))
    weekday = max(0, min(6, int(body.get("weekday") or 0)))
    cfg, _ = FleetAutoEmitConfig.objects.update_or_create(
        empresa=empresa,
        defaults={
            "enabled": bool(body.get("enabled")),
            "period_type": body.get("period_type") or "monthly",
            "day_of_month": day_of_month,
            "weekday": weekday,
            "auto_send_whatsapp": bool(body.get("auto_send_whatsapp")),
        },
    )
    return JsonResponse({"success": True, "config_id": cfg.id})


@login_required
@require_POST
def empresa_auto_emit_run_now(request, empresa_id):
    """Dispara o auto-emit imediatamente para esta empresa,
    independentemente do dia (mas continua idempotente).
    """
    from datetime import timedelta
    from django.utils import timezone
    from .models import FleetAutoEmitConfig
    import json as _json

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    cfg = FleetAutoEmitConfig.objects.filter(empresa=empresa).first()
    if not cfg or not cfg.enabled:
        return JsonResponse(
            {"success": False, "error":
             "Auto-emit não está activo. Activa primeiro."},
            status=400,
        )

    today = timezone.now().date()
    if cfg.period_type == "monthly":
        first_this = today.replace(day=1)
        period_to = first_this - timedelta(days=1)
        period_from = period_to.replace(day=1)
    else:  # weekly
        this_monday = today - timedelta(days=today.weekday())
        period_from = this_monday - timedelta(days=7)
        period_to = this_monday - timedelta(days=1)

    # Idempotência opcional — permitir override via query param ?force=1
    force = request.GET.get("force") == "1"
    if not force and (
        cfg.last_emitted_period_to
        and cfg.last_emitted_period_to >= period_to
    ):
        return JsonResponse({
            "success": False,
            "error": (
                f"Já emitido para o período {period_from} → {period_to}. "
                f"Adiciona ?force=1 ao URL para forçar."
            ),
        })

    # Disparar via empresa_lote_emit
    from settlements.views import empresa_lote_emit
    body = _json.dumps({
        "from": period_from.strftime("%Y-%m-%d"),
        "to": period_to.strftime("%Y-%m-%d"),
        "skip_overlap": True,
    }).encode("utf-8")
    from django.test import RequestFactory
    rf = RequestFactory()
    req = rf.post(
        f"/settlements/empresas/{empresa.id}/lote-emit/",
        data=body, content_type="application/json",
    )
    req.user = request.user
    try:
        resp = empresa_lote_emit(req, empresa.id)
        data = _json.loads(resp.content)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

    summary = data.get("summary", {})
    cfg.last_emitted_at = timezone.now()
    cfg.last_emitted_period_from = period_from
    cfg.last_emitted_period_to = period_to
    cfg.last_summary = summary
    cfg.save()

    return JsonResponse({
        "success": True,
        "summary": summary,
        "period": f"{period_from} → {period_to}",
        "message": (
            f"Período {period_from} → {period_to}: "
            f"{summary.get('n_created', 0)} PFs criadas, "
            f"€{summary.get('total_amount', 0)}. "
            f"Saltadas: {summary.get('n_skipped', 0)}"
        ),
    })


@login_required
def empresa_lancamento_create(request, empresa_id):
    """Cria um lançamento manual para a empresa."""
    from .models import EmpresaParceiraLancamento
    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)
    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    descricao = (data.get("descricao") or "").strip()
    if not descricao:
        return JsonResponse({"error": "Descrição obrigatória."}, status=400)
    try:
        from decimal import Decimal as D
        qtd_entregas = int(data.get("qtd_entregas") or 0)
        valor_por_entrega = D(str(data.get("valor_por_entrega", "0") or "0"))
        valor_base = D(str(data.get("valor_base", "0") or "0"))
        valor_bonus = D(str(data.get("valor_bonus", "0") or "0"))
        pacotes_perdidos = D(str(data.get("pacotes_perdidos", "0") or "0"))
        adiantamentos = D(str(data.get("adiantamentos", "0") or "0"))
    except Exception:
        return JsonResponse({"error": "Valores inválidos."}, status=400)
    try:
        from datetime import date
        periodo_inicio = date.fromisoformat(data["periodo_inicio"])
        periodo_fim = date.fromisoformat(data["periodo_fim"])
    except Exception:
        return JsonResponse({"error": "Datas inválidas (YYYY-MM-DD)."}, status=400)
    lc = EmpresaParceiraLancamento.objects.create(
        empresa=empresa,
        descricao=descricao,
        qtd_entregas=qtd_entregas,
        valor_por_entrega=valor_por_entrega,
        valor_base=valor_base,
        valor_bonus=valor_bonus,
        pacotes_perdidos=pacotes_perdidos,
        adiantamentos=adiantamentos,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        status=data.get("status", "RASCUNHO"),
        notas=data.get("notas", ""),
        created_by=request.user if request.user.is_authenticated else None,
    )
    return JsonResponse({
        "success": True,
        "lancamento": {
            "id": lc.id,
            "descricao": lc.descricao,
            "qtd_entregas": lc.qtd_entregas,
            "valor_por_entrega": str(lc.valor_por_entrega),
            "valor_base": str(lc.valor_base),
            "valor_bonus": str(lc.valor_bonus),
            "pacotes_perdidos": str(lc.pacotes_perdidos),
            "adiantamentos": str(lc.adiantamentos),
            "total_a_receber": str(lc.total_a_receber),
            "periodo": f"{lc.periodo_inicio.strftime('%d/%m')} – {lc.periodo_fim.strftime('%d/%m/%Y')}",
            "periodo_inicio_iso": lc.periodo_inicio.isoformat(),
            "periodo_fim_iso": lc.periodo_fim.isoformat(),
            "status": lc.status,
            "status_display": lc.get_status_display(),
            "notas": lc.notas,
        },
    })


@login_required
def empresa_lancamento_update(request, lancamento_id):
    """Actualiza um lançamento manual."""
    from .models import EmpresaParceiraLancamento
    lc = get_object_or_404(EmpresaParceiraLancamento, id=lancamento_id)
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)
    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    from decimal import Decimal as D
    from datetime import date
    lc.descricao = (data.get("descricao") or lc.descricao).strip()
    try:
        if "qtd_entregas" in data:
            lc.qtd_entregas = int(data.get("qtd_entregas") or 0)
        if "valor_por_entrega" in data:
            lc.valor_por_entrega = D(str(data.get("valor_por_entrega") or "0"))
        lc.valor_base = D(str(data.get("valor_base", lc.valor_base)))
        lc.valor_bonus = D(str(data.get("valor_bonus", lc.valor_bonus)))
        lc.pacotes_perdidos = D(str(data.get("pacotes_perdidos", lc.pacotes_perdidos)))
        lc.adiantamentos = D(str(data.get("adiantamentos", lc.adiantamentos)))
    except Exception:
        return JsonResponse({"error": "Valores inválidos."}, status=400)
    try:
        if data.get("periodo_inicio"):
            lc.periodo_inicio = date.fromisoformat(data["periodo_inicio"])
        if data.get("periodo_fim"):
            lc.periodo_fim = date.fromisoformat(data["periodo_fim"])
    except Exception:
        return JsonResponse({"error": "Datas inválidas."}, status=400)
    lc.status = data.get("status", lc.status)
    lc.notas = data.get("notas", lc.notas)
    lc.save()
    return JsonResponse({
        "success": True,
        "lancamento": {
            "id": lc.id,
            "descricao": lc.descricao,
            "qtd_entregas": lc.qtd_entregas,
            "valor_por_entrega": str(lc.valor_por_entrega),
            "valor_base": str(lc.valor_base),
            "valor_bonus": str(lc.valor_bonus),
            "pacotes_perdidos": str(lc.pacotes_perdidos),
            "adiantamentos": str(lc.adiantamentos),
            "total_a_receber": str(lc.total_a_receber),
            "periodo": f"{lc.periodo_inicio.strftime('%d/%m')} – {lc.periodo_fim.strftime('%d/%m/%Y')}",
            "periodo_inicio_iso": lc.periodo_inicio.isoformat(),
            "periodo_fim_iso": lc.periodo_fim.isoformat(),
            "status": lc.status,
            "status_display": lc.get_status_display(),
            "notas": lc.notas,
        },
    })


@login_required
def empresa_lancamento_delete(request, lancamento_id):
    """Elimina um lançamento manual."""
    from .models import EmpresaParceiraLancamento
    lc = get_object_or_404(EmpresaParceiraLancamento, id=lancamento_id)
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)
    lc.delete()
    return JsonResponse({"success": True})


@login_required
def empresa_parceira_prefatura_pdf(request, empresa_id):
    """Gera PDF da pré-fatura unificada de uma Empresa Parceira para um período."""
    from decimal import Decimal as D
    from settlements.models import DriverPreInvoice
    from .models import EmpresaParceiraLancamento
    from django.http import HttpResponse

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    try:
        mes = int(request.GET.get("mes", timezone.now().month))
        ano = int(request.GET.get("ano", timezone.now().year))
    except (ValueError, TypeError):
        mes = timezone.now().month
        ano = timezone.now().year

    # Pré-faturas dos motoristas
    driver_ids = list(empresa.motoristas.values_list("id", flat=True))
    pre_invoices_qs = (
        DriverPreInvoice.objects
        .filter(driver_id__in=driver_ids,
                periodo_inicio__month=mes, periodo_inicio__year=ano)
        .select_related("driver")
        .order_by("driver__nome_completo")
    )
    pre_invoices = []
    total_base = D("0")
    total_bonus = D("0")
    total_liquido = D("0")
    for pf in pre_invoices_qs:
        pre_invoices.append({
            "driver_nome": pf.driver.nome_completo,
            "driver_nif": pf.driver.nif,
            "periodo": f"{pf.periodo_inicio.strftime('%d/%m')} – {pf.periodo_fim.strftime('%d/%m/%Y')}",
            "base_entregas": str(pf.base_entregas),
            "total_bonus": str(pf.total_bonus),
            "total_a_receber": str(pf.total_a_receber),
            "status": pf.status,
            "status_display": pf.get_status_display(),
        })
        total_base += pf.base_entregas
        total_bonus += pf.total_bonus
        total_liquido += pf.total_a_receber

    # Lançamentos manuais
    lancamentos_qs = EmpresaParceiraLancamento.objects.filter(
        empresa=empresa,
        periodo_inicio__month=mes, periodo_inicio__year=ano,
    ).order_by("periodo_inicio")
    lancamentos = []
    total_lanc = D("0")
    for lc in lancamentos_qs:
        lancamentos.append({
            "descricao": lc.descricao,
            "qtd_entregas": lc.qtd_entregas,
            "valor_base": str(lc.valor_base),
            "valor_bonus": str(lc.valor_bonus),
            "pacotes_perdidos": str(lc.pacotes_perdidos),
            "adiantamentos": str(lc.adiantamentos),
            "total_a_receber": str(lc.total_a_receber),
            "periodo": f"{lc.periodo_inicio.strftime('%d/%m')} – {lc.periodo_fim.strftime('%d/%m/%Y')}",
            "status": lc.status,
            "status_display": lc.get_status_display(),
        })
        total_lanc += lc.total_a_receber

    grand_total = total_liquido + total_lanc
    iva_rate = empresa.taxa_iva / 100
    total_iva = grand_total * iva_rate
    total_com_iva = grand_total + total_iva

    totais = {
        "total_liquido": str(total_liquido),
        "total_lancamentos": str(total_lanc),
        "grand_total": str(grand_total),
        "total_iva": str(total_iva),
        "total_com_iva": str(total_com_iva),
    }

    try:
        from settlements.reports.pdf_generator import PDFGenerator
        gen = PDFGenerator()
        pdf_buf = gen.generate_empresa_parceira_pdf(
            empresa, mes, ano, pre_invoices, lancamentos, totais
        )
        nome_safe = empresa.nome.replace(" ", "_")[:40]
        month_names = ["Jan","Fev","Mar","Abr","Mai","Jun",
                       "Jul","Ago","Set","Out","Nov","Dez"]
        filename = f"PreFatura_{nome_safe}_{month_names[mes-1]}{ano}.pdf"
        response = HttpResponse(pdf_buf.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        return JsonResponse({"error": "ReportLab não está instalado."}, status=500)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def _empresa_to_dict(e):
    return {
        "id": e.id,
        "nome": e.nome,
        "nif": e.nif,
        "morada": e.morada,
        "codigo_postal": e.codigo_postal,
        "cidade": e.cidade,
        "email": e.email,
        "telefone": e.telefone,
        "contacto_nome": e.contacto_nome,
        "iban": e.iban,
        "taxa_iva": str(e.taxa_iva),
        "driver_default_price_per_package": (
            str(e.driver_default_price_per_package)
            if e.driver_default_price_per_package else ""
        ),
        "ativo": e.ativo,
        "notas": e.notas,
        "num_motoristas": e.num_motoristas,
    }


# ============================================================================
# RECLAMAÇÕES DE CLIENTES — API
# ============================================================================


def _complaint_to_dict(c):
    return {
        "id": c.id,
        "numero_pacote": c.numero_pacote,
        "tipo": c.tipo,
        "tipo_display": c.get_tipo_display(),
        "descricao": c.descricao,
        "nome_cliente": c.nome_cliente,
        "telefone_cliente": c.telefone_cliente,
        "email_cliente": c.email_cliente,
        "morada": c.morada,
        "codigo_postal": c.codigo_postal,
        "cidade": c.cidade,
        "status": c.status,
        "status_display": c.get_status_display(),
        "resposta_driver": c.resposta_driver,
        "notas": c.notas,
        "data_entrega": c.data_entrega.strftime("%Y-%m-%dT%H:%M") if c.data_entrega else "",
        "data_entrega_display": c.data_entrega.strftime("%d/%m/%Y %H:%M") if c.data_entrega else "",
        "deadline": c.deadline.strftime("%Y-%m-%dT%H:%M") if c.deadline else "",
        "deadline_display": c.deadline.strftime("%d/%m/%Y %H:%M") if c.deadline else "",
        "data_notificacao": c.data_notificacao.strftime("%d/%m/%Y %H:%M") if c.data_notificacao else "",
        "data_resposta": c.data_resposta.strftime("%d/%m/%Y %H:%M") if c.data_resposta else "",
        "data_fecho": c.data_fecho.strftime("%d/%m/%Y %H:%M") if c.data_fecho else "",
        "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
        "driver_id": c.driver.id if c.driver else None,
        "driver_nome": c.driver.nome_completo if c.driver else "",
        "whatsapp_text": c.whatsapp_text(),
        "attachments": [
            {
                "id": a.id,
                "tipo": a.tipo,
                "tipo_display": a.get_tipo_display(),
                "url": a.ficheiro.url,
                "descricao": a.descricao,
                "filename": a.ficheiro.name.split("/")[-1],
            }
            for a in c.attachments.all().order_by("created_at")
        ],
    }


@login_required
def admin_complaints_dashboard(request):
    """Painel global de reclamações de clientes."""
    from .models import CustomerComplaint
    from django.db.models import Count, Q
    from django.utils import timezone

    now = timezone.now()

    # KPI counts
    qs_all = CustomerComplaint.objects.all()
    kpis = {
        "abertas":      qs_all.filter(status="ABERTO").count(),
        "notificadas":  qs_all.filter(status="NOTIFICADO").count(),
        "respondidas":  qs_all.filter(status="RESPONDIDO").count(),
        "fechadas":     qs_all.filter(status="FECHADO").count(),
        "sem_resposta": qs_all.filter(
            deadline__lt=now,
            status__in=("ABERTO", "NOTIFICADO"),
        ).count(),
        "total":        qs_all.count(),
    }
    return render(request, "drivers_app/admin_complaints.html", {"kpis": kpis})


@login_required
def admin_complaints_list_api(request):
    """API JSON para o painel: lista filtrada + stats agregadas."""
    from .models import CustomerComplaint
    from django.db.models import Count, Q
    from django.utils import timezone

    now = timezone.now()
    qs = CustomerComplaint.objects.select_related("driver").prefetch_related("attachments")

    # ── Filtros ────────────────────────────────────────────────────────────
    status_f = request.GET.get("status", "")
    driver_f = request.GET.get("driver", "")
    cp4_f    = request.GET.get("cp4", "")
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    overdue_f = request.GET.get("overdue", "")

    if status_f:
        qs = qs.filter(status=status_f)
    if driver_f:
        qs = qs.filter(driver_id=driver_f)
    if cp4_f:
        qs = qs.filter(codigo_postal__startswith=cp4_f)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if overdue_f == "1":
        qs = qs.filter(deadline__lt=now, status__in=("ABERTO", "NOTIFICADO"))

    # ── Stats agregadas ────────────────────────────────────────────────────
    # Por dia (últimos 30)
    from django.db.models.functions import TruncDate
    by_day = (
        qs.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total=Count("id"))
        .order_by("-day")[:30]
    )
    # Por motorista (top 20)
    by_driver = (
        qs.values("driver__id", "driver__nome_completo")
        .annotate(total=Count("id"))
        .order_by("-total")[:20]
    )
    # Por CP4
    by_cp4 = (
        qs.values("codigo_postal")
        .annotate(total=Count("id"))
        .order_by("-total")[:20]
    )

    # ── Lista paginada (50 por página) ─────────────────────────────────────
    from django.core.paginator import Paginator
    page_num = int(request.GET.get("page", 1))
    paginator = Paginator(qs.order_by("-created_at"), 50)
    page_obj = paginator.get_page(page_num)

    return JsonResponse({
        "complaints": [_complaint_to_dict(c) for c in page_obj],
        "pagination": {
            "page":      page_obj.number,
            "num_pages": paginator.num_pages,
            "count":     paginator.count,
            "has_next":  page_obj.has_next(),
            "has_prev":  page_obj.has_previous(),
        },
        "stats": {
            "by_day":    [{"day": str(r["day"]), "total": r["total"]} for r in by_day],
            "by_driver": [{"id": r["driver__id"], "nome": r["driver__nome_completo"], "total": r["total"]} for r in by_driver],
            "by_cp4":    [{"cp4": (r["codigo_postal"] or "")[:4], "total": r["total"]} for r in by_cp4],
        },
    })


@login_required
def admin_complaints_driver_search(request):
    """Pesquisa de motoristas para o modal de nova reclamação."""
    q = (request.GET.get("q") or "").strip()
    qs = DriverProfile.objects.filter(status="ATIVO")
    if q:
        qs = qs.filter(
            Q(nome_completo__icontains=q) | Q(nif__icontains=q)
        )
    drivers = [
        {"id": d.id, "nome": d.nome_completo, "nif": d.nif or ""}
        for d in qs.order_by("nome_completo")[:20]
    ]
    return JsonResponse({"drivers": drivers})


@login_required
def driver_complaints_api(request, driver_id):
    """Retorna lista de reclamações de um motorista em JSON."""
    driver = get_object_or_404(DriverProfile, id=driver_id)
    complaints = (
        driver.customer_complaints
        .prefetch_related("attachments")
        .order_by("-created_at")
    )
    return JsonResponse({
        "complaints": [_complaint_to_dict(c) for c in complaints],
        "tipo_choices": [
            {"value": v, "label": l}
            for v, l in __import__("drivers_app.models", fromlist=["CustomerComplaint"]).CustomerComplaint.TIPO_CHOICES
        ],
        "status_choices": [
            {"value": v, "label": l}
            for v, l in __import__("drivers_app.models", fromlist=["CustomerComplaint"]).CustomerComplaint.STATUS_CHOICES
        ],
    })


@login_required
@require_http_methods(["POST"])
def driver_complaint_create(request, driver_id):
    """Cria uma nova reclamação de cliente para um motorista."""
    from .models import CustomerComplaint
    driver = get_object_or_404(DriverProfile, id=driver_id)

    if request.content_type and "multipart" in request.content_type:
        data = request.POST.dict()
    else:
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            data = request.POST.dict()

    required = ["numero_pacote", "tipo", "descricao", "nome_cliente",
                "telefone_cliente", "morada", "codigo_postal", "cidade"]
    for field in required:
        if not (data.get(field) or "").strip():
            return JsonResponse({"success": False, "error": f"Campo obrigatório: {field}"}, status=400)

    # Parse datetime fields
    from django.utils.dateparse import parse_datetime

    def _parse_dt(raw):
        raw = (raw or "").strip()
        return parse_datetime(raw) if raw else None

    deadline_val    = _parse_dt(data.get("deadline"))
    data_entrega_val = _parse_dt(data.get("data_entrega"))

    complaint = CustomerComplaint.objects.create(
        driver=driver,
        numero_pacote=data["numero_pacote"].strip(),
        tipo=data.get("tipo", "ENTREGA_FALSA"),
        descricao=data["descricao"].strip(),
        nome_cliente=data["nome_cliente"].strip(),
        telefone_cliente=data["telefone_cliente"].strip(),
        email_cliente=(data.get("email_cliente") or "").strip(),
        morada=data["morada"].strip(),
        codigo_postal=data["codigo_postal"].strip(),
        cidade=data["cidade"].strip(),
        notas=(data.get("notas") or "").strip(),
        deadline=deadline_val,
        data_entrega=data_entrega_val,
        created_by=request.user,
    )

    # Optional inline attachments (multi-file, one per list entry)
    attachment_files = request.FILES.getlist("attachment_files")
    if attachment_files:
        from .models import CustomerComplaintAttachment
        for i, f in enumerate(attachment_files):
            CustomerComplaintAttachment.objects.create(
                complaint=complaint,
                tipo=request.POST.get(f"attachment_tipo_{i}", "RECLAMACAO"),
                ficheiro=f,
                descricao=(request.POST.get(f"attachment_descricao_{i}") or "").strip(),
            )

    return JsonResponse({"success": True, "complaint": _complaint_to_dict(complaint)})


@login_required
@require_http_methods(["POST"])
def driver_complaint_update(request, complaint_id):
    """Actualiza status, resposta do motorista ou notas de uma reclamação."""
    from .models import CustomerComplaint
    complaint = get_object_or_404(CustomerComplaint, id=complaint_id)

    if request.content_type and "multipart" in request.content_type:
        data = request.POST.dict()
    else:
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            data = request.POST.dict()

    # Status transitions
    novo_status = (data.get("status") or "").strip().upper()
    if novo_status and novo_status != complaint.status:
        TRANSICOES = {
            "ABERTO":     ["NOTIFICADO", "CANCELADO"],
            "NOTIFICADO": ["RESPONDIDO", "CANCELADO", "ABERTO"],
            "RESPONDIDO": ["FECHADO", "NOTIFICADO"],
            "FECHADO":    [],
            "CANCELADO":  ["ABERTO"],
        }
        if novo_status not in TRANSICOES.get(complaint.status, []):
            return JsonResponse(
                {"success": False, "error": f"Transição {complaint.status} → {novo_status} não permitida."},
                status=400,
            )
        complaint.status = novo_status
        now = timezone.now()
        if novo_status == "NOTIFICADO" and not complaint.data_notificacao:
            complaint.data_notificacao = now
        elif novo_status == "RESPONDIDO" and not complaint.data_resposta:
            complaint.data_resposta = now
        elif novo_status == "FECHADO" and not complaint.data_fecho:
            complaint.data_fecho = now

    # Editable fields
    for field in ["resposta_driver", "notas", "descricao", "nome_cliente",
                  "telefone_cliente", "email_cliente", "morada", "codigo_postal",
                  "cidade", "numero_pacote", "tipo"]:
        if field in data:
            setattr(complaint, field, data[field])

    # Datetime fields update
    from django.utils.dateparse import parse_datetime as _pdt
    if "deadline" in data:
        raw = (data.get("deadline") or "").strip()
        complaint.deadline = _pdt(raw) if raw else None
    if "data_entrega" in data:
        raw = (data.get("data_entrega") or "").strip()
        complaint.data_entrega = _pdt(raw) if raw else None

    complaint.save()

    # Multiple files (attachment_files[]) — used by saveResposta()
    from .models import CustomerComplaintAttachment
    attachment_files = request.FILES.getlist("attachment_files")
    if attachment_files:
        for i, f in enumerate(attachment_files):
            CustomerComplaintAttachment.objects.create(
                complaint=complaint,
                tipo=request.POST.get(f"attachment_tipo_{i}", "RESPOSTA_DRIVER"),
                ficheiro=f,
                descricao=(request.POST.get(f"attachment_descricao_{i}") or "").strip(),
            )
    else:
        # Legacy single file support
        attachment_file = request.FILES.get("attachment_file")
        if attachment_file:
            CustomerComplaintAttachment.objects.create(
                complaint=complaint,
                tipo=data.get("attachment_tipo", "RESPOSTA_DRIVER"),
                ficheiro=attachment_file,
                descricao=(data.get("attachment_descricao") or "").strip(),
            )

    return JsonResponse({"success": True, "complaint": _complaint_to_dict(complaint)})


@login_required
@require_http_methods(["POST"])
def driver_complaint_add_attachment(request, complaint_id):
    """Adiciona um ficheiro a uma reclamação."""
    from .models import CustomerComplaint, CustomerComplaintAttachment
    complaint = get_object_or_404(CustomerComplaint, id=complaint_id)

    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    attachment = CustomerComplaintAttachment.objects.create(
        complaint=complaint,
        tipo=request.POST.get("tipo", "RECLAMACAO"),
        ficheiro=ficheiro,
        descricao=(request.POST.get("descricao") or "").strip(),
    )
    return JsonResponse({
        "success": True,
        "attachment": {
            "id": attachment.id,
            "tipo": attachment.tipo,
            "tipo_display": attachment.get_tipo_display(),
            "url": attachment.ficheiro.url,
            "descricao": attachment.descricao,
            "filename": attachment.ficheiro.name.split("/")[-1],
        },
    })


@login_required
@require_http_methods(["POST"])
def driver_complaint_delete_attachment(request, attachment_id):
    """Remove um ficheiro de uma reclamação."""
    from .models import CustomerComplaintAttachment
    att = get_object_or_404(CustomerComplaintAttachment, id=attachment_id)
    att.ficheiro.delete(save=False)
    att.delete()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def driver_complaint_delete(request, complaint_id):
    """Elimina uma reclamação (apenas ABERTO ou CANCELADO)."""
    from .models import CustomerComplaint
    complaint = get_object_or_404(CustomerComplaint, id=complaint_id)
    if complaint.status not in ("ABERTO", "CANCELADO"):
        return JsonResponse(
            {"success": False, "error": "Só é possível eliminar reclamações em estado Aberto ou Cancelado."},
            status=400,
        )
    complaint.delete()
    return JsonResponse({"success": True})


@login_required
def driver_complaint_pdf(request, complaint_id):
    """Gera PDF de uma reclamacao de cliente com logo, layout profissional e imagens dos anexos."""
    from .models import CustomerComplaint
    complaint = get_object_or_404(
        CustomerComplaint.objects.select_related("driver").prefetch_related("attachments"),
        id=complaint_id,
    )
    try:
        import os
        from io import BytesIO

        from django.conf import settings
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Image, KeepTogether, Paragraph,
            SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        # ── Palette ─────────────────────────────────────────────────────────
        INDIGO   = colors.HexColor("#4F46E5")
        INDIGO_D = colors.HexColor("#3730A3")
        SLATE    = colors.HexColor("#1E293B")
        GRAY     = colors.HexColor("#6B7280")
        LIGHT    = colors.HexColor("#F8FAFC")
        BORDER   = colors.HexColor("#E2E8F0")
        GREEN    = colors.HexColor("#059669")
        GREEN_L  = colors.HexColor("#F0FDF4")
        GREEN_B  = colors.HexColor("#10B981")
        AMBER    = colors.HexColor("#F59E0B")
        WHITE    = colors.white

        STATUS_COLORS = {
            "ABERTO":     colors.HexColor("#3B82F6"),
            "NOTIFICADO": AMBER,
            "RESPONDIDO": colors.HexColor("#8B5CF6"),
            "FECHADO":    GREEN,
            "CANCELADO":  colors.HexColor("#EF4444"),
        }

        # ── Page layout ──────────────────────────────────────────────────────
        page_w, page_h = A4
        margin = 1.8 * cm
        content_w = page_w - 2 * margin

        buffer = BytesIO()
        styles = getSampleStyleSheet()

        # ── Style helpers ────────────────────────────────────────────────────
        def ps(name, **kw):
            base = kw.pop("parent", styles["Normal"])
            return ParagraphStyle(name, parent=base, **kw)

        normal9 = ps("N9",  fontSize=9,  leading=13, textColor=SLATE)
        normal8 = ps("N8",  fontSize=8,  leading=12, textColor=GRAY)
        section = ps("SEC", fontSize=10, leading=14, textColor=INDIGO,
                     fontName="Helvetica-Bold")
        centered = ps("C",  fontSize=7,  leading=10, textColor=GRAY,
                      alignment=TA_CENTER)
        right7  = ps("R7",  fontSize=8,  leading=12, textColor=WHITE,
                     alignment=TA_RIGHT)
        white9  = ps("W9",  fontSize=9,  leading=13, textColor=WHITE)

        def section_header(label):
            t = Table([[Paragraph(label, section)]], colWidths=[content_w])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#EEF2FF")),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
                ("TOPPADDING",    (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBEFORE",    (0, 0), (0, -1), 3, INDIGO),
                ("BOX",           (0, 0), (-1, -1), 0.3, BORDER),
            ]))
            return t

        def kv_table(rows, col1=4.5 * cm):
            data = [[Paragraph("<b>" + k + "</b>", normal8),
                     Paragraph(str(v) if v else "---", normal9)]
                    for k, v in rows]
            t = Table(data, colWidths=[col1, content_w - col1])
            t.setStyle(TableStyle([
                ("BACKGROUND",     (0, 0), (0, -1), LIGHT),
                ("GRID",           (0, 0), (-1, -1), 0.3, BORDER),
                ("VALIGN",         (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",    (0, 0), (-1, -1), 8),
                ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
                ("TOPPADDING",     (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT, WHITE]),
            ]))
            return t

        def mini_kv(rows, col_w, col1=3.5 * cm):
            data = [[Paragraph("<b>" + k + "</b>", normal8),
                     Paragraph(str(v) if v else "---", normal9)]
                    for k, v in rows]
            t = Table(data, colWidths=[col1, col_w - col1])
            t.setStyle(TableStyle([
                ("BACKGROUND",     (0, 0), (0, -1), LIGHT),
                ("GRID",           (0, 0), (-1, -1), 0.3, BORDER),
                ("VALIGN",         (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING",    (0, 0), (-1, -1), 7),
                ("RIGHTPADDING",   (0, 0), (-1, -1), 7),
                ("TOPPADDING",     (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT, WHITE]),
            ]))
            return t

        def text_box(text, bg, border_color):
            t = Table([[Paragraph(text.replace("\n", "<br/>"), normal9)]],
                      colWidths=[content_w])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), bg),
                ("BOX",           (0, 0), (-1, -1), 0.5, border_color),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
                ("TOPPADDING",    (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]))
            return t

        elements = []

        # ── HEADER: logo (fundo branco, proporcional) + dados ref ────────────
        logo_path = os.path.join(settings.STATIC_ROOT, "img", "logo.png")
        try:
            if os.path.exists(logo_path):
                logo_el = Image(logo_path, width=3.5 * cm, height=3.5 * cm,
                                kind="proportional")
            else:
                raise FileNotFoundError
        except Exception:
            logo_el = Paragraph("<b>LÉGUAS FRANZINAS</b>",
                                ps("LF", fontSize=11, fontName="Helvetica-Bold",
                                   textColor=INDIGO))

        # Logo numa célula centrada (fundo branco — sem conflito de cor)
        logo_cell = Table([[logo_el]], colWidths=[4 * cm])
        logo_cell.setStyle(TableStyle([
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        # Dados de referência à direita
        ref_style = ps("REF", fontSize=8, leading=13, textColor=GRAY,
                       alignment=TA_RIGHT)
        ref_bold  = ps("RFB", fontSize=9, leading=14, textColor=SLATE,
                       fontName="Helvetica-Bold", alignment=TA_RIGHT)
        ref_col = [
            Paragraph("RECLAMAÇÃO DE CLIENTE", ref_bold),
            Paragraph(
                "Ref #" + str(complaint.id)
                + "  |  Criado em "
                + complaint.created_at.strftime("%d/%m/%Y %H:%M"),
                ref_style,
            ),
            Paragraph(
                "Gerado em " + datetime.now().strftime("%d/%m/%Y %H:%M"),
                ref_style,
            ),
        ]
        ref_cell = Table(
            [[p] for p in ref_col],
            colWidths=[content_w - 4 * cm - 0.3 * cm],
        )
        ref_cell.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ]))

        header_outer = Table(
            [[logo_cell, ref_cell]],
            colWidths=[4 * cm, content_w - 4 * cm],
        )
        header_outer.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("LINEAFTER",     (0, 0), (0, -1), 0.5, BORDER),
        ]))
        elements.append(header_outer)
        elements.append(Spacer(1, 0.25 * cm))

        # ── TÍTULO COLORIDO (separado do logo, fundo indigo) ─────────────────
        title_t = Table(
            [[Paragraph(
                "<font color='white'><b>RECLAMAÇÃO DE CLIENTE</b></font>",
                ps("TT", fontSize=13, leading=16, fontName="Helvetica-Bold",
                   textColor=WHITE, alignment=TA_CENTER),
            )]],
            colWidths=[content_w],
        )
        title_t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), INDIGO),
            ("TOPPADDING",    (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
            ("LINEBELOW",     (0, 0), (-1, -1), 2, INDIGO_D),
        ]))
        elements.append(title_t)
        elements.append(Spacer(1, 0.2 * cm))

        # ── STATUS BAR ───────────────────────────────────────────────────────
        sc = STATUS_COLORS.get(complaint.status, GRAY)
        deadline_str = ""
        if complaint.deadline:
            deadline_str = (
                "   |   Deadline: "
                + complaint.deadline.strftime("%d/%m/%Y %H:%M")
            )
        status_t = Table([[
            Paragraph(
                "<font color='white'><b>STATUS: "
                + complaint.get_status_display().upper() + "</b></font>",
                ps("SL", fontSize=9, leading=13, fontName="Helvetica-Bold",
                   textColor=WHITE),
            ),
            Paragraph(
                "<font color='white'>"
                + complaint.get_tipo_display() + deadline_str + "</font>",
                ps("SR", fontSize=9, leading=13, textColor=WHITE,
                   alignment=TA_RIGHT),
            ),
        ]], colWidths=[content_w * 0.4, content_w * 0.6])
        status_t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), sc),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(status_t)
        elements.append(Spacer(1, 0.45 * cm))

        # ── DADOS DO PACOTE ──────────────────────────────────────────────────
        elements.append(KeepTogether([
            section_header("Dados do Pacote"),
            Spacer(1, 0.15 * cm),
            kv_table([
                ("No Tracking",    complaint.numero_pacote),
                ("Tipo",           complaint.get_tipo_display()),
                ("Motorista",      complaint.driver.nome_completo),
                ("Data Entrega",   complaint.data_entrega.strftime("%d/%m/%Y %H:%M")
                                   if complaint.data_entrega else "---"),
                ("Data Registo",   complaint.created_at.strftime("%d/%m/%Y %H:%M")),
            ]),
        ]))
        elements.append(Spacer(1, 0.45 * cm))

        # ── RELATO ───────────────────────────────────────────────────────────
        elements.append(KeepTogether([
            section_header("Relato do Cliente"),
            Spacer(1, 0.15 * cm),
            text_box(complaint.descricao, LIGHT, BORDER),
        ]))
        elements.append(Spacer(1, 0.45 * cm))

        # ── CLIENTE + ENDERECO side-by-side ──────────────────────────────────
        col = (content_w - 0.4 * cm) / 2
        side_t = Table([[
            [section_header("Dados do Cliente"),
             Spacer(1, 0.1 * cm),
             mini_kv([
                 ("Nome",     complaint.nome_cliente),
                 ("Telefone", complaint.telefone_cliente),
                 ("Email",    complaint.email_cliente or "---"),
             ], col)],
            [section_header("Endereco de Entrega"),
             Spacer(1, 0.1 * cm),
             mini_kv([
                 ("Morada",        complaint.morada),
                 ("Codigo Postal", complaint.codigo_postal),
                 ("Cidade",        complaint.cidade),
             ], col)],
        ]], colWidths=[col, col])
        side_t.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (0, -1), 6),
            ("LEFTPADDING",   (1, 0), (-1, -1), 6),
        ]))
        elements.append(KeepTogether([side_t]))
        elements.append(Spacer(1, 0.45 * cm))

        # ── MENSAGEM WHATSAPP ────────────────────────────────────────────────
        elements.append(KeepTogether([
            section_header("Mensagem WhatsApp para o Motorista"),
            Spacer(1, 0.15 * cm),
            text_box(complaint.whatsapp_text(), GREEN_L, GREEN_B),
        ]))

        # ── RESPOSTA DO MOTORISTA ────────────────────────────────────────────
        if complaint.resposta_driver:
            elements.append(Spacer(1, 0.45 * cm))
            resp_label = "Resposta do Motorista"
            if complaint.data_resposta:
                resp_label += (
                    "  -  "
                    + complaint.data_resposta.strftime("%d/%m/%Y %H:%M")
                )
            elements.append(KeepTogether([
                section_header(resp_label),
                Spacer(1, 0.15 * cm),
                text_box(complaint.resposta_driver, GREEN_L, GREEN_B),
            ]))

        # ── ANEXOS (imagens em grelha 2x, PDFs como texto) ───────────────────
        attachments = list(complaint.attachments.all().order_by("created_at"))
        if attachments:
            elements.append(Spacer(1, 0.45 * cm))
            elements.append(
                section_header("Anexos (" + str(len(attachments)) + ")")
            )
            elements.append(Spacer(1, 0.2 * cm))

            IMG_MAX_W = (content_w - 0.4 * cm) / 2
            IMG_MAX_H = 7 * cm
            IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}

            img_cells = []
            non_img = []

            for att in attachments:
                try:
                    file_path = att.ficheiro.path
                    ext = os.path.splitext(file_path)[1].lower()
                    label = att.descricao or att.ficheiro.name.split("/")[-1]
                    caption = att.get_tipo_display() + "  |  " + label

                    if ext in IMAGE_EXTS and os.path.exists(file_path):
                        from PIL import Image as PILImage
                        with PILImage.open(file_path) as pil_img:
                            orig_w, orig_h = pil_img.size
                        ratio = min(
                            IMG_MAX_W / orig_w,
                            IMG_MAX_H / orig_h,
                            1.0,
                        )
                        rl_img = Image(
                            file_path,
                            width=orig_w * ratio,
                            height=orig_h * ratio,
                        )
                        rl_img.hAlign = "CENTER"
                        img_cells.append((rl_img, caption))
                    else:
                        non_img.append(caption)
                except Exception:
                    non_img.append(att.ficheiro.name.split("/")[-1])

            # 2-column image grid
            if img_cells:
                grid_rows = []
                for i in range(0, len(img_cells), 2):
                    row_imgs, row_caps = [], []
                    for j in range(2):
                        if i + j < len(img_cells):
                            rl_img, cap = img_cells[i + j]
                            row_imgs.append(rl_img)
                            row_caps.append(Paragraph(cap, centered))
                        else:
                            row_imgs.append("")
                            row_caps.append("")
                    grid_rows.append(row_imgs)
                    grid_rows.append(row_caps)

                grid_t = Table(
                    grid_rows,
                    colWidths=[IMG_MAX_W, IMG_MAX_W],
                    hAlign="CENTER",
                )
                grid_t.setStyle(TableStyle([
                    ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING",    (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
                    ("BOX",           (0, 0), (-1, -1), 0.3, BORDER),
                    ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
                    ("BACKGROUND",    (0, 0), (-1, -1), LIGHT),
                ]))
                elements.append(grid_t)

            # Non-image attachments listed as text
            if non_img:
                elements.append(Spacer(1, 0.15 * cm))
                for cap in non_img:
                    elements.append(Paragraph("- " + cap, normal9))

        # ── NOTAS INTERNAS ───────────────────────────────────────────────────
        if complaint.notas:
            elements.append(Spacer(1, 0.45 * cm))
            elements.append(KeepTogether([
                section_header("Notas Internas"),
                Spacer(1, 0.15 * cm),
                text_box(
                    complaint.notas,
                    colors.HexColor("#FFFBEB"),
                    AMBER,
                ),
            ]))

        # ── FOOTER ───────────────────────────────────────────────────────────
        elements.append(Spacer(1, 0.7 * cm))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
        elements.append(Spacer(1, 0.15 * cm))
        elements.append(Paragraph(
            "Leguas Franzinas - Unipessoal, Lda  |  "
            "Reclamacao #" + str(complaint.id) + "  |  "
            "Motorista: " + complaint.driver.nome_completo + "  |  "
            "Gerado em " + datetime.now().strftime("%d/%m/%Y %H:%M"),
            centered,
        ))

        # ── BUILD ────────────────────────────────────────────────────────────
        def on_page(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(GRAY)
            canvas.drawRightString(
                page_w - margin, 0.8 * cm,
                "Pag. " + str(doc.page),
            )
            canvas.restoreState()

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=margin, leftMargin=margin,
            topMargin=margin, bottomMargin=1.5 * cm,
        )
        doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
        buffer.seek(0)

        pkg_safe = complaint.numero_pacote.replace("/", "_")[:40]
        driver_safe = complaint.driver.nome_completo.replace(" ", "_")[:30]
        filename = (
            "Reclamacao_" + str(complaint.id)
            + "_" + pkg_safe + "_" + driver_safe + ".pdf"
        )
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="' + filename + '"'
        return response

    except ImportError as e:
        return JsonResponse(
            {"error": "Dependencia nao instalada: " + str(e)}, status=500
        )
    except Exception as e:
        import traceback
        return JsonResponse(
            {"error": str(e), "trace": traceback.format_exc()}, status=500
        )


# ============================================================================
# Bulk actions na página de Motoristas Ativos
# ============================================================================

@login_required
@require_POST
def bulk_block_drivers(request):
    """Bloqueia vários drivers de uma só vez. Body: ids=[int]"""
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    ids = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return JsonResponse(
            {"success": False, "error": "ids vazios"}, status=400,
        )
    n = DriverProfile.objects.filter(id__in=ids).update(
        status="BLOQUEADO", is_active=False,
    )
    return JsonResponse({"success": True, "updated": n})


@login_required
@require_POST
def bulk_unblock_drivers(request):
    """Desbloqueia (volta a ATIVO) vários drivers."""
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    ids = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return JsonResponse(
            {"success": False, "error": "ids vazios"}, status=400,
        )
    n = DriverProfile.objects.filter(id__in=ids).update(
        status="ATIVO", is_active=True,
    )
    return JsonResponse({"success": True, "updated": n})


@login_required
@require_POST
def bulk_whatsapp_drivers(request):
    """Envia mensagem WhatsApp a vários drivers via WPPConnect.

    Body: { ids: [int], message: str }
    Devolve sumário { sent, failed, no_phone }.
    """
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    ids = body.get("ids") or []
    message = (body.get("message") or "").strip()
    if not isinstance(ids, list) or not ids:
        return JsonResponse(
            {"success": False, "error": "ids vazios"}, status=400,
        )
    if not message:
        return JsonResponse(
            {"success": False, "error": "Mensagem vazia"}, status=400,
        )

    import requests
    from django.conf import settings as dj_settings
    api_url = (
        getattr(dj_settings, "WHATSAPP_API_URL", "")
        or "http://45.160.176.150:9090/message/sendText/leguasreports"
    )

    def _normalize_phone(p):
        digits = "".join(c for c in (p or "") if c.isdigit())
        # Adiciona indicativo PT se vier sem
        if digits and not digits.startswith("351") and len(digits) == 9:
            digits = "351" + digits
        return digits

    sent = 0
    failed = 0
    no_phone = 0
    drivers = DriverProfile.objects.filter(id__in=ids)
    for d in drivers:
        wa_num = _normalize_phone(d.telefone)
        if not wa_num:
            no_phone += 1
            continue
        try:
            r = requests.post(
                api_url,
                json={"number": wa_num, "text": message},
                timeout=10,
            )
            if r.status_code in (200, 201):
                sent += 1
            else:
                failed += 1
        except Exception:
            failed += 1
    return JsonResponse({
        "success": True,
        "sent": sent, "failed": failed, "no_phone": no_phone,
    })


@login_required
def bulk_export_drivers_csv(request):
    """Exporta drivers selecionados (ou filtrados) para CSV.

    Aceita ids=1,2,3 na querystring; se vazio, exporta a busca atual.
    """
    import csv
    ids_param = request.GET.get("ids", "")
    if ids_param:
        try:
            ids = [int(x) for x in ids_param.split(",") if x.strip()]
        except (ValueError, TypeError):
            ids = []
        qs = DriverProfile.objects.filter(id__in=ids)
    else:
        qs = DriverProfile.objects.exclude(
            status__in=["PENDENTE", "EM_ANALISE"],
        )
    qs = qs.select_related("empresa_parceira").order_by(
        "apelido", "nome_completo",
    )

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    fname = (
        f"motoristas_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{fname}"'
    )
    response.write("﻿")  # BOM para Excel abrir UTF-8
    writer = csv.writer(response, delimiter=";")
    writer.writerow([
        "NIF", "Apelido", "Nome Completo", "Email", "Telefone",
        "Tipo", "Status", "Frota", "Courier ID Cainiao",
        "Aprovado em",
    ])
    for d in qs:
        writer.writerow([
            d.nif or "",
            d.apelido or "",
            d.nome_completo or "",
            d.email or "",
            d.telefone or "",
            d.get_tipo_vinculo_display() if d.tipo_vinculo else "",
            d.get_status_display() if d.status else "",
            d.empresa_parceira.nome if d.empresa_parceira else "",
            d.courier_id_cainiao or "",
            d.approved_at.strftime("%Y-%m-%d") if d.approved_at else "",
        ])
    return response


@login_required
def driver_quickview(request, driver_id):
    """API para o drawer lateral. Retorna últimas 10 operações + meta."""
    driver = get_object_or_404(DriverProfile, id=driver_id)
    from settlements.models import CainiaoOperationTask

    last_ops = []
    if driver.courier_id_cainiao:
        qs = (
            CainiaoOperationTask.objects
            .filter(courier_id_cainiao=driver.courier_id_cainiao)
            .order_by("-task_date", "-id")[:10]
        )
        for op in qs:
            last_ops.append({
                "task_date": (
                    op.task_date.strftime("%Y-%m-%d")
                    if op.task_date else ""
                ),
                "waybill": op.waybill_number or "",
                "status": op.task_status or "",
                "city": op.destination_city or "",
                "zip": op.zip_code or "",
            })

    # Reclamações abertas
    from .models import CustomerComplaint
    open_status = ["ABERTO", "NOTIFICADO", "RESPONDIDO"]
    open_complaints = list(
        CustomerComplaint.objects.filter(
            driver=driver, status__in=open_status,
        ).values(
            "id", "numero_pacote", "tipo", "status", "created_at",
        )[:10]
    )
    for c in open_complaints:
        c["created_at"] = (
            c["created_at"].strftime("%Y-%m-%d")
            if c["created_at"] else ""
        )

    fleet_name = (
        driver.empresa_parceira.nome
        if driver.empresa_parceira else ""
    )
    phone = (driver.telefone or "").strip()
    # Normaliza para formato WhatsApp wa.me (só dígitos)
    wa_phone = "".join(c for c in phone if c.isdigit())

    return JsonResponse({
        "id": driver.id,
        "name": driver.apelido or driver.nome_completo,
        "full_name": driver.nome_completo or "",
        "nif": driver.nif or "",
        "email": driver.email or "",
        "phone": phone,
        "wa_phone": wa_phone,
        "fleet": fleet_name,
        "tipo_vinculo": driver.tipo_vinculo or "",
        "status": driver.status or "",
        "daily_capacity": driver.daily_capacity,
        "last_ops": last_ops,
        "open_complaints": open_complaints,
    })




# ─── Portal do Driver (Fase 1) ───
from .portal_views import (
    driver_portal,
    driver_portal_reports,
    driver_portal_invoices,
    driver_portal_profile,
    driver_pre_invoice_pdf,
    driver_pre_invoice_upload_recibo,
)


# ─── Admin: pedidos de alteração de perfil (Fase 2) ───
from .admin_change_requests import (
    change_requests_list,
    change_request_action,
)

# ─── Central de Motoristas (vista admin moderna) ───
from .central_views import drivers_central

# ─── Portal Admin: editar / docs / veículos / helpers / reclamações / logins / financeiro ───
from .portal_admin_views import (
    driver_admin_edit,
    driver_unify_search,
    driver_unify_preview,
    driver_unify_execute,
    driver_documents,
    driver_vehicles,
    driver_helpers,
    driver_complaints,
    driver_logins,
    driver_financeiro,
    driver_pre_invoice_detail,
)
