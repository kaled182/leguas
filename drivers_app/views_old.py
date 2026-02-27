from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.utils import timezone
from .models import DriverAccess, DriverRoute
from django.db import models
from django.contrib.auth import logout
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth import login as auth_login
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from geopy.geocoders import Nominatim
import folium
import re
from django.conf import settings


# Create your views here.
def driversmanager_login(request):
    """
    Renderiza a tela de login e autentica o motorista por email ou NIF.
    """
    if request.method == 'GET':
        return render(request, 'driversmanager-login.html')
    
    elif request.method == 'POST':
        identifier = request.POST.get('identifier')  # Pode ser email ou NIF
        password = request.POST.get('password')

        try:
            driver = DriverAccess.objects.get(Q(email=identifier) | Q(nif=identifier))
            if driver.check_password(password):
                # Login OK
                request.session['driver_id'] = driver.id
                # Retorna JSON se for AJAX, senão faz redirect normal
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'success', 'message': 'Login realizado com sucesso'})
                return redirect('driversmanager_view')
            else:
                # Senha incorreta
                return JsonResponse({'status': 'error', 'message': 'Invalid credentials'}, status=401)
        except DriverAccess.DoesNotExist:
            # Motorista não encontrado
            return JsonResponse({'status': 'error', 'message': 'Driver not found'}, status=404)

    return render(request, 'driversmanager-login.html')


def driversmanager_view(request):
    """
    Exibe os pedidos de todos os motoristas relacionados ao usuário autenticado,
    com filtros, cartões de resumo e paginação.
    """
    # Verifica se o motorista está autenticado via sessão manual
    driver_id = request.session.get('driver_id')
    if not driver_id:
        return redirect('driversmanager-login')

    # Se quiser buscar apenas o motorista autenticado:
    try:
        driver_access = DriverAccess.objects.get(id=driver_id)
    except DriverAccess.DoesNotExist:
        return redirect('driversmanager-login')

    # Aqui está o ajuste:
    driver = driver_access.driver  # Isso é uma instância de ordersmanager.Driver

    # Agora use driver normalmente:
    route = DriverRoute(driver)

    status = request.GET.get('status', '')
    date = request.GET.get('date', '')
    today = timezone.now().date()
    if not date:
        date = today

    all_orders_qs = []
    for driver in [driver]:
        route = DriverRoute(driver)
        qs = route.get_orders(date=date)
        if status:
            qs = qs.filter(simplified_order_status=status)
        all_orders_qs.append(qs)
    from itertools import chain
    orders_qs = list(chain.from_iterable(all_orders_qs))

    welcome_name = f"{driver_access.first_name} {driver_access.last_name}".strip() or 'Motorista'


    to_attempt_count = sum(route.get_orders(date=date).filter(simplified_order_status='to_attempt').count() for route in [DriverRoute(d) for d in [driver]])
    picked_up_count = sum(route.get_orders(date=date).filter(simplified_order_status='picked_up').count() for route in [DriverRoute(d) for d in [driver]])
    delivered_count = sum(route.get_orders(date=date).filter(simplified_order_status='delivered').count() for route in [DriverRoute(d) for d in [driver]])
    failed_count = sum(route.get_orders(date=date).filter(simplified_order_status='failed').count() for route in [DriverRoute(d) for d in [driver]])

    paginator = Paginator(orders_qs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    orders = [
        {
            'order_id': o.order_id,
            'retailer': o.retailer,
            'client_address': o.client_address,
            'intended_delivery_date': o.intended_delivery_date,
            'status': o.simplified_order_status,
        }
        for o in page_obj.object_list
    ]

    context = {
        'welcome_name': welcome_name,
        'orders': orders,
        'to_attempt_count': to_attempt_count,
        'delivered_count': delivered_count,
        'failed_count': failed_count,
        'page_obj': page_obj,
        'today': today,
        'request': request,
    }
    return render(request, 'driversmanager.html', context)

def driversmanager_logout(request):
    """
    Encerra a sessão do usuário e redireciona para a página de login.
    """
    logout(request)
    return redirect('authenticate')

def export_xlsx(request):
    """
    Exporta os endereços dos pedidos do motorista autenticado para um arquivo XLSX,
    pronto para importar no Circuit, incluindo Nº do volume, endereço, data prevista e recebimento (COD).
    """
    driver_id = request.session.get('driver_id')
    if not driver_id:
        return redirect('driversmanager-login')

    try:
        driver_access = DriverAccess.objects.get(id=driver_id)
    except DriverAccess.DoesNotExist:
        return redirect('driversmanager-login')

    driver = driver_access.driver
    route = DriverRoute(driver)

    date = request.GET.get('date', timezone.now().date())
    orders_qs = route.get_orders(date=date)

    # Cria o workbook e a planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endereços"

    # Cabeçalhos para o Circuit (ajuste conforme necessário)
    headers = ["Nº do Volume", "Endereço", "Data Prevista", "Recebimento (COD)"]
    ws.append(headers)

    # Preenche as linhas
    for idx, order in enumerate(orders_qs, start=1):
        ws.append([
            idx,  # Nº do Volume sequencial
            order.client_address,
            order.intended_delivery_date.strftime('%d/%m/%Y') if order.intended_delivery_date else '',
            getattr(order, 'cod', ''),  # Recebimento (COD), ajuste se necessário
        ])

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Prepara resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"enderecos_{date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def folium_map(request):
    """
    Exibe um mapa com os pedidos geolocalizados do motorista autenticado,
    focando apenas nos CPs de Viana do Castelo (começando por 49).
    """
    driver_id = request.session.get('driver_id')
    if not driver_id:
        return redirect('driversmanager_login')

    try:
        driver_access = DriverAccess.objects.get(id=driver_id)
    except DriverAccess.DoesNotExist:
        return redirect('driversmanager_login')

    driver = driver_access.driver
    route = DriverRoute(driver)

    date = request.GET.get('date', timezone.now().date())
    orders_qs = route.get_orders(date=date)

    geolocator = Nominatim(user_agent="leguas-monitoring", timeout=10)
    orders_with_coords = []

    def is_valid_postal_code(address):
        match = re.search(r'49\d{2}-\d{3}', address)
        return bool(match)

    def geocode_address(address):
        base_city = "Viana do Castelo, Portugal"
        full_address = f"{address}, {base_city}"
        return geolocator.geocode(full_address)

    for order in orders_qs:
        lat = getattr(order, 'client_latitude', None)
        lon = getattr(order, 'client_longitude', None)

        if not lat or not lon:
            if is_valid_postal_code(order.client_address):
                try:
                    location = geocode_address(order.client_address)
                    if location:
                        lat, lon = location.latitude, location.longitude
                except Exception:
                    lat, lon = None, None
            else:
                lat, lon = None, None

        if lat and lon:
            orders_with_coords.append({
                'order': order,
                'lat': lat,
                'lon': lon,
            })

    warehouse_lat, warehouse_lon = 41.504073, -8.761848
    m = folium.Map(location=[warehouse_lat, warehouse_lon], zoom_start=12)

    folium.Marker(
        location=[warehouse_lat, warehouse_lon],
        popup=folium.Popup("<b>Armazém</b>", max_width=200),
        tooltip="Armazém",
        icon=folium.Icon(color='red', icon='home')
    ).add_to(m)

    for item in orders_with_coords:
        order = item['order']
        lat = item['lat']
        lon = item['lon']
        gmaps_url = f"https://www.google.com/maps/dir/?api=1&destination={lat},{lon}"
        waze_url = f"https://waze.com/ul?ll={lat},{lon}&navigate=yes"

        popup_html = f"""
            <b>Pedido #{order.order_id}</b><br>
            <b>{order.retailer}</b><br>
            {order.client_address}<br>
            <a href="{gmaps_url}" target="_blank">Google Maps</a> |
            <a href="{waze_url}" target="_blank">Waze</a>
        """

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=order.client_address,
            icon=folium.Icon(color='blue', icon='info-sign')
        ).add_to(m)

    # Exporta o HTML do mapa
    map_html = m._repr_html_()

    context = {
        'map_html': map_html,
        'date': date,
        'driver_name': f"{driver_access.first_name} {driver_access.last_name}",
    }
    return render(request, 'folium_map.html', context)