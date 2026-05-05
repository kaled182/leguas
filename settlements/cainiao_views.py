"""Views para import de planilhas Cainiao."""
import io
from collections import defaultdict
from datetime import datetime as _dt, date as _date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_http_methods

from .models import CainiaoDelivery, CainiaoImportBatch, DriverCourierMapping, DriverHelper


@login_required
@require_http_methods(["POST"])
def cainiao_import_preview(request):
    """Parse do Excel Cainiao e devolve preview agrupado por driver."""
    try:
        import openpyxl
    except ImportError:
        return JsonResponse({"success": False, "error": "openpyxl nao instalado."}, status=500)

    from core.models import Partner

    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    date_from = request.POST.get("date_from", "").strip()
    date_to   = request.POST.get("date_to",   "").strip()
    if not date_from or not date_to:
        return JsonResponse({"success": False, "error": "Periodo obrigatorio."}, status=400)

    try:
        file_bytes = f.read()
        # Sem read_only para poder iterar múltiplas sheets e relê-las
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler Excel: {e}"}, status=400)

    # Procurar a sheet que contém dados Cainiao (pode não ser a sheet activa)
    def _find_cainiao_sheet(workbook):
        """Devolve (sheet, rows) da sheet que contém cabeçalho Cainiao."""
        for sname in workbook.sheetnames:
            sheet = workbook[sname]
            sheet_rows = list(sheet.iter_rows(values_only=True))
            for i, row in enumerate(sheet_rows):
                rv = [str(c).strip().lower() if c else "" for c in row]
                ne = [v for v in rv if v]
                if any("courier id" in v for v in rv) and len(ne) >= 3:
                    return sheet_rows
                if any("courier" in v for v in rv) and len(ne) >= 4:
                    return sheet_rows
        return None

    rows = _find_cainiao_sheet(wb)

    if not rows:
        # Debug: listar sheets e primeiras células de cada uma
        debug_sheets = []
        for sname in wb.sheetnames:
            sheet = wb[sname]
            first_vals = []
            for row in sheet.iter_rows(max_row=5, values_only=True):
                ne = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if ne:
                    first_vals.append(ne[:3])
                    break
            debug_sheets.append(f"{sname}:{first_vals}")
        return JsonResponse({
            "success": False,
            "error": f"Coluna 'Courier ID' nao encontrada. Sheets: {debug_sheets[:4]}"
        }, status=400)

    # Identificar linha exacta do cabeçalho dentro das rows da sheet encontrada
    header_row_vals = None
    header_idx = 0
    for i, row in enumerate(rows):
        rv = [str(c).strip().lower() if c else "" for c in row]
        ne = [v for v in rv if v]
        if (any("courier id" in v for v in rv) and len(ne) >= 3) or \
           (any("courier" in v for v in rv) and len(ne) >= 4):
            header_row_vals = rv
            header_idx = i
            break

    if header_row_vals is None:
        return JsonResponse({"success": False, "error": "Linha de cabecalho nao identificada."}, status=400)

    # Mapeamento case-insensitive: nome_normalizado → indice
    col = {v: idx for idx, v in enumerate(header_row_vals) if v}

    def get_col(name):
        """Encontra índice da coluna por substring case-insensitive."""
        name_lower = name.lower()
        for k, v in col.items():
            if name_lower in k:
                return v
        return None

    col_id      = get_col("courier id")
    col_name    = get_col("courier name")
    col_helper  = get_col("courier helper")
    col_status  = get_col("task status")
    col_lp      = get_col("lp no")
    col_waybill = get_col("waybill")

    if col_id is None or col_status is None:
        found = list(col.keys())
        return JsonResponse({
            "success": False,
            "error": f"Colunas obrigatorias em falta. Headers encontrados na linha {header_idx}: {found}"
        }, status=400)

    cainiao_partner = Partner.objects.filter(name__icontains="cainiao").first()

    driver_map = defaultdict(lambda: {"direct": 0, "helpers": defaultdict(int), "courier_name": ""})
    total_rows = 0
    total_delivered = 0

    for row in rows[header_idx + 1:]:
        if not row or not row[col_id]:
            continue
        total_rows += 1
        status = str(row[col_status] or "").strip()
        if status != "Delivered":
            continue
        total_delivered += 1

        courier_id = str(row[col_id] or "").strip()
        helper = str(row[col_helper] or "").strip() if col_helper is not None else ""
        cname  = str(row[col_name]   or "").strip() if col_name  is not None else ""

        entry = driver_map[courier_id]
        if cname and not entry["courier_name"]:
            entry["courier_name"] = cname
        if helper:
            entry["helpers"][helper] += 1
        else:
            entry["direct"] += 1

    result_rows = []
    for courier_id, counts in driver_map.items():
        mapping = None
        if cainiao_partner:
            mapping = DriverCourierMapping.objects.filter(
                partner=cainiao_partner, courier_id=courier_id
            ).select_related("driver").first()

        driver_id   = mapping.driver.id            if mapping else None
        driver_name = mapping.driver.nome_completo if mapping else None
        total_driver = counts["direct"] + sum(counts["helpers"].values())
        helpers_list = [{"name": k, "qty": v} for k, v in sorted(counts["helpers"].items())]

        result_rows.append({
            "courier_id":   courier_id,
            "courier_name": counts.get("courier_name", ""),
            "driver_id":    driver_id,
            "driver_name":  driver_name,
            "mapped":       mapping is not None,
            "total":        total_driver,
            "direct":       counts["direct"],
            "helpers":      helpers_list,
        })

    result_rows.sort(key=lambda r: (not r["mapped"], r["courier_id"]))

    return JsonResponse({
        "success":         True,
        "total_rows":      total_rows,
        "total_delivered": total_delivered,
        "date_from":       date_from,
        "date_to":         date_to,
        "filename":        f.name,
        "rows":            result_rows,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_import_confirm(request):
    """Confirma e persiste o import Cainiao apos preview.

    Idempotente: pacotes identificados por waybill_number (ou lp_number quando
    waybill vazio) nunca são inseridos duas vezes.
    O período é determinado automaticamente pelo min/max de Delivery Time.
    """
    import openpyxl
    from decimal import Decimal
    from datetime import date as _date, datetime as _dt

    from core.models import Partner
    from .models import DriverPreInvoice, PreInvoiceLine
    from .views import _gerar_numero_pre_fatura

    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"success": False, "error": "Ficheiro em falta."}, status=400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler Excel: {e}"}, status=400)

    # ── Localizar sheet e cabeçalho ─────────────────────────────────────
    rows = None
    for sname in wb.sheetnames:
        sheet = wb[sname]
        sheet_rows = list(sheet.iter_rows(values_only=True))
        for row in sheet_rows:
            rv = [str(c).strip().lower() if c else "" for c in row]
            ne = [v for v in rv if v]
            if (any("courier id" in v for v in rv) and len(ne) >= 3) or \
               (any("courier" in v for v in rv) and len(ne) >= 4):
                rows = sheet_rows
                break
        if rows is not None:
            break

    if rows is None:
        return JsonResponse({"success": False, "error": "Coluna 'Courier ID' nao encontrada."}, status=400)

    header_row_vals = None
    header_idx = 0
    for i, row in enumerate(rows):
        rv = [str(c).strip().lower() if c else "" for c in row]
        ne = [v for v in rv if v]
        if (any("courier id" in v for v in rv) and len(ne) >= 3) or \
           (any("courier" in v for v in rv) and len(ne) >= 4):
            header_row_vals = rv
            header_idx = i
            break

    if header_row_vals is None:
        return JsonResponse({"success": False, "error": "Linha de cabecalho nao identificada."}, status=400)

    col = {v: idx for idx, v in enumerate(header_row_vals) if v}

    def get_col(name):
        nl = name.lower()
        for k, v in col.items():
            if nl in k:
                return v
        return None

    col_id       = get_col("courier id")
    col_helper   = get_col("courier helper")
    col_status   = get_col("task status")
    col_lp       = get_col("lp no")
    col_waybill  = get_col("waybill")
    col_dtime    = get_col("delivery time")

    if col_id is None or col_status is None:
        return JsonResponse({"success": False, "error": "Colunas obrigatorias (Courier ID / Task Status) em falta."}, status=400)

    # ── Carregar mapeamentos driver ──────────────────────────────────────
    cainiao_partner = Partner.objects.filter(name__icontains="cainiao").first()
    mappings = {}
    if cainiao_partner:
        for m in DriverCourierMapping.objects.filter(partner=cainiao_partner).select_related("driver"):
            mappings[m.courier_id] = m.driver

    # ── Primeira passagem: recolher dados das linhas Delivered ───────────
    candidate_rows = []
    for row in rows[header_idx + 1:]:
        if not row or not row[col_id]:
            continue
        status = str(row[col_status] or "").strip()
        if status != "Delivered":
            continue

        courier_id = str(row[col_id]   or "").strip()
        helper     = str(row[col_helper]  or "").strip() if col_helper  is not None else ""
        lp         = str(row[col_lp]      or "").strip() if col_lp      is not None else ""
        waybill    = str(row[col_waybill] or "").strip() if col_waybill is not None else ""

        # Delivery Time — aceitar datetime nativo do openpyxl ou string
        dtime = None
        if col_dtime is not None:
            raw = row[col_dtime]
            if isinstance(raw, _dt):
                dtime = raw
            elif raw:
                try:
                    dtime = _dt.fromisoformat(str(raw).strip())
                except ValueError:
                    pass

        candidate_rows.append({
            "courier_id": courier_id,
            "helper": helper,
            "lp": lp,
            "waybill": waybill,
            "dtime": dtime,
        })

    if not candidate_rows:
        return JsonResponse({"success": False, "error": "Nenhuma entrega 'Delivered' encontrada no ficheiro."}, status=400)

    # ── Verificar duplicados: waybill e lp já existentes na BD ──────────
    all_waybills = {r["waybill"] for r in candidate_rows if r["waybill"]}
    all_lps      = {r["lp"]      for r in candidate_rows if r["lp"]}

    existing_waybills = set(
        CainiaoDelivery.objects.filter(waybill_number__in=all_waybills)
        .values_list("waybill_number", flat=True)
    ) if all_waybills else set()

    existing_lps = set(
        CainiaoDelivery.objects.filter(lp_number__in=all_lps)
        .values_list("lp_number", flat=True)
    ) if all_lps else set()

    new_rows = []
    skipped  = 0
    for r in candidate_rows:
        is_dup = (r["waybill"] and r["waybill"] in existing_waybills) or \
                 (r["lp"] and not r["waybill"] and r["lp"] in existing_lps)
        if is_dup:
            skipped += 1
        else:
            new_rows.append(r)

    if not new_rows:
        return JsonResponse({
            "success": False,
            "error":   f"Todas as {skipped} entregas deste ficheiro já foram importadas anteriormente.",
            "skipped": skipped,
        }, status=400)

    # ── Período automático via Delivery Time ─────────────────────────────
    dtimes = [r["dtime"] for r in new_rows if r["dtime"] is not None]
    if dtimes:
        periodo_inicio = min(dtimes).date()
        periodo_fim    = max(dtimes).date()
    else:
        # Fallback para datas manuais enviadas no form
        try:
            date_from = request.POST.get("date_from", "").strip()
            date_to   = request.POST.get("date_to",   "").strip()
            periodo_inicio = _date.fromisoformat(date_from)
            periodo_fim    = _date.fromisoformat(date_to)
        except (ValueError, AttributeError):
            return JsonResponse({"success": False, "error": "Nao foi possivel determinar o periodo. Preencha as datas manualmente."}, status=400)

    # ── Verificar que existe pelo menos um driver mapeado ────────────────
    unmapped_ids = {r["courier_id"] for r in new_rows if not mappings.get(r["courier_id"])}
    mapped_count = sum(1 for r in new_rows if mappings.get(r["courier_id"]))

    if mapped_count == 0:
        return JsonResponse({
            "success": False,
            "error": (
                f"Nenhum dos {len(new_rows)} pacotes tem driver mapeado. "
                f"Faça o Pré-visualizar e associe os couriers: {sorted(unmapped_ids)[:5]}"
            ),
            "unmapped_courier_ids": sorted(unmapped_ids),
        }, status=400)

    # ── Criar batch e persistir entregas ─────────────────────────────────
    batch = CainiaoImportBatch.objects.create(
        filename=f.name,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        created_by=request.user,
    )

    deliveries  = []
    helper_seen = defaultdict(set)

    for r in new_rows:
        driver = mappings.get(r["courier_id"])
        if not driver:
            continue
        deliveries.append(CainiaoDelivery(
            batch=batch, driver=driver,
            courier_id=r["courier_id"], helper_name=r["helper"],
            lp_number=r["lp"], waybill_number=r["waybill"],
            delivery_time=r["dtime"],
        ))
        if r["helper"]:
            helper_seen[driver.id].add((r["helper"], str(periodo_inicio)))

    CainiaoDelivery.objects.bulk_create(deliveries, batch_size=500)
    batch.total_delivered = len(candidate_rows)
    batch.save(update_fields=["total_delivered"])

    for driver_id, helpers in helper_seen.items():
        for helper_name, first_date in helpers:
            DriverHelper.objects.get_or_create(
                driver_id=driver_id, helper_name=helper_name,
                defaults={"first_seen": first_date},
            )

    # ── Gerar Pré-Faturas ────────────────────────────────────────────────
    try:
        taxa_str = (request.POST.get("taxa_por_entrega", "0") or "0").replace(",", ".")
        taxa = Decimal(taxa_str)
    except Exception:
        taxa = Decimal("0.00")

    pre_invoices_criadas = []

    if cainiao_partner and deliveries:
        driver_totals = {}
        for d in deliveries:
            if d.driver_id not in driver_totals:
                driver_totals[d.driver_id] = [d.driver, 0]
            driver_totals[d.driver_id][1] += 1

        for _did, (drv, total_pcts) in driver_totals.items():
            pi = DriverPreInvoice.objects.filter(
                driver=drv,
                periodo_inicio=periodo_inicio,
                periodo_fim=periodo_fim,
            ).first()

            if not pi:
                pi = DriverPreInvoice.objects.create(
                    numero=_gerar_numero_pre_fatura(),
                    driver=drv,
                    periodo_inicio=periodo_inicio,
                    periodo_fim=periodo_fim,
                    status="RASCUNHO",
                    api_source="cainiao",
                    created_by=request.user,
                )

            # Uma linha Cainiao por pré-fatura; atualiza se já existir
            existing_line = PreInvoiceLine.objects.filter(
                pre_invoice=pi, api_source="cainiao", parceiro=cainiao_partner
            ).first()

            if existing_line:
                existing_line.total_pacotes   += total_pcts
                existing_line.taxa_por_entrega = taxa
                existing_line.base_entregas    = Decimal(existing_line.total_pacotes) * taxa
                existing_line.save(update_fields=["total_pacotes", "taxa_por_entrega", "base_entregas"])
            else:
                PreInvoiceLine.objects.create(
                    pre_invoice=pi,
                    parceiro=cainiao_partner,
                    total_pacotes=total_pcts,
                    taxa_por_entrega=taxa,
                    base_entregas=Decimal(total_pcts) * taxa,
                    api_source="cainiao",
                )

            pi.recalcular()
            pre_invoices_criadas.append(pi.numero)

    return JsonResponse({
        "success":         True,
        "batch_id":        batch.id,
        "saved":           len(deliveries),
        "skipped":         skipped,
        "total_delivered": len(candidate_rows),
        "periodo_inicio":  str(periodo_inicio),
        "periodo_fim":     str(periodo_fim),
        "pre_invoices":    pre_invoices_criadas,
    })


@login_required
def cainiao_helpers_api(request, driver_id):
    """Devolve helpers e contagens de entregas organizados por mês."""
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    from drivers_app.models import DriverProfile

    driver = get_object_or_404(DriverProfile, id=driver_id)

    # Registered helpers (for status/first_seen info)
    registered = {
        h.helper_name: {"first_seen": h.first_seen.strftime("%d/%m/%Y") if h.first_seen else "", "is_active": h.is_active}
        for h in driver.helpers.all()
    }

    # Monthly aggregation: (month, helper_name) → count
    monthly_qs = (
        CainiaoDelivery.objects
        .filter(driver=driver)
        .annotate(month=TruncMonth("delivery_time"))
        .values("month", "helper_name")
        .annotate(count=Count("id"))
        .order_by("-month", "helper_name")
    )

    months = {}
    for row in monthly_qs:
        m = row["month"]
        if m:
            key = m.strftime("%Y-%m")
            label = m.strftime("%B %Y")
        else:
            key = "sem-data"
            label = "Sem data"

        if key not in months:
            months[key] = {"key": key, "label": label, "total": 0, "direct": 0, "helpers": []}

        months[key]["total"] += row["count"]
        if not row["helper_name"]:
            months[key]["direct"] += row["count"]
        else:
            info = registered.get(row["helper_name"], {})
            months[key]["helpers"].append({
                "name":       row["helper_name"],
                "qty":        row["count"],
                "is_active":  info.get("is_active", True),
                "first_seen": info.get("first_seen", ""),
            })

    months_list = sorted(months.values(), key=lambda x: x["key"], reverse=True)

    total  = CainiaoDelivery.objects.filter(driver=driver).count()
    direct = CainiaoDelivery.objects.filter(driver=driver, helper_name="").count()
    unique_helpers = list(registered.keys())

    return JsonResponse({
        "success":        True,
        "total":          total,
        "direct":         direct,
        "helpers_count":  len(unique_helpers),
        "months":         months_list,
    })


@login_required
def cainiao_drivers_list(request):
    """Lista todos os drivers activos para o dropdown de mapeamento."""
    from drivers_app.models import DriverProfile
    drivers = list(
        DriverProfile.objects.filter(status="ATIVO")
        .order_by("nome_completo")
        .values("id", "nome_completo")
    )
    return JsonResponse({"success": True, "drivers": drivers})


@login_required
@require_http_methods(["POST"])
def cainiao_link_driver(request):
    """Liga um courier_id a um DriverProfile existente (cria DriverCourierMapping)."""
    import json
    from core.models import Partner
    from drivers_app.models import DriverProfile

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "error": "JSON invalido."}, status=400)

    courier_id = body.get("courier_id", "").strip()
    driver_id  = body.get("driver_id")

    if not courier_id or not driver_id:
        return JsonResponse({"success": False, "error": "Dados incompletos."}, status=400)

    cainiao_partner = Partner.objects.filter(name__icontains="cainiao").first()
    if not cainiao_partner:
        return JsonResponse({"success": False, "error": "Parceiro Cainiao nao encontrado."}, status=400)

    driver = get_object_or_404(DriverProfile, id=driver_id)

    mapping, created = DriverCourierMapping.objects.get_or_create(
        partner=cainiao_partner,
        courier_id=courier_id,
        defaults={"driver": driver},
    )
    if not created and mapping.driver_id != driver.id:
        mapping.driver = driver
        mapping.save(update_fields=["driver"])

    return JsonResponse({
        "success":     True,
        "created":     created,
        "driver_id":   driver.id,
        "driver_name": driver.nome_completo,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_create_driver(request):
    """Cria DriverProfile mínimo + DriverCourierMapping a partir do import."""
    import json
    from core.models import Partner
    from drivers_app.models import DriverProfile

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "error": "JSON invalido."}, status=400)

    courier_id = body.get("courier_id", "").strip()
    nome       = body.get("nome", "").strip()
    nif        = body.get("nif", "").strip()
    telefone   = body.get("telefone", "").strip()
    email      = body.get("email", "").strip()

    if not courier_id or not nome:
        return JsonResponse({"success": False, "error": "Nome e Courier ID sao obrigatorios."}, status=400)

    cainiao_partner = Partner.objects.filter(name__icontains="cainiao").first()
    if not cainiao_partner:
        return JsonResponse({"success": False, "error": "Parceiro Cainiao nao encontrado."}, status=400)

    # Usar NIF placeholder se não fornecido
    if not nif:
        from .views import _gerar_nif_placeholder
        nif = _gerar_nif_placeholder() or f"999{courier_id[-6:].zfill(6)}"

    try:
        driver = DriverProfile.objects.create(
            nome_completo=nome,
            nif=nif,
            telefone=telefone or "000000000",
            email=email or f"driver_{courier_id}@cainiao.placeholder",
            status="ATIVO",
        )
        DriverCourierMapping.objects.create(
            partner=cainiao_partner,
            courier_id=courier_id,
            driver=driver,
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({
        "success":     True,
        "driver_id":   driver.id,
        "driver_name": driver.nome_completo,
    })


# ============================================================================
# CAINIAO — IMPORT PLANILHA (PREVISÃO / INBOUND / FINAL DO DIA)
# ============================================================================

def _parse_forecast_file(file_bytes):
    """
    Lê um ficheiro xlsx Cainiao e devolve (data_rows, idx, header_names).
    data_rows: lista de tuplos já sem a linha de cabeçalho.
    idx: dict com índices das colunas conhecidas.
    header_names: lista de strings com os nomes reais das colunas encontradas.
    Levanta ValueError com mensagem amigável em caso de erro.
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl não instalado.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Erro ao abrir ficheiro: {exc}")

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("Ficheiro vazio.")

    # Localizar linha do cabeçalho (pode não ser row 0)
    header = None
    header_row_idx = 0
    for i, row in enumerate(all_rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("tracking" in c for c in cells):
            header = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if header is None:
        debug = [[str(c) for c in r if c is not None] for r in all_rows[:3]]
        raise ValueError(
            f"Coluna 'Tracking No.' não encontrada. Primeiras linhas: {debug}"
        )

    def col(name):
        name_lower = name.lower()
        for i, h in enumerate(header):
            if h and name_lower in h.lower():
                return i
        return None

    idx = {
        "tracking":    col("Tracking No"),
        "lp":          col("LP No"),
        "site_code":   col("Site code"),
        "status":      col("Status"),
        "bigbag_id":   col("Inbound Bigbag ID"),
        "bigbag_no":   col("Bigbag No"),
        "sort_code":   col("Sort Code"),
        "order_type":  col("Order Type"),
        "sop_type":    col("SOP Typ"),
        "recv_name":   col("Receiver"),
        "recv_phone":  col("Contact"),
        "recv_region": col("Region"),
        "recv_city":   col("City"),
        "recv_zip":    col("Zip"),
        "recv_addr":   col("Detail Address"),
        "coords":      col("Receiver to"),
    }

    data_rows = all_rows[header_row_idx + 1:]
    return data_rows, idx, header


def _cell(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return str(v).strip() if v is not None else ""


@login_required
@require_http_methods(["POST"])
def cainiao_import_forecast_preview(request):
    """Analisa o ficheiro e devolve um resumo sem gravar nada na BD."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    operation_date = request.POST.get("operation_date", "").strip()
    if not operation_date:
        return JsonResponse({"success": False, "error": "Data da operação é obrigatória."}, status=400)

    from .models import CainiaoForecastPackage

    try:
        from datetime import date
        op_date = date.fromisoformat(operation_date)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data da operação inválida."}, status=400)

    try:
        data_rows, idx, header = _parse_forecast_file(ficheiro.read())
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    # Existing tracking numbers for this date (for upsert preview)
    existing_keys = set(
        CainiaoForecastPackage.objects.filter(operation_date=op_date)
        .values_list("tracking_number", flat=True)
    )

    status_counts = {}
    sample = []
    total_novos = 0
    total_atualizados = 0

    for row in data_rows:
        tracking = _cell(row, idx, "tracking")
        if not tracking:
            continue
        status_val = _cell(row, idx, "status") or CainiaoForecastPackage.STATUS_CREATE
        status_counts[status_val] = status_counts.get(status_val, 0) + 1
        if tracking in existing_keys:
            total_atualizados += 1
        else:
            total_novos += 1
            existing_keys.add(tracking)  # avoid double-count within file
        if len(sample) < 10:
            sample.append({
                "tracking": tracking,
                "status": status_val,
                "receiver": _cell(row, idx, "recv_name"),
                "city": _cell(row, idx, "recv_city"),
                "action": "update" if tracking in existing_keys and total_novos == 0 else "new",
            })

    return JsonResponse({
        "success": True,
        "total_novos": total_novos,
        "total_atualizados": total_atualizados,
        "total": total_novos + total_atualizados,
        "operation_date": operation_date,
        "status_counts": status_counts,
        "sample": sample,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_import_forecast(request):
    """
    Importa qualquer planilha Cainiao (Previsão, Inbound, Final do Dia).
    Suporta reimportação: pacotes existentes têm o status atualizado (upsert).
    """
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    operation_date = request.POST.get("operation_date", "").strip()
    import_type = request.POST.get("import_type", "FORECAST").strip().upper()

    if not operation_date:
        return JsonResponse({"success": False, "error": "Data da operação é obrigatória."}, status=400)

    from .models import CainiaoForecastBatch, CainiaoForecastPackage

    valid_types = {t[0] for t in CainiaoForecastBatch.IMPORT_TYPES}
    if import_type not in valid_types:
        import_type = CainiaoForecastBatch.TYPE_FORECAST

    try:
        from datetime import date
        op_date = date.fromisoformat(operation_date)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data da operação inválida."}, status=400)

    try:
        data_rows, idx, _header = _parse_forecast_file(ficheiro.read())
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)

    batch = CainiaoForecastBatch.objects.create(
        filename=ficheiro.name,
        import_type=import_type,
        operation_date=op_date,
        total_packages=0,
        updated_packages=0,
        created_by=request.user,
    )

    def parse_coords(raw):
        if not raw:
            return None, None
        parts = raw.split(",")
        if len(parts) == 2:
            try:
                from decimal import Decimal, InvalidOperation
                return Decimal(parts[0].strip()), Decimal(parts[1].strip())
            except InvalidOperation:
                pass
        return None, None

    existing = {
        pkg.tracking_number: pkg
        for pkg in CainiaoForecastPackage.objects.filter(operation_date=op_date)
    }

    to_create = []
    to_update = []
    update_fields = [
        "status", "last_import_batch",
        "lp_number", "site_code", "inbound_bigbag_id", "bigbag_number",
        "sort_code", "order_type", "sop_type",
        "receiver_name", "receiver_phone", "receiver_region",
        "receiver_city", "receiver_zip", "receiver_address",
        "latitude", "longitude",
    ]
    status_counts = {}

    for row in data_rows:
        tracking = _cell(row, idx, "tracking")
        if not tracking:
            continue
        status_val = _cell(row, idx, "status") or CainiaoForecastPackage.STATUS_CREATE
        status_counts[status_val] = status_counts.get(status_val, 0) + 1
        lat, lng = parse_coords(_cell(row, idx, "coords"))

        if tracking in existing:
            pkg = existing[tracking]
            pkg.status = status_val
            pkg.last_import_batch = batch
            pkg.lp_number = _cell(row, idx, "lp")
            pkg.site_code = _cell(row, idx, "site_code")
            pkg.inbound_bigbag_id = _cell(row, idx, "bigbag_id")
            pkg.bigbag_number = _cell(row, idx, "bigbag_no")
            pkg.sort_code = _cell(row, idx, "sort_code")
            pkg.order_type = _cell(row, idx, "order_type")
            pkg.sop_type = _cell(row, idx, "sop_type")
            pkg.receiver_name = _cell(row, idx, "recv_name")
            pkg.receiver_phone = _cell(row, idx, "recv_phone")
            pkg.receiver_region = _cell(row, idx, "recv_region")
            pkg.receiver_city = _cell(row, idx, "recv_city")
            pkg.receiver_zip = _cell(row, idx, "recv_zip")
            pkg.receiver_address = _cell(row, idx, "recv_addr")
            if lat is not None:
                pkg.latitude = lat
                pkg.longitude = lng
            to_update.append(pkg)
        else:
            to_create.append(CainiaoForecastPackage(
                operation_date=op_date,
                tracking_number=tracking,
                status=status_val,
                last_import_batch=batch,
                lp_number=_cell(row, idx, "lp"),
                site_code=_cell(row, idx, "site_code"),
                inbound_bigbag_id=_cell(row, idx, "bigbag_id"),
                bigbag_number=_cell(row, idx, "bigbag_no"),
                sort_code=_cell(row, idx, "sort_code"),
                order_type=_cell(row, idx, "order_type"),
                sop_type=_cell(row, idx, "sop_type"),
                receiver_name=_cell(row, idx, "recv_name"),
                receiver_phone=_cell(row, idx, "recv_phone"),
                receiver_region=_cell(row, idx, "recv_region"),
                receiver_city=_cell(row, idx, "recv_city"),
                receiver_zip=_cell(row, idx, "recv_zip"),
                receiver_address=_cell(row, idx, "recv_addr"),
                latitude=lat,
                longitude=lng,
            ))
            existing[tracking] = to_create[-1]

    if to_create:
        CainiaoForecastPackage.objects.bulk_create(to_create, batch_size=500)
    if to_update:
        CainiaoForecastPackage.objects.bulk_update(to_update, update_fields, batch_size=500)

    batch.total_packages = len(to_create)
    batch.updated_packages = len(to_update)
    batch.save(update_fields=["total_packages", "updated_packages"])

    return JsonResponse({
        "success": True,
        "import_type": import_type,
        "total_novos": len(to_create),
        "total_atualizados": len(to_update),
        "operation_date": operation_date,
        "batch_id": batch.id,
        "status_counts": status_counts,
    })


# ============================================================================
# FORECAST CP4 — API para o modal de previsão
# ============================================================================

@login_required
def cainiao_forecast_cp4(request):
    """Volume previsto por CP4 para uma data (default: amanhã).
    Aceita ?hub_id=N para filtrar apenas os CP4s do HUB seleccionado.

    Volume = pacotes de CainiaoPlanningPackage (do ficheiro forecast)
           + entradas manuais em CainiaoManualForecast (CP4+qty quando
             não há ficheiro disponível, apenas info verbal por código)
    """
    from datetime import timedelta
    from django.db.models import Count, Q
    from django.db.models.functions import Substr
    from .models import (
        CainiaoPlanningPackage, CainiaoHub, CainiaoManualForecast,
    )

    date_str = request.GET.get("date", "")
    try:
        target_date = _date.fromisoformat(date_str)
    except ValueError:
        target_date = _date.today() + timedelta(days=1)

    # HUB filter
    from .models import CainiaoHubCP4
    hub_id_str = request.GET.get("hub_id", "")
    selected_hub = None

    if hub_id_str:
        try:
            selected_hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(id=int(hub_id_str))
            hub_cp4_set = set(selected_hub.cp4_codes.values_list("cp4", flat=True))
        except (CainiaoHub.DoesNotExist, ValueError):
            # Fallback: all known CP4s
            hub_cp4_set = set(CainiaoHubCP4.objects.values_list("cp4", flat=True))
    else:
        # "Todos" — only count CP4s registered in any hub (exclude unknown CP4s)
        hub_cp4_set = set(CainiaoHubCP4.objects.values_list("cp4", flat=True))

    pkgs = CainiaoPlanningPackage.objects.filter(operation_date=target_date)

    # Always filter to known CP4s (union of all hubs or selected hub)
    if hub_cp4_set:
        hub_q = Q()
        for cp4 in hub_cp4_set:
            hub_q |= Q(receiver_zip__startswith=cp4)
        pkgs = pkgs.filter(hub_q)

    total = pkgs.count()

    cp4_rows = list(
        pkgs
        .exclude(receiver_zip="")
        .annotate(cp4=Substr("receiver_zip", 1, 4))
        .values("cp4", "receiver_city")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Collapse duplicates: same cp4 may appear with different city spellings
    seen = {}
    for row in cp4_rows:
        cp4 = row["cp4"]
        if cp4 not in seen:
            seen[cp4] = {"cp4": cp4, "city": row["receiver_city"] or "", "total": 0}
        seen[cp4]["total"] += row["total"]

    # ── Adicionar previsões MANUAIS (sem ficheiro, só CP4+qty) ──
    manual_qs = CainiaoManualForecast.objects.filter(
        operation_date=target_date,
    )
    if hub_cp4_set:
        manual_qs = manual_qs.filter(cp4__in=hub_cp4_set)
    manual_total = 0
    for mf in manual_qs:
        manual_total += mf.qty
        if mf.cp4 in seen:
            seen[mf.cp4]["total"] += mf.qty
            seen[mf.cp4]["manual_qty"] = (
                seen[mf.cp4].get("manual_qty", 0) + mf.qty
            )
            seen[mf.cp4]["manual_id"] = mf.id
            seen[mf.cp4]["manual_notes"] = mf.notes
        else:
            seen[mf.cp4] = {
                "cp4": mf.cp4,
                "city": "",
                "total": mf.qty,
                "manual_qty": mf.qty,
                "manual_id": mf.id,
                "manual_notes": mf.notes,
            }

    # Atualizar `total` global incluindo manuais
    total = total + manual_total

    cp4_final = sorted(seen.values(), key=lambda r: -r["total"])

    available = list(
        CainiaoPlanningPackage.objects
        .values_list("operation_date", flat=True)
        .order_by("-operation_date")
        .distinct()[:30]
    )

    # ── Backlog: pacotes não entregues do último dia ANTERIOR à data prevista ──
    # Filtramos task_date < target_date para que importações do próprio dia de
    # previsão não poluam o backlog (cada import do dia corrente mudaria a base).
    from .models import CainiaoOperationTask
    from django.db.models import Max, Case, When, IntegerField

    most_recent_op = CainiaoOperationTask.objects.filter(
        task_date__lt=target_date
    ).aggregate(d=Max("task_date"))["d"]
    backlog_total = 0
    backlog_rows = []
    backlog_date = None

    if most_recent_op:
        backlog_date = str(most_recent_op)
        bqs = CainiaoOperationTask.objects.filter(
            task_date=most_recent_op,
            task_status__in=["Driver_received", "Attempt Failure", "Unassign", "Assigned"],
        )
        if hub_cp4_set:
            bq = Q()
            for cp4 in hub_cp4_set:
                bq |= Q(zip_code__startswith=cp4)
            bqs = bqs.filter(bq)

        backlog_total = bqs.count()

        # Group by CP4 with breakdown
        raw_bl = (
            bqs.exclude(zip_code="")
            .annotate(cp4=Substr("zip_code", 1, 4))
            .values("cp4", "destination_city")
            .annotate(
                total=Count("id"),
                driver_received=Count(
                    Case(When(task_status="Driver_received", then=1), output_field=IntegerField())
                ),
                attempt_failure=Count(
                    Case(When(task_status="Attempt Failure", then=1), output_field=IntegerField())
                ),
                unassign=Count(
                    Case(When(task_status="Unassign", then=1), output_field=IntegerField())
                ),
                assigned=Count(
                    Case(When(task_status="Assigned", then=1), output_field=IntegerField())
                ),
            )
            .order_by("-total")
        )

        # Collapse duplicate city spellings per CP4
        bl_seen = {}
        for r in raw_bl:
            cp4 = r["cp4"]
            if cp4 not in bl_seen:
                bl_seen[cp4] = {
                    "cp4": cp4,
                    "city": r["destination_city"] or "",
                    "total": 0,
                    "driver_received": 0,
                    "attempt_failure": 0,
                    "unassign": 0,
                    "assigned": 0,
                }
            bl_seen[cp4]["total"] += r["total"]
            bl_seen[cp4]["driver_received"] += r["driver_received"]
            bl_seen[cp4]["attempt_failure"] += r["attempt_failure"]
            bl_seen[cp4]["unassign"] += r["unassign"]
            bl_seen[cp4]["assigned"] += r["assigned"]

        backlog_rows = sorted(bl_seen.values(), key=lambda r: -r["total"])

    return JsonResponse({
        "success": True,
        "date": str(target_date),
        "total": total,
        "cp4_rows": cp4_final,
        "available_dates": [str(d) for d in available],
        "selected_hub_id": selected_hub.id if selected_hub else None,
        "selected_hub_name": selected_hub.name if selected_hub else None,
        "backlog_total": backlog_total,
        "backlog_rows": backlog_rows,
        "backlog_date": backlog_date,
    })


# ============================================================================
# Forecast Manual (sem ficheiro — apenas CP4 + qty)
# ============================================================================

@login_required
def cainiao_manual_forecast_list(request):
    """Lista entradas manuais para uma data."""
    from .models import CainiaoManualForecast
    date_str = request.GET.get("date", "")
    try:
        target_date = _date.fromisoformat(date_str)
    except ValueError:
        from datetime import timedelta
        target_date = _date.today() + timedelta(days=1)
    rows = list(
        CainiaoManualForecast.objects.filter(operation_date=target_date)
        .order_by("cp4")
        .values("id", "cp4", "qty", "notes")
    )
    return JsonResponse({
        "success": True,
        "date": str(target_date),
        "rows": rows,
        "total": sum(r["qty"] for r in rows),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_manual_forecast_save(request):
    """Cria ou actualiza entrada manual (upsert por (date, cp4))."""
    import json
    from .models import CainiaoManualForecast

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        return JsonResponse(
            {"success": False, "error": "JSON inválido"}, status=400,
        )

    date_str = (body.get("date") or "").strip()
    cp4 = (body.get("cp4") or "").strip()
    qty_raw = body.get("qty")
    notes = (body.get("notes") or "").strip()

    try:
        target_date = _date.fromisoformat(date_str)
    except (ValueError, TypeError):
        return JsonResponse(
            {"success": False, "error": "Data inválida"}, status=400,
        )

    if not cp4 or not cp4.isdigit() or len(cp4) != 4:
        return JsonResponse(
            {"success": False, "error":
             "CP4 inválido (4 dígitos numéricos)"},
            status=400,
        )

    try:
        qty = int(qty_raw)
    except (TypeError, ValueError):
        return JsonResponse(
            {"success": False, "error":
             "Quantidade inválida"}, status=400,
        )

    if qty <= 0:
        # Se qty=0, apaga a entrada
        CainiaoManualForecast.objects.filter(
            operation_date=target_date, cp4=cp4,
        ).delete()
        return JsonResponse({"success": True, "deleted": True})

    obj, created = CainiaoManualForecast.objects.update_or_create(
        operation_date=target_date,
        cp4=cp4,
        defaults={"qty": qty, "notes": notes},
    )
    if created:
        obj.created_by = request.user
        obj.save(update_fields=["created_by"])

    return JsonResponse({
        "success": True,
        "created": created,
        "id": obj.id,
        "cp4": obj.cp4,
        "qty": obj.qty,
        "notes": obj.notes,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_manual_forecast_delete(request, manual_id):
    """Apaga uma entrada manual."""
    from .models import CainiaoManualForecast
    obj = get_object_or_404(CainiaoManualForecast, id=manual_id)
    obj.delete()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def cainiao_forecast_reset(request):
    """Apaga a previsão importada para uma data (e hub opcional).
    Mantém:
      - Backlog (vem de CainiaoOperationTask, intocado).
      - Previsões manuais (a menos que ?include_manual=1).
    Apaga:
      - CainiaoPlanningPackage (forecast tradicional)
      - CainiaoForecastPackage (formato PARCEL_LIST rico)
    """
    from .models import (
        CainiaoPlanningPackage, CainiaoForecastPackage,
        CainiaoHub, CainiaoManualForecast,
    )
    date_str = (request.POST.get("date") or "").strip()
    if not date_str:
        return JsonResponse(
            {"success": False, "error": "Data obrigatória."}, status=400,
        )
    try:
        target_date = _date.fromisoformat(date_str)
    except ValueError:
        return JsonResponse(
            {"success": False, "error": "Data inválida."}, status=400,
        )

    hub_id_str = (request.POST.get("hub_id") or "").strip()
    include_manual = (request.POST.get("include_manual") or "") in (
        "1", "true", "yes",
    )

    plan_qs = CainiaoPlanningPackage.objects.filter(
        operation_date=target_date,
    )
    fc_qs = CainiaoForecastPackage.objects.filter(
        operation_date=target_date,
    )
    manual_qs = CainiaoManualForecast.objects.filter(
        operation_date=target_date,
    )

    hub_label = None
    if hub_id_str:
        try:
            hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(
                id=int(hub_id_str),
            )
            hub_label = hub.name
            cp4s = list(hub.cp4_codes.values_list("cp4", flat=True))
            if cp4s:
                hub_q_plan = Q()
                hub_q_fc = Q()
                for cp4 in cp4s:
                    hub_q_plan |= Q(receiver_zip__startswith=cp4)
                    hub_q_fc |= Q(receiver_zip__startswith=cp4)
                plan_qs = plan_qs.filter(hub_q_plan)
                fc_qs = fc_qs.filter(hub_q_fc)
                manual_qs = manual_qs.filter(cp4__in=cp4s)
            else:
                # HUB sem CP4s configurados — não apaga nada
                plan_qs = plan_qs.none()
                fc_qs = fc_qs.none()
                manual_qs = manual_qs.none()
        except (CainiaoHub.DoesNotExist, ValueError):
            return JsonResponse(
                {"success": False, "error": "HUB inválido."}, status=400,
            )

    n_plan = plan_qs.count()
    n_fc = fc_qs.count()
    n_manual = manual_qs.count() if include_manual else 0

    plan_qs.delete()
    fc_qs.delete()
    if include_manual:
        manual_qs.delete()

    return JsonResponse({
        "success": True,
        "date": str(target_date),
        "hub": hub_label,
        "deleted_planning": n_plan,
        "deleted_forecast": n_fc,
        "deleted_manual": n_manual,
        "include_manual": include_manual,
    })


# ============================================================================
# Forecast PDF
# ============================================================================

@login_required
def cainiao_forecast_pdf(request):
    """Gera PDF moderno da Previsão de Volume (reutiliza a lógica do forecast_cp4)."""
    from io import BytesIO
    from datetime import datetime as _dt_now
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm, mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            HRFlowable, KeepTogether,
        )
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.graphics import renderPDF
    except ImportError:
        return HttpResponse("reportlab não instalado", status=500)

    # ── Parâmetros (mesmos do AJAX) ──────────────────────────────────────────
    from .models import (
        CainiaoPlanningPackage, CainiaoHub, CainiaoOperationTask,
        CainiaoManualForecast,
    )
    from django.db.models import Count, Q, Max, Case, When, IntegerField
    from django.db.models.functions import Substr

    date_str  = request.GET.get("date", "")
    hub_id    = request.GET.get("hub_id", "")
    cp4_filter = set(request.GET.getlist("cp4"))

    try:
        from datetime import date as _date
        target_date = _date.fromisoformat(date_str)
    except (ValueError, TypeError):
        from datetime import date as _date
        target_date = _date.today()

    selected_hub = None
    hub_cp4_set = set()
    if hub_id:
        try:
            selected_hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(id=int(hub_id))
            hub_cp4_set = {c.cp4 for c in selected_hub.cp4_codes.all()}
        except (CainiaoHub.DoesNotExist, ValueError):
            pass

    # ── Forecast rows ────────────────────────────────────────────────────────
    qs = CainiaoPlanningPackage.objects.filter(operation_date=target_date)
    if hub_cp4_set:
        hq = Q()
        for cp4 in hub_cp4_set:
            hq |= Q(receiver_zip__startswith=cp4)
        qs = qs.filter(hq)

    raw_fc = (
        qs.exclude(receiver_zip="")
        .annotate(cp4=Substr("receiver_zip", 1, 4))
        .values("cp4", "receiver_city")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    fc_map = {}
    for r in raw_fc:
        cp4 = r["cp4"]
        if cp4 not in fc_map:
            fc_map[cp4] = {
                "cp4": cp4,
                "city": r["receiver_city"] or "",
                "forecast": 0, "manual_qty": 0,
            }
        fc_map[cp4]["forecast"] += r["total"]

    # ── Adicionar previsões MANUAIS (CainiaoManualForecast) ─────────────
    manual_qs = CainiaoManualForecast.objects.filter(
        operation_date=target_date,
    )
    if hub_cp4_set:
        manual_qs = manual_qs.filter(cp4__in=hub_cp4_set)
    for mf in manual_qs:
        if mf.cp4 not in fc_map:
            fc_map[mf.cp4] = {
                "cp4": mf.cp4, "city": "",
                "forecast": 0, "manual_qty": 0,
            }
        fc_map[mf.cp4]["forecast"] += mf.qty
        fc_map[mf.cp4]["manual_qty"] = (
            fc_map[mf.cp4].get("manual_qty", 0) + mf.qty
        )

    total_forecast = sum(v["forecast"] for v in fc_map.values())

    # ── Backlog rows ─────────────────────────────────────────────────────────
    most_recent_op = CainiaoOperationTask.objects.filter(
        task_date__lt=target_date
    ).aggregate(d=Max("task_date"))["d"]
    backlog_total = 0
    backlog_rows = []
    backlog_date_str = ""
    if most_recent_op:
        backlog_date_str = most_recent_op.strftime("%d/%m/%Y")
        bqs = CainiaoOperationTask.objects.filter(
            task_date=most_recent_op,
            task_status__in=["Driver_received", "Attempt Failure", "Unassign", "Assigned"],
        )
        if hub_cp4_set:
            bq = Q()
            for cp4 in hub_cp4_set:
                bq |= Q(zip_code__startswith=cp4)
            bqs = bqs.filter(bq)
        bl_raw = (
            bqs.exclude(zip_code="")
            .annotate(cp4=Substr("zip_code", 1, 4))
            .values("cp4", "destination_city")
            .annotate(
                total=Count("id"),
                driver_received=Count(Case(When(task_status="Driver_received", then=1), output_field=IntegerField())),
                attempt_failure=Count(Case(When(task_status="Attempt Failure", then=1), output_field=IntegerField())),
                unassign=Count(Case(When(task_status="Unassign", then=1), output_field=IntegerField())),
                assigned=Count(Case(When(task_status="Assigned", then=1), output_field=IntegerField())),
            )
            .order_by("-total")
        )
        bl_seen = {}
        for r in bl_raw:
            cp4 = r["cp4"]
            if cp4 not in bl_seen:
                bl_seen[cp4] = {"cp4": cp4, "city": r["destination_city"] or "", "total": 0, "dr": 0, "af": 0, "ua": 0, "as_": 0}
            bl_seen[cp4]["total"]  += r["total"]
            bl_seen[cp4]["dr"]     += r["driver_received"]
            bl_seen[cp4]["af"]     += r["attempt_failure"]
            bl_seen[cp4]["ua"]     += r["unassign"]
            bl_seen[cp4]["as_"]    += r["assigned"]
        backlog_rows = sorted(bl_seen.values(), key=lambda r: -r["total"])
        backlog_total = sum(r["total"] for r in backlog_rows)

    # ── Merge & apply CP4 filter ─────────────────────────────────────────────
    cp4_map = {}
    for r in fc_map.values():
        cp4_map[r["cp4"]] = {
            "cp4": r["cp4"], "city": r["city"],
            "forecast": r["forecast"], "backlog": 0,
            "manual_qty": r.get("manual_qty", 0),
        }
    for r in backlog_rows:
        if r["cp4"] not in cp4_map:
            cp4_map[r["cp4"]] = {
                "cp4": r["cp4"], "city": r["city"],
                "forecast": 0, "backlog": 0, "manual_qty": 0,
            }
        cp4_map[r["cp4"]]["backlog"] = r["total"]
        if not cp4_map[r["cp4"]]["city"]:
            cp4_map[r["cp4"]]["city"] = r["city"]

    merged = sorted(cp4_map.values(), key=lambda r: -(r["forecast"] + r["backlog"]))
    if cp4_filter:
        merged = [r for r in merged if r["cp4"] in cp4_filter]
        backlog_rows = [r for r in backlog_rows if r["cp4"] in cp4_filter]

    sum_fc  = sum(r["forecast"] for r in merged)
    sum_bl  = sum(r["backlog"]  for r in merged)
    sum_tot = sum_fc + sum_bl

    # ── PDF colours & styles ─────────────────────────────────────────────────
    INDIGO    = colors.HexColor("#4F46E5")
    INDIGO_LT = colors.HexColor("#EEF2FF")
    INDIGO_MD = colors.HexColor("#818CF8")
    ORANGE    = colors.HexColor("#F97316")
    ORANGE_LT = colors.HexColor("#FFF7ED")
    DARK      = colors.HexColor("#111827")
    GRAY_900  = colors.HexColor("#111827")
    GRAY_700  = colors.HexColor("#374151")
    GRAY_500  = colors.HexColor("#6B7280")
    GRAY_200  = colors.HexColor("#E5E7EB")
    GRAY_100  = colors.HexColor("#F3F4F6")
    WHITE     = colors.white
    ROW_ALT   = colors.HexColor("#F9FAFB")

    def style(name, **kw):
        base = ParagraphStyle(name, fontName=kw.pop("font", "Helvetica"),
                              fontSize=kw.pop("size", 8), leading=kw.pop("leading", 10),
                              textColor=kw.pop("color", DARK),
                              alignment=kw.pop("align", TA_LEFT), **kw)
        return base

    s_title    = style("T",   font="Helvetica-Bold", size=18, color=WHITE, align=TA_LEFT, leading=22)
    s_sub      = style("Sub", size=9, color=colors.HexColor("#C7D2FE"), align=TA_LEFT, leading=12)
    s_kpi_num  = style("KN",  font="Helvetica-Bold", size=22, color=DARK, align=TA_CENTER, leading=26)
    s_kpi_lbl  = style("KL",  size=7, color=GRAY_500, align=TA_CENTER, leading=9)
    s_kpi_num_o = style("KNO", font="Helvetica-Bold", size=22, color=ORANGE, align=TA_CENTER, leading=26)
    s_kpi_num_i = style("KNI", font="Helvetica-Bold", size=22, color=INDIGO, align=TA_CENTER, leading=26)
    s_th       = style("TH",  font="Helvetica-Bold", size=7, color=WHITE, align=TA_CENTER, leading=9)
    s_th_l     = style("THL", font="Helvetica-Bold", size=7, color=WHITE, align=TA_LEFT, leading=9)
    s_td       = style("TD",  size=7, color=GRAY_700, align=TA_CENTER, leading=9)
    s_td_l     = style("TDL", size=7, color=GRAY_700, align=TA_LEFT, leading=9)
    s_td_cp4   = style("TC",  font="Helvetica-Bold", size=8, color=INDIGO, align=TA_CENTER, leading=10)
    s_td_fc    = style("TF",  font="Helvetica-Bold", size=8, color=INDIGO, align=TA_CENTER, leading=10)
    s_td_bl    = style("TB",  font="Helvetica-Bold", size=8, color=ORANGE, align=TA_CENTER, leading=10)
    s_td_tot   = style("TT",  font="Helvetica-Bold", size=8, color=DARK, align=TA_CENTER, leading=10)
    s_tfoot    = style("TF2", font="Helvetica-Bold", size=8, color=WHITE, align=TA_CENTER, leading=10)
    s_sec      = style("SEC", font="Helvetica-Bold", size=9, color=ORANGE, align=TA_LEFT, leading=11)
    s_th_o     = style("THO", font="Helvetica-Bold", size=7, color=WHITE, align=TA_CENTER, leading=9)
    s_td_dr    = style("DR",  font="Helvetica-Bold", size=8, color=colors.HexColor("#D97706"), align=TA_CENTER, leading=10)
    s_td_af    = style("AF",  font="Helvetica-Bold", size=8, color=colors.HexColor("#DC2626"), align=TA_CENTER, leading=10)

    # ── Mini bar helper ───────────────────────────────────────────────────────
    def make_bar(forecast, backlog, max_total, bar_w=3.5*cm, bar_h=5*mm):
        d = Drawing(bar_w, bar_h)
        if max_total == 0:
            return d
        fc_w = (forecast / max_total) * float(bar_w)
        bl_w = (backlog  / max_total) * float(bar_w)
        # background
        d.add(Rect(0, 0, bar_w, bar_h, fillColor=GRAY_200, strokeColor=None))
        if fc_w > 0:
            d.add(Rect(0, 0, fc_w, bar_h, fillColor=INDIGO_MD, strokeColor=None))
        if bl_w > 0:
            d.add(Rect(fc_w, 0, bl_w, bar_h, fillColor=ORANGE, strokeColor=None))
        return d

    # ── Build document ────────────────────────────────────────────────────────
    buffer = BytesIO()
    PAGE_W, PAGE_H = A4
    margin = 1.5 * cm
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin)

    usable_w = PAGE_W - 2 * margin
    elements = []

    # ── HEADER BAND ──────────────────────────────────────────────────────────
    hub_label = selected_hub.name if selected_hub else "Todos os HUBs"
    cp4_label = f" · CP4: {', '.join(sorted(cp4_filter))}" if cp4_filter else ""
    date_label = target_date.strftime("%d/%m/%Y")

    hdr = Table([[
        Paragraph("Previsão de Volume", s_title),
        Paragraph(f"{date_label} · {hub_label}{cp4_label}", s_sub),
    ]], colWidths=[usable_w * 0.6, usable_w * 0.4])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), INDIGO),
        ("TOPPADDING",    (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("LEFTPADDING",   (0, 0), (0, -1), 16),
        ("RIGHTPADDING",  (-1, 0), (-1, -1), 16),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [6, 6, 0, 0]),
    ]))
    elements.append(hdr)

    # ── KPI ROW ───────────────────────────────────────────────────────────────
    kpi_col = usable_w / 3
    kpi_data = [[
        Paragraph(str(sum_fc),  s_kpi_num),
        Paragraph(str(sum_bl),  s_kpi_num_o),
        Paragraph(str(sum_tot), s_kpi_num_i),
    ], [
        Paragraph("Previsão",         s_kpi_lbl),
        Paragraph("Backlog",          s_kpi_lbl),
        Paragraph("Estimativa Total", s_kpi_lbl),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=[kpi_col] * 3)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GRAY_100),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEAFTER",  (0, 0), (1, -1), 0.5, GRAY_200),
        ("LINEBELOW",  (0, -1), (-1, -1), 1.5, INDIGO),
        ("ROUNDEDCORNERS", [0, 0, 6, 6]),
    ]))
    elements.append(kpi_tbl)
    elements.append(Spacer(1, 0.4 * cm))

    # ── MAIN TABLE ────────────────────────────────────────────────────────────
    max_total = max((r["forecast"] + r["backlog"] for r in merged), default=1)
    COL_W = [0.7*cm, 1.5*cm, usable_w*0.25, 3.6*cm, 1.5*cm, 1.5*cm, 1.5*cm]

    header_row = [
        Paragraph("#",            s_th),
        Paragraph("CP4",          s_th),
        Paragraph("Cidade / Zona",s_th_l),
        Paragraph("Distribuição", s_th),
        Paragraph("Prev.",        s_th),
        Paragraph("Backlog",      s_th),
        Paragraph("Total",        s_th),
    ]
    table_data = [header_row]

    for i, row in enumerate(merged):
        row_total = row["forecast"] + row["backlog"]
        # PDF: não mostrar indicador manual (totais já incluem)
        table_data.append([
            Paragraph(str(i + 1),            s_td),
            Paragraph(row["cp4"],            s_td_cp4),
            Paragraph((row["city"] or "—").upper(), s_td_l),
            make_bar(row["forecast"], row["backlog"], max_total),
            Paragraph(str(row["forecast"]) if row["forecast"] else "—", s_td_fc),
            Paragraph(str(row["backlog"])  if row["backlog"]  else "—", s_td_bl),
            Paragraph(str(row_total),        s_td_tot),
        ])

    # Footer totals row
    table_data.append([
        Paragraph("Total", s_tfoot),
        Paragraph("",      s_tfoot),
        Paragraph("",      s_tfoot),
        Paragraph("",      s_tfoot),
        Paragraph(str(sum_fc),  s_tfoot),
        Paragraph(str(sum_bl),  s_tfoot),
        Paragraph(str(sum_tot), s_tfoot),
    ])

    main_tbl = Table(table_data, colWidths=COL_W, repeatRows=1)
    n_rows = len(table_data)
    style_cmds = [
        ("BACKGROUND",    (0, 0),  (-1, 0),       INDIGO),
        ("BACKGROUND",    (0, -1), (-1, -1),       GRAY_700),
        ("TOPPADDING",    (0, 0),  (-1, -1),       5),
        ("BOTTOMPADDING", (0, 0),  (-1, -1),       5),
        ("LEFTPADDING",   (0, 0),  (-1, -1),       4),
        ("RIGHTPADDING",  (0, 0),  (-1, -1),       4),
        ("ALIGN",         (0, 0),  (-1, -1),       "CENTER"),
        ("ALIGN",         (2, 1),  (2, -2),        "LEFT"),
        ("VALIGN",        (0, 0),  (-1, -1),       "MIDDLE"),
        ("GRID",          (0, 0),  (-1, -1),       0.3, GRAY_200),
        ("LINEBELOW",     (0, 0),  (-1, 0),        1,   INDIGO),
        ("LINEABOVE",     (0, -1), (-1, -1),       1.5, GRAY_500),
        ("ROWBACKGROUNDS",(0, 1),  (-1, -2),       [WHITE, ROW_ALT]),
    ]
    main_tbl.setStyle(TableStyle(style_cmds))
    elements.append(main_tbl)

    # ── BACKLOG DETAIL ────────────────────────────────────────────────────────
    if backlog_rows:
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=ORANGE, spaceAfter=6))
        elements.append(Paragraph(
            f"Detalhe do Backlog por CP4  ·  operação de {backlog_date_str}", s_sec))
        elements.append(Spacer(1, 0.2 * cm))

        BCOL_W = [0.6*cm, 1.3*cm, usable_w*0.25, 1.6*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.5*cm]
        s_td_as = style("AS", font="Helvetica-Bold", size=8, color=colors.HexColor("#2563EB"), align=TA_CENTER, leading=10)
        bl_header = [
            Paragraph("#",             s_th_o),
            Paragraph("CP4",           s_th_o),
            Paragraph("Cidade / Zona", s_th_o),
            Paragraph("Em Rota",       s_th_o),
            Paragraph("Falhou",        s_th_o),
            Paragraph("Atribuído",     s_th_o),
            Paragraph("Não Atrib.",    s_th_o),
            Paragraph("Total",         s_th_o),
        ]
        bl_data = [bl_header]
        for i, r in enumerate(backlog_rows):
            bl_data.append([
                Paragraph(str(i + 1),                 s_td),
                Paragraph(r["cp4"],                   style("TC2", font="Helvetica-Bold", size=8, color=ORANGE, align=TA_CENTER, leading=10)),
                Paragraph((r["city"] or "—").upper(), s_td_l),
                Paragraph(str(r["dr"]),               s_td_dr),
                Paragraph(str(r["af"]),               s_td_af),
                Paragraph(str(r["as_"]),              s_td_as),
                Paragraph(str(r["ua"]),               s_td),
                Paragraph(str(r["total"]),            style("BT", font="Helvetica-Bold", size=8, color=ORANGE, align=TA_CENTER, leading=10)),
            ])
        bl_data.append([
            Paragraph("Total", s_tfoot),
            Paragraph("", s_tfoot),
            Paragraph("", s_tfoot),
            Paragraph(str(sum(r["dr"]    for r in backlog_rows)), s_tfoot),
            Paragraph(str(sum(r["af"]    for r in backlog_rows)), s_tfoot),
            Paragraph(str(sum(r["as_"]   for r in backlog_rows)), s_tfoot),
            Paragraph(str(sum(r["ua"]    for r in backlog_rows)), s_tfoot),
            Paragraph(str(sum(r["total"] for r in backlog_rows)), s_tfoot),
        ])
        bl_tbl = Table(bl_data, colWidths=BCOL_W, repeatRows=1)
        bl_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0),  (-1, 0),  ORANGE),
            ("BACKGROUND",    (0, -1), (-1, -1), GRAY_700),
            ("ROWBACKGROUNDS",(0, 1),  (-1, -2), [WHITE, ORANGE_LT]),
            ("TOPPADDING",    (0, 0),  (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 5),
            ("LEFTPADDING",   (0, 0),  (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0),  (-1, -1), 4),
            ("ALIGN",         (0, 0),  (-1, -1), "CENTER"),
            ("ALIGN",         (2, 1),  (2, -2),  "LEFT"),
            ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
            ("GRID",          (0, 0),  (-1, -1), 0.3, GRAY_200),
            ("LINEBELOW",     (0, 0),  (-1, 0),  1,   ORANGE),
            ("LINEABOVE",     (0, -1), (-1, -1), 1.5, GRAY_500),
        ]))
        elements.append(bl_tbl)

    # ── FOOTER ────────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.5 * cm))
    generated = _dt_now.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(
        f"<font color='#9CA3AF'>Leguas Franzinas · Gerado em {generated}</font>",
        style("FT", size=7, align=TA_RIGHT, color=GRAY_500)))

    doc.build(elements)
    buffer.seek(0)
    fname = f"previsao_{date_label.replace('/', '-')}_{hub_label.replace(' ', '_')}.pdf"
    resp = HttpResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ============================================================================
# Helpers partilhados pelos 4 novos importadores
# ============================================================================

def _load_workbook_rows(file_bytes):
    """Abre xlsx e devolve todas as linhas da sheet activa."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _find_header_row(all_rows, required_col):
    """Procura a primeira linha que contenha `required_col` (case-insensitive).
    Devolve (header_list, header_idx) ou (None, -1)."""
    needle = required_col.lower()
    for i, row in enumerate(all_rows[:20]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(needle in c.lower() for c in cells):
            return cells, i
    return None, -1


def _col_idx(header, *names):
    """Devolve índice da coluna cujo nome corresponde a `names`.

    Estratégia:
    1. Match exacto (case/whitespace-insensitive) — PREFERIDO, evita que
       "Delivery Time" case com "Start Delivery Time".
    2. Fallback para substring match se não houver exacto.
    """
    # 1) exact (case-insensitive, stripped)
    for name in names:
        target = name.strip().lower()
        for i, h in enumerate(header):
            if h and str(h).strip().lower() == target:
                return i
    # 2) substring fallback
    for name in names:
        nl = name.strip().lower()
        for i, h in enumerate(header):
            if h and nl in str(h).lower():
                return i
    return None


def _parse_datetime(value):
    """Converte datetime nativo ou string ISO em datetime aware; None em caso de erro."""
    from django.utils.timezone import make_aware, is_aware
    if value is None:
        return None
    if isinstance(value, _dt):
        return value if is_aware(value) else make_aware(value)
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            naive = _dt.strptime(s, fmt)
            return make_aware(naive)
        except ValueError:
            pass
    return None


def _cell_str(row, idx):
    """Valor de uma célula como string limpa; '' se idx=None ou célula vazia."""
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _normalise_zip(z: str) -> str:
    """Normaliza código postal: '4935.0' → '4935', '3700-013' → '3700-013'."""
    if z.endswith(".0") and z[:-2].isdigit():
        return z[:-2]
    return z


def _cell_float(row, idx):
    """Valor de uma célula como float; None se inválido."""
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ============================================================================
# Forecast Excel export
# ============================================================================

@login_required
def cainiao_forecast_excel(request):
    """Gera XLSX da Previsão de Volume (mesma lógica do PDF/AJAX)."""
    from io import BytesIO
    from datetime import date as _date_cls
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl não instalado", status=500)

    from .models import (
        CainiaoPlanningPackage, CainiaoHub, CainiaoOperationTask,
        CainiaoManualForecast,
    )
    from django.db.models import Count, Q, Max, Case, When, IntegerField
    from django.db.models.functions import Substr

    date_str   = request.GET.get("date", "")
    hub_id     = request.GET.get("hub_id", "")
    cp4_filter = set(request.GET.getlist("cp4"))

    try:
        target_date = _date_cls.fromisoformat(date_str)
    except (ValueError, TypeError):
        target_date = _date_cls.today()

    selected_hub = None
    hub_cp4_set = set()
    if hub_id:
        try:
            selected_hub = CainiaoHub.objects.prefetch_related(
                "cp4_codes"
            ).get(id=int(hub_id))
            hub_cp4_set = {c.cp4 for c in selected_hub.cp4_codes.all()}
        except (CainiaoHub.DoesNotExist, ValueError):
            pass

    # ── Forecast (CainiaoPlanningPackage) ────────────────────────────────
    qs = CainiaoPlanningPackage.objects.filter(operation_date=target_date)
    if hub_cp4_set:
        hq = Q()
        for cp4 in hub_cp4_set:
            hq |= Q(receiver_zip__startswith=cp4)
        qs = qs.filter(hq)

    raw_fc = (
        qs.exclude(receiver_zip="")
        .annotate(cp4=Substr("receiver_zip", 1, 4))
        .values("cp4", "receiver_city")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    fc_map = {}
    for r in raw_fc:
        cp4 = r["cp4"]
        if cp4 not in fc_map:
            fc_map[cp4] = {
                "cp4": cp4,
                "city": r["receiver_city"] or "",
                "forecast": 0, "manual_qty": 0,
            }
        fc_map[cp4]["forecast"] += r["total"]

    # ── Adicionar previsões MANUAIS ──────────────────────────────────────
    manual_qs = CainiaoManualForecast.objects.filter(
        operation_date=target_date,
    )
    if hub_cp4_set:
        manual_qs = manual_qs.filter(cp4__in=hub_cp4_set)
    for mf in manual_qs:
        if mf.cp4 not in fc_map:
            fc_map[mf.cp4] = {
                "cp4": mf.cp4, "city": "",
                "forecast": 0, "manual_qty": 0,
            }
        fc_map[mf.cp4]["forecast"] += mf.qty
        fc_map[mf.cp4]["manual_qty"] = (
            fc_map[mf.cp4].get("manual_qty", 0) + mf.qty
        )

    # ── Backlog (CainiaoOperationTask, dia anterior ao target) ───────────
    most_recent_op = CainiaoOperationTask.objects.filter(
        task_date__lt=target_date
    ).aggregate(d=Max("task_date"))["d"]
    bl_map = {}
    backlog_date_str = ""
    if most_recent_op:
        backlog_date_str = most_recent_op.strftime("%d/%m/%Y")
        bqs = CainiaoOperationTask.objects.filter(
            task_date=most_recent_op,
            task_status__in=[
                "Driver_received", "Attempt Failure", "Unassign", "Assigned"
            ],
        )
        if hub_cp4_set:
            bq = Q()
            for cp4 in hub_cp4_set:
                bq |= Q(zip_code__startswith=cp4)
            bqs = bqs.filter(bq)
        bl_raw = (
            bqs.exclude(zip_code="")
            .annotate(cp4=Substr("zip_code", 1, 4))
            .values("cp4", "destination_city")
            .annotate(
                total=Count("id"),
                driver_received=Count(Case(When(task_status="Driver_received",
                                                then=1), output_field=IntegerField())),
                attempt_failure=Count(Case(When(task_status="Attempt Failure",
                                                then=1), output_field=IntegerField())),
                unassign=Count(Case(When(task_status="Unassign",
                                         then=1), output_field=IntegerField())),
                assigned=Count(Case(When(task_status="Assigned",
                                         then=1), output_field=IntegerField())),
            )
            .order_by("-total")
        )
        for r in bl_raw:
            cp4 = r["cp4"]
            if cp4 not in bl_map:
                bl_map[cp4] = {
                    "cp4": cp4,
                    "city": r["destination_city"] or "",
                    "total": 0, "driver_received": 0, "attempt_failure": 0,
                    "unassign": 0, "assigned": 0,
                }
            bl_map[cp4]["total"]            += r["total"]
            bl_map[cp4]["driver_received"]  += r["driver_received"]
            bl_map[cp4]["attempt_failure"]  += r["attempt_failure"]
            bl_map[cp4]["unassign"]         += r["unassign"]
            bl_map[cp4]["assigned"]         += r["assigned"]

    # ── Merge (CP4 union) e aplicar filtro CP4 visível ───────────────────
    all_cp4 = set(fc_map) | set(bl_map)
    if cp4_filter:
        all_cp4 &= cp4_filter
    rows = []
    for cp4 in sorted(all_cp4, key=lambda c: -(
        (fc_map.get(c, {}).get("forecast", 0)) +
        (bl_map.get(c, {}).get("total", 0))
    )):
        f = fc_map.get(cp4, {})
        b = bl_map.get(cp4, {})
        rows.append({
            "cp4": cp4,
            "city": (f.get("city") or b.get("city") or ""),
            "forecast":        f.get("forecast", 0),
            "manual_qty":      f.get("manual_qty", 0),
            "driver_received": b.get("driver_received", 0),
            "attempt_failure": b.get("attempt_failure", 0),
            "assigned":        b.get("assigned", 0),
            "unassign":        b.get("unassign", 0),
            "backlog":         b.get("total", 0),
            "estimativa":      f.get("forecast", 0) + b.get("total", 0),
        })

    total_forecast   = sum(r["forecast"] for r in rows)
    total_backlog    = sum(r["backlog"] for r in rows)
    total_estimativa = total_forecast + total_backlog

    # ── Build XLSX ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    # Estilos reutilizáveis
    title_font  = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill  = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    total_font  = Font(name="Calibri", size=11, bold=True)
    total_fill  = PatternFill("solid", fgColor="EEF2FF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ── Sheet 1: Resumo ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Previsão de Volume — Cainiao"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Data prevista:"
    ws["B3"] = target_date.strftime("%d/%m/%Y")
    ws["A4"] = "HUB:"
    ws["B4"] = selected_hub.name if selected_hub else "Todos"
    ws["A5"] = "Backlog do dia:"
    ws["B5"] = backlog_date_str or "—"
    ws["A6"] = "CP4 filtrados:"
    ws["B6"] = ", ".join(sorted(cp4_filter)) if cp4_filter else "Todos"
    for r in range(3, 7):
        ws[f"A{r}"].font = Font(bold=True)

    ws["A8"] = "Previsão (forecast)"
    ws["B8"] = total_forecast
    ws["A9"] = "Backlog"
    ws["B9"] = total_backlog
    ws["A10"] = "Estimativa total"
    ws["B10"] = total_estimativa
    for r in range(8, 11):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"].font = Font(bold=True, size=12)
        ws[f"B{r}"].alignment = right_align
    ws["B10"].fill = PatternFill("solid", fgColor="DBEAFE")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2: Detalhe por CP4 ────────────────────────────────────────
    ws2 = wb.create_sheet("Detalhe por CP4")
    headers = [
        "#", "CP4", "Cidade / Zona", "Previsão", "Manual",
        "Em Rota", "Falhou", "Atribuído", "Não Atrib.",
        "Backlog", "Estimativa Total",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    # Cor para destacar entradas com manual_qty > 0
    manual_fill = PatternFill("solid", fgColor="D1FAE5")  # emerald-100
    for r_idx, row in enumerate(rows, start=2):
        ws2.cell(row=r_idx, column=1, value=r_idx - 1).alignment = center
        ws2.cell(row=r_idx, column=2, value=row["cp4"]).font = Font(name="Consolas")
        ws2.cell(row=r_idx, column=3, value=row["city"]).alignment = left_align
        ws2.cell(row=r_idx, column=4, value=row["forecast"]).alignment = right_align
        # Coluna Manual — destaca em verde quando >0
        manual_cell = ws2.cell(row=r_idx, column=5, value=row.get("manual_qty", 0))
        manual_cell.alignment = right_align
        if row.get("manual_qty", 0) > 0:
            manual_cell.fill = manual_fill
            manual_cell.font = Font(bold=True, color="065F46")
        ws2.cell(row=r_idx, column=6, value=row["driver_received"]).alignment = right_align
        ws2.cell(row=r_idx, column=7, value=row["attempt_failure"]).alignment = right_align
        ws2.cell(row=r_idx, column=8, value=row["assigned"]).alignment = right_align
        ws2.cell(row=r_idx, column=9, value=row["unassign"]).alignment = right_align
        ws2.cell(row=r_idx, column=10, value=row["backlog"]).alignment = right_align
        ws2.cell(row=r_idx, column=11, value=row["estimativa"]).alignment = right_align
        for col_idx in range(1, 12):
            ws2.cell(row=r_idx, column=col_idx).border = border
    # Linha de totais
    if rows:
        tr = len(rows) + 2
        ws2.cell(row=tr, column=1, value="TOTAL").font = total_font
        ws2.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
        ws2.cell(row=tr, column=1).alignment = center
        ws2.cell(row=tr, column=4, value=total_forecast)
        ws2.cell(row=tr, column=5, value=sum(r.get("manual_qty", 0) for r in rows))
        ws2.cell(row=tr, column=6, value=sum(r["driver_received"] for r in rows))
        ws2.cell(row=tr, column=7, value=sum(r["attempt_failure"] for r in rows))
        ws2.cell(row=tr, column=8, value=sum(r["assigned"] for r in rows))
        ws2.cell(row=tr, column=9, value=sum(r["unassign"] for r in rows))
        ws2.cell(row=tr, column=10, value=total_backlog)
        ws2.cell(row=tr, column=11, value=total_estimativa)
        for col_idx in range(1, 12):
            cell = ws2.cell(row=tr, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            if col_idx >= 4:
                cell.alignment = right_align

    # Auto-fit width estimado
    widths = [5, 8, 28, 11, 9, 10, 10, 12, 12, 11, 16]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Stream em memória
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"forecast-cainiao-{target_date.isoformat()}"
    if selected_hub:
        fname += f"-{selected_hub.name.replace(' ', '_').lower()}"
    fname += ".xlsx"

    resp = HttpResponse(
        buf.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ============================================================================
# IMPORT — PLANILHA FORECAST (nova — Parcel ID, 27 colunas)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_planning_preview(request):
    """Preview da planilha Forecast (nova) — sem gravar na BD."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    operation_date_str = request.POST.get("operation_date", "").strip()
    if not operation_date_str:
        return JsonResponse({"success": False, "error": "Data da operação é obrigatória."}, status=400)

    try:
        op_date = _date.fromisoformat(operation_date_str)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data inválida."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Parcel ID")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Parcel ID' não encontrada."}, status=400)

    ci_parcel = _col_idx(header, "Parcel ID")
    ci_city   = _col_idx(header, "receiverCity")
    ci_zip    = _col_idx(header, "receiverZipCode")
    ci_name   = _col_idx(header, "receiverName")

    from .models import CainiaoPlanningPackage
    existing_keys = set(
        CainiaoPlanningPackage.objects.filter(operation_date=op_date)
        .values_list("parcel_id", flat=True)
    )

    total_novos = 0
    total_atualizados = 0
    sample = []
    seen = set()

    for row in all_rows[header_idx + 1:]:
        parcel_id = _cell_str(row, ci_parcel)
        if not parcel_id:
            continue
        if parcel_id in seen:
            continue
        seen.add(parcel_id)

        is_new = parcel_id not in existing_keys
        if is_new:
            total_novos += 1
        else:
            total_atualizados += 1

        if len(sample) < 10:
            sample.append({
                "parcel_id": parcel_id,
                "city": _cell_str(row, ci_city),
                "zip": _cell_str(row, ci_zip),
                "receiver": _cell_str(row, ci_name),
                "action": "new" if is_new else "update",
            })

    return JsonResponse({
        "success": True,
        "total_novos": total_novos,
        "total_atualizados": total_atualizados,
        "total": total_novos + total_atualizados,
        "operation_date": operation_date_str,
        "sample": sample,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_planning_import(request):
    """Import da planilha Forecast (nova) — upsert por (operation_date, parcel_id)."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    operation_date_str = request.POST.get("operation_date", "").strip()
    if not operation_date_str:
        return JsonResponse({"success": False, "error": "Data da operação é obrigatória."}, status=400)

    try:
        op_date = _date.fromisoformat(operation_date_str)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data inválida."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Parcel ID")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Parcel ID' não encontrada."}, status=400)

    ci = {
        "parcel":         _col_idx(header, "Parcel ID"),
        "lp_code":        _col_idx(header, "LP code"),
        "province":       _col_idx(header, "receiverProvince"),
        "city":           _col_idx(header, "receiverCity"),
        "zip":            _col_idx(header, "receiverZipCode"),
        "address":        _col_idx(header, "receiverDetailAddress"),
        "name":           _col_idx(header, "receiverName"),
        "phone":          _col_idx(header, "receiverPhoneNumber"),
        "email":          _col_idx(header, "recEmail"),
        "order_created":  _col_idx(header, "orderCreationTime"),
        "sc_consol":      _col_idx(header, "scConsolidationTime"),
        "sc_outbound":    _col_idx(header, "scOutboundTime"),
        "actual_inbound": _col_idx(header, "actualInboundTime"),
        "task_accept":    _col_idx(header, "Task aceptance"),
        "sign_time":      _col_idx(header, "signTime"),
        "exc_type":       _col_idx(header, "exceptionType"),
        "exc_reason":     _col_idx(header, "exceptionReason"),
        "last_exc":       _col_idx(header, "lastExceptionTime"),
        "hub":            _col_idx(header, "HUB"),
        "dsp":            _col_idx(header, "DSP"),
        "creation":       _col_idx(header, "Creation time"),
        "seller":         _col_idx(header, "Seller name"),
        "inbound":        _col_idx(header, "Inbound time"),
        "assign":         _col_idx(header, "Assign time"),
        "del_success":    _col_idx(header, "Delivery success"),
        "del_fail":       _col_idx(header, "Delivery fail"),
        "pickup":         _col_idx(header, "Pick-up time"),
    }

    from .models import CainiaoPlanningBatch, CainiaoPlanningPackage

    batch = CainiaoPlanningBatch.objects.create(
        filename=ficheiro.name,
        operation_date=op_date,
        created_by=request.user,
    )

    existing = {
        pkg.parcel_id: pkg
        for pkg in CainiaoPlanningPackage.objects.filter(operation_date=op_date)
    }

    to_create = []
    to_update = []
    seen = set()
    update_fields = [
        "lp_code", "receiver_province", "receiver_city", "receiver_zip",
        "receiver_address", "receiver_name", "receiver_phone", "receiver_email",
        "order_creation_time", "sc_consolidation_time", "sc_outbound_time",
        "actual_inbound_time", "task_acceptance_time", "sign_time",
        "exception_type", "exception_reason", "last_exception_time",
        "hub", "dsp", "creation_time", "seller_name", "inbound_time",
        "assign_time", "delivery_success_time", "delivery_fail_time",
        "pickup_time", "last_import_batch",
    ]

    for row in all_rows[header_idx + 1:]:
        parcel_id = _cell_str(row, ci["parcel"])
        if not parcel_id or parcel_id in seen:
            continue
        seen.add(parcel_id)

        fields = {
            "lp_code":              _cell_str(row, ci["lp_code"]),
            "receiver_province":    _cell_str(row, ci["province"]),
            "receiver_city":        _cell_str(row, ci["city"]),
            "receiver_zip":         _cell_str(row, ci["zip"]),
            "receiver_address":     _cell_str(row, ci["address"]),
            "receiver_name":        _cell_str(row, ci["name"]),
            "receiver_phone":       _cell_str(row, ci["phone"]),
            "receiver_email":       _cell_str(row, ci["email"]),
            "order_creation_time":  _parse_datetime(row[ci["order_created"]] if ci["order_created"] is not None else None),
            "sc_consolidation_time": _parse_datetime(row[ci["sc_consol"]] if ci["sc_consol"] is not None else None),
            "sc_outbound_time":     _parse_datetime(row[ci["sc_outbound"]] if ci["sc_outbound"] is not None else None),
            "actual_inbound_time":  _parse_datetime(row[ci["actual_inbound"]] if ci["actual_inbound"] is not None else None),
            "task_acceptance_time": _parse_datetime(row[ci["task_accept"]] if ci["task_accept"] is not None else None),
            "sign_time":            _parse_datetime(row[ci["sign_time"]] if ci["sign_time"] is not None else None),
            "exception_type":       _cell_str(row, ci["exc_type"]),
            "exception_reason":     _cell_str(row, ci["exc_reason"]),
            "last_exception_time":  _parse_datetime(row[ci["last_exc"]] if ci["last_exc"] is not None else None),
            "hub":                  _cell_str(row, ci["hub"]),
            "dsp":                  _cell_str(row, ci["dsp"]),
            "creation_time":        _parse_datetime(row[ci["creation"]] if ci["creation"] is not None else None),
            "seller_name":          _cell_str(row, ci["seller"]),
            "inbound_time":         _parse_datetime(row[ci["inbound"]] if ci["inbound"] is not None else None),
            "assign_time":          _parse_datetime(row[ci["assign"]] if ci["assign"] is not None else None),
            "delivery_success_time": _parse_datetime(row[ci["del_success"]] if ci["del_success"] is not None else None),
            "delivery_fail_time":   _parse_datetime(row[ci["del_fail"]] if ci["del_fail"] is not None else None),
            "pickup_time":          _parse_datetime(row[ci["pickup"]] if ci["pickup"] is not None else None),
            "last_import_batch":    batch,
        }

        if parcel_id in existing:
            pkg = existing[parcel_id]
            for k, v in fields.items():
                setattr(pkg, k, v)
            to_update.append(pkg)
        else:
            to_create.append(CainiaoPlanningPackage(
                operation_date=op_date, parcel_id=parcel_id, **fields
            ))
            existing[parcel_id] = to_create[-1]

    if to_create:
        CainiaoPlanningPackage.objects.bulk_create(to_create, batch_size=500)
    if to_update:
        CainiaoPlanningPackage.objects.bulk_update(to_update, update_fields, batch_size=500)

    batch.total_packages = len(to_create)
    batch.new_packages = len(to_create)
    batch.updated_packages = len(to_update)
    batch.save(update_fields=["total_packages", "new_packages", "updated_packages"])

    return JsonResponse({
        "success": True,
        "total_novos": len(to_create),
        "total_atualizados": len(to_update),
        "operation_date": operation_date_str,
        "batch_id": batch.id,
    })


# ============================================================================
# IMPORT — PLANILHA OPERATION UPDATE (EPOD Task List)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_operation_preview(request):
    """Preview da planilha Operation Update — data lida automaticamente de cada linha."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Waybill Number' não encontrada."}, status=400)

    ci_waybill  = _col_idx(header, "Waybill Number")
    ci_status   = _col_idx(header, "Task Status")
    ci_courier  = _col_idx(header, "Courier Name")
    ci_del_time = _col_idx(header, "Delivery Time")
    ci_taskdate = _col_idx(header, "Task Date")

    _STATUS_PRIORITY = {"Delivered": 4, "Attempt Failure": 3, "Driver_received": 2, "Assigned": 1, "Unassign": 0}

    def _parse_row_date(row):
        if ci_taskdate is None or ci_taskdate >= len(row):
            return None
        v = row[ci_taskdate]
        if isinstance(v, _date) and not isinstance(v, _dt):
            return v
        if isinstance(v, _dt):
            return v.date()
        s = str(v).strip()[:10] if v else ""
        try:
            return _date.fromisoformat(s)
        except ValueError:
            return None

    # Group by (task_date, waybill)
    from collections import defaultdict
    date_wb_rows = defaultdict(lambda: defaultdict(list))
    for row in all_rows[header_idx + 1:]:
        waybill = _cell_str(row, ci_waybill)
        task_date = _parse_row_date(row)
        if waybill and task_date:
            date_wb_rows[task_date][waybill].append(row)

    if not date_wb_rows:
        return JsonResponse({"success": False, "error": "Nenhuma linha com Waybill Number e Task Date válidos."}, status=400)

    from .models import CainiaoOperationTask
    all_dates = sorted(date_wb_rows.keys())
    existing_keys_by_date = {
        d: set(
            CainiaoOperationTask.objects.filter(task_date=d)
            .values_list("waybill_number", flat=True)
        )
        for d in all_dates
    }

    status_counts = {}
    courier_counts = {}
    date_counts = {}
    total_novos = 0
    total_atualizados = 0
    sample = []

    for task_date in all_dates:
        existing_keys = existing_keys_by_date[task_date]
        wb_rows = date_wb_rows[task_date]
        date_novos = date_atualizados = 0

        for waybill, r_list in wb_rows.items():
            # Sem cancel exclusion: cada linha (Cancel inclusive) é uma
            # assinatura legítima da cadeia. Best row = última cronologicamente.
            active_rows = r_list
            best_score = (-1, -1)
            best_row = None
            for row in active_rows:
                status = _cell_str(row, ci_status)
                priority = _STATUS_PRIORITY.get(status, 0)
                has_del_time = 1 if (ci_del_time is not None and row[ci_del_time]) else 0
                score = (priority, has_del_time)
                if score > best_score:
                    best_score = score
                    best_row = row

            status_val = _cell_str(best_row, ci_status)
            courier    = _cell_str(best_row, ci_courier)
            status_counts[status_val] = status_counts.get(status_val, 0) + 1
            if courier:
                courier_counts[courier] = courier_counts.get(courier, 0) + 1

            is_new = waybill not in existing_keys
            if is_new:
                total_novos += 1
                date_novos += 1
            else:
                total_atualizados += 1
                date_atualizados += 1

            if len(sample) < 10:
                sample.append({
                    "data": str(task_date),
                    "waybill": waybill,
                    "status": status_val,
                    "courier": courier,
                    "action": "new" if is_new else "update",
                })

        date_counts[str(task_date)] = {"novos": date_novos, "atualizados": date_atualizados}

    date_range = f"{all_dates[0]} → {all_dates[-1]}" if len(all_dates) > 1 else str(all_dates[0])

    return JsonResponse({
        "success": True,
        "total_novos": total_novos,
        "total_atualizados": total_atualizados,
        "total": total_novos + total_atualizados,
        "date_range": date_range,
        "dates_found": len(all_dates),
        "date_counts": date_counts,
        "status_counts": status_counts,
        "courier_counts": courier_counts,
        "sample": sample,
    })


@login_required
@require_http_methods(["POST"])
@login_required
@require_http_methods(["POST"])
def cainiao_operation_audit(request):
    """Audita uma planilha Operation Update SEM IMPORTAR — para o
    utilizador verificar antes o que vai entrar no sistema, e
    procurar waybills específicos.

    Body multipart:
      ficheiro: <xlsx>
      waybills_search: str opcional (csv ou linhas)
        Para cada waybill listado, mostra TODAS as linhas do
        ficheiro com aquele waybill, com status/data/courier.
    """
    from collections import defaultdict
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."},
            status=400,
        )
    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro ao ler ficheiro: {e}"},
            status=400,
        )
    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        return JsonResponse(
            {"success": False,
             "error": "Coluna 'Waybill Number' não encontrada."},
            status=400,
        )
    ci = {
        "waybill": _col_idx(header, "Waybill Number"),
        "task_status": _col_idx(header, "Task Status"),
        "courier_name": _col_idx(header, "Courier Name"),
        "taskdate": _col_idx(header, "Task Date"),
        "del_time": _col_idx(header, "Delivery Time"),
        "lp": _col_idx(header, "LP No."),
        "zip_code": _col_idx(header, "Zip Code"),
    }

    def _parse_row_date(row):
        ci_td = ci["taskdate"]
        if ci_td is None or ci_td >= len(row):
            return None
        v = row[ci_td]
        if isinstance(v, _date) and not isinstance(v, _dt):
            return v
        if isinstance(v, _dt):
            return v.date()
        s = str(v).strip()[:10] if v else ""
        try:
            return _date.fromisoformat(s)
        except ValueError:
            return None

    # Sumário geral
    total = 0
    by_status = defaultdict(int)
    by_date = defaultdict(int)
    waybills_seen = set()
    no_waybill = 0
    no_taskdate = 0

    # Procura por waybills específicos
    search_input = (request.POST.get("waybills_search") or "").strip()
    search_targets = set()
    if search_input:
        for s in search_input.replace(",", "\n").split("\n"):
            s = s.strip()
            if s:
                search_targets.add(s)
    search_results = defaultdict(list)

    for row in all_rows[header_idx + 1:]:
        total += 1
        wb = _cell_str(row, ci["waybill"])
        if not wb:
            no_waybill += 1
            continue
        td = _parse_row_date(row)
        status = _cell_str(row, ci["task_status"]) or "(vazio)"
        courier = _cell_str(row, ci["courier_name"])
        zip_code = _cell_str(row, ci["zip_code"])
        del_time = row[ci["del_time"]] if ci["del_time"] is not None and ci["del_time"] < len(row) else None
        if hasattr(del_time, "strftime"):
            del_time = del_time.strftime("%Y-%m-%d %H:%M")
        elif del_time:
            del_time = str(del_time)
        else:
            del_time = ""

        waybills_seen.add(wb)
        by_status[status] += 1
        if td:
            by_date[td.strftime("%Y-%m-%d")] += 1
        else:
            no_taskdate += 1

        if wb in search_targets:
            search_results[wb].append({
                "task_date": td.strftime("%Y-%m-%d") if td else "(sem data)",
                "task_status": status,
                "courier_name": courier,
                "delivery_time": del_time,
                "zip_code": zip_code,
            })

    # Para cada waybill procurado, devolve match (ou indicador de
    # "não encontrado")
    search_summary = []
    for wb in search_targets:
        rows = search_results.get(wb, [])
        search_summary.append({
            "waybill": wb,
            "n_rows": len(rows),
            "found": len(rows) > 0,
            "rows": rows,
        })

    return JsonResponse({
        "success": True,
        "filename": ficheiro.name,
        "total_data_rows": total,
        "skipped_no_waybill": no_waybill,
        "skipped_no_taskdate": no_taskdate,
        "n_unique_waybills": len(waybills_seen),
        "by_status": dict(by_status),
        "by_date": dict(sorted(by_date.items())),
        "search": search_summary,
    })


def cainiao_operation_import(request):
    """Import da planilha Operation Update.

    Estratégia:
      • A data efectiva (smart_date) é determinada por cascata:
        delivery_time → delivery_failure_time → start_delivery_time → Task Date.
        Alinha com a coluna 'Task Actual Date' do dashboard Cainiao.
      • Os pacotes são identificados pelo waybill_number e atualizados onde
        quer que estejam na BD — independentemente da task_date original.
      • Se o smart_date diferir da task_date guardada, o registo "move-se"
        de data (DELETE row antiga + INSERT na nova).
      • Status downgrade prevention: se o status novo tem prioridade
        inferior ao guardado, mantém-se o status antigo mas atualizam-se
        os outros campos.
    """
    try:
        return _cainiao_operation_import_impl(request)
    except Exception as e:
        import traceback, sys
        tb = traceback.format_exc()
        print("=== CAINIAO IMPORT ERROR ===", file=sys.stderr)
        print(tb, file=sys.stderr)
        return JsonResponse({
            "success": False,
            "error": f"{type(e).__name__}: {e}",
            "traceback": tb.splitlines()[-15:],
        }, status=500)


def _cainiao_operation_import_impl(request):
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Waybill Number' não encontrada."}, status=400)

    ci = {
        "waybill":       _col_idx(header, "Waybill Number"),
        "lp":            _col_idx(header, "LP No."),
        "task_status":   _col_idx(header, "Task Status"),
        "courier_name":  _col_idx(header, "Courier Name"),
        "dsp_name":      _col_idx(header, "DSP Name"),
        "delivery_type": _col_idx(header, "Delivery Type"),
        "order_type":    _col_idx(header, "Order Type"),
        "task_group":    _col_idx(header, "Task Group Order"),
        "sitecode":      _col_idx(header, "Sitecode"),
        "dest_country":  _col_idx(header, "Destination Country"),
        "dest_area":     _col_idx(header, "Area of destination"),
        "dest_city":     _col_idx(header, "destination city"),
        "zip_code":      _col_idx(header, "Zip Code"),
        "address":       _col_idx(header, "Detailed address"),
        "recv_lat":      _col_idx(header, "Receiver to Latitude"),
        "recv_lng":      _col_idx(header, "Receiver to Longitude"),
        "act_lat":       _col_idx(header, "Actual delivery Latitude"),
        "act_lng":       _col_idx(header, "Actual delivery  Longitude"),
        "gap_dist":      _col_idx(header, "Delivery Gap Distance"),
        "creation":      _col_idx(header, "Creation time"),
        "receipt":       _col_idx(header, "Receipt Time"),
        "outbound":      _col_idx(header, "Outbound time"),
        "start_del":     _col_idx(header, "Start Delivery Time"),
        "del_time":      _col_idx(header, "Delivery Time"),
        "del_fail":      _col_idx(header, "Delivery Failure Time"),
        "exc_type":      _col_idx(header, "Exception Type"),
        "exc_detail":    _col_idx(header, "Exception Detail"),
        "weight":        _col_idx(header, "Weight (g)"),
        "length":        _col_idx(header, "Long"),
        "width":         _col_idx(header, "Width"),
        "height":        _col_idx(header, "Height"),
        "seller":        _col_idx(header, "Seller Name"),
        "zone":          _col_idx(header, "Zone"),
        "orig_date":     _col_idx(header, "originalPlanTaskDate"),
        "sign_pod":      _col_idx(header, "Sign POD"),
        "pudo_addr":     _col_idx(header, "PUDO address"),
        "task_id":       _col_idx(header, "Task ID"),
        "taskdate":      _col_idx(header, "Task Date"),
        # Campos opcionais (Tracking + Planning combo)
        "priority":      _col_idx(header, "PRIORITYS", "Priority"),
        "pre_driver":    _col_idx(header, "Pre-assigned driver"),
        "pre_dsp":       _col_idx(header, "Pre-allocated DSP"),
        "wrong_parcel":  _col_idx(header, "Wrong HUB Parcel"),
        "arrive_wrong":  _col_idx(header, "Arrive Wrong Hub"),
        "comm_area":     _col_idx(header, "Commercial Area"),
        "hub_exc":       _col_idx(header, "HUB Exception Reason"),
        "bigbag":        _col_idx(header, "Bigbag No.", "Bigbag No"),
        "task_plan":     _col_idx(header, "Task Plan Date"),
    }

    def _yesno_to_bool(v):
        if v is None:
            return False
        s = str(v).strip().lower()
        return s in ("yes", "y", "true", "1", "sim", "s")

    def _to_date(v):
        if v is None or v == "":
            return None
        if isinstance(v, _date) and not isinstance(v, _dt):
            return v
        if isinstance(v, _dt):
            return v.date()
        s = str(v).strip()[:10]
        try:
            return _date.fromisoformat(s)
        except ValueError:
            return None

    def _planned_date(row):
        """Task Date (planeada) — usada como fallback."""
        ci_td = ci["taskdate"]
        if ci_td is None or ci_td >= len(row):
            return None
        return _to_date(row[ci_td])

    def _smart_date(row):
        """Task Date da linha = fonte de verdade.

        Regra do operador: se a planilha diz que o pacote é de Task Date X,
        o pacote pertence ao dia X. Fallback para Receipt/Creation apenas
        quando Task Date está vazia (raro)."""
        d = _planned_date(row)
        if d:
            return d
        # Fallback: receipt → creation
        for col in ("receipt", "creation", "del_time", "del_fail",
                    "start_del", "outbound"):
            idx = ci[col]
            if idx is None or idx >= len(row):
                continue
            dt = _parse_datetime(row[idx])
            if dt:
                return dt.date()
        return None

    STATUS_PRIORITY = {"Delivered": 4, "Attempt Failure": 3, "Driver_received": 2, "Assigned": 1, "Unassign": 0, "Cancel": -1}
    # Cancel é incluído para preservar histórico (driver pegou e devolveu/
    # passou para outro). Não é considerado nas KPIs de operação porque
    # core/views.py filtra explicitamente os 5 status operacionais.
    _VALID = {"Delivered", "Driver_received", "Attempt Failure", "Unassign", "Assigned", "Cancel"}

    from collections import defaultdict
    from .models import (
        CainiaoOperationBatch, CainiaoOperationTask,
        CainiaoOperationTaskHistory,
    )

    audit = {
        "total_rows_in_file": 0,
        "skipped_no_waybill": 0,
        "skipped_no_date": 0,
        "rows_by_status": defaultdict(int),
        "waybills_unique": 0,
        "waybills_skipped_cancel_only": 0,
        "waybills_no_status_change_due_downgrade": 0,
        "waybills_moved_date": 0,
        "skipped_waybill_examples": [],
        "cancelled_only_examples": [],
        "downgrade_examples": [],
        "moved_date_examples": [],
    }

    # ─── 1. Agrupar por waybill (não por data — data é decidida por waybill) ─
    wb_rows = defaultdict(list)
    audit["couriers_in_file"] = defaultdict(int)
    audit["dsp_in_file"] = defaultdict(int)
    audit["headers_detected"] = [
        {"col": chr(65 + i) if i < 26 else f"col{i}", "header": str(h)}
        for i, h in enumerate(header) if h is not None
    ]
    audit["courier_name_col_idx"] = ci.get("courier_name")
    audit["dsp_name_col_idx"] = ci.get("dsp_name")
    for row in all_rows[header_idx + 1:]:
        audit["total_rows_in_file"] += 1
        waybill = _cell_str(row, ci["waybill"])
        status = _cell_str(row, ci["task_status"])
        courier = _cell_str(row, ci["courier_name"])
        dsp = _cell_str(row, ci["dsp_name"])
        if status:
            audit["rows_by_status"][status] += 1
        if courier:
            audit["couriers_in_file"][courier] += 1
        if dsp:
            audit["dsp_in_file"][dsp] += 1
        if not waybill:
            audit["skipped_no_waybill"] += 1
            continue
        wb_rows[waybill].append(row)
    # Converter defaultdicts para dicts JSON-serializáveis
    audit["couriers_in_file"] = dict(audit["couriers_in_file"])
    audit["dsp_in_file"] = dict(audit["dsp_in_file"])

    if not wb_rows:
        return JsonResponse(
            {"success": False, "error": "Nenhuma linha com Waybill Number válido."},
            status=400,
        )

    # NOTE: courier_id_cainiao é EXCLUÍDO de update_fields porque é definido
    # apenas no INSERT (via resolver). Em re-imports queremos preservar
    # mappings manuais já feitos pelo utilizador.
    update_fields_all = [
        "lp_number", "task_status", "courier_name",
        "dsp_name", "delivery_type",
        "order_type", "task_group_order", "sitecode", "destination_country",
        "destination_area", "destination_city", "zip_code", "detailed_address",
        "receiver_latitude", "receiver_longitude", "actual_latitude", "actual_longitude",
        "delivery_gap_distance", "creation_time", "receipt_time", "outbound_time",
        "start_delivery_time", "delivery_time", "delivery_failure_time",
        "exception_type", "exception_detail", "weight_g", "length", "width", "height",
        "seller_name", "zone", "original_plan_task_date", "sign_pod", "pudo_address",
        "task_id", "last_import_batch",
        "is_priority_external", "pre_assigned_driver", "pre_allocated_dsp",
        "wrong_hub_parcel", "arrive_wrong_hub", "commercial_area",
        "hub_exception_reason", "bigbag_number", "task_plan_date",
    ]
    # Em downgrade: não tocar em status, courier nem nos seus timestamps.
    # Só actualizar campos genéricos (endereço, dimensões, etc.). Evita o
    # bug de mistura "Attempt Failure (de driver A) + courier=B".
    _status_time_fields = {
        "task_status",
        "delivery_time", "delivery_failure_time", "start_delivery_time",
        "courier_name", "courier_id_cainiao",
    }
    update_fields_no_status = [f for f in update_fields_all if f not in _status_time_fields]

    filename = ficheiro.name

    # Pre-load courier_name → courier_id resolver
    from .models import DriverCourierMapping, CourierNameAlias
    from drivers_app.models import DriverProfile
    from core.models import Partner

    _cainiao_partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    _name_to_courier_id = {}
    if _cainiao_partner:
        for m in DriverCourierMapping.objects.filter(partner=_cainiao_partner):
            if m.courier_name:
                _name_to_courier_id[m.courier_name] = m.courier_id
            _name_to_courier_id.setdefault(m.courier_id, m.courier_id)
        for a in CourierNameAlias.objects.filter(partner=_cainiao_partner):
            _name_to_courier_id.setdefault(a.courier_name, a.courier_id)
    for d in DriverProfile.objects.exclude(courier_id_cainiao="").exclude(apelido=""):
        _name_to_courier_id.setdefault(d.apelido, d.courier_id_cainiao)

    def _resolve_courier_id(courier_name):
        return _name_to_courier_id.get((courier_name or "").strip(), "")

    # Heurística: courier é placeholder de armazém/hub (não driver real).
    # Usado APENAS como tie-breaker final quando duas linhas têm exactamente
    # o mesmo timestamp e priority de status.
    _WAREHOUSE_KEYWORDS = ("ARMAZEM", "ARMAZÉM", "HUB", "WAREHOUSE", "DEPOSITO", "DEPÓSITO", "CENTRO_OPERACIONAL")

    def _is_warehouse_courier(courier_name):
        if not courier_name:
            return True
        cn_upper = courier_name.upper()
        return any(kw in cn_upper for kw in _WAREHOUSE_KEYWORDS)

    # Sentinela para rows sem nenhum timestamp — ficam abaixo de qualquer
    # row com timestamp real. Datetime aware com tzinfo UTC para comparações.
    from datetime import timezone as _tz
    _EPOCH = _dt(1970, 1, 1, tzinfo=_tz.utc)

    def _row_last_event_ts(row):
        """Timestamp do evento mais recente registado nesta linha.
        Reflecte 'a última coisa que aconteceu ao pacote nesta agente'."""
        latest = None
        for col in ("del_time", "del_fail", "start_del", "outbound", "receipt", "creation"):
            idx = ci[col]
            if idx is None or idx >= len(row):
                continue
            dt = _parse_datetime(row[idx])
            if dt and (latest is None or dt > latest):
                latest = dt
        return latest

    # ─── 2. Para cada waybill: best row + smart_date + signatures ─
    # Filosofia "cadeia de assinaturas":
    #   Cada linha da planilha = uma assinatura (courier+status+timestamp).
    #   - active_rows: linhas com status válido (NÃO excluímos por Cancel —
    #     Cancel também é uma assinatura legítima na timeline).
    #   - best_row: linha com timestamp mais recente = estado vigente actual.
    #   - all_signatures_per_wb: TODAS as linhas para irem à timeline.
    best_row_per_wb = {}
    all_signatures_per_wb = {}  # waybill → list[row] (ordenado por ts asc)
    for waybill, r_list in wb_rows.items():
        active_rows = [
            r for r in r_list
            if _cell_str(r, ci["task_status"]) in _VALID
        ]
        if not active_rows:
            audit["waybills_skipped_cancel_only"] += 1
            if len(audit["skipped_waybill_examples"]) < 20:
                audit["skipped_waybill_examples"].append({
                    "waybill": waybill,
                    "n_rows": len(r_list),
                    "statuses": list({
                        _cell_str(r, ci["task_status"]) for r in r_list
                    }),
                })
            continue

        # Algoritmo "Task Date mais recente vence" para estado vigente:
        # Score (tuple, decrescente):
        #   1. task_date_row — Task Date da linha (planilha = verdade)
        #   2. last_event_ts — timestamp do evento (desempate)
        #   3. priority — desempate quando ts iguais
        #   4. is_real_driver — prefere driver real sobre placeholder
        from datetime import date as _date_helper
        _DATE_MIN = _date_helper(1970, 1, 1)
        best_score = (_DATE_MIN, _EPOCH, -1, -1)
        best_row = None
        for row in active_rows:
            row_task_date = _planned_date(row) or _DATE_MIN
            last_ts = _row_last_event_ts(row) or _EPOCH
            status = _cell_str(row, ci["task_status"])
            courier = _cell_str(row, ci["courier_name"])
            priority = STATUS_PRIORITY.get(status, 0)
            is_real = 0 if _is_warehouse_courier(courier) else 1
            score = (row_task_date, last_ts, priority, is_real)
            if score > best_score:
                best_score = score
                best_row = row

        smart_dt = _smart_date(best_row)
        if not smart_dt:
            audit["skipped_no_date"] += 1
            continue

        best_row_per_wb[waybill] = (best_row, smart_dt)
        # Guardar todas as assinaturas em ordem cronológica para o history
        all_signatures_per_wb[waybill] = sorted(
            active_rows,
            key=lambda r: (_row_last_event_ts(r) or _EPOCH),
        )

    audit["waybills_unique"] = len(best_row_per_wb)

    if not best_row_per_wb:
        return JsonResponse(
            {"success": False, "error": "Nenhum waybill processável."},
            status=400,
        )

    # ─── 3. Lookup global por waybill (sem filtrar por data) ─
    # Pré-carregamos o estado COMPLETO de cada waybill já em BD, incluindo
    # courier e datas, para podermos detectar mudanças e registar timeline.
    existing_full = {}  # waybill → {task_date, task_status, courier_name, courier_id_cainiao}
    for r in CainiaoOperationTask.objects.filter(
        waybill_number__in=list(best_row_per_wb.keys())
    ).values(
        "waybill_number", "task_date", "task_status",
        "courier_name", "courier_id_cainiao",
    ):
        wb = r["waybill_number"]
        cur = existing_full.get(wb)
        # Em caso de múltiplas rows (datas diferentes), ficamos com a
        # de maior priority como representante do "estado vigente".
        if cur is None:
            existing_full[wb] = r
        else:
            cur_pri = STATUS_PRIORITY.get(cur["task_status"], 0)
            new_pri = STATUS_PRIORITY.get(r["task_status"], 0)
            if new_pri > cur_pri:
                existing_full[wb] = r

    # Atalho para o resto do código que só precisa de (date, status)
    existing_by_waybill = {
        wb: (r["task_date"], r["task_status"])
        for wb, r in existing_full.items()
    }

    # ─── 4. Decidir CREATE vs UPDATE vs MOVE-DATE vs SKIP por waybill ─
    # objs_by_date: smart_date → list[(obj, is_status_upgrade)]
    objs_upgrade_by_date = defaultdict(list)
    objs_no_status_by_date = defaultdict(list)
    waybills_to_delete = []  # [(waybill, old_date), ...] para MOVE-DATE
    # Entries de timeline a criar (preenche batch=None, atribuído depois)
    history_pending = []  # list[dict] — convertido em CainiaoOperationTaskHistory

    for waybill, (row, smart_dt) in best_row_per_wb.items():
        new_status = _cell_str(row, ci["task_status"])
        new_priority = STATUS_PRIORITY.get(new_status, 0)
        _cn = _cell_str(row, ci["courier_name"])
        _resolved_id = _resolve_courier_id(_cn)
        last_event_ts = _row_last_event_ts(row)

        tpd = None
        if ci["task_plan"] is not None and ci["task_plan"] < len(row):
            tpd = _to_date(row[ci["task_plan"]])

        # Decisão de data e tipo de operação
        is_status_upgrade = True
        target_date = smart_dt
        existing_main = existing_full.get(waybill)
        change_type = None  # se vai gerar history entry: 'created' | 'status_change' | 'courier_change' | 'date_move'
        if existing_main is None:
            change_type = "created"
        else:
            old_date = existing_main["task_date"]
            old_status = existing_main["task_status"]
            old_courier = existing_main["courier_name"]
            old_priority = STATUS_PRIORITY.get(old_status, 0)
            if new_priority < old_priority:
                # Downgrade: manter na data antiga e não tocar no status
                is_status_upgrade = False
                target_date = old_date
                audit["waybills_no_status_change_due_downgrade"] += 1
                if len(audit["downgrade_examples"]) < 20:
                    audit["downgrade_examples"].append({
                        "waybill": waybill,
                        "old_status": old_status,
                        "new_status_attempted": new_status,
                    })
                # Mesmo num downgrade, se o courier mudou, regista no history
                # (relevante para o caso "voltou ao armazém de outro driver")
                if (_cn or "") != (old_courier or ""):
                    change_type = "courier_change"
            elif old_date != smart_dt:
                # MOVE-DATE: status válido para upgrade mas data mudou
                # → apaga row antiga e insere em smart_date
                waybills_to_delete.append((waybill, old_date))
                audit["waybills_moved_date"] += 1
                if len(audit["moved_date_examples"]) < 20:
                    audit["moved_date_examples"].append({
                        "waybill": waybill,
                        "old_date": str(old_date),
                        "new_date": str(smart_dt),
                        "old_status": old_status,
                        "new_status": new_status,
                    })
                change_type = "date_move"
            elif old_status != new_status:
                change_type = "status_change"
            elif (_cn or "") != (old_courier or ""):
                change_type = "courier_change"

        if change_type:
            history_pending.append({
                "waybill_number": waybill,
                # Estado APÓS a mudança (vai ser persistido após bulk_create)
                "task_date": target_date,
                "task_status": (
                    new_status if is_status_upgrade
                    else (existing_main["task_status"] if existing_main else new_status)
                ),
                "courier_name": _cn,
                "courier_id_cainiao": _resolved_id,
                # Estado ANTES (None se for criação)
                "previous_task_date": existing_main["task_date"] if existing_main else None,
                "previous_task_status": existing_main["task_status"] if existing_main else "",
                "previous_courier_name": existing_main["courier_name"] if existing_main else "",
                "change_type": change_type,
                "event_timestamp": last_event_ts,
                "target_date_for_batch": target_date,  # usado para atribuir batch certo
            })

        # Signature entries: cada linha activa do waybill na planilha vira
        # uma entry no history (timeline completa "cadeia de assinaturas").
        # Permite reconstruir todos os couriers que tocaram o pacote.
        signatures = all_signatures_per_wb.get(waybill, [])
        for sig_row in signatures:
            sig_status = _cell_str(sig_row, ci["task_status"])
            sig_courier = _cell_str(sig_row, ci["courier_name"])
            sig_ts = _row_last_event_ts(sig_row)
            sig_date = _smart_date(sig_row)
            history_pending.append({
                "waybill_number": waybill,
                "task_date": sig_date or target_date,
                "task_status": sig_status,
                "courier_name": sig_courier,
                "courier_id_cainiao": _resolve_courier_id(sig_courier),
                "previous_task_date": None,
                "previous_task_status": "",
                "previous_courier_name": "",
                "change_type": "signature",
                "event_timestamp": sig_ts,
                "target_date_for_batch": target_date,
            })

        obj = CainiaoOperationTask(
            waybill_number=waybill,
            task_date=target_date,
            lp_number=_cell_str(row, ci["lp"]),
            task_status=new_status,
            courier_name=_cn,
            courier_id_cainiao=_resolve_courier_id(_cn),
            dsp_name=_cell_str(row, ci["dsp_name"]),
            delivery_type=_cell_str(row, ci["delivery_type"]),
            order_type=_cell_str(row, ci["order_type"]),
            task_group_order=_cell_str(row, ci["task_group"]),
            sitecode=_cell_str(row, ci["sitecode"]),
            destination_country=_cell_str(row, ci["dest_country"]),
            destination_area=_cell_str(row, ci["dest_area"]),
            destination_city=_cell_str(row, ci["dest_city"]),
            zip_code=_normalise_zip(_cell_str(row, ci["zip_code"])),
            detailed_address=_cell_str(row, ci["address"]),
            receiver_latitude=_cell_str(row, ci["recv_lat"]),
            receiver_longitude=_cell_str(row, ci["recv_lng"]),
            actual_latitude=_cell_str(row, ci["act_lat"]),
            actual_longitude=_cell_str(row, ci["act_lng"]),
            delivery_gap_distance=_cell_str(row, ci["gap_dist"]),
            creation_time=_parse_datetime(row[ci["creation"]] if ci["creation"] is not None else None),
            receipt_time=_parse_datetime(row[ci["receipt"]] if ci["receipt"] is not None else None),
            outbound_time=_parse_datetime(row[ci["outbound"]] if ci["outbound"] is not None else None),
            start_delivery_time=_parse_datetime(row[ci["start_del"]] if ci["start_del"] is not None else None),
            delivery_time=_parse_datetime(row[ci["del_time"]] if ci["del_time"] is not None else None),
            delivery_failure_time=_parse_datetime(row[ci["del_fail"]] if ci["del_fail"] is not None else None),
            exception_type=_cell_str(row, ci["exc_type"]),
            exception_detail=_cell_str(row, ci["exc_detail"]),
            weight_g=_cell_float(row, ci["weight"]),
            length=_cell_float(row, ci["length"]),
            width=_cell_float(row, ci["width"]),
            height=_cell_float(row, ci["height"]),
            seller_name=_cell_str(row, ci["seller"]),
            zone=_cell_str(row, ci["zone"]),
            original_plan_task_date=_cell_str(row, ci["orig_date"]),
            sign_pod=_cell_str(row, ci["sign_pod"]),
            pudo_address=_cell_str(row, ci["pudo_addr"]),
            task_id=_cell_str(row, ci["task_id"]),
            is_priority_external=_yesno_to_bool(
                row[ci["priority"]]
                if ci["priority"] is not None and ci["priority"] < len(row)
                else None
            ),
            pre_assigned_driver=_cell_str(row, ci["pre_driver"]),
            pre_allocated_dsp=_cell_str(row, ci["pre_dsp"]),
            wrong_hub_parcel=_yesno_to_bool(
                row[ci["wrong_parcel"]]
                if ci["wrong_parcel"] is not None and ci["wrong_parcel"] < len(row)
                else None
            ),
            arrive_wrong_hub=_yesno_to_bool(
                row[ci["arrive_wrong"]]
                if ci["arrive_wrong"] is not None and ci["arrive_wrong"] < len(row)
                else None
            ),
            commercial_area=_cell_str(row, ci["comm_area"]),
            hub_exception_reason=_cell_str(row, ci["hub_exc"]),
            bigbag_number=_cell_str(row, ci["bigbag"]),
            task_plan_date=tpd,
            last_import_batch=None,  # atribuído por data abaixo
        )

        if is_status_upgrade:
            objs_upgrade_by_date[target_date].append(obj)
        else:
            objs_no_status_by_date[target_date].append(obj)

    # ─── 5. Consolidação: cada waybill = 1 row vigente na smart_dt ─
    # Para todos os waybills que serão UPSERT em alguma smart_dt, apagar
    # quaisquer outras rows do mesmo waybill em datas != smart_dt. Resolve
    # o caso "pacote espalhado por várias datas" — mantém só a row actual.
    # A timeline está preservada no CainiaoOperationTaskHistory.
    waybill_to_target_date = {}
    for target_date, objs in objs_upgrade_by_date.items():
        for o in objs:
            waybill_to_target_date[o.waybill_number] = target_date
    for target_date, objs in objs_no_status_by_date.items():
        for o in objs:
            waybill_to_target_date.setdefault(o.waybill_number, target_date)

    if waybill_to_target_date:
        from django.db.models import Q
        from django.db import transaction
        # Procurar todas as rows existentes desses waybills em datas
        # diferentes da target_date, agrupar por (target_date, waybill) e apagar.
        existing_other_dates = list(
            CainiaoOperationTask.objects
            .filter(waybill_number__in=list(waybill_to_target_date.keys()))
            .values("waybill_number", "task_date")
        )
        # Agrupar por task_date as rows que precisam de ser apagadas
        deletes_by_date = defaultdict(list)
        consolidated_count = 0
        for r in existing_other_dates:
            wb = r["waybill_number"]
            target_dt = waybill_to_target_date.get(wb)
            if target_dt is None:
                continue
            if r["task_date"] != target_dt:
                deletes_by_date[r["task_date"]].append(wb)
                consolidated_count += 1

        if deletes_by_date:
            with transaction.atomic():
                for dt, wbs in deletes_by_date.items():
                    CainiaoOperationTask.objects.filter(
                        task_date=dt, waybill_number__in=wbs,
                    ).delete()
        audit["waybills_consolidated"] = consolidated_count

    # ─── 6. Criar batch + bulk_create por data ─
    all_target_dates = sorted(set(objs_upgrade_by_date.keys()) | set(objs_no_status_by_date.keys()))
    batch_ids = []
    batch_by_date = {}  # target_date → CainiaoOperationBatch (para atribuir ao history)
    total_novos = 0
    total_atualizados = 0
    existing_waybills_global = set(existing_by_waybill.keys())

    for target_date in all_target_dates:
        upgrade_objs = objs_upgrade_by_date.get(target_date, [])
        no_status_objs = objs_no_status_by_date.get(target_date, [])
        all_objs = upgrade_objs + no_status_objs
        if not all_objs:
            continue

        batch = CainiaoOperationBatch.objects.create(
            filename=filename,
            task_date=target_date,
            created_by=request.user,
        )
        batch_ids.append(batch.id)
        batch_by_date[target_date] = batch

        for o in all_objs:
            o.last_import_batch = batch

        # NOTA: unique_fields NÃO é passado — MySQL não suporta esse parâmetro
        # com update_conflicts. O motor MySQL usa automaticamente a unique key
        # via INSERT ... ON DUPLICATE KEY UPDATE.
        if upgrade_objs:
            CainiaoOperationTask.objects.bulk_create(
                upgrade_objs, batch_size=500,
                update_conflicts=True,
                update_fields=update_fields_all,
            )
        if no_status_objs:
            CainiaoOperationTask.objects.bulk_create(
                no_status_objs, batch_size=500,
                update_conflicts=True,
                update_fields=update_fields_no_status,
            )

        n_in_batch_existing = sum(1 for o in all_objs if o.waybill_number in existing_waybills_global)
        n_in_batch_new = len(all_objs) - n_in_batch_existing
        total_atualizados += n_in_batch_existing
        total_novos += n_in_batch_new

        batch.total_tasks = len(all_objs)
        batch.new_tasks = n_in_batch_new
        batch.updated_tasks = n_in_batch_existing
        batch.save(update_fields=["total_tasks", "new_tasks", "updated_tasks"])

    # ─── 7. Persistir entries de history (timeline) ─
    # Dedup prévio:
    #   - Para signatures: cada (waybill, courier, status, event_ts) é único.
    #     Evita duplicatas em re-imports da mesma planilha.
    #   - Para outros change_types: cada combinação no batch é única.
    # Fazemos esta verificação contra a BD para evitar entries idênticos
    # criados em imports anteriores.

    # 1. Pré-carregar signatures já existentes para os waybills tocados
    sig_keys_pending = set()  # (wb, courier, status, ts_iso)
    other_pending = []
    for h in history_pending:
        if h["change_type"] == "signature":
            ts = h["event_timestamp"]
            ts_iso = ts.isoformat() if ts else ""
            key = (h["waybill_number"], h["courier_name"] or "",
                   h["task_status"] or "", ts_iso)
            if key in sig_keys_pending:
                continue  # dedup intra-batch
            sig_keys_pending.add(key)
        else:
            other_pending.append(h)

    # 2. Existing signatures na BD para os mesmos waybills
    waybills_with_sigs = {h["waybill_number"] for h in history_pending}
    existing_sig_keys = set()
    if waybills_with_sigs:
        for r in CainiaoOperationTaskHistory.objects.filter(
            waybill_number__in=list(waybills_with_sigs),
            change_type="signature",
        ).values("waybill_number", "courier_name", "task_status", "event_timestamp"):
            ts = r["event_timestamp"]
            ts_iso = ts.isoformat() if ts else ""
            existing_sig_keys.add((
                r["waybill_number"], r["courier_name"] or "",
                r["task_status"] or "", ts_iso,
            ))

    history_objs = []
    audit["signatures_skipped_duplicate"] = 0
    for h in history_pending:
        if h["change_type"] == "signature":
            ts = h["event_timestamp"]
            ts_iso = ts.isoformat() if ts else ""
            key = (h["waybill_number"], h["courier_name"] or "",
                   h["task_status"] or "", ts_iso)
            if key in existing_sig_keys:
                audit["signatures_skipped_duplicate"] += 1
                continue
            existing_sig_keys.add(key)
        batch_obj = batch_by_date.get(h["target_date_for_batch"])
        history_objs.append(CainiaoOperationTaskHistory(
            waybill_number=h["waybill_number"],
            task_date=h["task_date"],
            task_status=h["task_status"],
            courier_name=h["courier_name"],
            courier_id_cainiao=h["courier_id_cainiao"],
            previous_task_date=h["previous_task_date"],
            previous_task_status=h["previous_task_status"],
            previous_courier_name=h["previous_courier_name"],
            change_type=h["change_type"],
            event_timestamp=h["event_timestamp"],
            batch=batch_obj,
        ))
    if history_objs:
        CainiaoOperationTaskHistory.objects.bulk_create(history_objs, batch_size=500)
    audit["history_entries_created"] = len(history_objs)

    most_recent_date = str(all_target_dates[-1]) if all_target_dates else None
    date_range = (
        f"{all_target_dates[0]} → {all_target_dates[-1]}"
        if len(all_target_dates) > 1
        else (str(all_target_dates[0]) if all_target_dates else "")
    )

    audit["rows_by_status"] = dict(audit["rows_by_status"])

    return JsonResponse({
        "success": True,
        "total_novos": total_novos,
        "total_atualizados": total_atualizados,
        "total_cleaned": 0,
        "dates_imported": len(all_target_dates),
        "date_range": date_range,
        "most_recent_date": most_recent_date,
        "batch_ids": batch_ids,
        "audit": audit,
    })


# ============================================================================
# IMPORT — PLANILHA DRIVER STATISTIC (8 colunas)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_driver_stat_preview(request):
    """Preview detalhado da planilha Driver Statistic.

    Para cada courier, retorna:
        - Sugestão de matching (já mapeado / por courier_id_cainiao /
          por apelido / sem match)
        - Sugestão de driver para vincular (com base no apelido)
        - Permite ao user escolher acção por linha antes do import
    """
    from .models import DriverCourierMapping
    from drivers_app.models import DriverProfile
    from core.models import Partner

    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."}, status=400,
        )

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400,
        )

    header, header_idx = _find_header_row(all_rows, "Courier ID")
    if header is None:
        return JsonResponse(
            {"success": False, "error": "Coluna 'Courier ID' não encontrada."},
            status=400,
        )

    ci_id    = _col_idx(header, "Courier ID")
    ci_name  = _col_idx(header, "Courier Name")
    ci_dsp   = _col_idx(header, "DSP Name")
    ci_total = _col_idx(header, "Total Parcels")
    ci_rate  = _col_idx(header, "Delivery Success Rate")
    ci_date  = _col_idx(header, "Dispatch Date")

    partner = Partner.objects.filter(name__iexact="CAINIAO").first()

    # Pre-load para matching
    mappings_by_id = {}
    if partner:
        mappings_by_id = {
            m.courier_id: m
            for m in DriverCourierMapping.objects.filter(partner=partner)
            .select_related("driver")
        }
    by_courier_id = {
        d.courier_id_cainiao: d for d in DriverProfile.objects.all()
        if d.courier_id_cainiao
    }
    by_apelido = {
        d.apelido.strip().lower(): d for d in DriverProfile.objects.all()
        if d.apelido
    }

    rows_data = []
    date_range = ""
    for row in all_rows[header_idx + 1:]:
        cid = _cell_str(row, ci_id)
        if not cid:
            continue
        cname = _cell_str(row, ci_name)
        if not date_range:
            date_range = _cell_str(row, ci_date)

        # Matching strategy: courier_name → APELIDO (não nome_completo)
        suggested_driver_id = None
        suggested_driver_name = None
        match_kind = "create"  # create | mapped | by_courier_id | by_apelido

        m = mappings_by_id.get(cid)
        if m:
            suggested_driver_id = m.driver_id
            suggested_driver_name = m.driver.nome_completo
            match_kind = "mapped"
        else:
            d = by_courier_id.get(cid)
            if d:
                suggested_driver_id = d.id
                suggested_driver_name = d.nome_completo
                match_kind = "by_courier_id"
            else:
                cname_lc = cname.strip().lower()
                d = by_apelido.get(cname_lc)
                if d:
                    suggested_driver_id = d.id
                    suggested_driver_name = d.nome_completo
                    match_kind = "by_apelido"

        rows_data.append({
            "courier_id":   cid,
            "courier_name": cname,
            "dsp_name":     _cell_str(row, ci_dsp),
            "total":        _cell_str(row, ci_total),
            "rate":         _cell_str(row, ci_rate),
            "match_kind":   match_kind,
            "suggested_driver_id":   suggested_driver_id,
            "suggested_driver_name": suggested_driver_name,
        })

    # All drivers para o dropdown de "vincular a outro"
    all_drivers = list(
        DriverProfile.objects.exclude(status="IRREGULAR")
        .values("id", "nome_completo", "apelido", "courier_id_cainiao")
        .order_by("nome_completo")
    )

    return JsonResponse({
        "success": True,
        "total": len(rows_data),
        "rows": rows_data,
        "date_range": date_range,
        "drivers": all_drivers,
        "summary": {
            "mapped":         sum(1 for r in rows_data if r["match_kind"] == "mapped"),
            "by_courier_id":  sum(1 for r in rows_data if r["match_kind"] == "by_courier_id"),
            "by_apelido":     sum(1 for r in rows_data if r["match_kind"] == "by_apelido"),
            "create":         sum(1 for r in rows_data if r["match_kind"] == "create"),
        },
    })


def _cleanup_stale_waybill_rows(waybills=None):
    """Para cada waybill, mantém APENAS a row com task_date mais recente.

    Apaga snapshots intermédios (Driver_received, Assigned, Attempt Failure,
    etc.) quando existe uma row posterior para o mesmo waybill — essa row
    posterior representa o estado actual do pacote.

    A semântica é: cada waybill = um pacote vivo no sistema; o estado mais
    recente é a verdade. Snapshots históricos podem ser consultados via
    backups/audit log se necessário.

    Args:
        waybills: lista opcional de waybill_numbers a limpar. Se None,
                  varre todos os waybills do DB. Use a forma scoped quando
                  chamado a partir de um import (mais rápido).

    Returns:
        dict {scanned_waybills, deleted_rows, kept_rows}
    """
    from django.db.models import OuterRef, Exists
    from .models import CainiaoOperationTask

    qs = CainiaoOperationTask.objects.all()
    if waybills:
        qs = qs.filter(waybill_number__in=list(waybills))

    later = CainiaoOperationTask.objects.filter(
        waybill_number=OuterRef("waybill_number"),
        task_date__gt=OuterRef("task_date"),
    )
    stale_ids = list(
        qs.annotate(has_later=Exists(later))
        .filter(has_later=True)
        .values_list("id", flat=True)
    )
    deleted = 0
    if stale_ids:
        # Apagar em lotes para evitar query gigante
        BATCH = 5000
        for i in range(0, len(stale_ids), BATCH):
            batch = stale_ids[i:i + BATCH]
            n, _ = CainiaoOperationTask.objects.filter(
                id__in=batch
            ).delete()
            deleted += n

    return {
        "scanned_waybills": qs.values("waybill_number").distinct().count(),
        "deleted_rows": deleted,
    }


@login_required
@require_http_methods(["POST"])
def cainiao_cleanup_stale_rows_view(request):
    """Endpoint admin para correr a limpeza de snapshots obsoletos.

    Body JSON opcional: { "dry_run": true } — só conta sem apagar.
    """
    import json
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    dry_run = bool(body.get("dry_run"))

    if dry_run:
        from django.db.models import OuterRef, Exists
        from .models import CainiaoOperationTask
        later = CainiaoOperationTask.objects.filter(
            waybill_number=OuterRef("waybill_number"),
            task_date__gt=OuterRef("task_date"),
        )
        n = (CainiaoOperationTask.objects
             .annotate(has_later=Exists(later))
             .filter(has_later=True).count())
        return JsonResponse({
            "success": True, "dry_run": True,
            "would_delete": n,
        })

    result = _cleanup_stale_waybill_rows()
    return JsonResponse({"success": True, **result})


@login_required
@require_http_methods(["POST"])
def cainiao_debug_headers_view(request):
    """Debug: recebe upload XLSX e devolve cabeçalho + 3 primeiras linhas.

    Útil para diagnosticar discrepâncias entre o que o user vê no Excel
    e o que o parser lê.
    """
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        # devolve as primeiras 5 linhas para ver
        return JsonResponse({
            "success": False,
            "error": "Cabeçalho não encontrado",
            "first_5_rows": [
                [str(c)[:50] for c in r] for r in all_rows[:5]
            ],
        })

    # Procurar especificamente onde está "LEGUAS-XPT" para identificar a coluna
    leguas_xpt_columns = []
    for row in all_rows[header_idx + 1:header_idx + 11]:
        for col_idx, cell in enumerate(row):
            s = str(cell or "")
            if "LEGUAS-XPT" in s.upper() or "LEGUAS XPT" in s.upper():
                col_header = str(header[col_idx]) if col_idx < len(header) else "(no header)"
                leguas_xpt_columns.append({
                    "col_idx": col_idx,
                    "col_letter": chr(65 + col_idx) if col_idx < 26 else f"AA+{col_idx-26}",
                    "header": col_header,
                    "sample_value": s[:50],
                })

    # Localizar todas as colunas com "Courier" no nome
    courier_cols = []
    for i, h in enumerate(header):
        if h and "courier" in str(h).lower():
            courier_cols.append({
                "col_idx": i,
                "col_letter": chr(65 + i) if i < 26 else f"AA+{i-26}",
                "header": str(h),
                "sample_row1": str(all_rows[header_idx + 1][i])[:50] if header_idx + 1 < len(all_rows) and i < len(all_rows[header_idx + 1]) else "",
            })

    # Devolver cabeçalho completo + 3 linhas
    return JsonResponse({
        "success": True,
        "filename": ficheiro.name,
        "total_rows": len(all_rows) - header_idx - 1,
        "header_row_idx": header_idx,
        "all_headers": [
            {"col_idx": i, "col_letter": chr(65 + i) if i < 26 else f"AA+{i-26}",
             "header": str(h)} for i, h in enumerate(header)
        ],
        "courier_columns_detected": courier_cols,
        "leguas_xpt_locations": leguas_xpt_columns[:10],
        "first_3_data_rows": [
            [str(c)[:60] for c in r]
            for r in all_rows[header_idx + 1:header_idx + 4]
        ],
    })


@login_required
@require_http_methods(["POST"])
def cainiao_dedup_signatures_view(request):
    """Remove duplicates de history signatures.

    Cada (waybill, courier, status, event_timestamp, change_type=signature)
    deve ser único. Quando re-imports criaram duplicates, este endpoint
    mantém apenas a entry mais antiga (primeira recorded_at) e apaga as
    restantes.
    """
    import json
    from django.db.models import Min
    from django.db import transaction
    from .models import CainiaoOperationTaskHistory

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    dry_run = bool(body.get("dry_run"))

    # Agrupar em python — mais simples e robusto que SQL group-by para
    # tuplas com NULL.
    qs = CainiaoOperationTaskHistory.objects.filter(change_type="signature")
    rows = list(qs.values(
        "id", "waybill_number", "courier_name", "task_status", "event_timestamp",
        "recorded_at",
    ))
    seen = {}  # key → id_to_keep
    to_delete = []
    for r in rows:
        key = (
            r["waybill_number"], r["courier_name"] or "",
            r["task_status"] or "",
            r["event_timestamp"].isoformat() if r["event_timestamp"] else "",
        )
        if key not in seen:
            seen[key] = r["id"]
        else:
            to_delete.append(r["id"])

    if dry_run:
        return JsonResponse({
            "success": True, "dry_run": True,
            "total_signatures": len(rows),
            "unique_signatures": len(seen),
            "duplicates_to_delete": len(to_delete),
        })

    if to_delete:
        with transaction.atomic():
            BATCH = 5000
            for i in range(0, len(to_delete), BATCH):
                batch = to_delete[i:i + BATCH]
                CainiaoOperationTaskHistory.objects.filter(
                    id__in=batch
                ).delete()

    return JsonResponse({
        "success": True, "dry_run": False,
        "total_signatures": len(rows),
        "unique_signatures": len(seen),
        "deleted": len(to_delete),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_consolidate_waybills_view(request):
    """Backfill: para cada waybill com múltiplas rows espalhadas em datas
    diferentes, mantém apenas a row 'vigente' (último evento mais recente)
    e apaga as outras. Cria entries no history para preservar timeline.

    Body JSON opcional:
        { "dry_run": true }

    Cenário: pacote teve linha em 01/05 (Driver_received), em 02/05 (Driver_received),
    em 30/04 (Attempt Failure) — após este backfill, fica só 1 row (a com
    timestamp mais recente).
    """
    import json
    from django.db.models import Count
    from django.db import transaction
    from .models import (
        CainiaoOperationTask, CainiaoOperationTaskHistory,
    )

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    dry_run = bool(body.get("dry_run"))

    # Encontrar waybills com >1 row
    duplicated_wbs = list(
        CainiaoOperationTask.objects
        .values("waybill_number")
        .annotate(n=Count("task_date", distinct=True))
        .filter(n__gt=1)
        .values_list("waybill_number", flat=True)
    )

    if not duplicated_wbs:
        return JsonResponse({
            "success": True, "dry_run": dry_run,
            "waybills_with_duplicates": 0, "rows_to_delete": 0,
        })

    # Para cada waybill duplicado, escolher a row vigente (timestamp mais recente)
    rows = list(
        CainiaoOperationTask.objects
        .filter(waybill_number__in=duplicated_wbs)
        .values(
            "id", "waybill_number", "task_date", "task_status", "courier_name",
            "courier_id_cainiao",
            "creation_time", "receipt_time", "outbound_time",
            "start_delivery_time", "delivery_time", "delivery_failure_time",
        )
    )

    def _row_event_ts(r):
        candidates = [
            r.get("delivery_time"), r.get("delivery_failure_time"),
            r.get("start_delivery_time"), r.get("outbound_time"),
            r.get("receipt_time"), r.get("creation_time"),
        ]
        return max((c for c in candidates if c), default=None)

    STATUS_PRIORITY = {
        "Delivered": 4, "Attempt Failure": 3,
        "Driver_received": 2, "Assigned": 1,
        "Unassign": 0, "Cancel": -1,
        "Stale_Armazem": -2,
    }

    by_wb = defaultdict(list)
    for r in rows:
        by_wb[r["waybill_number"]].append(r)

    keepers = {}     # waybill → row id a manter
    to_delete = []   # row ids a apagar
    history_to_create = []
    EPOCH = _dt(1970, 1, 1, tzinfo=_dt(2000, 1, 1).astimezone().tzinfo or None)
    from datetime import timezone as _tz_const
    EPOCH = _dt(1970, 1, 1, tzinfo=_tz_const.utc)

    for wb, rs in by_wb.items():
        # ordenar por (last_ts desc, priority desc) — primeiro é o keeper
        def sk(r):
            ts = _row_event_ts(r) or EPOCH
            pri = STATUS_PRIORITY.get(r["task_status"], 0)
            return (ts, pri)
        rs_sorted = sorted(rs, key=sk, reverse=True)
        keeper = rs_sorted[0]
        keepers[wb] = keeper["id"]
        # As outras rows são "consolidadas" — apagar e registar history
        for r in rs_sorted[1:]:
            to_delete.append(r["id"])
            history_to_create.append({
                "waybill_number": wb,
                "task_date": keeper["task_date"],
                "task_status": keeper["task_status"],
                "courier_name": keeper["courier_name"],
                "courier_id_cainiao": keeper["courier_id_cainiao"],
                "previous_task_date": r["task_date"],
                "previous_task_status": r["task_status"],
                "previous_courier_name": r["courier_name"],
                "change_type": "date_move",
                "event_timestamp": _row_event_ts(r),
            })

    if dry_run:
        return JsonResponse({
            "success": True, "dry_run": True,
            "waybills_with_duplicates": len(duplicated_wbs),
            "rows_to_delete": len(to_delete),
            "history_would_create": len(history_to_create),
        })

    # Executar: bulk delete + bulk create history
    with transaction.atomic():
        if to_delete:
            CainiaoOperationTask.objects.filter(id__in=to_delete).delete()
        if history_to_create:
            CainiaoOperationTaskHistory.objects.bulk_create([
                CainiaoOperationTaskHistory(
                    waybill_number=h["waybill_number"],
                    task_date=h["task_date"],
                    task_status=h["task_status"],
                    courier_name=h["courier_name"],
                    courier_id_cainiao=h["courier_id_cainiao"],
                    previous_task_date=h["previous_task_date"],
                    previous_task_status=h["previous_task_status"],
                    previous_courier_name=h["previous_courier_name"],
                    change_type=h["change_type"],
                    event_timestamp=h["event_timestamp"],
                )
                for h in history_to_create
            ], batch_size=1000)

    return JsonResponse({
        "success": True, "dry_run": False,
        "waybills_with_duplicates": len(duplicated_wbs),
        "rows_deleted": len(to_delete),
        "history_created": len(history_to_create),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_roll_forward_view(request):
    """Endpoint admin para mover pacotes Driver_received "vivos" (1-7 dias)
    para a data de HOJE.

    Body JSON opcional:
        { "dry_run": true, "threshold_days": 7 }

    Cria entries no CainiaoOperationTaskHistory para preservar timeline.
    """
    import json
    from .tasks import roll_forward_active_packages
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    dry_run = bool(body.get("dry_run"))
    threshold_days = body.get("threshold_days")

    result = roll_forward_active_packages(
        threshold_days=threshold_days,
        dry_run=dry_run,
    )
    return JsonResponse({"success": True, **result})


@login_required
@require_http_methods(["POST"])
def cainiao_mark_stale_armazem_view(request):
    """Endpoint admin para marcar pacotes presos no armazém como Stale.

    Body JSON opcional:
        { "dry_run": true, "threshold_days": 5 }

    - dry_run: só conta e devolve breakdown sem alterar BD.
    - threshold_days: dias após os quais um waybill é stale (default 5).

    Cria entries no CainiaoOperationTaskHistory para preservar timeline.
    Pode também ser chamado pelo Celery Beat às 03:00 diariamente.
    """
    import json
    from .tasks import mark_stale_armazem
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    dry_run = bool(body.get("dry_run"))
    threshold_days = body.get("threshold_days")

    result = mark_stale_armazem(
        threshold_days=threshold_days,
        dry_run=dry_run,
    )
    return JsonResponse({"success": True, **result})


@login_required
@require_http_methods(["POST"])
def cainiao_mark_deliveries_courier_id(request):
    """Marca entregas antigas com o Courier ID actual.

    O caso de uso: o courier_name na Cainiao foi renomeado (ex: 'Maria Freitas LF'
    → 'Maria_Freitas_LF'), mas o Courier ID permanece o mesmo. Entregas antigas
    no DB têm o nome antigo SEM courier_id_cainiao porque não havia mapping na
    altura. Esta função SÓ faz UPDATE no campo courier_id_cainiao das entregas
    com o courier_name antigo.

    NÃO cria DriverCourierMapping.
    NÃO modifica DriverProfile.
    NÃO toca em mais nada.

    Body JSON: { old_courier_name: str, target_courier_id: str }
    """
    import json
    from django.db import transaction
    from .models import (
        CainiaoOperationTask, DriverCourierMapping, CourierNameAlias,
    )

    try:
        body = json.loads(request.body or b"{}")
        old_name = (body.get("old_courier_name") or "").strip()
        target_id = (body.get("target_courier_id") or "").strip()
    except json.JSONDecodeError:
        return JsonResponse(
            {"success": False, "error": "JSON inválido"}, status=400,
        )

    if not old_name or not target_id:
        return JsonResponse(
            {"success": False, "error":
             "old_courier_name e target_courier_id obrigatórios"},
            status=400,
        )

    # Verifica se o target_courier_id existe num mapping CAINIAO
    mapping = DriverCourierMapping.objects.filter(
        partner__name__iexact="CAINIAO", courier_id=target_id
    ).select_related("driver", "partner").first()
    if not mapping:
        return JsonResponse(
            {"success": False, "error":
             f"Courier ID {target_id} não tem mapping CAINIAO."},
            status=404,
        )

    with transaction.atomic():
        updated = CainiaoOperationTask.objects.filter(
            courier_name=old_name,
            courier_id_cainiao="",
        ).update(courier_id_cainiao=target_id)

        # Persistir alias para que re-imports respeitem este mapping
        # automaticamente (sem ser necessário re-mapear manualmente).
        CourierNameAlias.objects.update_or_create(
            partner=mapping.partner,
            courier_name=old_name,
            defaults={"courier_id": target_id, "source": "manual"},
        )

    return JsonResponse({
        "success": True,
        "updated": updated,
        "target_courier_id": target_id,
        "target_courier_name": mapping.courier_name,
        "driver_name": mapping.driver.nome_completo,
    })


@login_required
def cainiao_existing_logins(request):
    """Lista todos os logins (DriverCourierMapping) CAINIAO para escolha
    no UI de "marcar entregas com courier_id existente".

    GET ?q=... filtra por courier_id, courier_name ou nome do driver.
    """
    from .models import DriverCourierMapping
    q = (request.GET.get("q") or "").strip().lower()
    qs = DriverCourierMapping.objects.filter(
        partner__name__iexact="CAINIAO"
    ).select_related("driver").order_by("courier_name")

    rows = []
    for m in qs:
        if q:
            haystack = " ".join([
                m.courier_id or "", m.courier_name or "",
                m.driver.nome_completo or "", m.driver.apelido or "",
            ]).lower()
            if q not in haystack:
                continue
        rows.append({
            "courier_id":   m.courier_id,
            "courier_name": m.courier_name,
            "driver_id":    m.driver_id,
            "driver_name":  m.driver.nome_completo,
            "apelido":      m.driver.apelido or "",
        })
        if len(rows) >= 50:
            break
    return JsonResponse({"success": True, "logins": rows})


@login_required
def cainiao_courier_backfill_status(request):
    """AJAX: estatística completa de entregas + sugestões automáticas.

    Devolve TODOS os courier_names pendentes (não TOP 30) e para cada
    um faz match (icontains + SOUNDEX) com DriverProfile para pré-popular
    a UI de mapeamento manual.
    """
    from django.db import connection
    from django.db.models import Count, Q
    from .models import CainiaoOperationTask
    from drivers_app.models import DriverProfile

    # Optional HUB filter: filtra entregas por CP4 do HUB
    hub_id = request.GET.get("hub_id", "").strip()
    qs = CainiaoOperationTask.objects.exclude(courier_name="")

    if hub_id:
        from .models import CainiaoHub
        try:
            hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(id=int(hub_id))
            cp4_set = list(hub.cp4_codes.values_list("cp4", flat=True))
            if cp4_set:
                hub_q = Q()
                for cp4 in cp4_set:
                    hub_q |= Q(zip_code__startswith=cp4)
                qs = qs.filter(hub_q)
            else:
                qs = qs.none()
        except (CainiaoHub.DoesNotExist, ValueError):
            pass

    total = qs.count()
    with_id = qs.exclude(courier_id_cainiao="").count()
    without_id = total - with_id

    # Audit: verifica se há IDs nas entregas sem mapping correspondente
    from .models import DriverCourierMapping
    ids_in_tasks = set(
        qs.exclude(courier_id_cainiao="")
        .values_list("courier_id_cainiao", flat=True).distinct()
    )
    ids_in_mappings = set(
        DriverCourierMapping.objects.filter(partner__name__iexact="CAINIAO")
        .values_list("courier_id", flat=True)
    )
    orphan_ids = sorted(ids_in_tasks - ids_in_mappings)

    # Lista TODOS os courier_names pendentes (no contexto do HUB filtrado)
    pending = list(
        qs.filter(courier_id_cainiao="")
        .values("courier_name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Tentar sugerir um driver para cada (apelido exact, nome exact, soundex)
    # Pre-load todos os drivers para lookup rápido
    drv_by_apelido = {
        d.apelido.strip().lower(): d
        for d in DriverProfile.objects.exclude(apelido="")
    }
    drv_by_name = {
        d.nome_completo.strip().lower(): d
        for d in DriverProfile.objects.all()
    }

    # SOUNDEX em batch
    soundex_drivers = {}
    if pending:
        with connection.cursor() as c:
            c.execute(
                "SELECT id, nome_completo, apelido, courier_id_cainiao, "
                "SOUNDEX(nome_completo) FROM drivers_app_driverprofile "
                "WHERE status != 'IRREGULAR'"
            )
            for did, name, apelido, cn_id, snd in c.fetchall():
                soundex_drivers.setdefault(snd, []).append({
                    "id": did, "name": name, "apelido": apelido or "",
                    "courier_id_cainiao": cn_id or "",
                })

    def _suggest(courier_name):
        cn = courier_name.strip().lower()
        # exact apelido
        d = drv_by_apelido.get(cn)
        if d:
            return {"id": d.id, "name": d.nome_completo,
                    "apelido": d.apelido or "",
                    "courier_id_cainiao": d.courier_id_cainiao or "",
                    "kind": "apelido"}
        # exact nome
        d = drv_by_name.get(cn)
        if d:
            return {"id": d.id, "name": d.nome_completo,
                    "apelido": d.apelido or "",
                    "courier_id_cainiao": d.courier_id_cainiao or "",
                    "kind": "nome"}
        # soundex
        with connection.cursor() as c:
            c.execute("SELECT SOUNDEX(%s)", [courier_name])
            snd = c.fetchone()[0]
        candidates = soundex_drivers.get(snd, [])
        if len(candidates) == 1:
            cd = candidates[0]
            return {**cd, "kind": "soundex"}
        return None

    enriched = []
    for p in pending:
        sug = _suggest(p["courier_name"])
        enriched.append({
            "courier_name": p["courier_name"],
            "total": p["total"],
            "suggestion": sug,
        })

    return JsonResponse({
        "success": True,
        "total": total,
        "with_id": with_id,
        "without_id": without_id,
        "pct_resolved": round(with_id / total * 100, 1) if total else 0,
        "unmapped_names": enriched,
        "orphan_ids_count": len(orphan_ids),
        "orphan_ids": orphan_ids[:20],
    })


@login_required
@require_http_methods(["POST"])
def cainiao_courier_backfill(request):
    """Backfill: para cada CainiaoOperationTask sem courier_id_cainiao,
    tenta resolver via DriverCourierMapping/DriverProfile actuais e gravar.

    Aceita hub_id (form ou query) para limitar o backfill ao HUB.
    """
    from django.db import transaction
    from django.db.models import Q
    from .models import (
        CainiaoOperationTask, DriverCourierMapping, CainiaoHub,
        CourierNameAlias,
    )
    from drivers_app.models import DriverProfile
    from core.models import Partner

    cainiao = Partner.objects.filter(name__iexact="CAINIAO").first()
    if not cainiao:
        return JsonResponse(
            {"success": False, "error": "Parceiro CAINIAO não configurado."},
            status=400,
        )

    # Optional HUB filter
    hub_id = (request.POST.get("hub_id")
              or request.GET.get("hub_id") or "").strip()
    hub_filter_q = None
    if hub_id:
        try:
            hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(id=int(hub_id))
            cp4_set = list(hub.cp4_codes.values_list("cp4", flat=True))
            if cp4_set:
                hub_filter_q = Q()
                for cp4 in cp4_set:
                    hub_filter_q |= Q(zip_code__startswith=cp4)
        except (CainiaoHub.DoesNotExist, ValueError):
            pass

    # Construir lookup name → courier_id (consulta também CourierNameAlias)
    # Prioridade: mapping actual > alias persistente > apelido do driver.
    # NÃO usa nome_completo (é o nome real da pessoa, não o alias do parceiro).
    name_to_id = {}
    for m in DriverCourierMapping.objects.filter(partner=cainiao):
        if m.courier_name:
            name_to_id[m.courier_name] = m.courier_id
        name_to_id.setdefault(m.courier_id, m.courier_id)
    for a in CourierNameAlias.objects.filter(partner=cainiao):
        name_to_id.setdefault(a.courier_name, a.courier_id)
    for d in DriverProfile.objects.exclude(courier_id_cainiao="").exclude(apelido=""):
        name_to_id.setdefault(d.apelido, d.courier_id_cainiao)

    # Iterate em batches para não estourar memória
    from django.db.models import Count
    base_filter = (
        CainiaoOperationTask.objects.filter(courier_id_cainiao="")
        .exclude(courier_name="")
    )
    if hub_filter_q is not None:
        base_filter = base_filter.filter(hub_filter_q)

    pending = (
        base_filter.values("courier_name")
        .annotate(n=Count("id"))
    )

    updated = 0
    not_found = []
    with transaction.atomic():
        for row in pending:
            cname = row["courier_name"]
            cid = name_to_id.get(cname)
            if not cid:
                not_found.append({"courier_name": cname, "count": row["n"]})
                continue
            update_qs = CainiaoOperationTask.objects.filter(
                courier_name=cname, courier_id_cainiao=""
            )
            if hub_filter_q is not None:
                update_qs = update_qs.filter(hub_filter_q)
            n = update_qs.update(courier_id_cainiao=cid)
            updated += n

    return JsonResponse({
        "success": True,
        "updated": updated,
        "not_found": not_found,
        "not_found_count": sum(r["count"] for r in not_found),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_driver_stat_guided_import(request):
    """Import guiado: aceita decisões por linha (vincular / criar / agrupar / ignorar).

    Body POST (multipart):
        decisions (JSON):
          [{ courier_id, courier_name, dsp_name, total_parcels, signed_parcels,
             rate, action: "link"|"create"|"group"|"skip",
             driver_id?: int (para link/group, recebe o link)
          }, ...]
        period_range: string com período (ex: "2026-04-01~2026-04-25")
    """
    import json
    from decimal import Decimal
    from django.db import transaction
    from .models import (
        CainiaoDriverStatBatch, CainiaoDriverStat, DriverCourierMapping,
    )
    from drivers_app.models import DriverProfile
    from core.models import Partner

    raw = request.POST.get("decisions", "")
    try:
        decisions = json.loads(raw or "[]")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "JSON inválido"}, status=400)

    partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    if not partner:
        return JsonResponse(
            {"success": False, "error": "Parceiro CAINIAO não configurado."},
            status=400,
        )

    date_range = (request.POST.get("period_range") or "").strip()
    drivers_created = 0
    mappings_created = 0
    mappings_updated = 0
    linked = 0
    skipped = 0

    with transaction.atomic():
        batch = CainiaoDriverStatBatch.objects.create(
            filename=request.POST.get("filename", "guided_import"),
            dispatch_date_range=date_range,
            created_by=request.user,
        )

        stat_objs = []

        for dec in decisions:
            cid = (dec.get("courier_id") or "").strip()
            if not cid:
                continue
            cname = (dec.get("courier_name") or "").strip()
            action = dec.get("action", "skip")

            try:
                total_parcels = int(dec.get("total_parcels") or 0)
            except (ValueError, TypeError):
                total_parcels = 0
            try:
                signed_parcels = int(dec.get("signed_parcels") or 0)
            except (ValueError, TypeError):
                signed_parcels = 0

            stat_objs.append(CainiaoDriverStat(
                batch=batch, courier_id=cid, courier_name=cname,
                dsp_name=dec.get("dsp_name") or "",
                total_parcels=total_parcels,
                delivery_success_rate=dec.get("rate") or "",
                signed_parcels=signed_parcels,
                courier_status="",
                dispatch_date=date_range,
            ))

            if action == "skip":
                skipped += 1
                continue

            driver = None
            if action in ("link", "group"):
                drv_id = dec.get("driver_id")
                if not drv_id:
                    skipped += 1
                    continue
                try:
                    driver = DriverProfile.objects.get(pk=int(drv_id))
                except DriverProfile.DoesNotExist:
                    skipped += 1
                    continue
                linked += 1
                # Sync apelido / courier_id_cainiao se vazios
                changed = []
                if not driver.courier_id_cainiao:
                    driver.courier_id_cainiao = cid
                    changed.append("courier_id_cainiao")
                if not driver.apelido and cname:
                    driver.apelido = cname
                    changed.append("apelido")
                if changed:
                    driver.save(update_fields=changed)

            elif action == "create":
                fake_nif = ("9" + cid)[-9:] if cid else "000000000"
                base_nif = fake_nif
                n = 0
                while DriverProfile.objects.filter(nif=fake_nif).exists():
                    n += 1
                    fake_nif = f"{base_nif[:7]}{n:02d}"
                    if n > 99:
                        break
                driver = DriverProfile.objects.create(
                    nif=fake_nif,
                    nome_completo=cname or f"Courier {cid}",
                    apelido=cname,
                    courier_id_cainiao=cid,
                    telefone="000000000",
                    email=f"driver.{cid}@import.local",
                    tipo_vinculo="DIRETO",
                    status="PENDENTE",
                    is_active=False,
                    importado_auto=True,
                    observacoes=(
                        f"Criado via import guiado Driver Statistic "
                        f"(Courier ID: {cid})."
                    ),
                )
                drivers_created += 1

            # Cria/actualiza mapping
            if driver:
                m, created_m = DriverCourierMapping.objects.update_or_create(
                    partner=partner, courier_id=cid,
                    defaults={"driver": driver, "courier_name": cname},
                )
                if created_m:
                    mappings_created += 1
                else:
                    mappings_updated += 1

        CainiaoDriverStat.objects.bulk_create(stat_objs, batch_size=500)
        batch.total_couriers = len(stat_objs)
        batch.save(update_fields=["total_couriers"])

    return JsonResponse({
        "success": True,
        "batch_id": batch.id,
        "drivers_created": drivers_created,
        "linked": linked,
        "mappings_created": mappings_created,
        "mappings_updated": mappings_updated,
        "skipped": skipped,
        "total": len(decisions),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_driver_stat_import(request):
    """Import da planilha Driver Statistic.

    Acções por courier (Courier ID):
      1. Guarda CainiaoDriverStat (histórico para comparação/auditoria)
      2. Cria/atualiza DriverCourierMapping (partner=CAINIAO, courier_id)
      3. Se ainda não há DriverProfile linkado → cria automaticamente
         (status=PENDENTE, importado_auto=True). Status na planilha
         "Disable" ⇒ DriverProfile.is_active=False.

    Os campos Delivery Success Rate / Signed Parcels NÃO são usados para
    gerar pré-faturas (essas usam CainiaoOperationTask). Servem apenas
    para comparação e exibição.
    """
    from django.db import transaction
    from .models import (
        CainiaoDriverStatBatch, CainiaoDriverStat, DriverCourierMapping,
        CourierNameAlias,
    )
    from drivers_app.models import DriverProfile, EmpresaParceira
    from core.models import Partner

    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."}, status=400,
        )

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400,
        )

    header, header_idx = _find_header_row(all_rows, "Courier ID")
    if header is None:
        return JsonResponse(
            {"success": False, "error": "Coluna 'Courier ID' não encontrada."},
            status=400,
        )

    ci = {
        "id":     _col_idx(header, "Courier ID"),
        "name":   _col_idx(header, "Courier Name"),
        "dsp":    _col_idx(header, "DSP Name"),
        "total":  _col_idx(header, "Total Parcels"),
        "rate":   _col_idx(header, "Delivery Success Rate"),
        "signed": _col_idx(header, "Signed Parcels"),
        "status": _col_idx(header, "Courier Status"),
        "date":   _col_idx(header, "Dispatch Date"),
    }

    # Cainiao partner (para os mappings). Optional partner_id from request
    # caso a importação venha do dashboard de outro parceiro no futuro.
    partner_id = request.POST.get("partner_id")
    if partner_id:
        try:
            partner = Partner.objects.get(pk=int(partner_id))
        except (Partner.DoesNotExist, ValueError):
            partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    else:
        partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    if not partner:
        return JsonResponse(
            {"success": False, "error": "Parceiro CAINIAO não configurado."},
            status=400,
        )

    # Detect date range from data
    dates = [_cell_str(row, ci["date"]) for row in all_rows[header_idx + 1:]
             if _cell_str(row, ci["date"])]
    date_range = dates[0] if dates else ""

    drivers_created = 0
    drivers_updated = 0
    mappings_created = 0
    mappings_updated = 0
    couriers_processed = 0
    rows_detail = []

    with transaction.atomic():
        batch = CainiaoDriverStatBatch.objects.create(
            filename=ficheiro.name,
            dispatch_date_range=date_range,
            created_by=request.user,
        )

        # Pre-load existing mappings to speed up
        existing_mappings = {
            m.courier_id: m
            for m in DriverCourierMapping.objects.filter(partner=partner)
            .select_related("driver")
        }
        # Pre-load existing drivers — só por courier_id_cainiao (chave Cainiao
        # estável) e por apelido (alias do parceiro).
        # NÃO usamos nome_completo porque é o nome real (pode estar formatado
        # diferente do alias do ficheiro).
        existing_by_courier_id = {
            d.courier_id_cainiao: d for d in DriverProfile.objects.all()
            if d.courier_id_cainiao
        }
        existing_by_apelido = {
            d.apelido: d for d in DriverProfile.objects.all()
            if d.apelido
        }

        stat_objs = []

        for row in all_rows[header_idx + 1:]:
            cid = _cell_str(row, ci["id"])
            if not cid:
                continue
            couriers_processed += 1

            courier_name = _cell_str(row, ci["name"])  # virá do "Courier Name"
            dsp_name = _cell_str(row, ci["dsp"])
            courier_status = _cell_str(row, ci["status"]).strip().lower()
            try:
                total_parcels = int(_cell_str(row, ci["total"]))
            except (ValueError, TypeError):
                total_parcels = 0
            try:
                signed_parcels = int(_cell_str(row, ci["signed"]))
            except (ValueError, TypeError):
                signed_parcels = 0

            stat_objs.append(CainiaoDriverStat(
                batch=batch,
                courier_id=cid,
                courier_name=courier_name,
                dsp_name=dsp_name,
                total_parcels=total_parcels,
                delivery_success_rate=_cell_str(row, ci["rate"]),
                signed_parcels=signed_parcels,
                courier_status=_cell_str(row, ci["status"]),
                dispatch_date=_cell_str(row, ci["date"]),
            ))

            # ── Driver linking ────────────────────────────────────────
            # Prioridade do match:
            #   1. DriverCourierMapping(partner, courier_id) — link explícito
            #   2. DriverProfile.courier_id_cainiao == cid — chave Cainiao única
            #   3. DriverProfile.apelido == courier_name (alias)
            #   4. Cria novo DriverProfile (status PENDENTE)
            driver = None
            mapping = existing_mappings.get(cid)
            action = ""

            if mapping:
                driver = mapping.driver
                # Actualiza apelido se vazio
                if not driver.apelido and courier_name:
                    driver.apelido = courier_name
                    driver.save(update_fields=["apelido"])
                # Garante courier_id_cainiao preenchido
                if not driver.courier_id_cainiao:
                    driver.courier_id_cainiao = cid
                    driver.save(update_fields=["courier_id_cainiao"])
                if mapping.courier_name != courier_name:
                    # Preserva o nome ANTIGO como alias antes de overwrite
                    # (assim re-imports e tasks históricas resolvem para o
                    # mesmo courier_id mesmo após o rename)
                    if mapping.courier_name:
                        CourierNameAlias.objects.update_or_create(
                            partner=partner,
                            courier_name=mapping.courier_name,
                            defaults={"courier_id": cid, "source": "auto"},
                        )
                    mapping.courier_name = courier_name
                    mapping.save(update_fields=["courier_name"])
                    mappings_updated += 1
                action = "mapped"
            else:
                # 2. Match pelo Courier ID (chave Cainiao única)
                driver = existing_by_courier_id.get(cid)
                if driver:
                    if not driver.apelido and courier_name:
                        driver.apelido = courier_name
                        driver.save(update_fields=["apelido"])
                    drivers_updated += 1
                    action = "linked_by_courier_id"
                else:
                    # 3. Match pelo apelido
                    driver = existing_by_apelido.get(courier_name)
                    if driver:
                        if not driver.courier_id_cainiao:
                            driver.courier_id_cainiao = cid
                            driver.save(update_fields=["courier_id_cainiao"])
                        drivers_updated += 1
                        action = "linked_by_apelido"
                    else:
                        # 4. Criar novo DriverProfile
                        fake_nif = ("9" + cid)[-9:] if cid else "000000000"
                        base_nif = fake_nif
                        n = 0
                        while DriverProfile.objects.filter(nif=fake_nif).exists():
                            n += 1
                            fake_nif = f"{base_nif[:7]}{n:02d}"
                            if n > 99:
                                break
                        driver = DriverProfile.objects.create(
                            nif=fake_nif,
                            nome_completo=courier_name or f"Courier {cid}",
                            apelido=courier_name,
                            courier_id_cainiao=cid,
                            telefone="000000000",
                            email=f"driver.{cid}@import.local",
                            tipo_vinculo="DIRETO",
                            status="PENDENTE",
                            is_active=False,
                            importado_auto=True,
                            observacoes=(
                                f"Criado automaticamente pelo import "
                                f"Driver Statistic Cainiao "
                                f"(Courier ID: {cid}, Apelido: {courier_name}, "
                                f"DSP: {dsp_name}). Pendente de aprovação. "
                                f"Edita NIF/nome reais ao aprovar."
                            ),
                        )
                        drivers_created += 1
                        action = "created"
                        existing_by_courier_id[cid] = driver
                        if courier_name:
                            existing_by_apelido[courier_name] = driver

                # Cria mapping
                mapping = DriverCourierMapping.objects.create(
                    partner=partner,
                    courier_id=cid,
                    courier_name=courier_name,
                    driver=driver,
                )
                existing_mappings[cid] = mapping
                mappings_created += 1

            # Persistir alias do nome actual (para que tasks futuras com este
            # nome resolvam para este courier_id mesmo após renames futuros)
            if courier_name:
                CourierNameAlias.objects.update_or_create(
                    partner=partner, courier_name=courier_name,
                    defaults={"courier_id": cid, "source": "import"},
                )

            # Status do courier na planilha → activa/desactiva driver
            if courier_status == "disable" and driver.is_active:
                driver.is_active = False
                driver.status = "BLOQUEADO"
                driver.save(update_fields=["is_active", "status"])

            rows_detail.append({
                "courier_id": cid,
                "courier_name": courier_name,
                "driver_id": driver.id if driver else None,
                "driver_name": driver.nome_completo if driver else None,
                "action": action,
                "total": total_parcels,
                "rate": _cell_str(row, ci["rate"]),
            })

        CainiaoDriverStat.objects.bulk_create(stat_objs, batch_size=500)
        batch.total_couriers = len(stat_objs)
        batch.save(update_fields=["total_couriers"])

    # Mapear sample no shape esperado pelo modal genérico (action: new/update)
    sample_for_modal = [
        {
            "courier": r["courier_name"],
            "id": r["courier_id"],
            "total": r["total"],
            "rate": r["rate"],
            "action": "new" if r["action"] == "created" else "update",
        }
        for r in rows_detail[:10]
    ]
    # Status counts agregados
    action_counts = {}
    for r in rows_detail:
        action_counts[r["action"]] = action_counts.get(r["action"], 0) + 1

    return JsonResponse({
        "success": True,
        "total": couriers_processed,
        "total_novos": drivers_created,
        "total_atualizados": drivers_updated + mappings_updated,
        "batch_id": batch.id,
        "date_range": date_range,
        # detalhe extra
        "drivers_created": drivers_created,
        "drivers_updated": drivers_updated,
        "mappings_created": mappings_created,
        "mappings_updated": mappings_updated,
        "courier_counts": action_counts,
        "sample": sample_for_modal,
        "rows": rows_detail[:50],  # cap for UI
    })


# ============================================================================
# IMPORT — PLANILHA DRIVER DETAIL INFO (16 colunas)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_driver_detail_preview(request):
    """Preview da planilha Driver Detail Info — sem gravar na BD."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    report_date_str = request.POST.get("report_date", "").strip()
    if not report_date_str:
        return JsonResponse({"success": False, "error": "Data do relatório é obrigatória."}, status=400)

    try:
        report_date = _date.fromisoformat(report_date_str)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data inválida."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Waybill Number' não encontrada."}, status=400)

    ci_waybill = _col_idx(header, "Waybill Number")
    ci_status  = _col_idx(header, "Task Status")
    ci_courier = _col_idx(header, "Courier Name")
    ci_courier_id = _col_idx(header, "Courier ID")

    from .models import CainiaoDriverDetailRecord, CainiaoDriverDetailBatch
    # Find existing batch for this date to check duplicates
    existing_waybills = set(
        CainiaoDriverDetailRecord.objects.filter(batch__report_date=report_date)
        .values_list("waybill_number", flat=True)
    )

    status_counts = {}
    courier_counts = {}
    total_novos = 0
    total_atualizados = 0
    sample = []
    seen = set()

    for row in all_rows[header_idx + 1:]:
        waybill = _cell_str(row, ci_waybill)
        if not waybill or waybill in seen:
            continue
        seen.add(waybill)

        status_val = _cell_str(row, ci_status)
        courier    = _cell_str(row, ci_courier)
        status_counts[status_val] = status_counts.get(status_val, 0) + 1
        if courier:
            courier_counts[courier] = courier_counts.get(courier, 0) + 1

        is_new = waybill not in existing_waybills
        if is_new:
            total_novos += 1
        else:
            total_atualizados += 1

        if len(sample) < 10:
            sample.append({
                "waybill": waybill,
                "status": status_val,
                "courier": courier,
                "action": "new" if is_new else "update",
            })

    return JsonResponse({
        "success": True,
        "total_novos": total_novos,
        "total_atualizados": total_atualizados,
        "total": total_novos + total_atualizados,
        "report_date": report_date_str,
        "status_counts": status_counts,
        "courier_counts": courier_counts,
        "sample": sample,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_driver_detail_import(request):
    """Import da planilha Driver Detail Info — upsert por (batch, waybill_number)."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    report_date_str = request.POST.get("report_date", "").strip()
    if not report_date_str:
        return JsonResponse({"success": False, "error": "Data do relatório é obrigatória."}, status=400)

    try:
        report_date = _date.fromisoformat(report_date_str)
    except ValueError:
        return JsonResponse({"success": False, "error": "Data inválida."}, status=400)

    try:
        all_rows = _load_workbook_rows(ficheiro.read())
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao ler ficheiro: {e}"}, status=400)

    header, header_idx = _find_header_row(all_rows, "Waybill Number")
    if header is None:
        return JsonResponse({"success": False, "error": "Coluna 'Waybill Number' não encontrada."}, status=400)

    ci = {
        "courier_id":    _col_idx(header, "Courier ID"),
        "courier_name":  _col_idx(header, "Courier Name"),
        "helper":        _col_idx(header, "Courier Helper"),
        "telephone":     _col_idx(header, "Courier Telephone"),
        "c_status":      _col_idx(header, "Courier Status"),
        "dsp":           _col_idx(header, "DSP Name"),
        "lp":            _col_idx(header, "LP No."),
        "waybill":       _col_idx(header, "Waybill Number"),
        "task_status":   _col_idx(header, "Task Status"),
        "del_time":      _col_idx(header, "Delivery Time"),
        "del_fail":      _col_idx(header, "Delivery Failure Time"),
        "exc_type":      _col_idx(header, "Exception Type"),
        "exc_detail":    _col_idx(header, "Exception Detail"),
        "weight":        _col_idx(header, "Weight (g)"),
        "del_type":      _col_idx(header, "Delivery Type"),
        "pudo":          _col_idx(header, "PUDO"),
    }

    from .models import CainiaoDriverDetailBatch, CainiaoDriverDetailRecord

    batch = CainiaoDriverDetailBatch.objects.create(
        filename=ficheiro.name,
        report_date=report_date,
        created_by=request.user,
    )

    to_create = []
    seen = set()

    for row in all_rows[header_idx + 1:]:
        waybill = _cell_str(row, ci["waybill"])
        if not waybill or waybill in seen:
            continue
        seen.add(waybill)

        to_create.append(CainiaoDriverDetailRecord(
            batch=batch,
            courier_id=_cell_str(row, ci["courier_id"]),
            courier_name=_cell_str(row, ci["courier_name"]),
            courier_helper=_cell_str(row, ci["helper"]),
            courier_telephone=_cell_str(row, ci["telephone"]),
            courier_status=_cell_str(row, ci["c_status"]),
            dsp_name=_cell_str(row, ci["dsp"]),
            lp_number=_cell_str(row, ci["lp"]),
            waybill_number=waybill,
            task_status=_cell_str(row, ci["task_status"]),
            delivery_time=_parse_datetime(row[ci["del_time"]] if ci["del_time"] is not None else None),
            delivery_failure_time=_parse_datetime(row[ci["del_fail"]] if ci["del_fail"] is not None else None),
            exception_type=_cell_str(row, ci["exc_type"]),
            exception_detail=_cell_str(row, ci["exc_detail"]),
            weight_g=_cell_float(row, ci["weight"]),
            delivery_type=_cell_str(row, ci["del_type"]),
            pudo_locker_id=_cell_str(row, ci["pudo"]),
        ))

    CainiaoDriverDetailRecord.objects.bulk_create(to_create, batch_size=500)
    batch.total_records = len(to_create)
    batch.new_records = len(to_create)
    batch.save(update_fields=["total_records", "new_records"])

    return JsonResponse({
        "success": True,
        "total": len(to_create),
        "batch_id": batch.id,
        "report_date": report_date_str,
    })


# ---------------------------------------------------------------------------
# HUBs API
# ---------------------------------------------------------------------------

def _hub_to_dict(hub):
    return {
        "id": hub.id,
        "name": hub.name,
        "address": hub.address,
        "cp4_codes": list(hub.cp4_codes.values_list("cp4", flat=True).order_by("cp4")),
    }


@login_required
def cainiao_hubs_list(request):
    """GET — lista todos os HUBs com os seus CP4s."""
    from .models import CainiaoHub
    hubs = CainiaoHub.objects.prefetch_related("cp4_codes").order_by("name")
    return JsonResponse({"success": True, "hubs": [_hub_to_dict(h) for h in hubs]})


@login_required
@require_http_methods(["POST"])
def cainiao_hub_create(request):
    """POST {name, address} — cria um novo HUB."""
    from .models import CainiaoHub
    import json
    try:
        body = json.loads(request.body)
    except Exception:
        body = request.POST
    name = (body.get("name") or "").strip()
    address = (body.get("address") or "").strip()
    if not name:
        return JsonResponse({"success": False, "error": "Nome obrigatório."}, status=400)
    if CainiaoHub.objects.filter(name__iexact=name).exists():
        return JsonResponse({"success": False, "error": "HUB com esse nome já existe."}, status=400)
    hub = CainiaoHub.objects.create(name=name, address=address)
    return JsonResponse({"success": True, "hub": _hub_to_dict(hub)})


@login_required
@require_http_methods(["POST"])
def cainiao_hub_update(request, hub_id):
    """POST {name, address} — actualiza nome/endereço do HUB."""
    from .models import CainiaoHub
    import json
    hub = get_object_or_404(CainiaoHub, id=hub_id)
    try:
        body = json.loads(request.body)
    except Exception:
        body = request.POST
    name = (body.get("name") or "").strip()
    address = (body.get("address") or "").strip()
    if not name:
        return JsonResponse({"success": False, "error": "Nome obrigatório."}, status=400)
    if CainiaoHub.objects.filter(name__iexact=name).exclude(id=hub_id).exists():
        return JsonResponse({"success": False, "error": "Outro HUB já tem esse nome."}, status=400)
    hub.name = name
    hub.address = address
    hub.save(update_fields=["name", "address"])
    return JsonResponse({"success": True, "hub": _hub_to_dict(hub)})


@login_required
@require_http_methods(["POST"])
def cainiao_hub_delete(request, hub_id):
    """POST — apaga o HUB (e os seus CP4s em cascata)."""
    from .models import CainiaoHub
    hub = get_object_or_404(CainiaoHub, id=hub_id)
    hub.delete()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def cainiao_hub_cp4_add(request, hub_id):
    """POST {cp4} — adiciona um código CP4 ao HUB."""
    from .models import CainiaoHub, CainiaoHubCP4
    import json
    hub = get_object_or_404(CainiaoHub, id=hub_id)
    try:
        body = json.loads(request.body)
    except Exception:
        body = request.POST
    cp4 = (body.get("cp4") or "").strip()
    if not cp4 or len(cp4) != 4 or not cp4.isdigit():
        return JsonResponse({"success": False, "error": "CP4 deve ter 4 dígitos."}, status=400)
    _, created = CainiaoHubCP4.objects.get_or_create(hub=hub, cp4=cp4)
    if not created:
        return JsonResponse({"success": False, "error": "CP4 já existe neste HUB."}, status=400)
    return JsonResponse({"success": True, "hub": _hub_to_dict(hub)})


@login_required
@require_http_methods(["POST"])
def cainiao_hub_cp4_remove(request, hub_id):
    """POST {cp4} — remove um código CP4 do HUB."""
    from .models import CainiaoHub, CainiaoHubCP4
    import json
    hub = get_object_or_404(CainiaoHub, id=hub_id)
    try:
        body = json.loads(request.body)
    except Exception:
        body = request.POST
    cp4 = (body.get("cp4") or "").strip()
    deleted, _ = CainiaoHubCP4.objects.filter(hub=hub, cp4=cp4).delete()
    if not deleted:
        return JsonResponse({"success": False, "error": "CP4 não encontrado neste HUB."}, status=404)
    return JsonResponse({"success": True, "hub": _hub_to_dict(hub)})


# ============================================================================
# DRILL-DOWN: DRIVERS / PACKAGES / WAYBILL DETAIL
# ============================================================================

_DRILL_VALID_STATUSES = ("Delivered", "Driver_received", "Attempt Failure", "Unassign", "Assigned")


def _apply_operation_filters(qs, request):
    """Aplica filtros comuns (date_from, date_to, hub_id, cp4[]) a um queryset
    de CainiaoOperationTask. Devolve (qs, info_dict) para uso em views."""
    from django.db.models import Q
    from .models import CainiaoHub, CainiaoHubCP4

    date_from_str = request.GET.get("date_from", "")
    date_to_str   = request.GET.get("date_to", "")
    op_date_str   = request.GET.get("op_date", "")
    if not date_from_str:
        date_from_str = op_date_str
    if not date_to_str:
        date_to_str = op_date_str

    try:
        date_from = _date.fromisoformat(date_from_str) if date_from_str else None
    except ValueError:
        date_from = None
    try:
        date_to = _date.fromisoformat(date_to_str) if date_to_str else None
    except ValueError:
        date_to = None
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    if date_from and date_to:
        qs = qs.filter(task_date__range=(date_from, date_to))
    elif date_from:
        qs = qs.filter(task_date__gte=date_from)
    elif date_to:
        qs = qs.filter(task_date__lte=date_to)

    hub_id_str = request.GET.get("hub_id", "")
    selected_hub = None
    hub_cp4_set = set()
    if hub_id_str:
        try:
            selected_hub = CainiaoHub.objects.prefetch_related("cp4_codes").get(id=int(hub_id_str))
            hub_cp4_set = set(selected_hub.cp4_codes.values_list("cp4", flat=True))
        except (CainiaoHub.DoesNotExist, ValueError):
            selected_hub = None

    if hub_cp4_set:
        hub_q = Q()
        for cp4 in hub_cp4_set:
            hub_q |= Q(zip_code__startswith=cp4)
        qs = qs.filter(hub_q)

    cp4s = [c for c in request.GET.getlist("cp4") if c]
    if cp4s:
        cp4_q = Q()
        for cp4 in cp4s:
            cp4_q |= Q(zip_code__startswith=cp4)
        qs = qs.filter(cp4_q)

    return qs, {
        "date_from": date_from,
        "date_to": date_to,
        "hub_id": hub_id_str,
        "selected_hub": selected_hub,
        "cp4s": cp4s,
    }


@login_required
def cainiao_drivers_by_status(request):
    """AJAX: lista drivers com pacotes num (ou mais) status, respeitando filtros.
    `status` pode ser um único valor ou vários separados por vírgula."""
    from django.db.models import Count, Max, Q
    from django.db.models.functions import Substr
    from .models import CainiaoOperationTask

    status_raw = request.GET.get("status", "")
    parts = [s.strip() for s in status_raw.split(",") if s.strip()]
    is_returned_filter = "returned" in parts
    focus_statuses = [s for s in parts if s in _DRILL_VALID_STATUSES]
    if not focus_statuses and not is_returned_filter:
        return JsonResponse({"success": False, "error": "status inválido"}, status=400)

    base_qs = CainiaoOperationTask.objects.filter(task_status__in=_DRILL_VALID_STATUSES)
    base_qs, info = _apply_operation_filters(base_qs, request)

    # Pacotes devolvidos: separar do queryset principal e tratar como
    # categoria própria. Para statuses normais (Driver_received, etc.),
    # devolvidos são excluídos. Quando status=returned, mostramos só eles.
    from .models import WaybillReturn as _WR
    _returned_set = set(
        _WR.objects.filter(
            return_status__in=(_WR.STATUS_RETURNED, _WR.STATUS_CLOSED),
        ).values_list("waybill_number", flat=True)
    )
    if is_returned_filter:
        base_qs = base_qs.filter(waybill_number__in=_returned_set)
    else:
        base_qs = base_qs.exclude(waybill_number__in=_returned_set)

    # ── RESOLUÇÃO POR ESTADO ACTUAL ─────────────────────────────────────
    # Para cada waybill candidato, pegamos a última row (estado real).
    # Depois filtramos por focus contra esse estado actual.
    candidate_wbs = list(
        base_qs.values_list("waybill_number", flat=True).distinct(),
    )
    period_objs = []
    if candidate_wbs:
        seen = set()
        for op in (
            CainiaoOperationTask.objects.filter(
                waybill_number__in=candidate_wbs,
                task_status__in=_DRILL_VALID_STATUSES,
            )
            .order_by("waybill_number", "-task_date", "-id")
        ):
            if op.waybill_number in seen:
                continue
            seen.add(op.waybill_number)
            period_objs.append(op)

    # Filtro focus aplicado ao estado actual
    if is_returned_filter:
        focused_objs = [
            op for op in period_objs
            if op.waybill_number in _returned_set
        ]
    else:
        focused_objs = [
            op for op in period_objs
            if op.task_status in focus_statuses
        ]

    # Agregar por courier_name (estado actual)
    from collections import defaultdict
    merged = defaultdict(lambda: {
        "courier_name": "", "dsp_name": "",
        "total": 0, "focus": 0,
        "delivered": 0, "failures": 0,
        "last_delivery": None,
        "cp4s": set(),
    })
    # Total/Delivered/Failures por courier — usar período total
    # (mostrar context: "este driver tem 81 pacotes, 3 em foco")
    for op in period_objs:
        if not op.courier_name:
            continue
        m = merged[op.courier_name]
        m["courier_name"] = op.courier_name
        if not m["dsp_name"]:
            m["dsp_name"] = op.dsp_name or ""
        m["total"] += 1
        if op.task_status == "Delivered":
            m["delivered"] += 1
            if op.delivery_time and (
                m["last_delivery"] is None
                or op.delivery_time > m["last_delivery"]
            ):
                m["last_delivery"] = op.delivery_time
        elif op.task_status == "Attempt Failure":
            m["failures"] += 1
    for op in focused_objs:
        if not op.courier_name:
            continue
        m = merged[op.courier_name]
        m["courier_name"] = op.courier_name
        m["focus"] += 1
        if op.zip_code:
            m["cp4s"].add(op.zip_code[:4])

    result = []
    for name, m in merged.items():
        if m["focus"] == 0:
            continue
        success_rate = round(
            m["delivered"] / m["total"] * 100, 1,
        ) if m["total"] else 0
        result.append({
            "courier_name": name,
            "dsp_name": m["dsp_name"],
            "focus_count": m["focus"],
            "total": m["total"],
            "delivered": m["delivered"],
            "failures": m["failures"],
            "success_rate": success_rate,
            "last_delivery": (
                m["last_delivery"].isoformat()
                if m["last_delivery"] else None
            ),
            "cp4s": sorted(m["cp4s"]),
        })

    result.sort(key=lambda r: -r["focus_count"])

    return JsonResponse({
        "success": True,
        "statuses": (
            ["returned"] if is_returned_filter else focus_statuses
        ),
        "total_drivers": len(result),
        "total_packages": sum(r["focus_count"] for r in result),
        "drivers": result,
    })


@login_required
def cainiao_driver_packages(request):
    """Página: lista de pacotes de um driver (respeita filtros do dashboard)."""
    from django.db.models import Count, Q
    from django.db.models.functions import Substr
    from .models import CainiaoOperationTask

    courier_name = request.GET.get("courier", "").strip()
    if not courier_name:
        return HttpResponse("Parâmetro 'courier' obrigatório.", status=400)

    status_filter = request.GET.get("status", "")
    _ALLOWED_FILTERS = _DRILL_VALID_STATUSES + ("returned",)
    if status_filter and status_filter not in _ALLOWED_FILTERS:
        status_filter = ""

    base_qs = CainiaoOperationTask.objects.filter(
        courier_name=courier_name,
        task_status__in=_DRILL_VALID_STATUSES,
    )
    base_qs, info = _apply_operation_filters(base_qs, request)

    # Pacotes com WaybillReturn fechada
    from .models import WaybillReturn as _WR
    candidate_wbs = list(
        base_qs.values_list("waybill_number", flat=True).distinct(),
    )
    returns_by_wb = {
        r.waybill_number: r
        for r in _WR.objects.filter(
            waybill_number__in=candidate_wbs,
            return_status__in=(_WR.STATUS_RETURNED, _WR.STATUS_CLOSED),
        )
    }

    # ── RESOLUÇÃO DE ESTADO ACTUAL ─────────────────────────────────────
    # Cada waybill (deste courier no período) → última row global.
    # O `current_obj` reflete o estado real actual; mas guardamos
    # também a "row deste courier" para exibir no template (data em
    # que ESTE courier tocou no pacote).
    latest_global = {}
    if candidate_wbs:
        for op in (
            CainiaoOperationTask.objects.filter(
                waybill_number__in=candidate_wbs,
                task_status__in=_DRILL_VALID_STATUSES,
            )
            .order_by("waybill_number", "-task_date", "-id")
        ):
            if op.waybill_number not in latest_global:
                latest_global[op.waybill_number] = op

    # Para cada waybill, qual a row mais recente DESTE courier (para mostrar)
    courier_row_by_wb = {}
    for op in (
        base_qs.order_by("waybill_number", "-task_date", "-id")
    ):
        if op.waybill_number not in courier_row_by_wb:
            courier_row_by_wb[op.waybill_number] = op

    # Status counts pelo estado ACTUAL (não pelo da row do courier)
    status_counts = {s: 0 for s in _DRILL_VALID_STATUSES}
    returned_count = 0
    for wb in candidate_wbs:
        if wb in returns_by_wb:
            returned_count += 1
            continue
        cur = latest_global.get(wb)
        if cur:
            status_counts[cur.task_status] = (
                status_counts.get(cur.task_status, 0) + 1
            )
    total_all = sum(status_counts.values())

    # Pre-built chip metadata
    _STATUS_META = [
        ("Delivered",       "Delivered",       "check-circle", "emerald"),
        ("Driver_received", "Driver Received", "truck",        "amber"),
        ("Assigned",        "Assigned",        "warehouse",    "blue"),
        ("Unassign",        "Unassign",        "package-2",    "slate"),
        ("Attempt Failure", "Attempt Failure", "x-circle",     "red"),
    ]
    status_chips = [
        {"key": k, "label": label, "icon": icon, "color": color,
         "count": status_counts.get(k, 0)}
        for k, label, icon, color in _STATUS_META
    ]
    if returned_count:
        status_chips.append({
            "key": "returned",
            "label": "Devolvidos",
            "icon": "undo-2",
            "color": "orange",
            "count": returned_count,
        })

    # ── Construção da lista de pacotes ─────────────────────────────────
    # Mostramos a row do CURRENT estado para o waybill (não a row antiga
    # do courier). Mas usamos o courier_row_by_wb para info auxiliar.
    if status_filter == "returned":
        target_wbs = list(returns_by_wb.keys())
    elif status_filter:
        target_wbs = [
            wb for wb, op in latest_global.items()
            if wb not in returns_by_wb
            and op.task_status == status_filter
        ]
    else:
        target_wbs = [
            wb for wb in candidate_wbs
            if wb not in returns_by_wb
        ]

    packages = []
    for wb in target_wbs[:500]:
        op = latest_global.get(wb) or courier_row_by_wb.get(wb)
        if op is None:
            continue
        ret = returns_by_wb.get(wb)
        op.return_record = ret
        op.is_returned = bool(ret)
        # row deste courier (informação contextual)
        op.courier_seen_row = courier_row_by_wb.get(wb)
        packages.append(op)
    packages.sort(
        key=lambda p: p.task_date, reverse=True,
    )

    # Stats header
    dsp_names = sorted({
        op.dsp_name for op in courier_row_by_wb.values() if op.dsp_name
    })
    cp4s_covered = sorted({
        op.zip_code[:4]
        for op in courier_row_by_wb.values() if op.zip_code
    })

    if status_filter == "returned":
        display_count = returned_count
    else:
        display_count = (
            status_counts.get(status_filter, 0)
            if status_filter else total_all
        )

    context = {
        "courier_name": courier_name,
        "dsp_names": dsp_names,
        "cp4s_covered": cp4s_covered,
        "status_filter": status_filter,
        "status_chips": status_chips,
        "total_all": total_all,
        "display_count": display_count,
        "packages": packages,
        "packages_shown": len(packages),
        "filters": info,
    }
    return render(request, "settlements/cainiao_driver_packages.html", context)


@login_required
def cainiao_waybill_detail(request, waybill):
    """Página: detalhe completo do pacote com timeline multi-dia."""
    from .models import (
        CainiaoOperationTask, CainiaoPlanningPackage,
        CainiaoForecastPackage, CainiaoDriverDetailRecord,
    )

    records = list(
        CainiaoOperationTask.objects
        .filter(waybill_number=waybill)
        .order_by("task_date")
    )
    if not records:
        return render(request, "settlements/cainiao_waybill_not_found.html", {
            "waybill": waybill,
        }, status=404)

    latest = records[-1]

    # Forecast (status simples) e detail record (helper / locker)
    forecast = (
        CainiaoForecastPackage.objects
        .filter(tracking_number=waybill)
        .order_by("-operation_date").first()
    )
    detail = (
        CainiaoDriverDetailRecord.objects
        .filter(waybill_number=waybill)
        .order_by("-created_at").first()
    )

    # Build timeline events across all records (multi-day)
    timeline = []
    for rec in records:
        events = [
            ("creation",        "Criação",              rec.creation_time),
            ("receipt",         "Recebido no hub",      rec.receipt_time),
            ("outbound",        "Saída do hub",         rec.outbound_time),
            ("start_delivery",  "Início da entrega",    rec.start_delivery_time),
            ("delivery",        "Entregue",             rec.delivery_time),
            ("failure",         "Falha de entrega",     rec.delivery_failure_time),
        ]
        day_events = [
            {"kind": k, "label": label, "time": t}
            for k, label, t in events if t
        ]
        day_events.sort(key=lambda e: e["time"])
        timeline.append({
            "record": rec,
            "task_date": rec.task_date,
            "events": day_events,
        })

    # Planning/forecast info if this waybill came from a planning import
    planning = (
        CainiaoPlanningPackage.objects
        .filter(parcel_id=waybill)
        .order_by("-operation_date")
        .first()
    )

    # Has map data?
    has_receiver_coords = bool(
        latest.receiver_latitude and latest.receiver_longitude
    )
    has_actual_coords = bool(
        latest.actual_latitude and latest.actual_longitude
    )

    # Cliente: prioriza planning (tem nome/telefone/email),
    # fallback para forecast (que pode ter nome/telefone)
    customer_name = ""
    customer_phone = ""
    customer_email = ""
    if planning:
        customer_name = planning.receiver_name or ""
        customer_phone = planning.receiver_phone or ""
        customer_email = planning.receiver_email or ""
    elif forecast:
        customer_name = forecast.receiver_name or ""
        customer_phone = forecast.receiver_phone or ""

    # Telefone normalizado para wa.me (só dígitos)
    wa_phone = "".join(c for c in customer_phone if c.isdigit())

    # Seller (planning > op task)
    seller_name = ""
    if planning and planning.seller_name:
        seller_name = planning.seller_name
    elif latest.seller_name:
        seller_name = latest.seller_name

    # Flag activo + resolução + devolução (se houver)
    from .models import (
        WaybillFlag, WaybillResolution, WaybillTag, WaybillWatch,
        WaybillReturn,
    )
    active_flag = WaybillFlag.objects.filter(
        waybill_number=waybill, cleared_at__isnull=True,
    ).select_related("flagged_by").first()
    resolution = WaybillResolution.objects.filter(
        waybill_number=waybill,
    ).select_related("resolved_by").first()
    return_record = WaybillReturn.objects.filter(
        waybill_number=waybill,
    ).select_related("marked_by", "batch").first()
    is_returned_closed = bool(
        return_record and not return_record.is_open
    )

    # Adicionar evento de devolução à timeline (se houver)
    if return_record:
        from datetime import datetime as _dtt, time as _tt
        from django.utils.timezone import (
            make_aware as _ma, is_aware as _ia,
        )
        ret_dt = _dtt.combine(return_record.return_date, _tt(0, 0))
        if not _ia(ret_dt):
            ret_dt = _ma(ret_dt)
        ret_label = (
            "Devolvido ao remetente"
            if is_returned_closed
            else f"Devolução: {return_record.get_return_status_display()}"
        )
        ret_event = {
            "kind": "return",
            "label": ret_label,
            "time": ret_dt,
            "reason": return_record.get_return_reason_display(),
            "tracking": return_record.return_tracking_number,
            "carrier": return_record.return_carrier,
            "cost": str(return_record.return_cost_eur),
            "notes": return_record.notes,
        }
        # Inserir no dia da devolução (cria novo bloco se não existir)
        ret_date = return_record.return_date
        slot = next(
            (t for t in timeline if t["task_date"] == ret_date),
            None,
        )
        if slot is None:
            timeline.append({
                "record": None,
                "task_date": ret_date,
                "events": [ret_event],
            })
            timeline.sort(key=lambda t: t["task_date"])
        else:
            slot["events"].append(ret_event)
            slot["events"].sort(key=lambda e: e["time"])

    # Priority computado + tentativas
    n_attempts = sum(
        1 for r in records if r.task_status == "Attempt Failure"
    )
    priority_info = _compute_package_priority(
        latest, n_attempts=n_attempts,
        is_returned=is_returned_closed,
    )

    # Trilha de drivers — quem pegou/tentou em cada momento
    drivers_seen = []  # ordem cronológica, sem duplicar consecutivos
    seen_set = set()
    for r in records:
        cn = (r.courier_name or "").strip()
        if cn and cn not in seen_set:
            drivers_seen.append({
                "name": cn,
                "first_date": r.task_date,
                "first_status": r.task_status,
                "dsp": (r.dsp_name or "").strip(),
            })
            seen_set.add(cn)
    unique_drivers_count = len(drivers_seen)
    attempt_records = [
        r for r in records if r.task_status == "Attempt Failure"
    ]

    # Tags + Watch (se utilizador autenticado)
    tags = list(WaybillTag.objects.filter(waybill_number=waybill))
    is_watching = False
    if request.user.is_authenticated:
        is_watching = WaybillWatch.objects.filter(
            waybill_number=waybill, user=request.user,
        ).exists()

    # Métricas de tempo (para o card de tempos)
    time_metrics = {}
    if records[0].creation_time and records[0].receipt_time:
        delta = records[0].receipt_time - records[0].creation_time
        time_metrics["creation_to_inbound"] = (
            f"{delta.days}d {delta.seconds // 3600}h"
        )
    if latest.receipt_time and latest.delivery_time:
        delta = latest.delivery_time - latest.receipt_time
        time_metrics["inbound_to_delivered"] = (
            f"{delta.days}d {delta.seconds // 3600}h"
        )
    if latest.start_delivery_time and latest.delivery_time:
        delta = latest.delivery_time - latest.start_delivery_time
        time_metrics["start_to_delivered"] = (
            f"{delta.seconds // 60}min"
        )

    # Histórico de mudanças (timeline operacional preservada)
    from .models import CainiaoOperationTaskHistory
    history_entries = list(
        CainiaoOperationTaskHistory.objects
        .filter(waybill_number=waybill)
        .order_by("recorded_at")
    )

    context = {
        "waybill": waybill,
        "records": records,
        "latest": latest,
        "timeline": timeline,
        "history_entries": history_entries,
        "planning": planning,
        "forecast": forecast,
        "detail": detail,
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "customer_email": customer_email,
        "wa_phone": wa_phone,
        "seller_name": seller_name,
        "has_receiver_coords": has_receiver_coords,
        "has_actual_coords": has_actual_coords,
        "total_attempts": len(records),
        "active_flag": active_flag,
        "resolution": resolution,
        "return_record": return_record,
        "priority_info": priority_info,
        "tags": tags,
        "is_watching": is_watching,
        "n_attempts": n_attempts,
        "time_metrics": time_metrics,
        "drivers_seen": drivers_seen,
        "unique_drivers_count": unique_drivers_count,
        "attempt_records": attempt_records,
    }
    return render(
        request, "settlements/cainiao_waybill_detail.html", context,
    )


@login_required
def cainiao_waybill_search(request):
    """AJAX: autocomplete de waybills por prefixo."""
    from .models import CainiaoOperationTask

    q = request.GET.get("q", "").strip()
    if len(q) < 4:
        return JsonResponse({"success": True, "results": []})

    rows = list(
        CainiaoOperationTask.objects
        .filter(waybill_number__icontains=q)
        .order_by("-task_date")
        .values("waybill_number", "task_status", "courier_name", "task_date")[:15]
    )

    seen = set()
    unique = []
    for r in rows:
        w = r["waybill_number"]
        if w in seen:
            continue
        seen.add(w)
        unique.append({
            "waybill": w,
            "status": r["task_status"],
            "courier": r["courier_name"] or "",
            "task_date": r["task_date"].isoformat() if r["task_date"] else None,
        })

    return JsonResponse({"success": True, "results": unique})


# ============================================================================
# MAP PAGE — geospatial visualisation of all deliveries
# ============================================================================

_MAP_PIN_LIMIT = 5000


def _coord(value):
    """Parse a CharField coord into a float; None if invalid or zero-ish."""
    if value is None or value == "":
        return None
    try:
        v = float(str(value).replace(",", ".").strip())
    except (TypeError, ValueError):
        return None
    if -0.0001 < v < 0.0001:
        return None
    return v


@login_required
def cainiao_map_pins(request):
    """AJAX: returns filtered pins for the map view."""
    from .models import CainiaoOperationTask

    base_qs = CainiaoOperationTask.objects.filter(task_status__in=_DRILL_VALID_STATUSES)
    base_qs, info = _apply_operation_filters(base_qs, request)

    status_raw = request.GET.get("status", "")
    statuses = [s.strip() for s in status_raw.split(",")
                if s.strip() in _DRILL_VALID_STATUSES]
    if statuses:
        base_qs = base_qs.filter(task_status__in=statuses)

    courier = request.GET.get("courier", "").strip()
    if courier:
        base_qs = base_qs.filter(courier_name=courier)

    source = request.GET.get("source", "actual")  # actual | receiver

    total_filtered = base_qs.count()

    fields = (
        "id", "waybill_number", "task_status", "courier_name", "dsp_name",
        "task_date", "zip_code", "destination_city", "detailed_address",
        "delivery_time", "delivery_failure_time", "start_delivery_time",
        "outbound_time", "receipt_time", "creation_time",
        "receiver_latitude", "receiver_longitude",
        "actual_latitude", "actual_longitude",
        "delivery_gap_distance", "exception_type",
    )

    rows = base_qs.values(*fields).order_by("-task_date")[:_MAP_PIN_LIMIT + 1]
    rows = list(rows)
    truncated = len(rows) > _MAP_PIN_LIMIT
    if truncated:
        rows = rows[:_MAP_PIN_LIMIT]

    pins = []
    missing_coords = 0
    for r in rows:
        rlat = _coord(r["receiver_latitude"])
        rlng = _coord(r["receiver_longitude"])
        alat = _coord(r["actual_latitude"])
        alng = _coord(r["actual_longitude"])

        if source == "receiver":
            lat, lng = rlat, rlng
        else:
            lat, lng = alat, alng
            if lat is None or lng is None:
                lat, lng = rlat, rlng

        if lat is None or lng is None:
            missing_coords += 1
            continue

        # `time` keeps the driver-interaction semantic (used by playback filter).
        # `order_time` has a broader fallback (adds outbound/receipt/creation)
        # so the route can be drawn even for packages that were not yet delivered.
        ts = r["delivery_time"] or r["delivery_failure_time"] or r["start_delivery_time"]
        ts_order = (
            r["delivery_time"] or r["delivery_failure_time"] or
            r["start_delivery_time"] or r["outbound_time"] or
            r["receipt_time"] or r["creation_time"]
        )
        pins.append({
            "id":         r["id"],
            "waybill":    r["waybill_number"],
            "status":     r["task_status"],
            "courier":    r["courier_name"] or "",
            "dsp":        r["dsp_name"] or "",
            "date":       r["task_date"].isoformat() if r["task_date"] else None,
            "zip":        r["zip_code"] or "",
            "city":       r["destination_city"] or "",
            "address":    r["detailed_address"] or "",
            "lat":        lat,
            "lng":        lng,
            "rlat":       rlat,
            "rlng":       rlng,
            "alat":       alat,
            "alng":       alng,
            "gap":        r["delivery_gap_distance"] or "",
            "exception":  r["exception_type"] or "",
            "time":       ts.isoformat() if ts else None,
            "order_time": ts_order.isoformat() if ts_order else None,
        })

    # Compute a "start point" for the route optimizer.
    # Preference: centroid of the selected HUB's pins if a hub is filtered,
    # otherwise centroid of all returned pins.
    start_lat = start_lng = None
    start_name = ""
    if pins:
        # If a hub was selected, try to locate its name and use pin centroid
        from .models import CainiaoHub
        hub_id = info.get("hub_id", "")
        if hub_id:
            try:
                h = CainiaoHub.objects.get(id=int(hub_id))
                start_name = h.name
            except (CainiaoHub.DoesNotExist, ValueError):
                pass
        # Use average of up to the first 200 pins to get a stable centroid
        sample = pins[:200]
        start_lat = sum(p["lat"] for p in sample) / len(sample)
        start_lng = sum(p["lng"] for p in sample) / len(sample)

    return JsonResponse({
        "success": True,
        "total_filtered": total_filtered,
        "returned": len(pins),
        "missing_coords": missing_coords,
        "truncated": truncated,
        "limit": _MAP_PIN_LIMIT,
        "pins": pins,
        "start": {"lat": start_lat, "lng": start_lng, "name": start_name} if start_lat else None,
    })


@login_required
def cainiao_map_view(request):
    """Page: full-screen delivery map with filters."""
    import json
    from django.utils.safestring import mark_safe
    from .models import CainiaoHub

    hubs = list(CainiaoHub.objects.prefetch_related("cp4_codes").order_by("name"))
    hub_meta = [{
        "id": h.id,
        "name": h.name,
        "latitude": _coord(h.latitude) if hasattr(h, "latitude") else None,
        "longitude": _coord(h.longitude) if hasattr(h, "longitude") else None,
        "cp4s": list(h.cp4_codes.values_list("cp4", flat=True)),
    } for h in hubs]

    context = {
        "hubs": hubs,
        "hub_meta_json": mark_safe(json.dumps(hub_meta)),
        "valid_statuses": list(_DRILL_VALID_STATUSES),
    }
    return render(request, "settlements/cainiao_map.html", context)


@login_required
def cainiao_map_compare(request):
    """Page: 2 synced maps side-by-side for day-to-day comparison."""
    from .models import CainiaoHub
    hubs = list(CainiaoHub.objects.order_by("name"))
    return render(request, "settlements/cainiao_map_compare.html", {
        "hubs": hubs,
        "valid_statuses": list(_DRILL_VALID_STATUSES),
    })


@login_required
def cainiao_map_missing_coords(request):
    """AJAX: list of packages without geo coords (given the filters)."""
    from .models import CainiaoOperationTask
    from django.db.models import Q

    base_qs = CainiaoOperationTask.objects.filter(
        task_status__in=_DRILL_VALID_STATUSES
    )
    base_qs, _ = _apply_operation_filters(base_qs, request)

    status_raw = request.GET.get("status", "")
    statuses = [s.strip() for s in status_raw.split(",")
                if s.strip() in _DRILL_VALID_STATUSES]
    if statuses:
        base_qs = base_qs.filter(task_status__in=statuses)
    courier = request.GET.get("courier", "").strip()
    if courier:
        base_qs = base_qs.filter(courier_name=courier)

    # No coords means all four lat/lng fields are empty OR parse to ~0
    missing = base_qs.filter(
        Q(actual_latitude="") | Q(actual_latitude__isnull=True),
        Q(actual_longitude="") | Q(actual_longitude__isnull=True),
        Q(receiver_latitude="") | Q(receiver_latitude__isnull=True),
        Q(receiver_longitude="") | Q(receiver_longitude__isnull=True),
    ).values(
        "waybill_number", "task_status", "courier_name", "task_date",
        "zip_code", "destination_city", "detailed_address",
    ).order_by("-task_date")[:500]

    rows = [{
        "waybill":  r["waybill_number"],
        "status":   r["task_status"],
        "courier":  r["courier_name"] or "",
        "date":     r["task_date"].isoformat() if r["task_date"] else None,
        "zip":      r["zip_code"] or "",
        "city":     r["destination_city"] or "",
        "address":  r["detailed_address"] or "",
    } for r in missing]

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "rows": rows,
    })


@login_required
def cainiao_map_overlays(request):
    """AJAX: aggregated overlays — HUB centroids, CP4 centroids, failure heatmap."""
    from .models import CainiaoOperationTask, CainiaoHub, CainiaoHubCP4

    base_qs = CainiaoOperationTask.objects.filter(
        task_status__in=_DRILL_VALID_STATUSES
    )
    base_qs, info = _apply_operation_filters(base_qs, request)

    # Fetch only what we need: zip_code, status, coords
    rows = list(
        base_qs.exclude(zip_code="")
        .values_list(
            "zip_code", "task_status",
            "receiver_latitude", "receiver_longitude",
            "actual_latitude", "actual_longitude",
        )
    )

    # CP4 → HUB mapping (for attribution of pins to HUBs)
    cp4_hub = {}
    for hub_id, cp4 in CainiaoHubCP4.objects.values_list("hub_id", "cp4"):
        cp4_hub.setdefault(cp4, hub_id)
    hub_names = dict(CainiaoHub.objects.values_list("id", "name"))

    # Aggregate centroids
    cp4_agg = {}     # cp4 → [sum_lat, sum_lng, n, city]
    hub_agg = {}     # hub_id → [sum_lat, sum_lng, n]
    heat_pts = []    # [[lat, lng, intensity=1], ...] Attempt Failure only
    heat_cap = 3000

    for zip_code, status, rlat_s, rlng_s, alat_s, alng_s in rows:
        # Prefer actual coords; fallback to receiver
        lat = _coord(alat_s) or _coord(rlat_s)
        lng = _coord(alng_s) or _coord(rlng_s)
        if lat is None or lng is None:
            continue
        cp4 = zip_code[:4] if zip_code else ""
        if cp4:
            entry = cp4_agg.setdefault(cp4, [0.0, 0.0, 0])
            entry[0] += lat; entry[1] += lng; entry[2] += 1
            hub_id = cp4_hub.get(cp4)
            if hub_id:
                h = hub_agg.setdefault(hub_id, [0.0, 0.0, 0])
                h[0] += lat; h[1] += lng; h[2] += 1
        if status == "Attempt Failure" and len(heat_pts) < heat_cap:
            heat_pts.append([lat, lng, 1])

    MIN_N = 5  # noise threshold for CP4 circles
    cp4_centroids = [
        {"cp4": cp4, "lat": s[0] / s[2], "lng": s[1] / s[2], "count": s[2]}
        for cp4, s in cp4_agg.items() if s[2] >= MIN_N
    ]
    hub_centroids = [
        {"id": hid, "name": hub_names.get(hid, f"HUB {hid}"),
         "lat": s[0] / s[2], "lng": s[1] / s[2], "count": s[2]}
        for hid, s in hub_agg.items() if s[2] > 0
    ]

    return JsonResponse({
        "success": True,
        "hubs": hub_centroids,
        "cp4_centroids": cp4_centroids,
        "failure_heat": heat_pts,
        "heat_capped": len(heat_pts) >= heat_cap,
    })


# ============================================================================
# Pacotes "Not Arrived" — atribuídos mas nunca picados pelo driver
# ============================================================================

def _business_days_between(start, end, holidays_set):
    """Conta dias úteis (seg-sex, excl. feriados) entre start e end inclusive.
    Ambos como datetime.date. Retorna 0 se end < start.
    """
    from datetime import timedelta
    if end < start:
        return 0
    n = 0
    cur = start
    while cur <= end:
        if cur.weekday() < 5 and cur not in holidays_set:
            n += 1
        cur += timedelta(days=1)
    return n


def _compute_not_arrived(
    hub_id=None, min_days=2, include_unregistered=False,
):
    """Calcula a lista de waybills 'Not Arrived'.

    Por defeito restringe a CP4s cadastrados em qualquer HUB
    (CainiaoHubCP4). `include_unregistered=True` desliga isso.
    """
    from datetime import timedelta
    from django.db.models import Min, Max, Count
    from django.utils import timezone
    from .models import (
        CainiaoOperationTask, CainiaoHubCP4, Holiday,
        CainiaoPlanningPackage, CainiaoForecastPackage,
        WaybillResolution, WaybillComment,
    )

    # CP4s a considerar:
    # - Se hub_id especificado, só desse HUB
    # - Senão, e include_unregistered=False, todos os CP4s registados
    # - Senão, sem filtro de CP4
    hub_cp4s = None
    if hub_id:
        hub_cp4s = list(
            CainiaoHubCP4.objects.filter(hub_id=hub_id)
            .values_list("cp4", flat=True)
        )
    elif not include_unregistered:
        all_registered = list(
            CainiaoHubCP4.objects.values_list(
                "cp4", flat=True,
            ).distinct()
        )
        if all_registered:
            hub_cp4s = all_registered

    # Status que indicam que o pacote SAIU do hub (foi físicamente
    # tocado pelo driver). Se algum snapshot tem um destes, NÃO é
    # Not Arrived.
    PICKED_STATUSES = (
        "Driver Received", "Driver_received",
        "Delivered", "Attempt Failure", "Attempt_Failure",
    )

    # Subquery: waybills que TÊM pelo menos um snapshot picked
    picked_waybills = set(
        CainiaoOperationTask.objects
        .filter(task_status__in=PICKED_STATUSES)
        .exclude(waybill_number="")
        .values_list("waybill_number", flat=True)
        .distinct()
    )

    # Waybills com resolução manual activa — também são excluídas
    resolved_waybills = set(
        WaybillResolution.objects
        .values_list("waybill_number", flat=True)
    )

    excluded = picked_waybills | resolved_waybills

    # Candidatos: waybills com Assigned (ou Unassign após Assigned)
    candidates_qs = (
        CainiaoOperationTask.objects
        .filter(task_status__in=["Assigned", "Unassign"])
        .exclude(waybill_number="")
        .exclude(waybill_number__in=excluded)
    )

    if hub_cp4s:
        from django.db.models import Q as _Q
        cp4_q = _Q()
        for cp4 in hub_cp4s:
            cp4_q |= _Q(zip_code__startswith=cp4)
        candidates_qs = candidates_qs.filter(cp4_q)

    # Agregar: por waybill, primeira aparição + última aparição
    aggregated = list(
        candidates_qs.values("waybill_number")
        .annotate(
            first_seen=Min("task_date"),
            last_seen=Max("task_date"),
        )
    )
    if not aggregated:
        return []

    today = timezone.now().date()
    holidays_set = set(
        Holiday.objects.filter(date__lte=today).values_list(
            "date", flat=True,
        )
    )

    # Filtra por antiguidade (≥ min_days dias úteis desde first_seen)
    eligible_waybills = []
    days_by_waybill = {}
    for r in aggregated:
        wb = r["waybill_number"]
        first = r["first_seen"]
        last = r["last_seen"]
        # business days entre first_seen+1 e hoje
        bd = _business_days_between(
            first + timedelta(days=1), today, holidays_set,
        )
        if bd >= min_days:
            eligible_waybills.append(wb)
            days_by_waybill[wb] = {
                "business_days": bd,
                "first_seen": first,
                "last_seen": last,
            }

    if not eligible_waybills:
        return []

    # Buscar o snapshot mais recente de cada waybill elegível,
    # com info do driver, CP4, cidade, peso, etc.
    latest_snapshots = {}
    snapshot_qs = (
        CainiaoOperationTask.objects
        .filter(waybill_number__in=eligible_waybills)
        .order_by("waybill_number", "-task_date", "-id")
    )
    for snap in snapshot_qs:
        if snap.waybill_number not in latest_snapshots:
            latest_snapshots[snap.waybill_number] = snap

    # Planning (cliente, seller) — bulk
    plannings = {
        p.parcel_id: p
        for p in CainiaoPlanningPackage.objects.filter(
            parcel_id__in=eligible_waybills,
        )
    }
    forecasts = {
        f.tracking_number: f
        for f in CainiaoForecastPackage.objects.filter(
            tracking_number__in=eligible_waybills,
        )
    }

    # Comments por waybill (count)
    comments_count = dict(
        WaybillComment.objects
        .filter(waybill_number__in=eligible_waybills)
        .values("waybill_number")
        .annotate(n=Count("id"))
        .values_list("waybill_number", "n")
    )

    # Preço Cainiao para custo estimado
    try:
        from core.finance import resolve_partner_price
        from core.models import Partner
        cainiao_partner = (
            Partner.objects.filter(name__iexact="CAINIAO").first()
        )
        partner_price = (
            float(resolve_partner_price(cainiao_partner) or 0)
            if cainiao_partner else 0.0
        )
    except Exception:
        partner_price = 0.0

    def _sla_level(days):
        """Cor do SLA: green <3, amber 3-4, orange 5-9, red ≥10."""
        if days >= 10:
            return "red"
        if days >= 5:
            return "orange"
        if days >= 3:
            return "amber"
        return "green"

    results = []
    for wb in eligible_waybills:
        snap = latest_snapshots.get(wb)
        if not snap:
            continue
        plan = plannings.get(wb)
        fc = forecasts.get(wb)
        cp4 = (snap.zip_code or "")[:4]
        customer_name = ""
        customer_phone = ""
        seller = ""
        if plan:
            customer_name = plan.receiver_name or ""
            customer_phone = plan.receiver_phone or ""
            seller = plan.seller_name or ""
        elif fc:
            customer_name = fc.receiver_name or ""
            customer_phone = fc.receiver_phone or ""
        if not seller and snap.seller_name:
            seller = snap.seller_name

        meta = days_by_waybill[wb]
        bd = meta["business_days"]
        results.append({
            "waybill": wb,
            "lp_number": snap.lp_number or "",
            "current_status": snap.task_status or "",
            "courier_name": snap.courier_name or "",
            "courier_id_cainiao": snap.courier_id_cainiao or "",
            "dsp_name": snap.dsp_name or "",
            "cp4": cp4,
            "city": snap.destination_city or "",
            "zip": snap.zip_code or "",
            "address": snap.detailed_address or "",
            "weight_g": snap.weight_g or 0,
            "seller": seller,
            "customer_name": customer_name,
            "customer_phone": customer_phone,
            "first_seen": meta["first_seen"].strftime("%Y-%m-%d"),
            "last_seen": meta["last_seen"].strftime("%Y-%m-%d"),
            "business_days_stuck": bd,
            "sla_level": _sla_level(bd),
            "should_escalate": bd >= 10,
            "estimated_cost_eur": round(partner_price, 4),
            "n_comments": comments_count.get(wb, 0),
        })

    # Ordena por dias parados desc, depois driver
    results.sort(
        key=lambda r: (
            -r["business_days_stuck"], r["courier_name"], r["waybill"],
        ),
    )
    return results


@login_required
def cainiao_not_arrived(request):
    """JSON: pacotes Not Arrived com enriquecimento completo.

    Query params:
      hub_id=N (opcional)
      min_days=2 (default)

    Devolve: rows, by_driver, by_cp4, summary (custo, escalação),
    history (30d), heatmap (DOW), spike alert.
    """
    hub_id = request.GET.get("hub_id")
    try:
        hub_id = int(hub_id) if hub_id else None
    except (ValueError, TypeError):
        hub_id = None
    try:
        min_days = int(request.GET.get("min_days") or 2)
    except (ValueError, TypeError):
        min_days = 2
    min_days = max(0, min_days)
    include_unregistered = (
        request.GET.get("include_unregistered") == "1"
    )

    rows = _compute_not_arrived(
        hub_id=hub_id, min_days=min_days,
        include_unregistered=include_unregistered,
    )

    # Sumário por driver
    by_driver = {}
    for r in rows:
        key = r["courier_name"] or "(sem driver)"
        by_driver.setdefault(key, {
            "driver": key, "n": 0, "weight_g": 0,
            "cp4s": set(),
        })
        by_driver[key]["n"] += 1
        by_driver[key]["weight_g"] += r["weight_g"] or 0
        if r["cp4"]:
            by_driver[key]["cp4s"].add(r["cp4"])
    by_driver_list = [
        {**v, "cp4s": sorted(list(v["cp4s"]))}
        for v in sorted(
            by_driver.values(), key=lambda x: -x["n"],
        )
    ]

    # Sumário por CP4
    by_cp4 = {}
    for r in rows:
        key = r["cp4"] or "(sem CP4)"
        by_cp4.setdefault(key, {
            "cp4": key, "n": 0, "weight_g": 0, "drivers": set(),
        })
        by_cp4[key]["n"] += 1
        by_cp4[key]["weight_g"] += r["weight_g"] or 0
        if r["courier_name"]:
            by_cp4[key]["drivers"].add(r["courier_name"])
    by_cp4_list = [
        {**v, "drivers": sorted(list(v["drivers"]))}
        for v in sorted(by_cp4.values(), key=lambda x: -x["n"])
    ]

    # Custo total estimado (só faz sentido se há rows)
    total_cost = sum(r.get("estimated_cost_eur", 0) for r in rows)
    n_to_escalate = sum(1 for r in rows if r.get("should_escalate"))

    # Histórico: contagem de pacotes Not Arrived que entraram nos
    # últimos 30 dias (por first_seen). Útil para line chart.
    from datetime import timedelta
    from collections import Counter
    from django.utils import timezone as _tz
    today_d = _tz.now().date()
    history_30d = []
    counter = Counter(r["first_seen"] for r in rows)
    for i in range(29, -1, -1):
        d = today_d - timedelta(days=i)
        ds = d.strftime("%Y-%m-%d")
        history_30d.append({"date": ds, "n": counter.get(ds, 0)})

    # Heatmap por dia da semana (de first_seen)
    dow_counter = Counter()
    for r in rows:
        from datetime import datetime
        try:
            d = datetime.strptime(r["first_seen"], "%Y-%m-%d").date()
            dow_counter[d.weekday()] += 1
        except (ValueError, TypeError):
            pass
    heatmap_dow = [
        {"dow": i, "label": label, "n": dow_counter.get(i, 0)}
        for i, label in enumerate([
            "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom",
        ])
    ]

    # Spike alert: comparar últimos 7 dias vs 7 dias anteriores
    last7 = sum(
        1 for r in rows
        if r["first_seen"] >= (
            today_d - timedelta(days=7)
        ).strftime("%Y-%m-%d")
    )
    prev7 = sum(
        1 for r in rows
        if (today_d - timedelta(days=14)).strftime("%Y-%m-%d")
        <= r["first_seen"]
        < (today_d - timedelta(days=7)).strftime("%Y-%m-%d")
    )
    spike = None
    if prev7 > 0 and last7 >= 3 * prev7:
        spike = {
            "level": "high",
            "message": (
                f"⚠ Últimos 7 dias têm {last7} pacotes Not Arrived "
                f"(vs {prev7} na semana anterior). Spike de "
                f"{round(last7/prev7, 1)}x"
            ),
        }
    elif prev7 == 0 and last7 >= 5:
        spike = {
            "level": "medium",
            "message": (
                f"⚠ {last7} pacotes Not Arrived nesta semana "
                "(semana passada teve 0)."
            ),
        }

    # Hot drivers: 5+ stuck no resultado actual
    hot_drivers = [d for d in by_driver_list if d["n"] >= 5]
    hot_cp4s = [c for c in by_cp4_list if c["n"] >= 5]

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "min_days": min_days,
        "hub_id": hub_id,
        "rows": rows,
        "by_driver": by_driver_list,
        "by_cp4": by_cp4_list,
        "summary": {
            "total_packages": len(rows),
            "total_cost_eur": round(total_cost, 2),
            "n_to_escalate": n_to_escalate,
            "n_drivers_affected": len(by_driver_list),
            "n_cp4s_affected": len(by_cp4_list),
            "last7_days": last7,
            "prev7_days": prev7,
        },
        "history_30d": history_30d,
        "heatmap_dow": heatmap_dow,
        "spike": spike,
        "hot_drivers": hot_drivers,
        "hot_cp4s": hot_cp4s,
    })


@login_required
def cainiao_not_arrived_xlsx(request):
    """Exporta Not Arrived para Excel pronto a enviar aos superiores."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from django.utils import timezone

    hub_id = request.GET.get("hub_id")
    try:
        hub_id = int(hub_id) if hub_id else None
    except (ValueError, TypeError):
        hub_id = None
    try:
        min_days = int(request.GET.get("min_days") or 2)
    except (ValueError, TypeError):
        min_days = 2

    rows = _compute_not_arrived(hub_id=hub_id, min_days=min_days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Not Arrived"

    # Estilo cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        "solid", fgColor="1E3A8A",
    )
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Título
    ws["A1"] = "Pacotes Not Arrived (atribuídos mas nunca picados)"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells("A1:M1")

    today = timezone.now().date().strftime("%d/%m/%Y")
    ws["A2"] = (
        f"Gerado: {today} · Antiguidade ≥ {min_days} dias úteis"
        f" · Total: {len(rows)} pacotes"
    )
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:M2")

    headers = [
        "Waybill", "LP No.", "Estado actual", "Driver",
        "Courier ID", "DSP", "CP4", "Cidade", "Morada",
        "Peso (g)", "Seller", "1ª aparição", "Última aparição",
        "Dias úteis parado",
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=r["waybill"])
        ws.cell(row=i, column=2, value=r["lp_number"])
        ws.cell(row=i, column=3, value=r["current_status"])
        ws.cell(row=i, column=4, value=r["courier_name"])
        ws.cell(row=i, column=5, value=r["courier_id_cainiao"])
        ws.cell(row=i, column=6, value=r["dsp_name"])
        ws.cell(row=i, column=7, value=r["cp4"])
        ws.cell(row=i, column=8, value=r["city"])
        ws.cell(row=i, column=9, value=r["address"])
        ws.cell(row=i, column=10, value=r["weight_g"])
        ws.cell(row=i, column=11, value=r["seller"])
        ws.cell(row=i, column=12, value=r["first_seen"])
        ws.cell(row=i, column=13, value=r["last_seen"])
        days_cell = ws.cell(
            row=i, column=14, value=r["business_days_stuck"],
        )
        # Cor: rosa intenso se ≥ 5 dias parado
        if r["business_days_stuck"] >= 5:
            days_cell.fill = PatternFill(
                "solid", fgColor="FECACA",
            )
            days_cell.font = Font(bold=True, color="991B1B")
        elif r["business_days_stuck"] >= 3:
            days_cell.fill = PatternFill(
                "solid", fgColor="FED7AA",
            )

        for col_idx in range(1, 15):
            ws.cell(row=i, column=col_idx).border = border
            ws.cell(row=i, column=col_idx).alignment = left

    # Largura colunas
    widths = [28, 18, 14, 22, 14, 22, 8, 18, 40, 10, 22, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A5"

    # Sheet 2: agregado por driver
    ws2 = wb.create_sheet("Por Driver")
    ws2["A1"] = "Pacotes Not Arrived agregados por Driver"
    ws2["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws2.merge_cells("A1:D1")

    by_driver = {}
    for r in rows:
        key = r["courier_name"] or "(sem driver)"
        by_driver.setdefault(key, {"n": 0, "w": 0, "cp4s": set()})
        by_driver[key]["n"] += 1
        by_driver[key]["w"] += r["weight_g"] or 0
        if r["cp4"]:
            by_driver[key]["cp4s"].add(r["cp4"])

    for col_idx, h in enumerate(
        ["Driver", "Pacotes", "Peso total (g)", "CP4s afectados"], 1,
    ):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, (drv, d) in enumerate(
        sorted(by_driver.items(), key=lambda x: -x[1]["n"]), start=4,
    ):
        ws2.cell(row=i, column=1, value=drv)
        ws2.cell(row=i, column=2, value=d["n"])
        ws2.cell(row=i, column=3, value=d["w"])
        ws2.cell(
            row=i, column=4, value=", ".join(sorted(list(d["cp4s"]))),
        )
        for col_idx in range(1, 5):
            ws2.cell(row=i, column=col_idx).border = border

    for i, w in enumerate([28, 12, 16, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    fname = (
        f"not_arrived_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{fname}"'
    )
    wb.save(response)
    return response


# ============================================================================
# Resolução manual e comentários
# ============================================================================

@login_required
@require_http_methods(["POST"])
def waybill_resolve(request, waybill):
    """Marca um waybill como resolvido (entregue noutro HUB, etc.).

    Body JSON: {
      resolution_type, other_hub, other_courier, other_delivery_time,
      notes
    }
    """
    import json
    from .models import WaybillResolution
    waybill = (waybill or "").strip()
    if not waybill:
        return JsonResponse(
            {"success": False, "error": "Waybill vazio"}, status=400,
        )

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    rtype = (body.get("resolution_type") or "").strip()
    valid_types = [t[0] for t in WaybillResolution.TYPE_CHOICES]
    if rtype not in valid_types:
        rtype = WaybillResolution.TYPE_DELIVERED_OTHER_HUB

    delivery_time = None
    raw_dt = (body.get("other_delivery_time") or "").strip()
    if raw_dt:
        from django.utils.dateparse import parse_datetime
        delivery_time = parse_datetime(raw_dt)

    obj, created = WaybillResolution.objects.update_or_create(
        waybill_number=waybill,
        defaults={
            "resolution_type": rtype,
            "other_hub": (body.get("other_hub") or "").strip()[:100],
            "other_courier": (
                body.get("other_courier") or ""
            ).strip()[:200],
            "other_delivery_time": delivery_time,
            "notes": (body.get("notes") or "").strip(),
            "resolved_by": (
                request.user if request.user.is_authenticated
                else None
            ),
        },
    )
    return JsonResponse({
        "success": True,
        "created": created,
        "resolution": {
            "waybill": obj.waybill_number,
            "resolution_type": obj.resolution_type,
            "resolution_type_display": obj.get_resolution_type_display(),
            "other_hub": obj.other_hub,
            "other_courier": obj.other_courier,
            "other_delivery_time": (
                obj.other_delivery_time.strftime("%Y-%m-%d %H:%M")
                if obj.other_delivery_time else ""
            ),
            "notes": obj.notes,
            "resolved_at": obj.resolved_at.strftime("%Y-%m-%d %H:%M"),
            "resolved_by": (
                obj.resolved_by.username if obj.resolved_by else ""
            ),
        },
    })


@login_required
@require_http_methods(["POST"])
def waybill_unresolve(request, waybill):
    """Apaga a resolução de um waybill (volta a aparecer em Not Arrived)."""
    from .models import WaybillResolution
    n, _ = WaybillResolution.objects.filter(
        waybill_number=waybill,
    ).delete()
    return JsonResponse({"success": True, "deleted": n})


@login_required
def waybill_get_resolution(request, waybill):
    """Devolve a resolução actual de um waybill (se existir)."""
    from .models import WaybillResolution
    obj = WaybillResolution.objects.filter(
        waybill_number=waybill,
    ).first()
    if not obj:
        return JsonResponse({"success": True, "resolution": None})
    return JsonResponse({
        "success": True,
        "resolution": {
            "waybill": obj.waybill_number,
            "resolution_type": obj.resolution_type,
            "resolution_type_display": obj.get_resolution_type_display(),
            "other_hub": obj.other_hub,
            "other_courier": obj.other_courier,
            "other_delivery_time": (
                obj.other_delivery_time.strftime("%Y-%m-%dT%H:%M")
                if obj.other_delivery_time else ""
            ),
            "notes": obj.notes,
            "resolved_at": obj.resolved_at.strftime("%Y-%m-%d %H:%M"),
            "resolved_by": (
                obj.resolved_by.username if obj.resolved_by else ""
            ),
        },
    })


@login_required
@require_http_methods(["POST"])
def waybill_bulk_resolve(request):
    """Resolve vários waybills de uma vez.

    Body: {
      waybills: [str], resolution_type, other_hub, other_courier,
      other_delivery_time (opcional, aplica a todos), notes
    }
    """
    import json
    from .models import WaybillResolution
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    waybills = body.get("waybills") or []
    if not isinstance(waybills, list) or not waybills:
        return JsonResponse(
            {"success": False, "error": "waybills vazios"}, status=400,
        )

    rtype = (body.get("resolution_type") or "").strip()
    valid_types = [t[0] for t in WaybillResolution.TYPE_CHOICES]
    if rtype not in valid_types:
        rtype = WaybillResolution.TYPE_DELIVERED_OTHER_HUB

    delivery_time = None
    raw_dt = (body.get("other_delivery_time") or "").strip()
    if raw_dt:
        from django.utils.dateparse import parse_datetime
        delivery_time = parse_datetime(raw_dt)

    user = request.user if request.user.is_authenticated else None
    n_created = 0
    n_updated = 0
    for wb in waybills:
        wb = str(wb).strip()
        if not wb:
            continue
        _, created = WaybillResolution.objects.update_or_create(
            waybill_number=wb,
            defaults={
                "resolution_type": rtype,
                "other_hub": (body.get("other_hub") or "").strip()[:100],
                "other_courier": (
                    body.get("other_courier") or ""
                ).strip()[:200],
                "other_delivery_time": delivery_time,
                "notes": (body.get("notes") or "").strip(),
                "resolved_by": user,
            },
        )
        if created:
            n_created += 1
        else:
            n_updated += 1

    return JsonResponse({
        "success": True,
        "n_created": n_created,
        "n_updated": n_updated,
        "total": n_created + n_updated,
    })


# ─── Comments ───────────────────────────────────────────────────────────

@login_required
def waybill_comments_list(request, waybill):
    """Lista os comentários de um waybill (ordem: mais recente primeiro)."""
    from .models import WaybillComment
    comments = list(
        WaybillComment.objects
        .filter(waybill_number=waybill)
        .select_related("author")
        .order_by("-created_at")[:100]
    )
    data = [{
        "id": c.id,
        "body": c.body,
        "author": (
            c.author.get_username() if c.author else "(sistema)"
        ),
        "created_at": c.created_at.strftime("%Y-%m-%d %H:%M"),
    } for c in comments]
    return JsonResponse({"success": True, "comments": data})


@login_required
@require_http_methods(["POST"])
def waybill_comments_add(request, waybill):
    """Adiciona um comentário a um waybill."""
    import json
    from .models import WaybillComment
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    text = (body.get("body") or "").strip()
    if not text:
        return JsonResponse(
            {"success": False, "error": "Comentário vazio"},
            status=400,
        )
    c = WaybillComment.objects.create(
        waybill_number=waybill,
        body=text,
        author=(
            request.user if request.user.is_authenticated else None
        ),
    )
    return JsonResponse({
        "success": True,
        "comment": {
            "id": c.id,
            "body": c.body,
            "author": (
                c.author.get_username() if c.author else "(sistema)"
            ),
            "created_at": c.created_at.strftime("%Y-%m-%d %H:%M"),
        },
    })


@login_required
@require_http_methods(["POST"])
def waybill_comments_delete(request, comment_id):
    """Apaga um comentário (só autor ou superuser)."""
    from .models import WaybillComment
    c = get_object_or_404(WaybillComment, id=comment_id)
    if (
        request.user.is_authenticated
        and (
            request.user.is_superuser
            or c.author_id == request.user.id
        )
    ):
        c.delete()
        return JsonResponse({"success": True})
    return JsonResponse(
        {"success": False, "error": "Sem permissão"}, status=403,
    )


# ============================================================================
# Scan de planilha de outro HUB (em memória)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_not_arrived_scan_xlsx(request):
    """Recebe um Excel de outro HUB e procura os waybills da lista
    actual de Not Arrived. Devolve os matches em memória — não persiste.

    POST: multipart/form-data
      file: <xlsx>
      hub_id, min_days: (para recalcular a lista actual)
    """
    import openpyxl
    from io import BytesIO

    upload = request.FILES.get("file")
    if not upload:
        return JsonResponse(
            {"success": False, "error": "Ficheiro em falta"},
            status=400,
        )

    hub_id = request.POST.get("hub_id")
    try:
        hub_id = int(hub_id) if hub_id else None
    except (ValueError, TypeError):
        hub_id = None
    try:
        min_days = int(request.POST.get("min_days") or 2)
    except (ValueError, TypeError):
        min_days = 2

    rows = _compute_not_arrived(hub_id=hub_id, min_days=min_days)
    target_waybills = {r["waybill"] for r in rows}
    if not target_waybills:
        return JsonResponse({
            "success": True, "matches": [], "total_scanned": 0,
            "message": "Nenhum pacote Not Arrived activo.",
        })

    # Parse xlsx — sem read_only para evitar quirks com files
    # que têm freeze panes / merged cells / múltiplas sheets
    try:
        workbook = openpyxl.load_workbook(
            BytesIO(upload.read()), data_only=True,
        )
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"XLSX inválido: {e}"},
            status=400,
        )

    HEADER_ALIASES = {
        "waybill_number": [
            "Waybill Number", "Waybill No.", "Waybill_no",
            "Waybill", "WaybillNumber", "Tracking Number",
            "Tracking No.", "Tracking",
        ],
        "lp_number": [
            "LP No.", "LP Number", "LP_no", "LP", "LP Code",
            "LP No",
        ],
        "task_status": [
            "Task Status", "Status", "Estado", "Delivery Status",
        ],
        "task_date": [
            "Task Date", "Data", "Date", "Operation Date",
        ],
        "courier_name": [
            "Courier Name", "Driver", "Motorista", "Driver Name",
            "Courier", "Driver_name",
        ],
        "courier_id_cainiao": [
            "Courier ID", "Courier_id", "Courier ID Cainiao",
            "Driver ID", "Courier Id",
        ],
        "dsp_name": [
            "DSP Name", "DSP", "Frota", "DSP_name",
        ],
        "delivery_time": [
            "Delivery Time", "Entrega", "Delivered At",
            "Delivery_time", "Sign Time", "Signed At",
        ],
        "destination_city": [
            "Destination City", "City", "Cidade", "Receiver City",
        ],
        "zip_code": [
            "Zip Code", "Zip", "CP", "Código Postal", "Postal Code",
            "Receiver Zip",
        ],
        "exception_type": [
            "Exception Type", "Exception", "Exceção",
        ],
        "exception_detail": [
            "Exception Detail", "Detalhe", "Razão",
            "Exception Reason",
        ],
    }

    def _detect_header(ws):
        """Tenta encontrar a linha de cabeçalho.
        Estratégia em 3 passes:
          1. Match exacto (case-insensitive) em até 10 linhas
          2. Match por substring (ex: célula contém 'waybill')
          3. Devolve None se nada encontrado
        """
        # Capturar primeiras 15 linhas para análise
        sample_rows = []
        for idx, row in enumerate(
            ws.iter_rows(values_only=True),
        ):
            if idx >= 15:
                break
            sample_rows.append(row)

        # Pass 1: match exacto
        for row in sample_rows:
            if not row:
                continue
            non_empty = sum(1 for c in row if c is not None)
            if non_empty < 3:
                continue
            candidate = {}
            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                v_str = str(val).strip()
                v_low = v_str.lower()
                for key, aliases in HEADER_ALIASES.items():
                    if any(v_low == a.lower() for a in aliases):
                        candidate[key] = col_idx
                        break
            if "waybill_number" in candidate:
                return row, candidate

        # Pass 2: match por substring (mais permissivo)
        for row in sample_rows:
            if not row:
                continue
            non_empty = sum(1 for c in row if c is not None)
            if non_empty < 3:
                continue
            candidate = {}
            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                v_low = str(val).strip().lower()
                for key, aliases in HEADER_ALIASES.items():
                    if key in candidate:
                        continue
                    for a in aliases:
                        a_low = a.lower()
                        # Substring de qualquer lado
                        if a_low in v_low or v_low in a_low:
                            candidate[key] = col_idx
                            break
            if "waybill_number" in candidate:
                return row, candidate

        return None, {}

    # Procura em todas as sheets
    detected_ws = None
    header = None
    header_map = {}
    for sheet in workbook.worksheets:
        h, hm = _detect_header(sheet)
        if hm.get("waybill_number") is not None:
            detected_ws = sheet
            header = h
            header_map = hm
            break

    if header_map.get("waybill_number") is None:
        # Diagnóstico: mostrar o conteúdo das primeiras linhas das
        # primeiras sheets para o utilizador perceber o que está a
        # ser lido
        sample_info = []
        for sheet in workbook.worksheets[:3]:
            sheet_sample = []
            for idx, row in enumerate(
                sheet.iter_rows(values_only=True),
            ):
                if idx >= 3:
                    break
                vals = [
                    str(v).strip() if v is not None else ""
                    for v in row[:10]
                ]
                sheet_sample.append(" | ".join(
                    v for v in vals if v
                )[:200])
            sample_info.append(
                f"Sheet '{sheet.title}': "
                + (
                    " // ".join(sheet_sample)
                    if sheet_sample else "(vazia)"
                )
            )
        return JsonResponse({
            "success": False,
            "error": (
                "Não consegui identificar a coluna 'Waybill Number'"
                " em nenhuma sheet do ficheiro.\n\n"
                + "\n".join(sample_info)
            ),
        }, status=400)

    # Pré-carrega "drivers nossos" — qualquer driver activo OU
    # registado como CourierNameAlias do Cainiao é considerado
    # "nosso" (mesmo que esteja inactivo).
    from drivers_app.models import DriverProfile
    from .models import DriverCourierMapping, CourierNameAlias
    from core.models import Partner
    our_courier_names = set()
    our_courier_ids = set()
    cainiao_partner = (
        Partner.objects.filter(name__iexact="CAINIAO").first()
    )
    for d in DriverProfile.objects.exclude(apelido=""):
        our_courier_names.add(d.apelido.strip().lower())
        if d.courier_id_cainiao:
            our_courier_ids.add(d.courier_id_cainiao.strip())
    if cainiao_partner:
        for m in DriverCourierMapping.objects.filter(
            partner=cainiao_partner,
        ):
            if m.courier_name:
                our_courier_names.add(m.courier_name.strip().lower())
            if m.courier_id:
                our_courier_ids.add(m.courier_id.strip())
        for a in CourierNameAlias.objects.filter(
            partner=cainiao_partner,
        ):
            if a.courier_name:
                our_courier_names.add(a.courier_name.strip().lower())
            if a.courier_id:
                our_courier_ids.add(a.courier_id.strip())

    matches = []
    total_scanned = 0
    wb_col = header_map["waybill_number"]

    # Itera todas as linhas e procura matches
    for row in detected_ws.iter_rows(values_only=True):
        if not row or len(row) <= wb_col:
            continue
        wb_value = row[wb_col]
        if not wb_value:
            continue
        wb_str = str(wb_value).strip()
        # Salta a própria linha de cabeçalho
        if wb_str.lower() in (
            "waybill number", "waybill no.", "waybill",
            "tracking number",
        ):
            continue
        total_scanned += 1
        if wb_str not in target_waybills:
            continue
        # Match!
        m = {"waybill": wb_str}
        for key, col_idx in header_map.items():
            if key == "waybill_number":
                continue
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                m[key] = ""
                continue
            if hasattr(val, "strftime"):
                m[key] = val.strftime("%Y-%m-%d %H:%M")
            else:
                m[key] = str(val).strip()

        # Classificação automática do match
        cn = (m.get("courier_name") or "").strip().lower()
        cid = (m.get("courier_id_cainiao") or "").strip()
        is_our_driver = bool(
            (cn and cn in our_courier_names)
            or (cid and cid in our_courier_ids)
        )
        m["is_our_driver"] = is_our_driver
        # Tipo sugerido
        if is_our_driver:
            m["suggested_resolution"] = "DELIVERED_LATE"
            m["suggested_label"] = (
                "Driver nosso — actualização perdida"
            )
        else:
            m["suggested_resolution"] = "DELIVERED_OTHER_HUB"
            m["suggested_label"] = (
                "Outro HUB entregou"
            )
        matches.append(m)

    # Sumário por classificação
    n_our = sum(1 for x in matches if x["is_our_driver"])
    n_other = len(matches) - n_our

    return JsonResponse({
        "success": True,
        "filename": upload.name,
        "sheet": detected_ws.title,
        "total_scanned": total_scanned,
        "n_matches": len(matches),
        "n_targets": len(target_waybills),
        "n_our_driver": n_our,
        "n_other_hub": n_other,
        "matches": matches,
        "header_map": list(header_map.keys()),
    })


# ============================================================================
# Forecast cross-reference (pacotes com forecast sem operação)
# ============================================================================

@login_required
def cainiao_waybill_diagnose(request, waybill):
    """Diagnóstico: para um waybill, mostra o que está na DB e o que
    foi visto nos últimos N imports (batches). Permite perceber
    porque um pacote ficou parado num estado.
    """
    from .models import (
        CainiaoOperationTask, CainiaoOperationBatch,
        WaybillResolution, WaybillComment,
    )
    waybill = (waybill or "").strip()
    if not waybill:
        return JsonResponse(
            {"success": False, "error": "Waybill vazio"}, status=400,
        )

    # Estado actual na DB
    db_rows = []
    for op in (
        CainiaoOperationTask.objects
        .filter(waybill_number=waybill)
        .select_related("last_import_batch")
        .order_by("-task_date")
    ):
        db_rows.append({
            "id": op.id,
            "task_date": op.task_date.strftime("%Y-%m-%d"),
            "task_status": op.task_status,
            "courier_name": op.courier_name,
            "courier_id_cainiao": op.courier_id_cainiao,
            "delivery_time": (
                op.delivery_time.strftime("%Y-%m-%d %H:%M")
                if op.delivery_time else ""
            ),
            "last_import_batch_id": op.last_import_batch_id,
            "last_import_batch_filename": (
                op.last_import_batch.filename
                if op.last_import_batch else ""
            ),
            "last_import_batch_at": (
                op.last_import_batch.created_at.strftime(
                    "%Y-%m-%d %H:%M",
                )
                if op.last_import_batch else ""
            ),
        })

    # Resolução manual (se existir)
    res = WaybillResolution.objects.filter(
        waybill_number=waybill,
    ).select_related("resolved_by").first()
    resolution = None
    if res:
        resolution = {
            "type": res.resolution_type,
            "type_display": res.get_resolution_type_display(),
            "resolved_at": res.resolved_at.strftime("%Y-%m-%d %H:%M"),
            "resolved_by": (
                res.resolved_by.get_username()
                if res.resolved_by else "(sistema)"
            ),
            "other_hub": res.other_hub,
            "other_courier": res.other_courier,
            "notes": res.notes,
        }

    # Comments
    n_comments = WaybillComment.objects.filter(
        waybill_number=waybill,
    ).count()

    # Últimos 20 batches Operation Update (para contexto)
    recent_batches = []
    for b in (
        CainiaoOperationBatch.objects.order_by("-created_at")[:20]
    ):
        # Verifica se o waybill passou neste batch
        seen_in_batch = (
            CainiaoOperationTask.objects.filter(
                waybill_number=waybill,
                last_import_batch=b,
            ).exists()
        )
        recent_batches.append({
            "id": b.id,
            "filename": b.filename,
            "task_date": b.task_date.strftime("%Y-%m-%d"),
            "created_at": b.created_at.strftime("%Y-%m-%d %H:%M"),
            "total_tasks": b.total_tasks,
            "new_tasks": b.new_tasks,
            "updated_tasks": b.updated_tasks,
            "seen_in_batch": seen_in_batch,
        })

    # Sumário
    has_db = len(db_rows) > 0
    last_seen_batch = next(
        (b for b in recent_batches if b["seen_in_batch"]),
        None,
    )

    # Diagnóstico textual
    if not has_db:
        diagnosis = (
            "Este waybill NUNCA foi importado para o sistema. "
            "Verifica se foi gerado correctamente no Cainiao."
        )
    elif resolution:
        diagnosis = (
            f"Resolvido manualmente como '{resolution['type_display']}' "
            f"em {resolution['resolved_at']}."
        )
    elif last_seen_batch:
        diagnosis = (
            f"Última vez visto num import: {last_seen_batch['filename']}"
            f" em {last_seen_batch['created_at']}."
        )
    else:
        diagnosis = (
            "O waybill está na DB mas NÃO apareceu em nenhum dos "
            "últimos 20 imports. Provavelmente saiu do EPOD do nosso "
            "HUB (transferido para outro HUB ou fechado)."
        )

    return JsonResponse({
        "success": True,
        "waybill": waybill,
        "db_rows": db_rows,
        "resolution": resolution,
        "n_comments": n_comments,
        "recent_batches": recent_batches,
        "diagnosis": diagnosis,
        "in_db": has_db,
        "in_recent_imports": last_seen_batch is not None,
    })


@login_required
def cainiao_resolutions_list(request):
    """Lista todas as WaybillResolution (pacotes Not Arrived que foram
    resolvidos manualmente ou automaticamente).

    Query params:
      type=ALL|<TYPE>  (default ALL)
      days_back=N      (default 90, máx 365)
    """
    from datetime import timedelta
    from django.utils import timezone as _tz
    from .models import WaybillResolution, CainiaoOperationTask

    rtype = (request.GET.get("type") or "ALL").strip().upper()
    try:
        days_back = int(request.GET.get("days_back") or 90)
    except (ValueError, TypeError):
        days_back = 90
    days_back = max(1, min(days_back, 365))

    cutoff = _tz.now() - timedelta(days=days_back)
    qs = (
        WaybillResolution.objects
        .select_related("resolved_by")
        .filter(resolved_at__gte=cutoff)
        .order_by("-resolved_at")
    )
    valid_types = [t[0] for t in WaybillResolution.TYPE_CHOICES]
    if rtype != "ALL" and rtype in valid_types:
        qs = qs.filter(resolution_type=rtype)

    resolutions = list(qs[:1000])

    # Bulk fetch de info útil dos snapshots (cidade, peso, driver
    # original, CP4) para mostrar contexto na tabela
    waybills = [r.waybill_number for r in resolutions]
    snapshots = {}
    if waybills:
        for op in (
            CainiaoOperationTask.objects
            .filter(waybill_number__in=waybills)
            .order_by("waybill_number", "-task_date", "-id")
        ):
            if op.waybill_number not in snapshots:
                snapshots[op.waybill_number] = op

    rows = []
    for r in resolutions:
        snap = snapshots.get(r.waybill_number)
        cp4 = (snap.zip_code or "")[:4] if snap else ""
        rows.append({
            "waybill": r.waybill_number,
            "resolution_type": r.resolution_type,
            "resolution_type_display": r.get_resolution_type_display(),
            "resolved_at": r.resolved_at.strftime("%Y-%m-%d %H:%M"),
            "resolved_by": (
                r.resolved_by.get_username()
                if r.resolved_by else "(sistema)"
            ),
            "other_hub": r.other_hub or "",
            "other_courier": r.other_courier or "",
            "other_delivery_time": (
                r.other_delivery_time.strftime("%Y-%m-%d %H:%M")
                if r.other_delivery_time else ""
            ),
            "notes": r.notes or "",
            "cp4": cp4,
            "city": snap.destination_city if snap else "",
            "weight_g": snap.weight_g if snap and snap.weight_g else 0,
            "original_courier": (
                snap.courier_name if snap else ""
            ),
        })

    # Sumário por tipo
    by_type = {}
    for r in resolutions:
        key = r.resolution_type
        by_type.setdefault(key, 0)
        by_type[key] += 1
    by_type_list = [
        {
            "type": k,
            "label": dict(WaybillResolution.TYPE_CHOICES).get(k, k),
            "n": v,
        }
        for k, v in sorted(by_type.items(), key=lambda x: -x[1])
    ]

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "days_back": days_back,
        "rows": rows,
        "by_type": by_type_list,
    })


@login_required
def cainiao_forecast_orphans(request):
    """Pacotes que TINHAM forecast (planning) mas nunca apareceram em
    CainiaoOperationTask. Tipo 2 de "perdido": não chegou ao hub.

    Query: hub_id, days_back (default 30)
    """
    from .models import (
        CainiaoOperationTask, CainiaoPlanningPackage,
        CainiaoForecastPackage,
    )
    from datetime import timedelta
    from django.utils import timezone as _tz

    try:
        days_back = int(request.GET.get("days_back") or 30)
    except (ValueError, TypeError):
        days_back = 30
    days_back = max(1, min(days_back, 180))

    today_d = _tz.now().date()
    cutoff = today_d - timedelta(days=days_back)

    # Waybills que apareceram em Operation Task no período
    op_waybills = set(
        CainiaoOperationTask.objects
        .filter(task_date__gte=cutoff)
        .exclude(waybill_number="")
        .values_list("waybill_number", flat=True)
        .distinct()
    )

    # Planning packages que NÃO estão em op_waybills
    plannings = (
        CainiaoPlanningPackage.objects
        .filter(operation_date__gte=cutoff)
        .exclude(parcel_id__in=op_waybills)
        .order_by("-operation_date")[:5000]
    )

    rows = []
    for p in plannings:
        rows.append({
            "waybill": p.parcel_id,
            "planning_date": p.operation_date.strftime("%Y-%m-%d"),
            "hub": p.hub or "",
            "dsp": p.dsp or "",
            "city": p.receiver_city or "",
            "zip": p.receiver_zip or "",
            "customer_name": p.receiver_name or "",
            "customer_phone": p.receiver_phone or "",
            "seller": p.seller_name or "",
        })

    # Agregado por HUB
    by_hub = {}
    for r in rows:
        key = r["hub"] or "(sem HUB)"
        by_hub.setdefault(key, 0)
        by_hub[key] += 1
    by_hub_list = sorted(
        [{"hub": k, "n": v} for k, v in by_hub.items()],
        key=lambda x: -x["n"],
    )

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "days_back": days_back,
        "rows": rows,
        "by_hub": by_hub_list,
    })


# ============================================================================
# WhatsApp ao driver com lista de stuck
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_whatsapp_driver_stuck(request):
    """Envia WhatsApp ao driver com a lista de pacotes Not Arrived
    atribuídos a ele.

    Body JSON: { driver_name: str, hub_id: int, min_days: int }
    """
    import json
    import requests
    from django.conf import settings as dj_settings
    from drivers_app.models import DriverProfile

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    driver_name = (body.get("driver_name") or "").strip()
    if not driver_name:
        return JsonResponse(
            {"success": False, "error": "driver_name vazio"},
            status=400,
        )

    hub_id = body.get("hub_id")
    try:
        hub_id = int(hub_id) if hub_id else None
    except (ValueError, TypeError):
        hub_id = None
    try:
        min_days = int(body.get("min_days") or 2)
    except (ValueError, TypeError):
        min_days = 2

    rows = _compute_not_arrived(hub_id=hub_id, min_days=min_days)
    drv_rows = [
        r for r in rows if r["courier_name"] == driver_name
    ]
    if not drv_rows:
        return JsonResponse({
            "success": False,
            "error": "Sem pacotes para este driver",
        }, status=404)

    # Tenta achar o driver na DB pelo apelido / nome
    driver = (
        DriverProfile.objects.filter(apelido=driver_name).first()
        or DriverProfile.objects.filter(
            nome_completo__icontains=driver_name,
        ).first()
    )
    if not driver or not driver.telefone:
        return JsonResponse({
            "success": False,
            "error": (
                f"Driver {driver_name} sem telefone registado."
            ),
        }, status=400)

    # Normalizar telefone
    phone_digits = "".join(
        c for c in (driver.telefone or "") if c.isdigit()
    )
    if phone_digits and not phone_digits.startswith("351"):
        if len(phone_digits) == 9:
            phone_digits = "351" + phone_digits

    # Construir mensagem
    lines = [
        f"Olá {driver.apelido or driver.nome_completo} 👋",
        "",
        f"Tens {len(drv_rows)} pacote(s) que constam atribuídos a "
        "ti mas que nunca foram lidos:",
        "",
    ]
    for r in drv_rows[:30]:
        lines.append(
            f"• {r['waybill']} (CP4 {r['cp4']}, "
            f"{r['business_days_stuck']}d parado)"
        )
    if len(drv_rows) > 30:
        lines.append(f"... e mais {len(drv_rows) - 30} pacotes")
    lines += [
        "",
        "Por favor verifica se ainda os tens contigo ou se "
        "estão no hub. Obrigado!",
    ]
    message = "\n".join(lines)

    # Enviar via API WhatsApp
    api_url = (
        getattr(dj_settings, "WHATSAPP_API_URL", "")
        or "http://45.160.176.150:9090/message/sendText/leguasreports"
    )
    try:
        r = requests.post(
            api_url,
            json={"number": phone_digits, "text": message},
            timeout=15,
        )
        ok = r.status_code in (200, 201)
    except Exception as e:
        return JsonResponse({
            "success": False, "error": f"Erro de envio: {e}",
        }, status=500)

    return JsonResponse({
        "success": ok,
        "n_packages": len(drv_rows),
        "phone": phone_digits,
        "preview": message[:300],
    })


# ============================================================================
# Camada 1 — WaybillFlag: sinalização persistente
# ============================================================================

def _serialize_flag(f):
    return {
        "id": f.id,
        "waybill": f.waybill_number,
        "flag_type": f.flag_type,
        "flag_type_display": f.get_flag_type_display(),
        "reason": f.reason,
        "reason_display": f.get_reason_display(),
        "notes": f.notes or "",
        "flagged_at": f.flagged_at.strftime("%Y-%m-%d %H:%M"),
        "flagged_by": (
            f.flagged_by.get_username() if f.flagged_by else "(sistema)"
        ),
        "is_active": f.is_active,
        "cleared_at": (
            f.cleared_at.strftime("%Y-%m-%d %H:%M")
            if f.cleared_at else ""
        ),
        "cleared_reason": f.cleared_reason or "",
        "auto_cleared": f.auto_cleared,
        "cleared_by": (
            f.cleared_by.get_username() if f.cleared_by else ""
        ),
    }


@login_required
@require_http_methods(["POST"])
def waybill_flag_create(request, waybill):
    """Sinaliza um waybill (cria WaybillFlag activo).

    Body JSON: { flag_type, reason, notes }
    Se já existe flag activo, é actualizado em vez de duplicado.
    """
    import json
    from .models import WaybillFlag

    waybill = (waybill or "").strip()
    if not waybill:
        return JsonResponse(
            {"success": False, "error": "Waybill vazio"}, status=400,
        )

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    flag_type = (body.get("flag_type") or "").strip()
    valid_types = [t[0] for t in WaybillFlag.TYPE_CHOICES]
    if flag_type not in valid_types:
        flag_type = WaybillFlag.TYPE_NOT_ARRIVED

    reason = (body.get("reason") or "other").strip()
    valid_reasons = [r[0] for r in WaybillFlag.REASON_CHOICES]
    if reason not in valid_reasons:
        reason = "other"

    user = request.user if request.user.is_authenticated else None

    # Reusa flag activo se existir, senão cria novo
    existing = WaybillFlag.objects.filter(
        waybill_number=waybill, cleared_at__isnull=True,
    ).first()
    if existing:
        existing.flag_type = flag_type
        existing.reason = reason
        existing.notes = (body.get("notes") or "").strip()
        existing.flagged_by = user or existing.flagged_by
        existing.save()
        return JsonResponse({
            "success": True, "created": False,
            "flag": _serialize_flag(existing),
        })

    flag = WaybillFlag.objects.create(
        waybill_number=waybill,
        flag_type=flag_type,
        reason=reason,
        notes=(body.get("notes") or "").strip(),
        flagged_by=user,
    )
    return JsonResponse({
        "success": True, "created": True,
        "flag": _serialize_flag(flag),
    })


@login_required
@require_http_methods(["POST"])
def waybill_flag_clear(request, waybill):
    """Limpa o flag activo de um waybill (manual).

    Body JSON: { cleared_reason: str opcional }
    """
    import json
    from .models import WaybillFlag
    from django.utils import timezone as _tz

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    user = request.user if request.user.is_authenticated else None
    n = WaybillFlag.objects.filter(
        waybill_number=waybill, cleared_at__isnull=True,
    ).update(
        cleared_at=_tz.now(),
        cleared_reason=(
            body.get("cleared_reason")
            or "Limpo manualmente"
        ),
        cleared_by=user,
        auto_cleared=False,
    )
    return JsonResponse({"success": True, "cleared": n})


@login_required
def waybill_flag_get(request, waybill):
    """Devolve flag activo (se existir) e histórico de flags
    deste waybill."""
    from .models import WaybillFlag
    qs = WaybillFlag.objects.filter(
        waybill_number=waybill,
    ).select_related("flagged_by", "cleared_by").order_by(
        "-flagged_at",
    )
    flags = [_serialize_flag(f) for f in qs]
    active = next((f for f in flags if f["is_active"]), None)
    return JsonResponse({
        "success": True,
        "active": active,
        "history": flags,
    })


@login_required
def waybill_flag_list(request):
    """Lista flags activos com filtros + sumário."""
    from .models import WaybillFlag, CainiaoOperationTask

    flag_type = (request.GET.get("flag_type") or "").strip()
    qs = WaybillFlag.objects.filter(
        cleared_at__isnull=True,
    ).select_related("flagged_by")
    if flag_type:
        qs = qs.filter(flag_type=flag_type)
    flags = list(qs.order_by("-flagged_at"))

    # Enriquecer com info do snapshot mais recente
    waybills = [f.waybill_number for f in flags]
    snapshots = {}
    if waybills:
        for op in (
            CainiaoOperationTask.objects
            .filter(waybill_number__in=waybills)
            .order_by("waybill_number", "-task_date", "-id")
        ):
            if op.waybill_number not in snapshots:
                snapshots[op.waybill_number] = op

    rows = []
    for f in flags:
        snap = snapshots.get(f.waybill_number)
        d = _serialize_flag(f)
        d["cp4"] = (snap.zip_code or "")[:4] if snap else ""
        d["city"] = snap.destination_city if snap else ""
        d["courier_name"] = snap.courier_name if snap else ""
        d["current_status"] = snap.task_status if snap else ""
        rows.append(d)

    # Sumário
    by_type = {}
    by_reason = {}
    for f in flags:
        by_type[f.flag_type] = by_type.get(f.flag_type, 0) + 1
        by_reason[f.reason] = by_reason.get(f.reason, 0) + 1

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "rows": rows,
        "by_type": by_type,
        "by_reason": by_reason,
    })


@login_required
@require_http_methods(["POST"])
def waybill_flag_bulk(request):
    """Sinaliza vários waybills de uma vez.

    Body: { waybills: [str], flag_type, reason, notes }
    """
    import json
    from .models import WaybillFlag

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    waybills = body.get("waybills") or []
    if not isinstance(waybills, list) or not waybills:
        return JsonResponse(
            {"success": False, "error": "waybills vazios"},
            status=400,
        )

    flag_type = (body.get("flag_type") or "").strip()
    valid_types = [t[0] for t in WaybillFlag.TYPE_CHOICES]
    if flag_type not in valid_types:
        flag_type = WaybillFlag.TYPE_NOT_ARRIVED

    reason = (body.get("reason") or "other").strip()
    valid_reasons = [r[0] for r in WaybillFlag.REASON_CHOICES]
    if reason not in valid_reasons:
        reason = "other"

    notes = (body.get("notes") or "").strip()
    user = request.user if request.user.is_authenticated else None

    n_created = 0
    n_updated = 0
    for wb in waybills:
        wb = str(wb).strip()
        if not wb:
            continue
        existing = WaybillFlag.objects.filter(
            waybill_number=wb, cleared_at__isnull=True,
        ).first()
        if existing:
            existing.flag_type = flag_type
            existing.reason = reason
            if notes:
                existing.notes = notes
            existing.save()
            n_updated += 1
        else:
            WaybillFlag.objects.create(
                waybill_number=wb,
                flag_type=flag_type, reason=reason,
                notes=notes, flagged_by=user,
            )
            n_created += 1
    return JsonResponse({
        "success": True,
        "n_created": n_created,
        "n_updated": n_updated,
        "total": n_created + n_updated,
    })


# ============================================================================
# Camada 2 — Snapshots diários
# ============================================================================

def _build_snapshot(*, hub_id=None, automatic=True, user=None):
    """Cria um snapshot da lista Not Arrived actual.

    Retorna o objecto NotArrivedSnapshot criado (ou existente se já
    havia um para hoje + hub).
    """
    from django.utils import timezone as _tz
    from .models import (
        NotArrivedSnapshot, NotArrivedSnapshotRow, CainiaoHub,
    )
    from decimal import Decimal

    hub = None
    if hub_id:
        try:
            hub = CainiaoHub.objects.filter(id=int(hub_id)).first()
        except (ValueError, TypeError):
            hub = None

    today_d = _tz.now().date()
    rows = _compute_not_arrived(
        hub_id=(hub.id if hub else None), min_days=2,
    )

    # update_or_create — uma snapshot por (data, hub) por dia
    snap, created = NotArrivedSnapshot.objects.update_or_create(
        snapshot_date=today_d, hub=hub,
        defaults={
            "is_automatic": automatic,
            "created_by": user,
            "total_packages": len(rows),
            "total_cost_eur": Decimal(
                str(sum(r.get("estimated_cost_eur", 0) for r in rows))
            ),
            "n_to_escalate": sum(
                1 for r in rows if r.get("should_escalate")
            ),
            "n_drivers_affected": len({
                r.get("courier_name") for r in rows
                if r.get("courier_name")
            }),
            "n_cp4s_affected": len({
                r.get("cp4") for r in rows if r.get("cp4")
            }),
        },
    )

    # Re-criar rows (idempotente)
    if not created:
        snap.rows.all().delete()

    bulk = []
    from datetime import datetime as _dt2
    for r in rows:
        first_seen_d = None
        if r.get("first_seen"):
            try:
                first_seen_d = _dt2.strptime(
                    r["first_seen"], "%Y-%m-%d",
                ).date()
            except (ValueError, TypeError):
                pass
        bulk.append(NotArrivedSnapshotRow(
            snapshot=snap,
            waybill_number=r["waybill"],
            courier_name=r.get("courier_name") or "",
            courier_id_cainiao=r.get("courier_id_cainiao") or "",
            cp4=r.get("cp4") or "",
            city=r.get("city") or "",
            customer_name=r.get("customer_name") or "",
            seller=r.get("seller") or "",
            weight_g=r.get("weight_g") or None,
            business_days_stuck=r.get("business_days_stuck") or 0,
            sla_level=r.get("sla_level") or "",
            first_seen=first_seen_d,
            estimated_cost_eur=Decimal(
                str(r.get("estimated_cost_eur", 0))
            ),
        ))
    if bulk:
        NotArrivedSnapshotRow.objects.bulk_create(
            bulk, batch_size=500,
        )

    return snap


@login_required
@require_http_methods(["POST"])
def cainiao_snapshot_create(request):
    """Cria snapshot manual da Not Arrived agora (sem esperar pelo
    Celery)."""
    user = request.user if request.user.is_authenticated else None
    snap = _build_snapshot(automatic=False, user=user)
    return JsonResponse({
        "success": True,
        "snapshot_id": snap.id,
        "snapshot_date": snap.snapshot_date.strftime("%Y-%m-%d"),
        "total_packages": snap.total_packages,
    })


@login_required
def cainiao_snapshot_list(request):
    """Lista snapshots (cabeçalhos)."""
    from .models import NotArrivedSnapshot
    qs = NotArrivedSnapshot.objects.select_related(
        "hub", "created_by",
    ).order_by("-snapshot_date")[:60]
    return JsonResponse({
        "success": True,
        "snapshots": [{
            "id": s.id,
            "snapshot_date": s.snapshot_date.strftime("%Y-%m-%d"),
            "created_at": s.created_at.strftime("%Y-%m-%d %H:%M"),
            "is_automatic": s.is_automatic,
            "hub": s.hub.name if s.hub else "",
            "total_packages": s.total_packages,
            "total_cost_eur": float(s.total_cost_eur),
            "n_to_escalate": s.n_to_escalate,
            "n_drivers_affected": s.n_drivers_affected,
            "n_cp4s_affected": s.n_cp4s_affected,
            "created_by": (
                s.created_by.get_username()
                if s.created_by else "(automático)"
            ),
        } for s in qs],
    })


@login_required
def cainiao_snapshot_detail(request, snapshot_id):
    """Detalhe completo de um snapshot."""
    from .models import NotArrivedSnapshot, WaybillResolution
    snap = get_object_or_404(NotArrivedSnapshot, id=snapshot_id)
    rows = list(snap.rows.all().order_by(
        "-business_days_stuck", "courier_name",
    ))

    # Para cada row, indicar se entretanto foi resolvido (status hoje)
    waybills = [r.waybill_number for r in rows]
    resolved = set(
        WaybillResolution.objects.filter(
            waybill_number__in=waybills,
        ).values_list("waybill_number", flat=True)
    )

    return JsonResponse({
        "success": True,
        "snapshot": {
            "id": snap.id,
            "snapshot_date": snap.snapshot_date.strftime("%Y-%m-%d"),
            "created_at": snap.created_at.strftime("%Y-%m-%d %H:%M"),
            "is_automatic": snap.is_automatic,
            "total_packages": snap.total_packages,
            "total_cost_eur": float(snap.total_cost_eur),
            "n_to_escalate": snap.n_to_escalate,
            "n_drivers_affected": snap.n_drivers_affected,
            "n_cp4s_affected": snap.n_cp4s_affected,
            "hub": snap.hub.name if snap.hub else "",
        },
        "rows": [{
            "waybill": r.waybill_number,
            "courier_name": r.courier_name,
            "courier_id_cainiao": r.courier_id_cainiao,
            "cp4": r.cp4,
            "city": r.city,
            "customer_name": r.customer_name,
            "seller": r.seller,
            "weight_g": r.weight_g or 0,
            "business_days_stuck": r.business_days_stuck,
            "sla_level": r.sla_level,
            "first_seen": (
                r.first_seen.strftime("%Y-%m-%d")
                if r.first_seen else ""
            ),
            "estimated_cost_eur": float(r.estimated_cost_eur),
            "resolved_since": r.waybill_number in resolved,
        } for r in rows],
    })


@login_required
def cainiao_snapshot_diff(request):
    """Compara dois snapshots e devolve added/removed/maintained.

    Query: ?from=<id>&to=<id>
    """
    from .models import NotArrivedSnapshot
    try:
        f_id = int(request.GET.get("from") or 0)
        t_id = int(request.GET.get("to") or 0)
    except (ValueError, TypeError):
        return JsonResponse(
            {"success": False, "error": "IDs inválidos"}, status=400,
        )
    sf = NotArrivedSnapshot.objects.filter(id=f_id).first()
    st = NotArrivedSnapshot.objects.filter(id=t_id).first()
    if not sf or not st:
        return JsonResponse(
            {"success": False, "error": "Snapshot não encontrado"},
            status=404,
        )

    from_rows = {r.waybill_number: r for r in sf.rows.all()}
    to_rows = {r.waybill_number: r for r in st.rows.all()}

    added_keys = set(to_rows) - set(from_rows)
    removed_keys = set(from_rows) - set(to_rows)
    maintained_keys = set(from_rows) & set(to_rows)

    def _row_dict(r):
        return {
            "waybill": r.waybill_number,
            "courier_name": r.courier_name,
            "cp4": r.cp4,
            "city": r.city,
            "business_days_stuck": r.business_days_stuck,
        }

    return JsonResponse({
        "success": True,
        "from": {
            "id": sf.id, "date": sf.snapshot_date.strftime("%Y-%m-%d"),
            "total": sf.total_packages,
        },
        "to": {
            "id": st.id, "date": st.snapshot_date.strftime("%Y-%m-%d"),
            "total": st.total_packages,
        },
        "added": [_row_dict(to_rows[k]) for k in sorted(added_keys)],
        "removed": [
            _row_dict(from_rows[k]) for k in sorted(removed_keys)
        ],
        "maintained_count": len(maintained_keys),
        "n_added": len(added_keys),
        "n_removed": len(removed_keys),
    })


@login_required
def cainiao_waybill_lifecycle(request, waybill):
    """Ciclo de vida de um waybill em snapshots + flags + resolutions.

    Mostra timeline cronológica:
      - Quando apareceu pela 1ª vez em snapshot
      - Quando foi sinalizado/limpo (flags)
      - Quando foi resolvido
    """
    from .models import (
        NotArrivedSnapshotRow, WaybillFlag, WaybillResolution,
    )

    # Snapshots em que o waybill apareceu
    snap_rows = list(
        NotArrivedSnapshotRow.objects.filter(
            waybill_number=waybill,
        ).select_related("snapshot").order_by("snapshot__snapshot_date")
    )
    flags = list(WaybillFlag.objects.filter(
        waybill_number=waybill,
    ).select_related("flagged_by", "cleared_by").order_by("flagged_at"))
    resolution = WaybillResolution.objects.filter(
        waybill_number=waybill,
    ).select_related("resolved_by").first()

    timeline = []
    if snap_rows:
        timeline.append({
            "kind": "first_snapshot",
            "date": snap_rows[0].snapshot.snapshot_date.strftime(
                "%Y-%m-%d",
            ),
            "label": "Primeira aparição em snapshot diário",
            "details": {
                "courier": snap_rows[0].courier_name,
                "cp4": snap_rows[0].cp4,
                "days_stuck": snap_rows[0].business_days_stuck,
            },
        })
        timeline.append({
            "kind": "last_snapshot",
            "date": snap_rows[-1].snapshot.snapshot_date.strftime(
                "%Y-%m-%d",
            ),
            "label": (
                f"Última aparição em snapshot ({len(snap_rows)} "
                "snapshots no total)"
            ),
            "details": {
                "courier": snap_rows[-1].courier_name,
                "days_stuck": snap_rows[-1].business_days_stuck,
            },
        })
    for f in flags:
        timeline.append({
            "kind": "flag_created",
            "date": f.flagged_at.strftime("%Y-%m-%d %H:%M"),
            "label": f"Sinalizado: {f.get_flag_type_display()}",
            "details": {
                "reason": f.get_reason_display(),
                "by": (
                    f.flagged_by.get_username()
                    if f.flagged_by else "(sistema)"
                ),
                "notes": f.notes,
            },
        })
        if f.cleared_at:
            timeline.append({
                "kind": "flag_cleared",
                "date": f.cleared_at.strftime("%Y-%m-%d %H:%M"),
                "label": (
                    "Flag limpo automaticamente"
                    if f.auto_cleared else "Flag limpo manualmente"
                ),
                "details": {
                    "reason": f.cleared_reason,
                    "by": (
                        f.cleared_by.get_username()
                        if f.cleared_by else ""
                    ),
                },
            })
    if resolution:
        timeline.append({
            "kind": "resolution",
            "date": resolution.resolved_at.strftime("%Y-%m-%d %H:%M"),
            "label": (
                f"Resolvido: {resolution.get_resolution_type_display()}"
            ),
            "details": {
                "by": (
                    resolution.resolved_by.get_username()
                    if resolution.resolved_by else "(sistema)"
                ),
                "other_hub": resolution.other_hub,
                "other_courier": resolution.other_courier,
                "notes": resolution.notes,
            },
        })

    timeline.sort(key=lambda x: x["date"])

    return JsonResponse({
        "success": True,
        "waybill": waybill,
        "n_snapshots": len(snap_rows),
        "snapshot_dates": [
            r.snapshot.snapshot_date.strftime("%Y-%m-%d")
            for r in snap_rows
        ],
        "n_flags": len(flags),
        "active_flag": next(
            (
                {"type": f.flag_type, "reason": f.reason}
                for f in flags if f.is_active
            ), None,
        ),
        "is_resolved": resolution is not None,
        "timeline": timeline,
    })


@login_required
def cainiao_snapshot_xlsx(request, snapshot_id):
    """Exporta um snapshot histórico para Excel (auditoria)."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from .models import NotArrivedSnapshot

    snap = get_object_or_404(NotArrivedSnapshot, id=snapshot_id)
    rows = list(snap.rows.all().order_by(
        "-business_days_stuck", "courier_name",
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Snapshot {snap.snapshot_date}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="7C2D12")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = (
        f"Snapshot Histórico Not Arrived — {snap.snapshot_date}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="7C2D12")
    ws.merge_cells("A1:K1")
    ws["A2"] = (
        f"Total: {snap.total_packages} pacotes · "
        f"€{snap.total_cost_eur:.2f} em risco · "
        f"{snap.n_to_escalate} a escalar · "
        f"Gerado: {snap.created_at:%Y-%m-%d %H:%M}"
    )
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:K2")

    headers = [
        "Waybill", "Driver", "Courier ID", "CP4", "Cidade",
        "Cliente", "Seller", "Peso (g)", "Dias úteis parado",
        "SLA", "1ª aparição",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for ridx, r in enumerate(rows, start=5):
        ws.cell(row=ridx, column=1, value=r.waybill_number)
        ws.cell(row=ridx, column=2, value=r.courier_name)
        ws.cell(row=ridx, column=3, value=r.courier_id_cainiao)
        ws.cell(row=ridx, column=4, value=r.cp4)
        ws.cell(row=ridx, column=5, value=r.city)
        ws.cell(row=ridx, column=6, value=r.customer_name)
        ws.cell(row=ridx, column=7, value=r.seller)
        ws.cell(row=ridx, column=8, value=r.weight_g or 0)
        ws.cell(row=ridx, column=9, value=r.business_days_stuck)
        ws.cell(row=ridx, column=10, value=r.sla_level)
        ws.cell(
            row=ridx, column=11,
            value=r.first_seen.strftime("%Y-%m-%d") if r.first_seen else "",
        )
        for ci in range(1, 12):
            ws.cell(row=ridx, column=ci).border = border

    widths = [28, 22, 14, 8, 18, 22, 22, 10, 12, 8, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    fname = (
        f"snapshot_not_arrived_{snap.snapshot_date}.xlsx"
    )
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{fname}"'
    )
    wb.save(response)
    return response


# ============================================================================
# Camada 3 — Lógica de Prioridade + Settings + Página global
# ============================================================================

def _compute_package_priority(
    snapshot, n_attempts=None, settings=None, is_returned=None,
):
    """Determina se um pacote é prioritário e devolve as razões.

    NÃO marca como prioritário se:
      - o pacote já está entregue (task_status == 'Delivered'), OU
      - foi devolvido ao remetente (WaybillReturn fechada).
    Em ambos os casos não há nada a fazer.
    """
    from django.utils import timezone as _tz
    from .models import (
        PrioritySettings, CainiaoOperationTask, WaybillReturn,
    )

    if settings is None:
        settings = PrioritySettings.load()

    reasons = []
    is_priority = False
    is_delivered = (
        (snapshot.task_status or "").strip() == "Delivered"
    )

    # Verificar se foi devolvido (apenas se não passou explicitamente)
    if is_returned is None:
        ret = WaybillReturn.objects.filter(
            waybill_number=snapshot.waybill_number,
            return_status__in=(
                WaybillReturn.STATUS_RETURNED,
                WaybillReturn.STATUS_CLOSED,
            ),
        ).only("id").first()
        is_returned = bool(ret)

    # Calcular days_in_hub mesmo para Delivered (mostra info histórica)
    days_in_hub = 0
    ref_time = snapshot.receipt_time or snapshot.creation_time
    if ref_time:
        try:
            # Se entregue, conta até à entrega; senão, até hoje
            end_time = (
                snapshot.delivery_time if is_delivered
                else _tz.now()
            ) or _tz.now()
            delta = end_time - ref_time
            days_in_hub = max(0, delta.days)
        except Exception:
            days_in_hub = 0

    # Calcular tentativas
    if n_attempts is None:
        n_attempts = (
            CainiaoOperationTask.objects.filter(
                waybill_number=snapshot.waybill_number,
                task_status="Attempt Failure",
            ).count()
        )

    # CRÍTICO: pacote entregue ou devolvido nunca é prioritário
    if is_delivered or is_returned:
        return {
            "is_priority": False,
            "reasons": [],
            "days_in_hub": days_in_hub,
            "n_attempts": n_attempts,
            "is_delivered": is_delivered,
            "is_returned": is_returned,
        }

    # Flag externa do Cainiao
    if getattr(snapshot, "is_priority_external", False):
        is_priority = True
        reasons.append({
            "type": "external",
            "label": "Marcado PRIORITYS=Yes pelo Cainiao",
        })

    if not settings.auto_apply:
        return {
            "is_priority": is_priority,
            "reasons": reasons,
            "days_in_hub": days_in_hub,
            "n_attempts": n_attempts,
            "is_delivered": False,
            "is_returned": False,
        }

    if (
        settings.min_days_in_hub > 0
        and days_in_hub >= settings.min_days_in_hub
    ):
        is_priority = True
        reasons.append({
            "type": "days_in_hub",
            "label": (
                f"Há {days_in_hub} dia(s) no HUB "
                f"(limite: {settings.min_days_in_hub})"
            ),
        })

    if (
        settings.min_attempts > 0
        and n_attempts >= settings.min_attempts
    ):
        is_priority = True
        reasons.append({
            "type": "attempts",
            "label": (
                f"{n_attempts} tentativa(s) de entrega "
                f"(limite: {settings.min_attempts})"
            ),
        })

    return {
        "is_priority": is_priority,
        "reasons": reasons,
        "days_in_hub": days_in_hub,
        "n_attempts": n_attempts,
        "is_delivered": False,
        "is_returned": False,
    }


@login_required
def priority_settings_get(request):
    from .models import PrioritySettings
    s = PrioritySettings.load()
    return JsonResponse({
        "success": True,
        "settings": {
            "min_days_in_hub": s.min_days_in_hub,
            "min_attempts": s.min_attempts,
            "auto_apply": s.auto_apply,
            "updated_at": s.updated_at.strftime("%Y-%m-%d %H:%M"),
        },
    })


@login_required
@require_http_methods(["POST"])
def priority_settings_save(request):
    import json
    from .models import PrioritySettings
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    s = PrioritySettings.load()
    if "min_days_in_hub" in body:
        try:
            s.min_days_in_hub = max(0, int(body["min_days_in_hub"]))
        except (ValueError, TypeError):
            pass
    if "min_attempts" in body:
        try:
            s.min_attempts = max(0, int(body["min_attempts"]))
        except (ValueError, TypeError):
            pass
    if "auto_apply" in body:
        s.auto_apply = bool(body["auto_apply"])
    s.save()
    return JsonResponse({
        "success": True,
        "settings": {
            "min_days_in_hub": s.min_days_in_hub,
            "min_attempts": s.min_attempts,
            "auto_apply": s.auto_apply,
        },
    })


@login_required
def cainiao_packages_search(request):
    """Busca avançada global de pacotes.

    Por defeito, restringe a pacotes cujo CP4 está cadastrado em algum
    HUB do sistema (CainiaoHubCP4). Usar `include_unregistered=1` para
    desactivar esse filtro e ver todos os pacotes.
    """
    from django.db.models import Q, Count
    from django.core.paginator import Paginator
    from .models import (
        CainiaoOperationTask, CainiaoHub, CainiaoHubCP4,
        WaybillTag, WaybillFlag, WaybillResolution,
        PrioritySettings,
    )

    from django.utils.dateparse import parse_date
    settings_obj = PrioritySettings.load()
    qs = CainiaoOperationTask.objects.all()

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(waybill_number__icontains=q)
            | Q(lp_number__icontains=q)
            | Q(courier_name__icontains=q)
            | Q(seller_name__icontains=q)
            | Q(zip_code__icontains=q)
            | Q(destination_city__icontains=q)
            | Q(detailed_address__icontains=q)
        )

    # NOTA: filtros de status (status_csv, open_only) são aplicados
    # DEPOIS do dedupe (mais abaixo). Aplicar aqui causaria pacotes
    # com row Driver_received antiga + Delivered actual a aparecerem
    # como "Driver_received" mesmo já estando entregues.
    status_csv = (request.GET.get("status") or "").strip()
    statuses_filter = [s for s in status_csv.split(",") if s]
    open_only = request.GET.get("open_only") == "1"

    hub_id = request.GET.get("hub_id")
    include_unregistered = request.GET.get(
        "include_unregistered",
    ) == "1"

    n_registered_cp4s = 0
    if hub_id:
        # Filtro por HUB específico
        try:
            hub = CainiaoHub.objects.filter(id=int(hub_id)).first()
            if hub:
                cp4s = list(
                    CainiaoHubCP4.objects.filter(hub=hub)
                    .values_list("cp4", flat=True)
                )
                if cp4s:
                    cp4_q = Q()
                    for cp4 in cp4s:
                        cp4_q |= Q(zip_code__startswith=cp4)
                    qs = qs.filter(cp4_q)
        except (ValueError, TypeError):
            pass
    elif not include_unregistered:
        # Default: restringe a TODOS os CP4s cadastrados em qualquer HUB
        all_cp4s = list(
            CainiaoHubCP4.objects.values_list(
                "cp4", flat=True,
            ).distinct()
        )
        n_registered_cp4s = len(all_cp4s)
        if all_cp4s:
            cp4_q = Q()
            for cp4 in all_cp4s:
                cp4_q |= Q(zip_code__startswith=cp4)
            qs = qs.filter(cp4_q)

    cp4_csv = (request.GET.get("cp4") or "").strip()
    if cp4_csv:
        cp4_q = Q()
        for cp4 in cp4_csv.split(","):
            cp4 = cp4.strip()
            if cp4:
                cp4_q |= Q(zip_code__startswith=cp4)
        qs = qs.filter(cp4_q)

    driver = (request.GET.get("driver") or "").strip()
    if driver:
        qs = qs.filter(courier_name__icontains=driver)
    seller = (request.GET.get("seller") or "").strip()
    if seller:
        qs = qs.filter(seller_name__icontains=seller)

    if request.GET.get("wrong_hub") == "1":
        qs = qs.filter(arrive_wrong_hub=True)
    if request.GET.get("has_exception") == "1":
        qs = qs.exclude(exception_type="")

    df = (request.GET.get("date_from") or "").strip()
    if df:
        try:
            d = parse_date(df)
            if d:
                qs = qs.filter(task_date__gte=d)
        except (ValueError, TypeError):
            pass
    dt_ = (request.GET.get("date_to") or "").strip()
    if dt_:
        try:
            d = parse_date(dt_)
            if d:
                qs = qs.filter(task_date__lte=d)
        except (ValueError, TypeError):
            pass

    priority_filter = (request.GET.get("priority") or "").strip()

    qs = qs.order_by("waybill_number", "-task_date", "-id")
    seen = set()
    latest = []
    for op in qs[:5000]:
        if op.waybill_number in seen:
            continue
        seen.add(op.waybill_number)
        latest.append(op)

    # Aplicar filtros de status APÓS dedupe — contra o estado actual.
    # Isto garante que filtrar por "Driver_received" só mostra pacotes
    # actualmente nesse estado, e não os que tiveram essa row no passado
    # mas foram entregues entretanto.
    if statuses_filter:
        latest = [
            op for op in latest if op.task_status in statuses_filter
        ]
    if open_only:
        latest = [
            op for op in latest if op.task_status != "Delivered"
        ]

    waybills = [op.waybill_number for op in latest]
    attempts_count = dict(
        CainiaoOperationTask.objects.filter(
            waybill_number__in=waybills,
            task_status="Attempt Failure",
        ).values("waybill_number").annotate(
            n=Count("id"),
        ).values_list("waybill_number", "n")
    )

    tags_by_wb = {}
    for t in WaybillTag.objects.filter(
        waybill_number__in=waybills,
    ):
        tags_by_wb.setdefault(t.waybill_number, []).append(t.tag)
    flags_set = set(
        WaybillFlag.objects.filter(
            waybill_number__in=waybills,
            cleared_at__isnull=True,
        ).values_list("waybill_number", flat=True)
    )
    resolved_set = set(
        WaybillResolution.objects.filter(
            waybill_number__in=waybills,
        ).values_list("waybill_number", flat=True)
    )
    from .models import WaybillReturn as _WR
    returns_by_wb = {
        r.waybill_number: r
        for r in _WR.objects.filter(waybill_number__in=waybills)
    }
    returned_set = {
        wb for wb, r in returns_by_wb.items()
        if r.return_status in (_WR.STATUS_RETURNED, _WR.STATUS_CLOSED)
    }

    rows = []
    for op in latest:
        n_att = attempts_count.get(op.waybill_number, 0)
        is_ret = op.waybill_number in returned_set
        prio = _compute_package_priority(
            op, n_attempts=n_att, settings=settings_obj,
            is_returned=is_ret,
        )
        if priority_filter == "yes" and not prio["is_priority"]:
            continue
        if priority_filter == "no" and prio["is_priority"]:
            continue
        rows.append({
            "waybill": op.waybill_number,
            "lp_number": op.lp_number,
            "task_date": op.task_date.strftime("%Y-%m-%d"),
            "task_status": op.task_status,
            "courier_name": op.courier_name,
            "dsp_name": op.dsp_name,
            "cp4": (op.zip_code or "")[:4],
            "city": op.destination_city,
            "zip_code": op.zip_code,
            "address": op.detailed_address,
            "seller": op.seller_name,
            "weight_g": op.weight_g or 0,
            "creation_time": (
                op.creation_time.strftime("%Y-%m-%d %H:%M")
                if op.creation_time else ""
            ),
            "receipt_time": (
                op.receipt_time.strftime("%Y-%m-%d %H:%M")
                if op.receipt_time else ""
            ),
            # Só mostrar delivery_time se o status for Delivered (caso
            # contrário a coluna marca falsamente como entregue um pacote
            # que ainda está em Attempt Failure / Driver_received).
            "delivery_time": (
                op.delivery_time.strftime("%Y-%m-%d %H:%M")
                if op.delivery_time
                and op.task_status == "Delivered" else ""
            ),
            "failure_time": (
                op.delivery_failure_time.strftime("%Y-%m-%d %H:%M")
                if op.delivery_failure_time
                and op.task_status == "Attempt Failure" else ""
            ),
            "is_priority": prio["is_priority"],
            "priority_reasons": prio["reasons"],
            "days_in_hub": prio["days_in_hub"],
            "n_attempts": n_att,
            "is_priority_external": op.is_priority_external,
            "wrong_hub_parcel": op.wrong_hub_parcel,
            "arrive_wrong_hub": op.arrive_wrong_hub,
            "exception_type": op.exception_type,
            "tags": tags_by_wb.get(op.waybill_number, []),
            "is_flagged": op.waybill_number in flags_set,
            "is_resolved": op.waybill_number in resolved_set,
            "is_returned": is_ret,
            "return_date": (
                returns_by_wb[op.waybill_number].return_date.strftime(
                    "%Y-%m-%d",
                )
                if op.waybill_number in returns_by_wb else ""
            ),
            "return_status": (
                returns_by_wb[op.waybill_number].return_status
                if op.waybill_number in returns_by_wb else ""
            ),
        })

    try:
        page_size = max(10, min(int(
            request.GET.get("page_size") or 50,
        ), 200))
    except (ValueError, TypeError):
        page_size = 50
    paginator = Paginator(rows, page_size)
    try:
        page_number = int(request.GET.get("page") or 1)
    except (ValueError, TypeError):
        page_number = 1
    page_obj = paginator.get_page(page_number)

    n_priority = sum(1 for r in rows if r["is_priority"])
    n_open = sum(
        1 for r in rows
        if r["task_status"] != "Delivered" and not r["is_returned"]
    )
    n_wrong_hub = sum(1 for r in rows if r["arrive_wrong_hub"])

    return JsonResponse({
        "success": True,
        "total": len(rows),
        "page": page_obj.number,
        "n_pages": paginator.num_pages,
        "page_size": page_size,
        "rows": list(page_obj.object_list),
        "kpis": {
            "n_total": len(rows),
            "n_priority": n_priority,
            "n_open": n_open,
            "n_wrong_hub": n_wrong_hub,
        },
        "settings": {
            "min_days_in_hub": settings_obj.min_days_in_hub,
            "min_attempts": settings_obj.min_attempts,
            "auto_apply": settings_obj.auto_apply,
        },
        "filter_info": {
            "include_unregistered": include_unregistered,
            "n_registered_cp4s": n_registered_cp4s,
            "hub_filtered": bool(hub_id),
        },
    })


@login_required
def cainiao_packages_page(request):
    """Renderiza a página HTML de pesquisa global de pacotes."""
    from .models import CainiaoHub
    return render(
        request,
        "settlements/cainiao_packages.html",
        {"hubs": CainiaoHub.objects.all().order_by("name")},
    )


@login_required
def cainiao_hubs_with_counts(request):
    """Lista HUBs com contagens (total pacotes, abertos, prioritários).

    Usa filtros opcionais (priority, open_only) consistente com a
    pesquisa global. Útil para os quick-chips HUB.
    """
    from django.db.models import Q, Count
    from .models import (
        CainiaoHub, CainiaoHubCP4, CainiaoOperationTask,
        PrioritySettings,
    )

    settings_obj = PrioritySettings.load()
    hubs = list(CainiaoHub.objects.all().order_by("name"))
    open_only = request.GET.get("open_only") == "1"
    priority_only = request.GET.get("priority") == "yes"

    result = []
    for hub in hubs:
        cp4s = list(
            CainiaoHubCP4.objects.filter(hub=hub)
            .values_list("cp4", flat=True).distinct()
        )
        if not cp4s:
            result.append({
                "id": hub.id, "name": hub.name,
                "n_cp4s": 0, "n_total": 0, "n_open": 0,
                "n_priority": 0, "cp4_codes": [],
            })
            continue
        cp4_q = Q()
        for cp4 in cp4s:
            cp4_q |= Q(zip_code__startswith=cp4)
        # Apenas snapshots mais recentes por waybill
        latest_qs = CainiaoOperationTask.objects.filter(
            cp4_q,
        ).order_by("waybill_number", "-task_date", "-id")
        seen = set()
        n_total = 0
        n_open = 0
        n_priority = 0
        # Para evitar carregar 50k objectos só para contar,
        # usamos amostragem rápida via SQL: dedupe em Python
        # com LIMIT alto.
        for op in latest_qs[:30000]:
            if op.waybill_number in seen:
                continue
            seen.add(op.waybill_number)
            n_total += 1
            if op.task_status != "Delivered":
                n_open += 1
                # Priority compute (rápido)
                prio = _compute_package_priority(
                    op, settings=settings_obj,
                )
                if prio["is_priority"]:
                    n_priority += 1
        result.append({
            "id": hub.id, "name": hub.name,
            "n_cp4s": len(cp4s),
            "n_total": n_total,
            "n_open": n_open,
            "n_priority": n_priority,
            "cp4_codes": sorted(cp4s),
        })

    return JsonResponse({
        "success": True,
        "hubs": result,
    })


# ============================================================================
# Tags + Watch + SavedFilter — endpoints CRUD
# ============================================================================

@login_required
def waybill_tags_list(request, waybill):
    from .models import WaybillTag
    tags = list(
        WaybillTag.objects.filter(waybill_number=waybill)
        .select_related("created_by")
    )
    return JsonResponse({
        "success": True,
        "tags": [{
            "id": t.id, "tag": t.tag, "notes": t.notes,
            "created_at": t.created_at.strftime("%Y-%m-%d %H:%M"),
            "by": t.created_by.username if t.created_by else "",
        } for t in tags],
    })


@login_required
@require_http_methods(["POST"])
def waybill_tag_add(request, waybill):
    import json
    from .models import WaybillTag
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    tag = (body.get("tag") or "").strip()[:50]
    if not tag:
        return JsonResponse(
            {"success": False, "error": "tag vazia"}, status=400,
        )
    obj, created = WaybillTag.objects.get_or_create(
        waybill_number=waybill, tag=tag,
        defaults={
            "notes": (body.get("notes") or "")[:255],
            "created_by": (
                request.user if request.user.is_authenticated
                else None
            ),
        },
    )
    return JsonResponse({
        "success": True, "created": created, "id": obj.id,
    })


@login_required
@require_http_methods(["POST"])
def waybill_tag_delete(request, tag_id):
    from .models import WaybillTag
    n, _ = WaybillTag.objects.filter(id=tag_id).delete()
    return JsonResponse({"success": True, "deleted": n})


@login_required
@require_http_methods(["POST"])
def waybill_watch_toggle(request, waybill):
    from .models import WaybillWatch
    user = request.user
    existing = WaybillWatch.objects.filter(
        waybill_number=waybill, user=user,
    ).first()
    if existing:
        existing.delete()
        return JsonResponse({"success": True, "watching": False})
    WaybillWatch.objects.create(waybill_number=waybill, user=user)
    return JsonResponse({"success": True, "watching": True})


@login_required
def waybill_watch_list(request):
    from .models import WaybillWatch
    qs = WaybillWatch.objects.filter(user=request.user).order_by(
        "-created_at",
    )
    return JsonResponse({
        "success": True,
        "watches": [{
            "waybill": w.waybill_number,
            "notes": w.notes,
            "created_at": w.created_at.strftime("%Y-%m-%d %H:%M"),
        } for w in qs],
    })


@login_required
@require_http_methods(["GET", "POST", "DELETE"])
def saved_filters(request):
    """GET: lista; POST: cria/atualiza; DELETE ?id=X: apaga."""
    import json
    from .models import SavedSearchFilter
    from django.db.models import Q
    user = request.user
    if request.method == "GET":
        qs = SavedSearchFilter.objects.filter(
            Q(user=user) | Q(is_shared=True),
        )
        return JsonResponse({
            "success": True,
            "filters": [{
                "id": f.id, "name": f.name,
                "filter_data": f.filter_data,
                "is_shared": f.is_shared,
                "is_mine": f.user_id == user.id,
                "owner": f.user.username,
            } for f in qs],
        })
    if request.method == "POST":
        try:
            body = json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            body = {}
        name = (body.get("name") or "").strip()[:100]
        if not name:
            return JsonResponse(
                {"success": False, "error": "name vazio"},
                status=400,
            )
        obj, created = SavedSearchFilter.objects.update_or_create(
            user=user, name=name,
            defaults={
                "filter_data": body.get("filter_data") or {},
                "is_shared": bool(body.get("is_shared")),
            },
        )
        return JsonResponse({
            "success": True, "id": obj.id, "created": created,
        })
    fid = request.GET.get("id")
    try:
        n, _ = SavedSearchFilter.objects.filter(
            id=int(fid), user=user,
        ).delete()
        return JsonResponse({"success": True, "deleted": n})
    except (ValueError, TypeError):
        return JsonResponse(
            {"success": False, "error": "id inválido"}, status=400,
        )


# ============================================================================
# Devoluções (Returns) — endpoints
# ============================================================================

def _serialize_return(r):
    return {
        "id": r.id,
        "waybill": r.waybill_number,
        "return_status": r.return_status,
        "return_status_display": r.get_return_status_display(),
        "return_reason": r.return_reason,
        "return_reason_display": r.get_return_reason_display(),
        "return_date": (
            r.return_date.strftime("%Y-%m-%d") if r.return_date
            else ""
        ),
        "return_tracking_number": r.return_tracking_number or "",
        "return_carrier": r.return_carrier or "",
        "return_cost_eur": float(r.return_cost_eur or 0),
        "batch_id": r.batch_id,
        "batch_name": r.batch.name if r.batch else "",
        "customer_notified": r.customer_notified,
        "notification_method": r.notification_method or "",
        "notes": r.notes or "",
        "marked_by": (
            r.marked_by.get_username() if r.marked_by else ""
        ),
        "created_at": r.created_at.strftime("%Y-%m-%d %H:%M"),
        "closed_at": (
            r.closed_at.strftime("%Y-%m-%d %H:%M")
            if r.closed_at else ""
        ),
        "is_open": r.is_open,
    }


@login_required
@require_http_methods(["POST"])
def waybill_return_create(request, waybill):
    """Cria/actualiza uma devolução para um waybill.

    Body JSON: {return_status, return_reason, return_tracking_number,
    return_carrier, return_cost_eur, return_date, notes,
    customer_notified, notification_method}
    """
    import json
    from .models import WaybillReturn
    from django.utils.dateparse import parse_date as _pd
    from django.utils import timezone as _tz

    waybill = (waybill or "").strip()
    if not waybill:
        return JsonResponse(
            {"success": False, "error": "Waybill vazio"}, status=400,
        )

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    rstatus = (body.get("return_status") or "").strip()
    valid_statuses = [s[0] for s in WaybillReturn.STATUS_CHOICES]
    if rstatus not in valid_statuses:
        rstatus = WaybillReturn.STATUS_PREPARING

    rreason = (body.get("return_reason") or "").strip()
    valid_reasons = [r[0] for r in WaybillReturn.REASON_CHOICES]
    if rreason not in valid_reasons:
        rreason = "address_wrong"

    rdate = body.get("return_date")
    parsed_date = _pd(rdate) if rdate else _tz.now().date()

    user = request.user if request.user.is_authenticated else None

    defaults = {
        "return_status": rstatus,
        "return_reason": rreason,
        "return_date": parsed_date,
        "return_tracking_number": (
            body.get("return_tracking_number") or ""
        ).strip()[:100],
        "return_carrier": (
            body.get("return_carrier") or ""
        ).strip()[:100],
        "return_cost_eur": body.get("return_cost_eur") or 0,
        "customer_notified": bool(body.get("customer_notified")),
        "notification_method": (
            body.get("notification_method") or ""
        ).strip()[:30],
        "notes": (body.get("notes") or "").strip(),
        "marked_by": user,
    }
    if rstatus in (
        WaybillReturn.STATUS_RETURNED, WaybillReturn.STATUS_CLOSED,
    ):
        defaults["closed_at"] = _tz.now()

    obj, created = WaybillReturn.objects.update_or_create(
        waybill_number=waybill, defaults=defaults,
    )
    return JsonResponse({
        "success": True, "created": created,
        "return": _serialize_return(obj),
    })


@login_required
def waybill_return_get(request, waybill):
    """Devolve a devolução activa (se existir)."""
    from .models import WaybillReturn
    obj = WaybillReturn.objects.filter(
        waybill_number=waybill,
    ).select_related("marked_by", "batch").first()
    if not obj:
        return JsonResponse({"success": True, "return": None})
    return JsonResponse({
        "success": True, "return": _serialize_return(obj),
    })


@login_required
@require_http_methods(["POST"])
def waybill_return_delete(request, waybill):
    from .models import WaybillReturn
    n, _ = WaybillReturn.objects.filter(
        waybill_number=waybill,
    ).delete()
    return JsonResponse({"success": True, "deleted": n})


@login_required
def waybill_return_list(request):
    """Lista devoluções com filtros + KPIs."""
    from .models import WaybillReturn
    from django.db.models import Q

    qs = WaybillReturn.objects.select_related(
        "marked_by", "batch",
    )

    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(return_status=status)
    reason = (request.GET.get("reason") or "").strip()
    if reason:
        qs = qs.filter(return_reason=reason)
    if request.GET.get("open_only") == "1":
        qs = qs.exclude(return_status__in=[
            WaybillReturn.STATUS_RETURNED,
            WaybillReturn.STATUS_CLOSED,
        ])

    df = (request.GET.get("date_from") or "").strip()
    if df:
        try:
            d = parse_date(df)
            if d:
                qs = qs.filter(return_date__gte=d)
        except (ValueError, TypeError):
            pass
    dt_ = (request.GET.get("date_to") or "").strip()
    if dt_:
        try:
            d = parse_date(dt_)
            if d:
                qs = qs.filter(return_date__lte=d)
        except (ValueError, TypeError):
            pass

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(waybill_number__icontains=q)
            | Q(return_tracking_number__icontains=q)
            | Q(notes__icontains=q)
        )

    qs = qs.order_by("-return_date", "-created_at")
    rows = [_serialize_return(r) for r in qs[:1000]]

    # KPIs
    by_reason = {}
    by_status = {}
    total_cost = 0
    n_open = 0
    for r in rows:
        by_reason[r["return_reason"]] = by_reason.get(
            r["return_reason"], 0,
        ) + 1
        by_status[r["return_status"]] = by_status.get(
            r["return_status"], 0,
        ) + 1
        total_cost += r["return_cost_eur"]
        if r["is_open"]:
            n_open += 1

    return JsonResponse({
        "success": True, "total": len(rows), "rows": rows,
        "kpis": {
            "n_total": len(rows),
            "n_open": n_open,
            "by_reason": by_reason,
            "by_status": by_status,
            "total_cost_eur": round(total_cost, 2),
        },
    })


@login_required
@require_http_methods(["POST"])
def waybill_return_bulk(request):
    """Marca múltiplos waybills como devolvidos.

    Body: { waybills: [str], return_status, return_reason,
            return_carrier, batch_name (opcional, cria/usa batch),
            notes }
    """
    import json
    from .models import WaybillReturn, ReturnBatch

    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    waybills = body.get("waybills") or []
    if not isinstance(waybills, list) or not waybills:
        return JsonResponse(
            {"success": False, "error": "waybills vazios"},
            status=400,
        )

    rstatus = (body.get("return_status") or "").strip()
    valid_statuses = [s[0] for s in WaybillReturn.STATUS_CHOICES]
    if rstatus not in valid_statuses:
        rstatus = WaybillReturn.STATUS_PREPARING

    rreason = (body.get("return_reason") or "").strip()
    valid_reasons = [r[0] for r in WaybillReturn.REASON_CHOICES]
    if rreason not in valid_reasons:
        rreason = "address_wrong"

    user = request.user if request.user.is_authenticated else None

    # Cria batch se nome fornecido
    batch = None
    batch_name = (body.get("batch_name") or "").strip()
    if batch_name:
        batch = ReturnBatch.objects.create(
            name=batch_name, created_by=user,
            notes=(body.get("notes") or ""),
        )

    from django.utils import timezone as _tz
    today = _tz.now().date()
    n_created = 0
    n_updated = 0
    for wb in waybills:
        wb = str(wb).strip()
        if not wb:
            continue
        defaults = {
            "return_status": rstatus,
            "return_reason": rreason,
            "return_date": today,
            "return_carrier": (
                body.get("return_carrier") or ""
            ).strip()[:100],
            "notes": (body.get("notes") or "").strip(),
            "marked_by": user,
            "batch": batch,
        }
        _, created = WaybillReturn.objects.update_or_create(
            waybill_number=wb, defaults=defaults,
        )
        if created:
            n_created += 1
        else:
            n_updated += 1

    return JsonResponse({
        "success": True,
        "n_created": n_created,
        "n_updated": n_updated,
        "total": n_created + n_updated,
        "batch_id": batch.id if batch else None,
        "batch_name": batch.name if batch else "",
    })


@login_required
def return_batch_list(request):
    """Lista lotes de devolução."""
    from .models import ReturnBatch
    qs = ReturnBatch.objects.select_related(
        "created_by",
    ).order_by("-created_at")[:200]
    return JsonResponse({
        "success": True,
        "batches": [{
            "id": b.id,
            "name": b.name,
            "created_at": b.created_at.strftime("%Y-%m-%d %H:%M"),
            "created_by": (
                b.created_by.get_username()
                if b.created_by else ""
            ),
            "n_returns": b.returns.count(),
            "total_cost_eur": float(b.total_cost_eur or 0),
            "closed_at": (
                b.closed_at.strftime("%Y-%m-%d %H:%M")
                if b.closed_at else ""
            ),
            "notes": b.notes or "",
        } for b in qs],
    })


@login_required
def return_batch_xlsx(request, batch_id):
    """Exporta um lote de devolução para Excel (lista p/ imprimir)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from .models import ReturnBatch

    batch = get_object_or_404(ReturnBatch, id=batch_id)
    returns = list(batch.returns.all().order_by("waybill_number"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Lote #{batch.id}"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="92400E")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws["A1"] = (
        f"Lote de Devolução #{batch.id} — {batch.name}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="92400E")
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        f"{len(returns)} pacote(s) · "
        f"€{batch.total_cost_eur:.2f} · "
        f"Criado em {batch.created_at:%Y-%m-%d %H:%M}"
    )
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:H2")

    headers = [
        "Waybill", "Razão", "Status", "Tracking retorno",
        "Carrier", "Custo €", "Notas", "Marcado por",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for ridx, r in enumerate(returns, start=5):
        ws.cell(row=ridx, column=1, value=r.waybill_number)
        ws.cell(row=ridx, column=2, value=r.get_return_reason_display())
        ws.cell(row=ridx, column=3, value=r.get_return_status_display())
        ws.cell(row=ridx, column=4, value=r.return_tracking_number)
        ws.cell(row=ridx, column=5, value=r.return_carrier)
        ws.cell(row=ridx, column=6, value=float(r.return_cost_eur or 0))
        ws.cell(row=ridx, column=7, value=r.notes)
        ws.cell(
            row=ridx, column=8,
            value=r.marked_by.get_username() if r.marked_by else "",
        )
        for ci in range(1, 9):
            ws.cell(row=ridx, column=ci).border = border

    widths = [28, 22, 22, 20, 16, 10, 30, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    fname = f"lote_devolucoes_{batch.id}_{batch.created_at:%Y%m%d}.xlsx"
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{fname}"'
    )
    wb.save(response)
    return response


# ============================================================================
# PARCEL_LIST — formato Cainiao mais rico (105 colunas)
# ============================================================================

def _parse_parcel_list_file(file_bytes):
    """Lê _PARCEL_LIST xlsx, devolve (data_rows, idx, header)."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl não instalado.")
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True,
        )
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Erro ao abrir ficheiro: {exc}")

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("Ficheiro vazio.")

    header = None
    header_row_idx = 0
    for i, row in enumerate(all_rows[:15]):
        cells = [
            str(c).strip().lower() if c is not None else ""
            for c in row
        ]
        if any("tracking no" in c for c in cells) and any(
            "site code" in c for c in cells
        ):
            header = [
                str(c).strip() if c is not None else "" for c in row
            ]
            header_row_idx = i
            break

    if header is None:
        debug = [
            [str(c) for c in r if c is not None]
            for r in all_rows[:3]
        ]
        raise ValueError(
            "Cabeçalhos esperados não encontrados (Tracking No. + "
            f"Site code). Primeiras linhas: {debug}"
        )

    def col(*names):
        for name in names:
            n_low = name.lower()
            for i, h in enumerate(header):
                if h and n_low == h.strip().lower():
                    return i
        for name in names:
            n_low = name.lower()
            for i, h in enumerate(header):
                if h and n_low in h.lower():
                    return i
        return None

    idx = {
        "tracking":     col("Tracking No.", "Tracking No"),
        "lp":           col("LP No.", "LP No"),
        "site_code":    col("Site code"),
        "status":       col("Status"),
        "bigbag_id":    col("Inbound Bigbag ID"),
        "bigbag_no":    col("Bigbag No.", "Bigbag No"),
        "sort_code":    col("Sort Code"),
        "order_type":   col("Order Type"),
        "sop_type":     col("SOP Type"),
        "recv_name":    col("Receiver's Name", "Receiver Name"),
        "recv_contact": col(
            "Receiver's Contact Number", "Contact Number",
        ),
        "recv_phone":   col(
            "Receiver's Phone Number", "Phone Number",
        ),
        "recv_email":   col("Receiver’s Email", "Receiver's Email"),
        "addr_abnormal": col("Address Abnormal"),
        "recv_country": col("Receiver Country"),
        "recv_region":  col(
            "Receiver's Region/Province", "Region/Province",
        ),
        "recv_city":    col("Receiver City"),
        "recv_zip":     col("Receiver's Zip Code", "Zip Code"),
        "recv_addr":    col(
            "Receiver's Detail Address", "Detail Address",
        ),
        "mod_city":     col("Modified City"),
        "mod_zip":      col("Modified Zip Code"),
        "mod_addr":     col("Modified Detail Address"),
        "mod_phone":    col("Modified Phone Number"),
        "remarks":      col("Remarks"),
        "coords":       col(
            "Receiver to (Latitude,Longitude)",
            "Latitude,Longitude",
        ),
        "exc_type":     col("Exception Type"),
        "exc_detail":   col("Exception Detail"),
        "exc_time":     col("Last Exception Time"),
        "exc_by":       col("Last Exception Registed By"),
        "create_time":  col("Create Time"),
        "act_inbound":  col("Actual Inbound Time"),
        "act_outbound": col("Actual Outbound Time"),
        "act_delivery": col("Actual Delivery Time"),
        "task_id":      col("Task ID"),
        "pod":          col("POD"),
        "is_consol":    col("isConsolidation?", "isConsolidation"),
        "weight":       col("Weight"),
        "length":       col("Length"),
        "width":        col("Width"),
        "height":       col("Height"),
        "dsp_name":     col("DSP Name"),
        "locker_id":    col("Locker ID"),
        "delivery_mode": col("delivery Mode", "Delivery Mode"),
        "seller_name":  col("Seller Name"),
        "pin_code":     col("PINCode", "PIN Code"),
        "has_pin_code": col("HasPINCode", "Has PIN Code"),
        "zone":         col("Zone"),
        "task_plan":    col("Last Task Plan Date"),
        "wrong_hub":    col("Arriving at the wrong hub"),
        "wrong_hub_name": col("Wrong hub name"),
        "comm_area":    col(
            "hasCommercialAreaTag", "Commercial Area",
        ),
    }

    data_rows = all_rows[header_row_idx + 1:]
    return data_rows, idx, header


def _extract_date_from_filename(fn):
    """Extrai YYYYMMDD do nome do ficheiro Cainiao. Retorna ISO date
    string (YYYY-MM-DD) ou None se não encontrar.

    Nomes típicos: "_PARCEL_LIST 20260430052115.xlsx" → "2026-04-30"
    """
    import re
    if not fn:
        return None
    m = re.search(r"(\d{8})", fn)
    if not m:
        return None
    s = m.group(1)
    try:
        y, mo, d = int(s[:4]), int(s[4:6]), int(s[6:8])
        if not (2020 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31):
            return None
        return f"{y:04d}-{mo:02d}-{d:02d}"
    except (ValueError, TypeError):
        return None


def cainiao_parcel_list_preview(request):
    """Preview do PARCEL_LIST sem importar."""
    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."},
            status=400,
        )
    try:
        data_rows, idx, header = _parse_parcel_list_file(
            ficheiro.read(),
        )
    except ValueError as e:
        return JsonResponse(
            {"success": False, "error": str(e)}, status=400,
        )

    status_counts = {}
    n_rows = 0
    n_no_tracking = 0
    n_wrong_hub = 0
    n_with_modified_addr = 0
    n_with_exception = 0
    sample = []
    for row in data_rows:
        tracking = _cell(row, idx, "tracking")
        if not tracking:
            n_no_tracking += 1
            continue
        n_rows += 1
        st = _cell(row, idx, "status") or "(vazio)"
        status_counts[st] = status_counts.get(st, 0) + 1
        if _cell(row, idx, "wrong_hub"):
            n_wrong_hub += 1
        if _cell(row, idx, "mod_addr") or _cell(row, idx, "mod_zip"):
            n_with_modified_addr += 1
        if _cell(row, idx, "exc_type"):
            n_with_exception += 1
        if len(sample) < 10:
            sample.append({
                "tracking": tracking,
                "lp": _cell(row, idx, "lp"),
                "status": st,
                "city": _cell(row, idx, "recv_city"),
                "zip": _cell(row, idx, "recv_zip"),
                "seller": _cell(row, idx, "seller_name"),
                "dsp": _cell(row, idx, "dsp_name"),
                "wrong_hub": _cell(row, idx, "wrong_hub"),
                "exception": _cell(row, idx, "exc_type"),
                "action": "—",
            })

    # Sugestão de data baseada no nome do ficheiro (YYYYMMDD)
    suggested_date = _extract_date_from_filename(ficheiro.name)

    return JsonResponse({
        "success": True,
        "filename": ficheiro.name,
        "suggested_date": suggested_date,  # ex: "2026-04-30" ou None
        "n_columns": len(header),
        "total": n_rows,
        "skipped_no_tracking": n_no_tracking,
        "n_wrong_hub": n_wrong_hub,
        "n_with_modified_addr": n_with_modified_addr,
        "n_with_exception": n_with_exception,
        "status_counts": status_counts,
        "sample": sample,
    })


@login_required
@require_http_methods(["POST"])
def cainiao_parcel_list_import(request):
    """Importa PARCEL_LIST para CainiaoForecastPackage."""
    from datetime import date as _d, datetime as _dt2
    from decimal import Decimal, InvalidOperation
    from .models import CainiaoForecastBatch, CainiaoForecastPackage

    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."},
            status=400,
        )
    operation_date = (
        request.POST.get("operation_date") or ""
    ).strip()
    if not operation_date:
        return JsonResponse(
            {"success": False,
             "error": "Data da operação é obrigatória."}, status=400,
        )
    try:
        op_date = _d.fromisoformat(operation_date)
    except ValueError:
        return JsonResponse(
            {"success": False,
             "error": "Data da operação inválida."}, status=400,
        )

    # ─── Salvaguardas anti-duplicação ─────────────────────────────────
    # Permite bypass via flag force=true (após user confirmar no UI).
    force = request.POST.get("force", "").lower() in ("1", "true", "yes")

    # 1. Filename traz YYYYMMDD que não bate com operation_date
    suggested = _extract_date_from_filename(ficheiro.name)
    if suggested and not force and suggested != operation_date:
        return JsonResponse({
            "success": False,
            "warning_type": "filename_date_mismatch",
            "error": (
                f"O nome do ficheiro indica data {suggested} mas estás "
                f"a importar como {operation_date}. Confirma que está "
                "correcto antes de prosseguir."
            ),
            "suggested_date": suggested,
            "operation_date": operation_date,
        }, status=409)

    # 2. Já existe batch com MESMO filename mas outra operation_date
    if not force:
        existing_other = CainiaoForecastBatch.objects.filter(
            filename=ficheiro.name,
        ).exclude(operation_date=op_date).order_by("-created_at").first()
        if existing_other:
            return JsonResponse({
                "success": False,
                "warning_type": "filename_already_imported",
                "error": (
                    f"Este ficheiro '{ficheiro.name}' já foi importado "
                    f"em {existing_other.operation_date} (batch "
                    f"#{existing_other.id}). Importá-lo agora como "
                    f"{operation_date} vai criar pacotes duplicados em "
                    "datas diferentes. Confirma que é mesmo este o "
                    "ficheiro correcto."
                ),
                "existing_batch_id": existing_other.id,
                "existing_operation_date": str(existing_other.operation_date),
                "operation_date": operation_date,
            }, status=409)

    try:
        data_rows, idx, header = _parse_parcel_list_file(
            ficheiro.read(),
        )
    except ValueError as e:
        return JsonResponse(
            {"success": False, "error": str(e)}, status=400,
        )

    batch = CainiaoForecastBatch.objects.create(
        filename=ficheiro.name,
        import_type=CainiaoForecastBatch.TYPE_FORECAST,
        operation_date=op_date,
        total_packages=0,
        updated_packages=0,
        created_by=request.user,
    )

    def _yesno_bool(v):
        if not v:
            return False
        return str(v).strip().lower() in (
            "yes", "y", "true", "1", "sim", "s",
        )

    def _parse_coords(raw):
        if not raw:
            return None, None
        parts = str(raw).split(",")
        if len(parts) == 2:
            try:
                return (
                    Decimal(parts[0].strip()),
                    Decimal(parts[1].strip()),
                )
            except InvalidOperation:
                pass
        return None, None

    def _parse_dt(v):
        from django.utils.timezone import make_aware, is_aware
        if v is None or v == "":
            return None
        if hasattr(v, "year") and hasattr(v, "hour"):
            return v if is_aware(v) else make_aware(v)
        s = str(v).strip()
        if not s:
            return None
        for fmt in (
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
            "%Y-%m-%d", "%Y/%m/%d %H:%M:%S",
        ):
            try:
                naive = _dt2.strptime(s, fmt)
                return make_aware(naive)
            except ValueError:
                pass
        return None

    def _parse_d(v):
        dt = _parse_dt(v)
        if dt is None:
            return None
        try:
            return dt.date()
        except AttributeError:
            return dt

    def _parse_float(v):
        if v is None or v == "":
            return None
        try:
            return float(str(v).replace(",", "."))
        except (ValueError, TypeError):
            return None

    def _row_at(row, key):
        i = idx.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    existing = {
        p.tracking_number: p
        for p in CainiaoForecastPackage.objects.filter(
            operation_date=op_date,
        )
    }

    update_fields = [
        "status", "last_import_batch",
        "lp_number", "site_code", "inbound_bigbag_id",
        "bigbag_number", "sort_code", "order_type", "sop_type",
        "receiver_name", "receiver_phone", "receiver_region",
        "receiver_city", "receiver_zip", "receiver_address",
        "latitude", "longitude",
        "receiver_contact_number", "receiver_email",
        "address_abnormal", "modified_zip_code", "modified_address",
        "modified_phone", "exception_type", "exception_detail",
        "last_exception_time", "create_time", "actual_inbound_time",
        "actual_outbound_time", "actual_delivery_time", "weight_g",
        "dimensions_lwh", "dsp_name", "seller_name",
        "last_task_plan_date", "arriving_at_wrong_hub",
        "wrong_hub_name", "has_commercial_area_tag",
        "zone", "task_id", "is_consolidation",
        "pin_code", "has_pin_code", "locker_id",
        "delivery_mode", "pod_url",
    ]

    to_create = []
    to_update = []
    status_counts = {}

    for row in data_rows:
        tracking = _cell(row, idx, "tracking")
        if not tracking:
            continue
        status_val = (
            _cell(row, idx, "status")
            or CainiaoForecastPackage.STATUS_CREATE
        )
        status_counts[status_val] = status_counts.get(
            status_val, 0,
        ) + 1
        lat, lng = _parse_coords(_cell(row, idx, "coords"))
        dims_parts = [
            _cell(row, idx, "length") or "—",
            _cell(row, idx, "width") or "—",
            _cell(row, idx, "height") or "—",
        ]
        dims = (
            "×".join(dims_parts) if any(p != "—" for p in dims_parts)
            else ""
        )

        kwargs = {
            "status": status_val,
            "last_import_batch": batch,
            "lp_number": _cell(row, idx, "lp"),
            "site_code": _cell(row, idx, "site_code"),
            "inbound_bigbag_id": _cell(row, idx, "bigbag_id"),
            "bigbag_number": _cell(row, idx, "bigbag_no"),
            "sort_code": _cell(row, idx, "sort_code"),
            "order_type": _cell(row, idx, "order_type"),
            "sop_type": _cell(row, idx, "sop_type"),
            "receiver_name": _cell(row, idx, "recv_name"),
            "receiver_phone": _cell(row, idx, "recv_phone"),
            "receiver_region": _cell(row, idx, "recv_region"),
            "receiver_city": _cell(row, idx, "recv_city"),
            "receiver_zip": _cell(row, idx, "recv_zip"),
            "receiver_address": _cell(row, idx, "recv_addr"),
            "receiver_contact_number": _cell(row, idx, "recv_contact"),
            "receiver_email": _cell(row, idx, "recv_email"),
            "address_abnormal": _cell(row, idx, "addr_abnormal"),
            "modified_zip_code": _cell(row, idx, "mod_zip"),
            "modified_address": _cell(row, idx, "mod_addr"),
            "modified_phone": _cell(row, idx, "mod_phone"),
            "exception_type": _cell(row, idx, "exc_type"),
            "exception_detail": _cell(row, idx, "exc_detail"),
            "last_exception_time": _parse_dt(
                _row_at(row, "exc_time"),
            ),
            "create_time": _parse_dt(_row_at(row, "create_time")),
            "actual_inbound_time": _parse_dt(
                _row_at(row, "act_inbound"),
            ),
            "actual_outbound_time": _parse_dt(
                _row_at(row, "act_outbound"),
            ),
            "actual_delivery_time": _parse_dt(
                _row_at(row, "act_delivery"),
            ),
            "weight_g": _parse_float(_cell(row, idx, "weight")),
            "dimensions_lwh": dims[:50],
            "dsp_name": _cell(row, idx, "dsp_name"),
            "seller_name": _cell(row, idx, "seller_name"),
            "last_task_plan_date": _parse_d(
                _row_at(row, "task_plan"),
            ),
            "arriving_at_wrong_hub": _yesno_bool(
                _cell(row, idx, "wrong_hub"),
            ),
            "wrong_hub_name": _cell(row, idx, "wrong_hub_name"),
            "has_commercial_area_tag": _yesno_bool(
                _cell(row, idx, "comm_area"),
            ),
            "zone": _cell(row, idx, "zone"),
            "task_id": _cell(row, idx, "task_id"),
            "is_consolidation": _yesno_bool(
                _cell(row, idx, "is_consol"),
            ),
            "pin_code": _cell(row, idx, "pin_code"),
            "has_pin_code": _yesno_bool(
                _cell(row, idx, "has_pin_code"),
            ),
            "locker_id": _cell(row, idx, "locker_id"),
            "delivery_mode": _cell(row, idx, "delivery_mode"),
            "pod_url": _cell(row, idx, "pod"),
        }
        if lat is not None:
            kwargs["latitude"] = lat
            kwargs["longitude"] = lng

        if tracking in existing:
            pkg = existing[tracking]
            for k, v in kwargs.items():
                setattr(pkg, k, v)
            to_update.append(pkg)
        else:
            to_create.append(CainiaoForecastPackage(
                operation_date=op_date,
                tracking_number=tracking,
                **kwargs,
            ))

    if to_create:
        CainiaoForecastPackage.objects.bulk_create(
            to_create, batch_size=500,
        )
    if to_update:
        CainiaoForecastPackage.objects.bulk_update(
            to_update, fields=update_fields, batch_size=500,
        )

    batch.total_packages = len(to_create) + len(to_update)
    batch.updated_packages = len(to_update)
    batch.save(
        update_fields=["total_packages", "updated_packages"],
    )

    # ── Espelhar em CainiaoPlanningPackage para o dashboard de Forecast ──
    # PARCEL_LIST é a previsão de volume do dia seguinte; o dashboard
    # de "Previsão de Volume" lê de CainiaoPlanningPackage.
    from .models import CainiaoPlanningBatch, CainiaoPlanningPackage
    planning_batch = CainiaoPlanningBatch.objects.create(
        filename=ficheiro.name,
        operation_date=op_date,
        created_by=request.user,
    )
    existing_plan = {
        p.parcel_id: p
        for p in CainiaoPlanningPackage.objects.filter(
            operation_date=op_date,
        )
    }
    plan_update_fields = [
        "lp_code", "receiver_province", "receiver_city",
        "receiver_zip", "receiver_address", "receiver_name",
        "receiver_phone", "receiver_email",
        "actual_inbound_time", "sc_outbound_time", "sign_time",
        "exception_type", "exception_reason", "last_exception_time",
        "hub", "dsp", "creation_time", "seller_name",
        "inbound_time", "delivery_success_time", "last_import_batch",
    ]
    plan_to_create = []
    plan_to_update = []
    plan_seen = set()

    for row in data_rows:
        tracking = _cell(row, idx, "tracking")
        if not tracking or tracking in plan_seen:
            continue
        plan_seen.add(tracking)

        # Endereço/telefone preferem versão modificada (Cainiao corrige)
        addr = (
            _cell(row, idx, "mod_addr")
            or _cell(row, idx, "recv_addr")
        )
        zip_code = (
            _cell(row, idx, "mod_zip")
            or _cell(row, idx, "recv_zip")
        )
        phone = (
            _cell(row, idx, "mod_phone")
            or _cell(row, idx, "recv_phone")
        )
        create_dt = _parse_dt(_row_at(row, "create_time"))
        inbound_dt = _parse_dt(_row_at(row, "act_inbound"))
        outbound_dt = _parse_dt(_row_at(row, "act_outbound"))
        delivery_dt = _parse_dt(_row_at(row, "act_delivery"))
        status_val_p = _cell(row, idx, "status") or ""
        is_delivered = "deliver" in str(status_val_p).lower()

        plan_fields = {
            "lp_code": _cell(row, idx, "lp"),
            "receiver_province": _cell(row, idx, "recv_region"),
            "receiver_city": _cell(row, idx, "recv_city"),
            "receiver_zip": zip_code,
            "receiver_address": addr,
            "receiver_name": _cell(row, idx, "recv_name"),
            "receiver_phone": phone,
            "receiver_email": _cell(row, idx, "recv_email"),
            "actual_inbound_time": inbound_dt,
            "sc_outbound_time": outbound_dt,
            "sign_time": delivery_dt if is_delivered else None,
            "exception_type": _cell(row, idx, "exc_type"),
            "exception_reason": _cell(row, idx, "exc_detail"),
            "last_exception_time": _parse_dt(
                _row_at(row, "exc_time"),
            ),
            "hub": _cell(row, idx, "site_code"),
            "dsp": _cell(row, idx, "dsp_name"),
            "creation_time": create_dt,
            "seller_name": _cell(row, idx, "seller_name"),
            "inbound_time": inbound_dt,
            "delivery_success_time": (
                delivery_dt if is_delivered else None
            ),
            "last_import_batch": planning_batch,
        }

        if tracking in existing_plan:
            pkg = existing_plan[tracking]
            for k, v in plan_fields.items():
                setattr(pkg, k, v)
            plan_to_update.append(pkg)
        else:
            plan_to_create.append(CainiaoPlanningPackage(
                operation_date=op_date,
                parcel_id=tracking,
                **plan_fields,
            ))

    if plan_to_create:
        CainiaoPlanningPackage.objects.bulk_create(
            plan_to_create, batch_size=500,
        )
    if plan_to_update:
        CainiaoPlanningPackage.objects.bulk_update(
            plan_to_update, plan_update_fields, batch_size=500,
        )

    planning_batch.total_packages = (
        len(plan_to_create) + len(plan_to_update)
    )
    planning_batch.new_packages = len(plan_to_create)
    planning_batch.updated_packages = len(plan_to_update)
    planning_batch.save(update_fields=[
        "total_packages", "new_packages", "updated_packages",
    ])

    return JsonResponse({
        "success": True,
        "total": batch.total_packages,
        "total_novos": len(to_create),
        "total_atualizados": len(to_update),
        "status_counts": status_counts,
        "operation_date": str(op_date),
        "filename": ficheiro.name,
        "batch_id": batch.id,
        "planning_batch_id": planning_batch.id,
        "planning_novos": len(plan_to_create),
        "planning_atualizados": len(plan_to_update),
        "format": "PARCEL_LIST",
        "n_columns": len(header),
    })


# ============================================================================
# Auto-detecção do tipo de planilha
# ============================================================================

@login_required
@require_http_methods(["POST"])
def cainiao_detect_xlsx_type(request):
    """Lê os primeiros bytes do xlsx e tenta identificar o tipo
    (Forecast / Planning / Operation Update / Driver Stat /
    Driver Detail / Parcel List)."""
    import openpyxl
    from io import BytesIO

    ficheiro = request.FILES.get("ficheiro")
    if not ficheiro:
        return JsonResponse(
            {"success": False, "error": "Nenhum ficheiro enviado."},
            status=400,
        )

    try:
        # NOTE: read_only=True trunca linhas com dimensões mal-definidas
        # em alguns ficheiros Cainiao (PARCEL_LIST). Lemos só 15 linhas.
        wb = openpyxl.load_workbook(
            BytesIO(ficheiro.read()), data_only=True,
        )
        ws = wb.active
        first_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 15:
                break
            first_rows.append(row)
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"XLSX inválido: {e}"},
            status=400,
        )

    all_text = " ".join(
        str(c).strip().lower()
        for r in first_rows
        for c in r if c is not None
    )

    has_tracking_no = "tracking no" in all_text
    has_waybill = "waybill number" in all_text
    has_parcel_id = "parcel id" in all_text
    has_inbound_bigbag = "inbound bigbag id" in all_text
    has_site_code = "site code" in all_text
    has_sop_type = "sop type" in all_text
    has_courier_name = "courier name" in all_text
    has_task_status = "task status" in all_text
    has_modified_addr = "modified detail address" in all_text
    has_pre_assigned = "pre-assigned" in all_text
    has_courier_id_perf = (
        "courier id" in all_text
        and "performance" not in all_text
    )
    has_driver_stat_marker = (
        "delivered (today)" in all_text
        or "attempt failure (today)" in all_text
        or "average delivery time" in all_text
    )

    detected = None
    confidence = "low"

    # PARCEL_LIST: tracking + site_code + inbound_bigbag + modified_addr
    if (
        has_tracking_no and has_site_code
        and has_inbound_bigbag and has_modified_addr
    ):
        detected = "parcel-list"
        confidence = "high"
    elif has_waybill and has_task_status and has_courier_name:
        detected = "operation"
        confidence = "high"
    elif has_parcel_id and "receiver name" in all_text:
        detected = "planning"
        confidence = "high"
    elif has_tracking_no and has_inbound_bigbag and has_sop_type:
        detected = "forecast"
        confidence = "high"
    elif has_driver_stat_marker:
        detected = "driver-stat"
        confidence = "high"
    elif has_courier_id_perf and has_pre_assigned:
        detected = "driver-detail"
        confidence = "medium"
    elif has_tracking_no:
        detected = "forecast"
        confidence = "low"
    elif has_waybill:
        detected = "operation"
        confidence = "low"

    headers_sample = []
    for row in first_rows:
        non_empty = sum(
            1 for c in row if c is not None and str(c).strip()
        )
        if non_empty >= 3:
            headers_sample = [
                str(c).strip() if c is not None else ""
                for c in row[:30]
            ]
            break

    return JsonResponse({
        "success": True,
        "filename": ficheiro.name,
        "detected": detected,
        "confidence": confidence,
        "headers_sample": [h for h in headers_sample if h][:20],
        "type_label": {
            "parcel-list": "Parcel List (formato rico, 105 cols)",
            "operation": "Operation Update (EPOD Task List)",
            "planning": "Planning",
            "forecast": "Forecast (simplificado)",
            "driver-stat": "Driver Statistic",
            "driver-detail": "Driver Detail Info",
        }.get(detected, "Desconhecido"),
    })

