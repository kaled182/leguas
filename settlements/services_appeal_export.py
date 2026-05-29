"""Exportação de recursos para o parceiro (Cainiao).

- build_appeals_xlsx: tabela com as colunas oficiais da Cainiao (em espanhol).
- helpers: sugestão de motivo (1 dos 12), info da fatura de dedução (data/mês
  e DSP) e nomes dos justificantes.
- batch_filename: "<DSP> <mês_es>" (mês = fatura de dedução), para nomear o
  e-mail e o anexo conforme as regras 3.1/3.2.
"""
from __future__ import annotations

import io
import os
from collections import Counter

MONTHS_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre",
    12: "diciembre",
}

# Colunas oficiais da Cainiao (espanhol — não traduzir)
XLSX_HEADERS = [
    "Número de LP",
    "Número de seguimiento",
    "Fecha de la factura",
    "Importe de la deducción",
    "Confirmación la penalización - Y/N",
    "Motivos de la reclamación",
    "Observacion de DSP",
    "Hay justificante - Y/N",
    "Nombre de Justificante",
]


def suggest_appeal_reason(claim):
    """Sugere um dos 12 motivos oficiais a partir do tipo do claim/reclamação."""
    R_FALTA = "El cliente reclama faltan artículos/productos incorrectos"
    R_ROBO = "Casos de robo/penalización/pérdida/paquete dañado"
    R_RECIBIDO = "Cliente ha recibido"
    R_URGENTE = (
        "El cliente se llama por entrega urgente/cambio de dirección/error de "
        "dirección que provoca un error de entrega/el cliente no está en casa, etc."
    )

    cc = getattr(claim, "customer_complaint", None)
    tipo = (cc.tipo if cc else "") or ""
    ct = claim.claim_type or ""

    by_tipo = {
        "ITEM_FALTANDO": R_FALTA,
        "PACOTE_DANIFICADO": R_ROBO,
        "ENTREGA_FALSA": R_RECIBIDO,
        "ENTREGA_ATRASADA": R_URGENTE,
    }
    if tipo in by_tipo:
        return by_tipo[tipo]

    by_claim_type = {
        "FAKE_DELIVERY": R_RECIBIDO,
        "ORDER_LOSS": R_ROBO,
        "ORDER_DAMAGE": R_ROBO,
        "CUSTOMER_COMPLAINT": R_RECIBIDO,
        "LATE_DELIVERY": R_URGENTE,
    }
    return by_claim_type.get(ct, "")


def claim_proof_names(claim):
    """Nomes dos ficheiros de prova do recurso (evidência + anexos não-notificação)."""
    names = []
    if getattr(claim, "evidence_file", None):
        try:
            names.append(os.path.basename(claim.evidence_file.name))
        except Exception:
            pass
    cc = getattr(claim, "customer_complaint", None)
    if cc:
        for att in cc.attachments.exclude(tipo="RECLAMACAO"):
            try:
                names.append(os.path.basename(att.ficheiro.name))
            except Exception:
                continue
    return names


def claim_billing_info(claim):
    """Data/mês da fatura de dedução, DSP e LP para um claim Cainiao."""
    from .models import CainiaoBillingLine, CainiaoOperationTask

    invoice_date = None
    dsp = ""
    lp = ""

    line = (
        CainiaoBillingLine.objects
        .filter(claim=claim)
        .select_related("import_session", "import_session__partner_invoice", "task")
        .first()
    )
    if line:
        sess = line.import_session
        pi = getattr(sess, "partner_invoice", None) if sess else None
        if pi and pi.issue_date:
            invoice_date = pi.issue_date
        elif sess and getattr(sess, "period_to", None):
            invoice_date = sess.period_to
        if line.task:
            dsp = line.task.dsp_name or ""
            lp = line.task.lp_number or ""

    # Fallbacks
    if not invoice_date:
        invoice_date = claim.operation_task_date or (
            claim.occurred_at.date() if claim.occurred_at else None
        )
    if (not dsp or not lp) and claim.waybill_number:
        t = (
            CainiaoOperationTask.objects
            .filter(waybill_number=claim.waybill_number)
            .order_by("-task_date")
            .first()
        )
        if t:
            dsp = dsp or (t.dsp_name or "")
            lp = lp or (t.lp_number or "")
    if not dsp:
        try:
            from system_config.models import SystemConfiguration
            dsp = (SystemConfiguration.get_config().company_name or "").strip()
        except Exception:
            dsp = ""
        dsp = dsp or "LÉGUAS FRANZINAS - UNIPESSOAL LDA"

    return {"invoice_date": invoice_date, "dsp": dsp, "lp": lp}


def _safe_filename(s):
    keep = "-_. "
    return "".join(c for c in (s or "") if c.isalnum() or c in keep).strip() or "recursos"


def batch_filename(claims):
    """Nome base do ficheiro: '<DSP> <mês_es>' (mês da fatura de dedução)."""
    infos = [claim_billing_info(c) for c in claims]
    dsp = next((i["dsp"] for i in infos if i["dsp"]), "DSP")
    months = [
        (i["invoice_date"].year, i["invoice_date"].month)
        for i in infos if i["invoice_date"]
    ]
    if months:
        (_, m), _ = Counter(months).most_common(1)[0]
        mes = MONTHS_ES.get(m, "")
    else:
        mes = ""
    base = f"{dsp} {mes}".strip()
    return _safe_filename(base)


def build_appeals_xlsx(claims):
    """Gera o xlsx com as colunas oficiais da Cainiao. Devolve (bytes, filename)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reclamaciones"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(XLSX_HEADERS)
    for col, _ in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for claim in claims:
        info = claim_billing_info(claim)
        proofs = claim_proof_names(claim)
        ws.append([
            info["lp"] or "",
            claim.waybill_number or "",
            info["invoice_date"].strftime("%d/%m/%Y") if info["invoice_date"] else "",
            float(claim.amount or 0),
            claim.penalty_confirmed or "N",
            claim.appeal_reason or suggest_appeal_reason(claim),
            claim.dsp_observation or claim.justification or "",
            "Y" if proofs else "N",
            ", ".join(proofs),
        ])

    widths = [22, 28, 14, 14, 16, 50, 32, 10, 32]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), batch_filename(claims)
