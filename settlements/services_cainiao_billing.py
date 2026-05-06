"""Importação idempotente da pré-fatura Cainiao (XLSX).

A Cainiao envia mensalmente (ou quinzenalmente) um ficheiro `.xlsx`
com 2 folhas:

  Folha 1 (`合并账单_主单_xxx`): linhas de facturação
    Cabeçalho: Fee type, Biz time, imp de fac, moneda, REFs, CP name,
               Ciudad, staff id, billing id, FB1, FB2

  Folha 2 (`Sheet1`): pivot table de sumário (ignorada — redundante)

Idempotência:
  - file_hash (SHA256) bloqueia reimport do mesmo ficheiro físico.
  - Para conteúdo: unique_together(waybill, fee_type, billing_id) em
    CainiaoBillingLine garante que reimports parciais (quinzenas
    sobrepostas) não criam duplicados — usamos update_or_create.

Resolução de FKs no save:
  - waybill_number → CainiaoOperationTask (qualquer task com esse waybill)
  - staff_id → DriverProfile via courier_id_cainiao OU
               DriverCourierMapping.courier_id
  - Se staff_id vazio (claims) → driver via task.courier_id_cainiao

PartnerInvoice:
  - Criada/atualizada automaticamente no fim com totais agregados,
    associada ao Partner "CAINIAO".
"""
from __future__ import annotations

import hashlib
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Iterator, Optional

from django.db import transaction
from django.utils import timezone

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────
#  Constantes
# ─────────────────────────────────────────────────────────────────────

# Cabeçalhos da Folha 1 — usados para localizar a folha correcta
EXPECTED_HEADERS = (
    "Fee type", "Biz time", "imp de fac", "moneda", "REFs",
    "CP name", "Ciudad", "staff id", "billing id", "FB1", "FB2",
)

# Mapeamento header → chave do dict de saída (lowercase para comparar)
HEADER_TO_KEY = {
    "fee type": "fee_type",
    "biz time": "biz_time",
    "imp de fac": "amount",
    "moneda": "moneda",
    "refs": "waybill_number",
    "cp name": "cp_name",
    "ciudad": "ciudad",
    "staff id": "staff_id",
    "billing id": "cainiao_billing_id",
    "fb1": "fb1",
    "fb2": "fb2",
}


# ─────────────────────────────────────────────────────────────────────
#  Hashing + parsing
# ─────────────────────────────────────────────────────────────────────

def compute_file_hash(file_obj) -> str:
    """SHA256 do ficheiro. Posiciona o cursor de volta a 0 após ler."""
    h = hashlib.sha256()
    file_obj.seek(0)
    for chunk in iter(lambda: file_obj.read(8192), b""):
        h.update(chunk)
    file_obj.seek(0)
    return h.hexdigest()


def _detect_billing_sheet(workbook):
    """Devolve a folha que contém os cabeçalhos esperados (Folha 1)."""
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row < 2 or ws.max_column < 5:
            continue
        # Lê a primeira linha não-vazia
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            row_lower = [
                str(c).strip().lower() if c else "" for c in row
            ]
            if "fee type" in row_lower and "refs" in row_lower:
                return ws, row_lower
    raise ValueError(
        "Folha de pré-fatura Cainiao não encontrada — esperava "
        "cabeçalhos 'Fee type' e 'REFs' na linha 1-3."
    )


def parse_xlsx_to_rows(file_obj) -> Iterator[dict]:
    """Itera linhas (excl. cabeçalho) como dicts com as chaves do modelo.

    Ignora linhas vazias e normaliza tipos:
      - amount: Decimal (suporta vírgula PT)
      - biz_time: datetime
      - staff_id, cainiao_billing_id: str (mantém zeros à esquerda)
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws, header_row = _detect_billing_sheet(wb)

    # Mapear cada coluna do header para a chave do modelo
    col_map = {}
    for idx, header in enumerate(header_row):
        key = HEADER_TO_KEY.get(header.lower().strip())
        if key:
            col_map[idx] = key

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        rec = {}
        for idx, val in enumerate(row):
            key = col_map.get(idx)
            if key is None:
                continue
            rec[key] = val

        # Saltar linhas que não têm fee_type nem waybill (lixo)
        if not rec.get("fee_type") and not rec.get("waybill_number"):
            continue

        # Normalizar tipos
        rec["fee_type"] = (rec.get("fee_type") or "").strip()
        rec["waybill_number"] = (rec.get("waybill_number") or "").strip()
        rec["staff_id"] = _to_str(rec.get("staff_id"))
        rec["cainiao_billing_id"] = _to_str(rec.get("cainiao_billing_id"))
        rec["cp_name"] = (rec.get("cp_name") or "").strip()
        rec["ciudad"] = (rec.get("ciudad") or "").strip()
        rec["fb1"] = _to_text(rec.get("fb1"))
        rec["fb2"] = _to_text(rec.get("fb2"))
        rec["moneda"] = (rec.get("moneda") or "EUR").strip().upper()
        rec["amount"] = _to_decimal(rec.get("amount"))
        rec["biz_time"] = _to_datetime(rec.get("biz_time"))

        if rec["amount"] is None or rec["biz_time"] is None:
            continue
        if not rec["waybill_number"]:
            continue
        yield rec


def _to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _to_text(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _to_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


# ─────────────────────────────────────────────────────────────────────
#  Resolução de FKs (driver / task)
# ─────────────────────────────────────────────────────────────────────

def _build_resolution_caches():
    """Constrói caches em memória para evitar 30k queries no import.

    Retorna:
      cid_to_driver:   {courier_id: DriverProfile}
      cname_to_driver: {courier_name_lower: DriverProfile}
    """
    from drivers_app.models import DriverProfile
    from .models import DriverCourierMapping

    cid_to_driver = {}
    cname_to_driver = {}

    drivers = DriverProfile.objects.all()
    for d in drivers:
        if d.courier_id_cainiao:
            cid_to_driver[d.courier_id_cainiao] = d
        if d.apelido:
            cname_to_driver[d.apelido.strip().lower()] = d

    for m in DriverCourierMapping.objects.select_related("driver"):
        if not m.driver_id:
            continue
        if m.courier_id and m.courier_id not in cid_to_driver:
            d = drivers.filter(id=m.driver_id).first()
            if d:
                cid_to_driver[m.courier_id] = d
        if m.courier_name:
            key = m.courier_name.strip().lower()
            if key not in cname_to_driver:
                d = drivers.filter(id=m.driver_id).first()
                if d:
                    cname_to_driver[key] = d

    return cid_to_driver, cname_to_driver


def _resolve_task(waybill: str):
    """Devolve a CainiaoOperationTask 'Delivered' mais recente para o
    waybill, ou qualquer outra se não houver Delivered.

    O waybill da planilha Cainiao (REFs, ex: LP88020041701064070) pode
    estar guardado em CainiaoOperationTask como `lp_number` em vez de
    `waybill_number` — depende da fase em que foi importado o EPOD
    Task List. Procuramos nos dois campos.
    """
    from django.db.models import Q
    from .models import CainiaoOperationTask
    if not waybill:
        return None
    qs = CainiaoOperationTask.objects.filter(
        Q(waybill_number=waybill) | Q(lp_number=waybill)
    )
    delivered = qs.filter(task_status="Delivered").order_by(
        "-task_date",
    ).first()
    if delivered:
        return delivered
    return qs.order_by("-task_date").first()


def _resolve_driver(staff_id: str, task, cid_to_driver,
                    cname_to_driver):
    """Resolve driver por staff_id (cache) ou via courier da task."""
    if staff_id and staff_id in cid_to_driver:
        return cid_to_driver[staff_id]
    if task is not None:
        if task.courier_id_cainiao and task.courier_id_cainiao in cid_to_driver:
            return cid_to_driver[task.courier_id_cainiao]
        if task.courier_name:
            key = task.courier_name.strip().lower()
            if key in cname_to_driver:
                return cname_to_driver[key]
    return None


# ─────────────────────────────────────────────────────────────────────
#  Importação principal
# ─────────────────────────────────────────────────────────────────────

CHUNK_SIZE = 500


def import_cainiao_billing(file_obj, file_name: str, user=None):
    """Importa o XLSX e devolve a CainiaoBillingImport.

    Idempotente:
      - Se já existe import com o mesmo file_hash, devolve esse (sem
        criar nada novo).
      - Linhas existentes (waybill+fee+billing_id) são actualizadas
        em vez de duplicadas.

    Performance:
      - Sem @transaction.atomic global: 30k linhas numa única
        transação geram MySQL "Lock wait timeout exceeded" (errno
        1205) quando o servidor tem outras queries concorrentes.
      - bulk_create(update_conflicts=True) em chunks de 500 linhas,
        cada chunk em transação curta independente.
      - Resolução de driver em batch: pré-fetch de tasks por waybill
        de cada chunk, evita 30k queries individuais.
    """
    from .models import (
        CainiaoBillingImport, CainiaoBillingLine, CainiaoOperationTask,
        PartnerInvoice,
    )
    from core.models import Partner

    file_hash = compute_file_hash(file_obj)
    existing = CainiaoBillingImport.objects.filter(
        file_hash=file_hash,
    ).first()
    if existing:
        if existing.status == "FAILED":
            # Limpa o import anterior falhado para permitir nova tentativa
            log.info(
                "Cainiao billing import anterior FAILED (hash %s) — "
                "removendo para reimport.", file_hash[:12],
            )
            if existing.partner_invoice_id:
                existing.partner_invoice.delete()
            existing.delete()
        else:
            log.info(
                "Cainiao billing import já existe (hash %s)", file_hash[:12],
            )
            return existing, "already_imported"

    # Parsear todas as linhas em memória
    rows = list(parse_xlsx_to_rows(file_obj))
    if not rows:
        raise ValueError("Ficheiro vazio ou sem linhas válidas.")

    period_from = min(r["biz_time"] for r in rows).date()
    period_to = max(r["biz_time"] for r in rows).date()

    # Sessão criada num transaction curto (commit imediato)
    with transaction.atomic():
        session = CainiaoBillingImport.objects.create(
            file_name=file_name[:255],
            file_hash=file_hash,
            period_from=period_from,
            period_to=period_to,
            status="PROCESSING",
            imported_by=(
                user if user and user.is_authenticated else None
            ),
        )

    cid_to_driver, cname_to_driver = _build_resolution_caches()

    n_envio = 0
    n_compensacion = 0
    total_envio = Decimal("0.00")
    total_comp = Decimal("0.00")
    billing_ids = set()
    staff_ids = set()

    try:
        # Processar em chunks (cada chunk = transaction curta)
        for chunk_start in range(0, len(rows), CHUNK_SIZE):
            chunk = rows[chunk_start:chunk_start + CHUNK_SIZE]

            # Pre-fetch de tasks por waybill OU lp_number (batch — 1 query).
            # No XLSX Cainiao, REFs (waybill_number da billing) é tipicamente
            # o LP público (LP88...) — que no CainiaoOperationTask pode estar
            # guardado em lp_number. Indexamos por ambos para resolver.
            from django.db.models import Q as _Q
            wbs = [r["waybill_number"] for r in chunk if r["waybill_number"]]
            tasks_by_wb = {}
            if wbs:
                qs_tasks = CainiaoOperationTask.objects.filter(
                    _Q(waybill_number__in=wbs) | _Q(lp_number__in=wbs)
                ).order_by("-task_date")  # mais recente primeiro
                # Indexar por ambos os campos. Prioridade: Delivered > outros.
                for t in qs_tasks:
                    for key in (t.waybill_number, t.lp_number):
                        if not key:
                            continue
                        existing = tasks_by_wb.get(key)
                        if existing is None:
                            tasks_by_wb[key] = t
                        elif (
                            existing.task_status != "Delivered"
                            and t.task_status == "Delivered"
                        ):
                            tasks_by_wb[key] = t

            objs = []
            for r in chunk:
                task = tasks_by_wb.get(r["waybill_number"])
                driver = _resolve_driver(
                    r["staff_id"], task, cid_to_driver, cname_to_driver,
                )
                objs.append(CainiaoBillingLine(
                    import_session=session,
                    fee_type=r["fee_type"],
                    biz_time=r["biz_time"],
                    amount=r["amount"],
                    moneda=r["moneda"],
                    waybill_number=r["waybill_number"],
                    cp_name=r["cp_name"],
                    ciudad=r["ciudad"],
                    staff_id=r["staff_id"],
                    cainiao_billing_id=r["cainiao_billing_id"],
                    fb1=r["fb1"],
                    fb2=r["fb2"],
                    task=task,
                    driver=driver,
                ))

                if r["fee_type"] == "envio fee":
                    n_envio += 1
                    total_envio += r["amount"]
                elif r["fee_type"] == "compensacion":
                    n_compensacion += 1
                    total_comp += r["amount"]
                if r["cainiao_billing_id"]:
                    billing_ids.add(r["cainiao_billing_id"])
                if r["staff_id"]:
                    staff_ids.add(r["staff_id"])

            # bulk_create com update_conflicts (Django 4.1+).
            # Reimports de quinzena que sobreponham → faz UPDATE em
            # vez de duplicar (graças ao unique_together).
            #
            # Nota: em MySQL/MariaDB/SQLite, o argumento unique_fields=
            # NÃO pode ser passado (o backend usa automaticamente os
            # unique constraints da tabela). Apenas no PostgreSQL é
            # que ele é obrigatório.
            with transaction.atomic():
                CainiaoBillingLine.objects.bulk_create(
                    objs,
                    update_conflicts=True,
                    update_fields=[
                        "import_session", "biz_time", "amount", "moneda",
                        "cp_name", "ciudad", "staff_id", "fb1", "fb2",
                        "task", "driver", "updated_at",
                    ],
                )

        # PartnerInvoice agregada (transaction curta)
        partner_invoice = None
        with transaction.atomic():
            cainiao_partner = Partner.objects.filter(
                name__iexact="CAINIAO",
            ).first()
            if cainiao_partner:
                invoice_number = (
                    f"CAINIAO-{period_from.strftime('%Y%m%d')}-"
                    f"{period_to.strftime('%Y%m%d')}-{session.id}"
                )
                gross = total_envio + total_comp
                partner_invoice = PartnerInvoice.objects.create(
                    partner=cainiao_partner,
                    invoice_number=invoice_number,
                    external_reference=(
                        "; ".join(sorted(billing_ids))[:200]
                    ),
                    period_start=period_from,
                    period_end=period_to,
                    gross_amount=gross,
                    tax_amount=Decimal("0.00"),
                    net_amount=gross,
                    status="PENDING",
                    issue_date=timezone.now().date(),
                    due_date=period_to + timedelta(days=30),
                )

            session.partner_invoice = partner_invoice
            session.total_lines = n_envio + n_compensacion
            session.total_envio = total_envio
            session.total_compensacion = total_comp
            session.n_billing_ids = len(billing_ids)
            session.n_staff_ids = len(staff_ids)
            session.status = "COMPLETED"
            session.save(update_fields=[
                "partner_invoice", "total_lines", "total_envio",
                "total_compensacion", "n_billing_ids", "n_staff_ids",
                "status",
            ])

        return session, "created"

    except Exception as e:
        log.exception("Erro a importar Cainiao billing")
        try:
            session.status = "FAILED"
            session.error_message = str(e)[:2000]
            session.save(update_fields=["status", "error_message"])
        except Exception:
            pass
        raise


# ─────────────────────────────────────────────────────────────────────
#  Re-resolução de matching (sem reimportar)
# ─────────────────────────────────────────────────────────────────────

def reresolve_matching(session):
    """Re-corre a resolução de task/driver para um import existente.

    Útil quando:
      - o EPOD Task List foi importado depois da billing (tasks
        agora existem mas no momento do import billing não existiam);
      - o algoritmo de resolução foi melhorado;
      - DriverCourierMapping novos foram adicionados.

    Não duplica dados — apenas faz UPDATE em CainiaoBillingLine
    nos campos task_id e driver_id em chunks.
    """
    from django.db.models import Q as _Q
    from .models import CainiaoBillingLine, CainiaoOperationTask

    cid_to_driver, cname_to_driver = _build_resolution_caches()

    qs = session.lines.all().only(
        "id", "waybill_number", "staff_id", "task_id", "driver_id",
    )
    total = qs.count()
    n_resolved_task = 0
    n_resolved_driver = 0

    chunk = []
    chunk_size = CHUNK_SIZE

    def _flush(lines_to_update):
        if not lines_to_update:
            return
        with transaction.atomic():
            CainiaoBillingLine.objects.bulk_update(
                lines_to_update, fields=["task_id", "driver_id"],
            )

    # Pré-buscar tasks em batch para todo o conjunto (1 query única
    # com __in que cobre todos os waybills do session)
    wbs = list(set(qs.exclude(waybill_number="").values_list(
        "waybill_number", flat=True,
    )))
    tasks_by_wb = {}
    # Em chunks de 5k para não estourar query
    for i in range(0, len(wbs), 5000):
        sub = wbs[i:i + 5000]
        for t in CainiaoOperationTask.objects.filter(
            _Q(waybill_number__in=sub) | _Q(lp_number__in=sub),
        ).order_by("-task_date"):
            for key in (t.waybill_number, t.lp_number):
                if not key:
                    continue
                existing = tasks_by_wb.get(key)
                if existing is None:
                    tasks_by_wb[key] = t
                elif (
                    existing.task_status != "Delivered"
                    and t.task_status == "Delivered"
                ):
                    tasks_by_wb[key] = t

    for line in qs.iterator(chunk_size=chunk_size):
        old_task = line.task_id
        old_driver = line.driver_id
        task = tasks_by_wb.get(line.waybill_number)
        driver = _resolve_driver(
            line.staff_id, task, cid_to_driver, cname_to_driver,
        )
        new_task = task.id if task else None
        new_driver = driver.id if driver else None
        if new_task != old_task or new_driver != old_driver:
            line.task_id = new_task
            line.driver_id = new_driver
            chunk.append(line)
            if new_task and not old_task:
                n_resolved_task += 1
            if new_driver and not old_driver:
                n_resolved_driver += 1
        if len(chunk) >= chunk_size:
            _flush(chunk)
            chunk = []
    _flush(chunk)

    return {
        "total": total,
        "newly_resolved_task": n_resolved_task,
        "newly_resolved_driver": n_resolved_driver,
    }


# ─────────────────────────────────────────────────────────────────────
#  Análise / queries auxiliares
# ─────────────────────────────────────────────────────────────────────

def diagnose_delivered_no_billing(session):
    """Categoriza tasks Delivered no período sem linha de envio nesta
    pré-fatura, identificando a causa provável de cada uma.

    Categorias (mutuamente exclusivas, processadas por ordem):
      1. billed_other_import: aparece como envio fee em OUTRO
         CainiaoBillingImport (border effect — Cainiao facturou no
         mês seguinte ou anterior).
      2. compensation_this_import: aparece como compensación nesta
         pré-fatura (foi descontado, não pago).
      3. compensation_other_import: aparece como compensación noutro
         import.
      4. duplicate_locally: o mesmo waybill_number tem várias linhas
         em CainiaoOperationTask com Delivered em datas diferentes
         (suspeita de import duplicado ou re-entrega).
      5. genuinely_unpaid: nenhuma das anteriores → potencial perda
         genuína.
    """
    from collections import defaultdict
    from django.db.models import Q as _Q
    from .models import CainiaoBillingLine, CainiaoOperationTask

    # 1. Lista base: Delivered no período, sem linha de envio neste import
    billed_wbs_this = set(
        session.lines.filter(fee_type="envio fee").values_list(
            "waybill_number", flat=True,
        )
    )
    delivered_qs = (
        CainiaoOperationTask.objects
        .filter(
            task_date__range=(session.period_from, session.period_to),
            task_status="Delivered",
        )
        .exclude(waybill_number__in=billed_wbs_this)
    )

    # 2. Recolher os waybills em estudo (de waybill_number e lp_number)
    delivered_keys = set()
    delivered_list = list(delivered_qs.values(
        "id", "waybill_number", "lp_number", "task_date",
        "courier_name", "courier_id_cainiao", "destination_city",
    ))
    for t in delivered_list:
        if t["waybill_number"]:
            delivered_keys.add(t["waybill_number"])
        if t["lp_number"]:
            delivered_keys.add(t["lp_number"])

    if not delivered_list:
        return {
            "total": 0,
            "billed_other_import": [],
            "compensation_this_import": [],
            "compensation_other_import": [],
            "duplicate_locally": [],
            "genuinely_unpaid": [],
            "counts": {
                "total": 0, "billed_other_import": 0,
                "compensation_this_import": 0,
                "compensation_other_import": 0,
                "duplicate_locally": 0, "genuinely_unpaid": 0,
            },
        }

    # 3. Lookup em todos os CainiaoBillingLine cujos waybills coincidam
    cross_lines = (
        CainiaoBillingLine.objects
        .filter(waybill_number__in=delivered_keys)
        .select_related("import_session")
        .values(
            "waybill_number", "fee_type", "import_session_id",
            "import_session__period_from", "import_session__period_to",
        )
    )
    by_wb = defaultdict(list)
    for cl in cross_lines:
        by_wb[cl["waybill_number"]].append(cl)

    # 4. Detectar duplicados locais (mesmo waybill_number Delivered em
    #    >= 2 datas diferentes)
    dup_counts = defaultdict(set)
    for t in CainiaoOperationTask.objects.filter(
        task_date__range=(session.period_from, session.period_to),
        task_status="Delivered",
        waybill_number__in=[t["waybill_number"] for t in delivered_list],
    ).values("waybill_number", "task_date"):
        dup_counts[t["waybill_number"]].add(t["task_date"])
    duplicates = {wb for wb, dates in dup_counts.items() if len(dates) > 1}

    # 5. Categorizar
    billed_other = []
    comp_this = []
    comp_other = []
    dup_local = []
    genuine = []

    for t in delivered_list:
        wb = t["waybill_number"]
        lp = t["lp_number"] or ""
        # Lookup combinando waybill e lp
        candidates = list(by_wb.get(wb, [])) + list(by_wb.get(lp, []))

        # 5.1 envio fee em outro import
        billed_other_match = next(
            (c for c in candidates
             if c["fee_type"] == "envio fee"
             and c["import_session_id"] != session.id),
            None,
        )
        if billed_other_match:
            t["_reason"] = (
                f"Facturado em import #{billed_other_match['import_session_id']} "
                f"({billed_other_match['import_session__period_from']}→"
                f"{billed_other_match['import_session__period_to']})"
            )
            billed_other.append(t)
            continue

        # 5.2 compensación neste import
        comp_this_match = next(
            (c for c in candidates
             if c["fee_type"] == "compensacion"
             and c["import_session_id"] == session.id),
            None,
        )
        if comp_this_match:
            t["_reason"] = "Aparece como compensación neste import"
            comp_this.append(t)
            continue

        # 5.3 compensación em outro import
        comp_other_match = next(
            (c for c in candidates if c["fee_type"] == "compensacion"),
            None,
        )
        if comp_other_match:
            t["_reason"] = (
                f"Compensación em import "
                f"#{comp_other_match['import_session_id']}"
            )
            comp_other.append(t)
            continue

        # 5.4 duplicate local
        if wb in duplicates:
            dates = sorted(dup_counts[wb])
            t["_reason"] = (
                f"Delivered em {len(dates)} datas: "
                f"{', '.join(d.strftime('%d/%m') for d in dates)}"
            )
            dup_local.append(t)
            continue

        # 5.5 genuíno
        t["_reason"] = "Sem facturação em nenhum import — confirmar com Cainiao"
        genuine.append(t)

    return {
        "total": len(delivered_list),
        "billed_other_import": billed_other,
        "compensation_this_import": comp_this,
        "compensation_other_import": comp_other,
        "duplicate_locally": dup_local,
        "genuinely_unpaid": genuine,
        "counts": {
            "total": len(delivered_list),
            "billed_other_import": len(billed_other),
            "compensation_this_import": len(comp_this),
            "compensation_other_import": len(comp_other),
            "duplicate_locally": len(dup_local),
            "genuinely_unpaid": len(genuine),
        },
    }


def reconciliation_for_import(session):
    """Compara CainiaoBillingLine com CainiaoOperationTask local.

    Devolve dict com 4 secções:
      - paid_no_task: linhas pagas pela Cainiao mas sem task local
      - delivered_no_billing: tasks Delivered no período sem linha de envio
      - both: contagem das que batem certo
      - special_prices: linhas envio fee com amount != 1.60
    """
    from .models import CainiaoOperationTask

    lines = session.lines.all()

    paid_no_task = lines.filter(
        fee_type="envio fee", task__isnull=True,
    )

    billed_waybills = set(
        lines.filter(fee_type="envio fee").values_list(
            "waybill_number", flat=True,
        )
    )

    delivered_no_billing = (
        CainiaoOperationTask.objects
        .filter(
            task_date__range=(session.period_from, session.period_to),
            task_status="Delivered",
        )
        .exclude(waybill_number__in=billed_waybills)
        .order_by("task_date")
    )

    matched_count = (
        lines.filter(fee_type="envio fee", task__isnull=False).count()
    )

    special_prices = lines.filter(
        fee_type="envio fee",
    ).exclude(amount=Decimal("1.60"))

    return {
        "paid_no_task": paid_no_task,
        "paid_no_task_count": paid_no_task.count(),
        "delivered_no_billing": delivered_no_billing,
        "delivered_no_billing_count": delivered_no_billing.count(),
        "matched_count": matched_count,
        "special_prices": special_prices,
        "special_prices_count": special_prices.count(),
    }
