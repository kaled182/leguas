import csv
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_http_methods

from .models import (
    CainiaoDelivery,
    CainiaoImportBatch,
    DriverClaim,
    DriverCourierMapping,
    DriverHelper,
    DriverPreInvoice,
    DriverSettlement,
    PartnerInvoice,
    PreInvoiceAdvance,
    PreInvoiceBonus,
    PreInvoiceLine,
    PreInvoiceLostPackage,
    SettlementRun,
)
from .reports.pdf_generator import PDFGenerator
from .services import compute_payouts


def _parse_dates(request):
    dfrom = (
        parse_date(request.GET.get("date_from"))
        if request.GET.get("date_from")
        else None
    )
    dto = parse_date(request.GET.get("date_to")) if request.GET.get("date_to") else None
    return dfrom, dto


def summary(request):
    dfrom, dto = _parse_dates(request)
    qs = SettlementRun.objects.all()
    if dfrom:
        qs = qs.filter(run_date__gte=dfrom)
    if dto:
        qs = qs.filter(run_date__lte=dto)
    if request.GET.get("driver"):
        qs = qs.filter(driver__name=request.GET["driver"])
    if request.GET.get("client"):
        qs = qs.filter(client=request.GET["client"])

    agg = qs.aggregate(
        runs=Sum(1),
        qtd_pact=Sum("qtd_pact"),
        qtd_entregue=Sum("qtd_entregue"),
        bruto=Sum("total_pct"),
        gasoleo=Sum("gasoleo"),
        desconto_tickets=Sum("desconto_tickets"),
        rec_liq_tickets=Sum("rec_liq_tickets"),
        outros=Sum("outros"),
        liquido=Sum("vl_final"),
    )
    qtd_pact = float(agg.get("qtd_pact") or 0)
    qtd_ent = float(agg.get("qtd_entregue") or 0)
    taxa = round((qtd_ent / qtd_pact) * 100, 2) if qtd_pact > 0 else 0.0
    media = round((float(agg.get("liquido") or 0) / qtd_ent), 2) if qtd_ent > 0 else 0.0

    # Para requisições AJAX, retorna JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(
            {
                "runs": int(agg.get("runs") or 0),
                "qtd_pact": int(agg.get("qtd_pact") or 0),
                "qtd_entregue": int(agg.get("qtd_entregue") or 0),
                "bruto": float(agg.get("bruto") or 0),
                "gasoleo": float(agg.get("gasoleo") or 0),
                "desconto_tickets": float(agg.get("desconto_tickets") or 0),
                "rec_liq_tickets": float(agg.get("rec_liq_tickets") or 0),
                "outros": float(agg.get("outros") or 0),
                "liquido": float(agg.get("liquido") or 0),
                "taxa_sucesso_pct": taxa,
                "avg_liq_por_pacote": media,
            }
        )

    # Para requisições normais, renderiza o template
    return render(request, "settlements/summary.html")


def drivers_rank(request):
    dfrom, dto = _parse_dates(request)
    qs = SettlementRun.objects.all()
    if dfrom:
        qs = qs.filter(run_date__gte=dfrom)
    if dto:
        qs = qs.filter(run_date__lte=dto)

    data = qs.values("driver__name").annotate(
        entregues=Sum("qtd_entregue"),
        qtd_pact=Sum("qtd_pact"),
        liquido=Sum("vl_final"),
    )
    result = []
    for row in data:
        entregues = int(row["entregues"] or 0)
        pact = int(row["qtd_pact"] or 0)
        taxa = round((entregues / pact) * 100, 2) if pact else 0.0
        result.append(
            {
                "driver": row["driver__name"],
                "entregues": entregues,
                "taxa_media": taxa,
                "liquido": float(row["liquido"] or 0),
            }
        )
    result.sort(key=lambda r: (-r["liquido"], r["driver"]))

    # Para requisições AJAX, retorna JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(result, safe=False)

    # Para requisições normais, renderiza o template
    return render(request, "settlements/drivers_rank.html")


def runs_list(request):
    dfrom, dto = _parse_dates(request)
    qs = SettlementRun.objects.select_related("driver").all()
    if dfrom:
        qs = qs.filter(run_date__gte=dfrom)
    if dto:
        qs = qs.filter(run_date__lte=dto)
    if request.GET.get("driver"):
        qs = qs.filter(driver__name=request.GET["driver"])
    if request.GET.get("client"):
        qs = qs.filter(client=request.GET["client"])
    if request.GET.get("area"):
        qs = qs.filter(area_code=request.GET["area"])

    data = list(
        qs.values(
            "run_date",
            "client",
            "area_code",
            "driver__name",
            "qtd_saida",
            "qtd_pact",
            "qtd_entregue",
            "vl_pct",
            "total_pct",
            "gasoleo",
            "desconto_tickets",
            "rec_liq_tickets",
            "outros",
            "vl_final",
            "notes",
        ).order_by("-run_date")[:1000]
    )

    # Para requisições AJAX, retorna JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(data, safe=False)

    # Para requisições normais, renderiza o template
    return render(request, "settlements/runs_list.html")


def payouts(request):
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    # Para requisições normais sem parâmetros de data, apenas renderiza o template
    if (
        not (date_from and date_to)
        and request.headers.get("X-Requested-With") != "XMLHttpRequest"
    ):
        return render(request, "settlements/payouts.html")

    # Para requisições AJAX, exige os parâmetros de data
    if not (date_from and date_to):
        return JsonResponse(
            {"error": "date_from & date_to são obrigatórios (YYYY-MM-DD)"},
            status=400,
        )

    client = request.GET.get("client")
    area = request.GET.get("area")
    pf = datetime.strptime(date_from, "%Y-%m-%d").date()
    pt = datetime.strptime(date_to, "%Y-%m-%d").date()
    data = compute_payouts(pf, pt, client, area)

    # Para requisições AJAX, retorna JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse(data, safe=False)

    # Para requisições normais, renderiza o template
    return render(request, "settlements/payouts.html")


def payouts_csv(request):
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    if not (date_from and date_to):
        return JsonResponse(
            {"error": "date_from & date_to são obrigatórios (YYYY-MM-DD)"},
            status=400,
        )
    client = request.GET.get("client")
    area = request.GET.get("area")
    pf = datetime.strptime(date_from, "%Y-%m-%d").date()
    pt = datetime.strptime(date_to, "%Y-%m-%d").date()
    data = compute_payouts(pf, pt, client, area)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"payouts_{pf}_{pt}.csv"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    w = csv.writer(resp, delimiter=";")
    w.writerow(
        [
            "driver",
            "period_from",
            "period_to",
            "entregues",
            "bruto_pkg",
            "bonus",
            "fixo",
            "bruto_total",
            "descontos",
            "liquido",
            "media_liq_por_pacote",
        ]
    )
    for r in data:
        w.writerow(
            [
                r["driver"],
                r["period_from"],
                r["period_to"],
                r["entregues"],
                r["bruto_pkg"],
                r["bonus"],
                r["fixo"],
                r["bruto_total"],
                r["descontos"],
                r["liquido"],
                r["media_liq_por_pacote"],
            ]
        )
    return resp


# ==========================================
# FINANCIAL SYSTEM VIEWS (Fase 6)
# ==========================================


@login_required
def financial_dashboard(request):
    """Dashboard principal do sistema financeiro"""
    today = timezone.now().date()
    first_day_month = today.replace(day=1)

    # Total counts
    total_invoices = PartnerInvoice.objects.count()
    total_settlements = DriverPreInvoice.objects.count()
    total_claims = DriverClaim.objects.count()

    # KPIs - Invoices
    invoices_stats = {
        "total": total_invoices,
        "pending": PartnerInvoice.objects.filter(
            status__in=["DRAFT", "PENDING"]
        ).count(),
        "paid": PartnerInvoice.objects.filter(status="PAID").count(),
        "overdue": PartnerInvoice.objects.filter(status="OVERDUE").count(),
    }

    # KPIs - Liquidações de Motoristas (Pré-Faturas)
    # Antes contava DriverSettlement (legacy); agora conta DriverPreInvoice
    # que é o sistema actual de pagamento.
    settlements_stats = {
        "total": total_settlements,
        "paid": DriverPreInvoice.objects.filter(status="PAGO").count(),
        "pending": DriverPreInvoice.objects.filter(
            status__in=["RASCUNHO", "CALCULADO", "APROVADO",
                        "PENDENTE", "CONTESTADO"],
        ).count(),
    }

    # KPIs - Claims
    claims_stats = {
        "total": total_claims,
        "pending": DriverClaim.objects.filter(status="PENDING").count(),
        "approved": DriverClaim.objects.filter(status="APPROVED").count(),
        "rejected": DriverClaim.objects.filter(status="REJECTED").count(),
    }

    # KPIs - Reembolsos a sócios (terceiros)
    from decimal import Decimal as _Dec
    from django.db.models import Sum
    from .models import ThirdPartyReimbursement
    pending_reimb = ThirdPartyReimbursement.objects.filter(status="PENDENTE")
    reimbursement_stats = {
        "pending_count": pending_reimb.count(),
        "pending_total": (
            pending_reimb.aggregate(t=Sum("valor"))["t"] or _Dec("0.00")
        ),
    }

    # KPIs - Conta-Corrente Motoristas (Pré-faturas Advance pendentes)
    pending_cash = PreInvoiceAdvance.objects.filter(status="PENDENTE")
    cash_entry_stats = {
        "pending_count": pending_cash.count(),
        "pending_total": (
            pending_cash.aggregate(t=Sum("valor"))["t"] or _Dec("0.00")
        ),
    }

    # Recent invoices - formatted for financial_recent_items.html
    recent_invoices_qs = PartnerInvoice.objects.select_related("partner").order_by(
        "-created_at"
    )[:5]
    recent_invoices = []
    for inv in recent_invoices_qs:
        status_class = {
            "PAID": "bg-emerald-100 text-emerald-800 dark:bg-emerald-900/30 dark:text-emerald-400",
            "PENDING": "bg-amber-100 text-amber-800 dark:bg-amber-900/30 dark:text-amber-400",
            "OVERDUE": "bg-red-100 text-red-800 dark:bg-red-900/30 dark:text-red-400",
            "DRAFT": "bg-gray-100 text-gray-800 dark:bg-gray-700 dark:text-gray-300",
        }.get(
            inv.status,
            "bg-gray-100 text-gray-800 dark:bg-gray-700 dark:text-gray-300",
        )

        recent_invoices.append(
            {
                "title": f'{inv.partner.name if inv.partner else "N/A"} - {inv.invoice_number}',
                "description": f'{inv.period_start.strftime("%d/%m/%Y")} - {inv.period_end.strftime("%d/%m/%Y")}',
                "value": f"€{inv.net_amount:,.2f}",
                "badge": inv.get_status_display(),
                "badge_class": status_class,
            }
        )

    # Pending claims - formatted for financial_recent_items.html
    pending_claims_qs = (
        DriverClaim.objects.select_related("driver")
        .filter(status="PENDING")
        .order_by("-created_at")[:5]
    )
    pending_claims = []
    for claim in pending_claims_qs:
        pending_claims.append(
            {
                "title": f'{claim.driver.nome_completo if claim.driver else "N/A"} - {claim.get_claim_type_display()}',
                "description": claim.description[:80]
                + ("..." if len(claim.description) > 80 else ""),
                "value": f"€{claim.amount:,.2f}" if claim.amount else "—",
                "badge": f"#{claim.id}",
                "badge_class": "bg-amber-100 text-amber-800 dark:bg-amber-900/30 dark:text-amber-400",
            }
        )

    context = {
        "invoices_stats": invoices_stats,
        "settlements_stats": settlements_stats,
        "claims_stats": claims_stats,
        "reimbursement_stats": reimbursement_stats,
        "cash_entry_stats": cash_entry_stats,
        "recent_invoices": recent_invoices,
        "pending_claims": pending_claims,
    }

    return render(request, "settlements/financial_dashboard_v2.html", context)


@login_required
def invoice_list(request):
    """Lista de invoices de partners"""
    from django.db.models import Sum
    invoices = PartnerInvoice.objects.select_related(
        "partner",
    ).prefetch_related("cainiao_import").all()

    # Filtros
    status = request.GET.get("status")
    if status:
        invoices = invoices.filter(status=status)

    partner_id = request.GET.get("partner")
    if partner_id:
        invoices = invoices.filter(partner_id=partner_id)

    # Filtros de data
    date_from = request.GET.get("date_from")
    if date_from:
        invoices = invoices.filter(period_start__gte=date_from)

    date_to = request.GET.get("date_to")
    if date_to:
        invoices = invoices.filter(period_end__lte=date_to)

    # Busca
    search = request.GET.get("search")
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search)
            | Q(external_reference__icontains=search)
            | Q(partner__name__icontains=search)
        )

    # Ordenação
    invoices = invoices.order_by("-created_at")

    # Paginação
    paginator = Paginator(invoices, 25)  # 25 items por página
    page = request.GET.get("page", 1)

    try:
        invoices = paginator.page(page)
    except PageNotAnInteger:
        invoices = paginator.page(1)
    except EmptyPage:
        invoices = paginator.page(paginator.num_pages)

    # KPIs (sobre o queryset completo, antes da paginação)
    base_qs = PartnerInvoice.objects.all()
    total_pending = base_qs.filter(
        status__in=["PENDING", "DRAFT"],
    ).aggregate(s=Sum("net_amount"))["s"] or 0
    total_overdue = base_qs.filter(status="OVERDUE").aggregate(
        s=Sum("net_amount"),
    )["s"] or 0
    total_paid_year = base_qs.filter(
        status="PAID",
        paid_date__year=timezone.now().year,
    ).aggregate(s=Sum("net_amount"))["s"] or 0
    count_cainiao_imports = base_qs.filter(
        cainiao_import__isnull=False,
    ).count()

    context = {
        "invoices": invoices,
        "status_choices": PartnerInvoice.STATUS_CHOICES,
        "kpi_total_pending": total_pending,
        "kpi_total_overdue": total_overdue,
        "kpi_total_paid_year": total_paid_year,
        "kpi_count_cainiao_imports": count_cainiao_imports,
    }

    return render(request, "settlements/invoice_list_v2.html", context)


@login_required
def invoice_detail(request, invoice_id):
    """Detalhes de uma invoice"""
    invoice = get_object_or_404(
        PartnerInvoice.objects.select_related("partner"), id=invoice_id
    )

    # Pedidos relacionados (sample - pode precisar de ajuste conforme seu modelo Order)
    from orders_manager.models import Order

    related_orders = Order.objects.filter(
        partner=invoice.partner,
        created_at__gte=invoice.period_start,
        created_at__lte=invoice.period_end,
        current_status="DELIVERED",
    ).select_related("assigned_driver")[:50]

    context = {
        "invoice": invoice,
        "related_orders": related_orders,
    }

    return render(request, "settlements/invoice_detail.html", context)


@login_required
def invoice_download_pdf(request, invoice_id):
    """Download do PDF da invoice"""
    invoice = get_object_or_404(
        PartnerInvoice.objects.select_related("partner"), id=invoice_id
    )

    pdf_gen = PDFGenerator()
    pdf_file = pdf_gen.generate_invoice_pdf(invoice)

    response = HttpResponse(pdf_file.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
    )

    return response


@login_required
def settlement_list(request):
    """Liquidações de Motoristas — lista de Pré-Faturas (DriverPreInvoice).

    Por defeito mostra apenas PFs emitidas no mês atual (filtro sobre
    `created_at`). O utilizador pode ajustar `date_from`/`date_to` para
    janelas maiores ou menores.

    Esta página é a vista global de todas as PFs (pagamentos a motoristas)
    no sistema. É independente da página de backoffice
    `/settlements/pre-invoices/` (que serve para gestão por estado e tem UI
    própria de bulk actions).
    """
    from datetime import timedelta
    from decimal import Decimal as _Dec
    from django.db.models import Sum

    # Default: mês actual (primeiro ao último dia)
    today = timezone.now().date()
    first_day = today.replace(day=1)
    next_month_first = (first_day + timedelta(days=32)).replace(day=1)
    last_day = next_month_first - timedelta(days=1)

    # Sempre aplicar mês atual quando o campo está vazio ou ausente.
    # O utilizador deve preencher datas explícitas para sair do mês actual.
    date_from = request.GET.get("date_from") or first_day.strftime("%Y-%m-%d")
    date_to = request.GET.get("date_to") or last_day.strftime("%Y-%m-%d")

    qs = DriverPreInvoice.objects.select_related("driver").all()

    status = request.GET.get("status")
    if status:
        qs = qs.filter(status=status)

    driver_search = request.GET.get("driver")
    if driver_search:
        qs = qs.filter(driver__nome_completo__icontains=driver_search)

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    qs = qs.order_by("-periodo_fim", "-id")

    # KPIs (sobre o queryset filtrado)
    aggs = qs.aggregate(
        total_geral=Sum("total_a_receber"),
        total_pago=Sum(
            "total_a_receber",
            filter=Q(status="PAGO"),
        ),
        total_pendente=Sum(
            "total_a_receber",
            filter=Q(status__in=["RASCUNHO", "CALCULADO", "APROVADO",
                                 "PENDENTE", "CONTESTADO"]),
        ),
    )
    kpis = {
        "n_total": qs.count(),
        "n_pago": qs.filter(status="PAGO").count(),
        "n_pendente": qs.filter(
            status__in=["RASCUNHO", "CALCULADO", "APROVADO",
                        "PENDENTE", "CONTESTADO"],
        ).count(),
        "total_geral": aggs["total_geral"] or _Dec("0.00"),
        "total_pago": aggs["total_pago"] or _Dec("0.00"),
        "total_pendente": aggs["total_pendente"] or _Dec("0.00"),
    }

    paginator = Paginator(qs, 25)
    page = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    _meses_pt = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
    }
    context = {
        "pre_invoices": page_obj,
        "kpis": kpis,
        "status_choices": DriverPreInvoice.STATUS_CHOICES,
        # Defaults para os inputs do formulário
        "filter_date_from": date_from or "",
        "filter_date_to": date_to or "",
        "month_label": f"{_meses_pt[today.month]}/{today.year}",
    }
    return render(request, "settlements/settlement_list_v2.html", context)


@login_required
def claim_list(request):
    """Lista e criação de claims de motoristas"""
    from drivers_app.models import DriverProfile

    # POST: criar novo claim
    if request.method == "POST":
        driver_id = request.POST.get("driver_id")
        claim_type = request.POST.get("claim_type")
        amount = request.POST.get("amount")
        description = request.POST.get("description")
        occurred_at = request.POST.get("occurred_at")
        tracking_code = request.POST.get("tracking_code", "").strip()

        errors = []
        if not driver_id:
            errors.append("Motorista é obrigatório.")
        if not claim_type:
            errors.append("Tipo é obrigatório.")
        if not amount:
            errors.append("Valor é obrigatório.")
        if not description:
            errors.append("Descrição é obrigatória.")

        if not errors:
            from orders_manager.models import Order
            order = None
            if tracking_code:
                order = Order.objects.filter(tracking_code__iexact=tracking_code).first()

            claim = DriverClaim.objects.create(
                driver_id=driver_id,
                claim_type=claim_type,
                amount=amount,
                description=description,
                occurred_at=occurred_at or timezone.now(),
                order=order,
                created_by=request.user,
                status="PENDING",
            )
            if request.FILES.get("evidence_file"):
                claim.evidence_file = request.FILES["evidence_file"]
                claim.save()

            from django.contrib import messages
            messages.success(request, f"Reclamação #{claim.id} criada com sucesso.")
            from django.shortcuts import redirect
            return redirect("claim-list")

    claims = DriverClaim.objects.select_related(
        "driver", "settlement", "order", "vehicle_incident"
    ).all()

    # Filtros
    status = request.GET.get("status")
    if status:
        claims = claims.filter(status=status)

    claim_type = request.GET.get("claim_type")
    if claim_type:
        claims = claims.filter(claim_type=claim_type)

    driver_name = request.GET.get("driver", "").strip()
    if driver_name:
        claims = claims.filter(driver__nome_completo__icontains=driver_name)

    date_filter = request.GET.get("date", "").strip()
    if date_filter:
        parsed = parse_date(date_filter)
        if parsed:
            claims = claims.filter(occurred_at__date=parsed)

    # Ordenação
    claims = claims.order_by("-created_at")

    # Paginação
    paginator = Paginator(claims, 25)
    page = request.GET.get("page", 1)

    try:
        claims = paginator.page(page)
    except PageNotAnInteger:
        claims = paginator.page(1)
    except EmptyPage:
        claims = paginator.page(paginator.num_pages)

    drivers = DriverProfile.objects.filter(status="ATIVO").order_by("nome_completo")

    context = {
        "claims": claims,
        "status_choices": DriverClaim.STATUS_CHOICES,
        "type_choices": DriverClaim.CLAIM_TYPES,
        "drivers": drivers,
    }

    return render(request, "settlements/claim_list_v2.html", context)


@login_required
@require_http_methods(["POST"])
def claim_update(request, claim_id):
    """Edita campos do DriverClaim antes de aprovar/rejeitar.

    Permite ao operador ajustar o valor a deduzir (a Cainiao pode
    descontar €30 mas a empresa pode decidir repassar só parte ao
    driver) e o tipo da reclamação.

    Bloqueia se já foi descontado num settlement fechado.
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    from decimal import Decimal, InvalidOperation

    claim = get_object_or_404(DriverClaim, id=claim_id)

    if claim.settlement_id:
        messages.error(
            request,
            f"Não é possível editar — já foi descontado no Acerto "
            f"#{claim.settlement_id}. Desvincule primeiro.",
        )
        return redirect("claim-detail", claim_id=claim.id)

    # amount
    raw_amount = (request.POST.get("amount") or "").replace(",", ".").strip()
    if raw_amount:
        try:
            new_amount = Decimal(raw_amount)
            if new_amount < 0:
                raise ValueError("negativo")
            claim.amount = new_amount
        except (InvalidOperation, ValueError):
            messages.error(request, f"Valor inválido: {raw_amount}")
            return redirect("claim-detail", claim_id=claim.id)

    # claim_type
    new_type = (request.POST.get("claim_type") or "").strip()
    if new_type and new_type in dict(DriverClaim.CLAIM_TYPES):
        claim.claim_type = new_type

    # description
    new_desc = request.POST.get("description")
    if new_desc is not None:
        claim.description = new_desc.strip()[:2000]

    claim.save(update_fields=[
        "amount", "claim_type", "description", "updated_at",
    ])

    # Se o claim já estava APPROVED e o valor mudou, propagar para
    # qualquer PreInvoiceLostPackage gerado a partir deste claim.
    if claim.status == "APPROVED":
        from .models import PreInvoiceLostPackage
        marker = f"auto:driver_claim:{claim.id}"
        affected_pfs = []
        for pkg in PreInvoiceLostPackage.objects.filter(api_source=marker):
            pkg.valor = claim.amount
            pkg.save(update_fields=["valor"])
            affected_pfs.append(pkg.pre_invoice)
        for pf in {p.id: p for p in affected_pfs}.values():
            pf.recalcular()

    messages.success(
        request,
        f"Reclamação #{claim.id} atualizada para €{claim.amount:.2f}.",
    )
    return redirect("claim-detail", claim_id=claim.id)


@login_required
@require_http_methods(["POST"])
def claim_delete(request, claim_id):
    """Apaga uma reclamação (DriverClaim).

    Bloqueia se já estiver associada a um DriverSettlement (foi
    descontada num acerto fechado) ou a uma CainiaoBillingLine
    confirmada — nesses casos é preciso desfazer o link primeiro.
    """
    from django.contrib import messages
    from django.shortcuts import redirect

    claim = get_object_or_404(DriverClaim, id=claim_id)

    if claim.settlement_id:
        messages.error(
            request,
            f"Não é possível apagar — Reclamação #{claim.id} já foi "
            f"descontada no Acerto #{claim.settlement_id}. "
            "Desvincule do acerto primeiro.",
        )
        return redirect("claim-detail", claim_id=claim.id)

    cainiao_lines = claim.cainiao_billing_lines.all()
    if cainiao_lines.exists():
        # Apenas desliga as billing lines (não apaga as billing lines)
        cainiao_lines.update(claim=None)

    summary = (
        f"#{claim.id} {claim.driver.nome_completo} "
        f"€{claim.amount} ({claim.get_claim_type_display()})"
    )
    claim.delete()
    messages.success(request, f"Reclamação eliminada: {summary}")
    return redirect("claim-list")


@login_required
def claim_detail(request, claim_id):
    """Detalhes de um claim, com ações de aprovação/rejeição"""
    from django.contrib import messages
    from django.shortcuts import redirect

    claim = get_object_or_404(
        DriverClaim.objects.select_related(
            "driver", "settlement", "order", "vehicle_incident"
        ),
        id=claim_id,
    )

    if request.method == "POST":
        action = request.POST.get("action")
        notes = request.POST.get("review_notes", "").strip()

        if action == "approve" and claim.status == "PENDING":
            claim.approve(request.user, notes)
            messages.success(request, f"Reclamação #{claim.id} aprovada.")
        elif action == "reject" and claim.status in ("PENDING", "APPEALED"):
            claim.reject(request.user, notes)
            messages.success(request, f"Reclamação #{claim.id} rejeitada.")
        else:
            messages.error(request, "Ação inválida ou status incompatível.")

        return redirect("claim-detail", claim_id=claim.id)

    context = {
        "claim": claim,
    }

    return render(request, "settlements/claim_detail.html", context)


# ============================================================================
# PRÉ-FATURAS DE MOTORISTA — GESTÃO / BACKOFFICE
# ============================================================================


@login_required
def pre_invoice_admin_list(request):
    """Lista unificada de pré-faturas (motoristas + empresas parceiras)."""
    import calendar
    from decimal import Decimal
    from django.shortcuts import redirect
    from django.utils import timezone
    from django.core.paginator import Paginator
    from django.db.models import Count, Sum
    from core.models import Partner
    from drivers_app.models import EmpresaParceiraLancamento

    # Garantir que status tem sempre valor padrão explícito no URL
    # Redirecionar se status ausente ou vazio (não foi escolhido explicitamente)
    if not request.GET.get("status"):
        params = request.GET.copy()
        params["status"] = "NAO_PAGO"
        return redirect(f"{request.path}?{params.urlencode()}")

    qs = (
        DriverPreInvoice.objects.select_related("driver")
        .prefetch_related("linhas__parceiro")
        .order_by("-periodo_fim", "-created_at")
    )

    # Filtros
    status_f   = request.GET.get("status", "").strip()
    driver_f   = request.GET.get("driver", "").strip()
    month_f    = request.GET.get("month", "").strip()   # formato YYYY-MM
    parceiro_f = request.GET.get("parceiro", "").strip()
    tipo_f     = request.GET.get("tipo", "").strip()    # "motorista"|"empresa"|""

    # Converter mês em date_from / date_to para os filtros
    date_from = date_to = ""
    if month_f:
        try:
            y, m = int(month_f[:4]), int(month_f[5:7])
            date_from = f"{y}-{m:02d}-01"
            date_to   = f"{y}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
        except (ValueError, IndexError):
            pass

    if status_f == "NAO_PAGO":
        qs = qs.exclude(status="PAGO")
    elif status_f and status_f != "TODOS":
        qs = qs.filter(status=status_f)
    if driver_f:
        qs = qs.filter(driver__nome_completo__icontains=driver_f)
    if date_from:
        qs = qs.filter(periodo_inicio__gte=date_from)
    if date_to:
        qs = qs.filter(periodo_fim__lte=date_to)
    if parceiro_f:
        qs = qs.filter(linhas__parceiro_id=parceiro_f).distinct()

    # KPIs — driver pre-invoices
    agg = DriverPreInvoice.objects.aggregate(
        pendente_valor=Sum("total_a_receber", filter=Q(status="PENDENTE")),
        pago_valor=Sum("total_a_receber", filter=Q(status="PAGO")),
        aprovado_valor=Sum("total_a_receber", filter=Q(status="APROVADO")),
    )
    # KPIs — empresa lancamentos
    lagg = EmpresaParceiraLancamento.objects.aggregate(
        pendente_valor=Sum("valor_base", filter=Q(status="PENDENTE")),
        pago_valor=Sum("valor_base", filter=Q(status="PAGO")),
        aprovado_valor=Sum("valor_base", filter=Q(status="APROVADO")),
    )

    def _d(v):
        return Decimal(v or 0)

    kpis = {
        "total_a_pagar": _d(agg["aprovado_valor"]) + _d(agg["pendente_valor"])
                       + _d(lagg["aprovado_valor"]) + _d(lagg["pendente_valor"]),
        "pendente": _d(agg["pendente_valor"]) + _d(lagg["pendente_valor"]),
        "pago": _d(agg["pago_valor"]) + _d(lagg["pago_valor"]),
    }

    # Empresa lancamentos for the second section of the unified table
    lqs = (
        EmpresaParceiraLancamento.objects
        .select_related("empresa", "created_by")
        .order_by("-periodo_inicio", "-created_at")
    )
    if status_f == "NAO_PAGO":
        lqs = lqs.exclude(status="PAGO")
    elif status_f and status_f != "TODOS":
        lqs = lqs.filter(status=status_f)
    if driver_f:
        lqs = lqs.filter(empresa__nome__icontains=driver_f)
    if date_from:
        lqs = lqs.filter(periodo_inicio__gte=date_from)
    if date_to:
        lqs = lqs.filter(periodo_fim__lte=date_to)

    # Paginate driver pre-invoices (primary table)
    show_motoristas = tipo_f != "empresa"
    show_empresas   = tipo_f != "motorista"

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get("page"))

    parceiros = list(Partner.objects.filter(is_active=True).values("id", "name").order_by("name"))

    return render(request, "settlements/pre_invoice_admin_list.html", {
        "pre_invoices": page,
        "lancamentos": lqs if show_empresas else [],
        "status_choices": DriverPreInvoice.STATUS_CHOICES,
        "kpis": kpis,
        "parceiros": parceiros,
        "show_motoristas": show_motoristas,
        "show_empresas": show_empresas,
        "tipo_f": tipo_f,
        "month_f": month_f,
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_change_status(request, pre_invoice_id):
    """Muda o status de uma pré-fatura com validação de transição."""
    import json

    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    if request.content_type and "multipart" in request.content_type:
        body = request.POST.dict()
    else:
        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            body = request.POST.dict()

    novo_status = (body.get("status") or "").strip().upper()
    nota = (body.get("nota") or "").strip()

    # Acção especial: calcular recalcula e define CALCULADO automaticamente
    if novo_status == "CALCULAR":
        pf.recalcular()  # sets status=CALCULADO e save() internamente
        return JsonResponse({
            "success": True,
            "status": pf.status,
            "status_display": pf.get_status_display(),
            "total_a_receber": str(pf.total_a_receber),
        })

    permitidos = DriverPreInvoice.TRANSICOES.get(pf.status, [])
    if novo_status not in permitidos:
        return JsonResponse(
            {"success": False, "error": f"Transição {pf.status} → {novo_status} não permitida."},
            status=400,
        )

    pf.status = novo_status

    # Acções automáticas por estado
    if novo_status == "PAGO":
        from datetime import date
        pf.data_pagamento = pf.data_pagamento or date.today()
        ref = body.get("referencia_pagamento", "").strip()
        if ref:
            pf.referencia_pagamento = ref
        # Comprovante (multipart — só presente se enviado via FormData)
        if request.FILES.get("comprovante"):
            pf.comprovante_pagamento = request.FILES["comprovante"]

    if nota:
        pf.observacoes = (pf.observacoes + f"\n[{novo_status}] {nota}").strip()

    pf.save()
    return JsonResponse({
        "success": True,
        "status": pf.status,
        "status_display": pf.get_status_display(),
        "comprovante_url": pf.comprovante_pagamento.url if pf.comprovante_pagamento else "",
    })


# ============================================================================
# PRÉ-FATURAS DE MOTORISTA — API JSON (consumida pelo modal do drivers_app)
# ============================================================================


def _get_cainiao_helpers_breakdown(pf, linha):
    """Retorna breakdown de entregas por helper/courier para linhas Cainiao."""
    if linha.api_source != "cainiao":
        return []
    from django.db.models import Count as _Count
    qs = (
        CainiaoDelivery.objects
        .filter(
            driver=pf.driver,
            delivery_time__date__gte=pf.periodo_inicio,
            delivery_time__date__lte=pf.periodo_fim,
        )
        .values("courier_id", "helper_name")
        .annotate(count=_Count("id"))
        .order_by("-count")
    )
    return [
        {
            "courier_id": h["courier_id"],
            "helper_name": h["helper_name"] or "Principal",
            "count": h["count"],
        }
        for h in qs
    ]


@login_required
def driver_fleet_invoice_lines_api(request, driver_id):
    """Lista as linhas FleetInvoiceDriverLine onde este motorista
    aparece — usado pelo modal do motorista quando ele pertence a uma
    frota, para mostrar 'PFs da Frota onde estou incluído'.
    """
    from drivers_app.models import DriverProfile
    from .models import FleetInvoiceDriverLine
    driver = get_object_or_404(DriverProfile, id=driver_id)
    qs = (
        FleetInvoiceDriverLine.objects
        .select_related("fleet_invoice", "fleet_invoice__empresa")
        .filter(driver=driver)
        .order_by("-fleet_invoice__periodo_fim", "-fleet_invoice_id")
    )
    rows = []
    for ln in qs:
        fi = ln.fleet_invoice
        rows.append({
            "line_id": ln.id,
            "fleet_invoice_id": fi.id,
            "fleet_invoice_numero": fi.numero,
            "empresa_id": fi.empresa_id,
            "empresa_nome": fi.empresa.nome if fi.empresa_id else "",
            "periodo_inicio": fi.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": fi.periodo_fim.strftime("%Y-%m-%d"),
            "status": fi.status,
            "status_display": fi.get_status_display(),
            "deliveries": ln.deliveries,
            "price_per_package": str(ln.price_per_package),
            "price_source": ln.price_source,
            "base_amount": str(ln.base_amount),
            "bonus_days_count": ln.bonus_days_count,
            "bonus_amount": str(ln.bonus_amount),
            "claims_count": ln.claims_count,
            "claims_amount": str(ln.claims_amount),
            "subtotal": str(ln.subtotal),
        })
    return JsonResponse({
        "success": True,
        "driver_id": driver.id,
        "driver_nome": driver.nome_completo,
        "is_fleet_driver": bool(driver.empresa_parceira_id),
        "empresa_parceira_id": driver.empresa_parceira_id,
        "empresa_parceira_nome": (
            driver.empresa_parceira.nome
            if driver.empresa_parceira_id else ""
        ),
        "lines": rows,
        "n_lines": len(rows),
    })


@login_required
def driver_pre_invoices_api(request, driver_id):
    """Retorna lista de pré-faturas de um motorista em JSON."""
    from core.models import Partner
    from drivers_app.models import DriverProfile
    driver = get_object_or_404(DriverProfile, id=driver_id)

    pre_invoices = (
        DriverPreInvoice.objects.filter(driver=driver)
        .prefetch_related(
            "linhas__parceiro",
            "bonificacoes",
            "pacotes_perdidos",
            "adiantamentos",
        )
        .order_by("-periodo_fim")
    )

    parceiros = list(
        Partner.objects.filter(is_active=True)
        .values("id", "name")
        .order_by("name")
    )

    data = []
    for pf in pre_invoices:
        data.append({
            "id": pf.id,
            "numero": pf.numero,
            "driver_id": pf.driver_id,
            "periodo_inicio": pf.periodo_inicio.strftime("%d/%m/%Y"),
            "periodo_fim": pf.periodo_fim.strftime("%d/%m/%Y"),
            "base_entregas": str(pf.base_entregas),
            "total_bonus": str(pf.total_bonus),
            "ajuste_manual": str(pf.ajuste_manual),
            "penalizacoes_gerais": str(pf.penalizacoes_gerais),
            "total_pacotes_perdidos": str(pf.total_pacotes_perdidos),
            "total_adiantamentos": str(pf.total_adiantamentos),
            "subtotal_bruto": str(pf.subtotal_bruto),
            "total_a_receber": str(pf.total_a_receber),
            "status": pf.status,
            "status_display": pf.get_status_display(),
            "data_pagamento": pf.data_pagamento.strftime("%d/%m/%Y") if pf.data_pagamento else "",
            "referencia_pagamento": pf.referencia_pagamento,
            "comprovante_url": pf.comprovante_pagamento.url if pf.comprovante_pagamento else "",
            "fatura_url": pf.fatura_ficheiro.url if pf.fatura_ficheiro else "",
            "observacoes": pf.observacoes,
            "linhas": [
                {
                    "id": l.id,
                    "parceiro_id": l.parceiro_id,
                    "parceiro_nome": l.parceiro.name if l.parceiro else "—",
                    "courier_id": l.courier_id,
                    "total_pacotes": l.total_pacotes,
                    "taxa_por_entrega": str(l.taxa_por_entrega),
                    "dsr_percentual": str(l.dsr_percentual),
                    "base_entregas": str(l.base_entregas),
                    "observacoes": l.observacoes,
                    "api_source": l.api_source,
                    "helpers": _get_cainiao_helpers_breakdown(pf, l),
                }
                for l in pf.linhas.all()
            ],
            "bonificacoes": [
                {
                    "id": b.id,
                    "data": b.data.strftime("%d/%m/%Y"),
                    "tipo_display": b.get_tipo_display(),
                    "qtd_entregas_elegiveis": b.qtd_entregas_elegiveis,
                    "bonus_calculado": str(b.bonus_calculado),
                    "observacoes": b.observacoes,
                }
                for b in pf.bonificacoes.all()
            ],
            "pacotes_perdidos": [
                {
                    "id": p.id,
                    "data": p.data.strftime("%d/%m/%Y") if p.data else "",
                    "numero_pacote": p.numero_pacote,
                    "descricao": p.descricao,
                    "valor": str(p.valor),
                    "iva_percentual": str(p.iva_percentual),
                    "valor_com_iva": str(p.valor_com_iva),
                }
                for p in pf.pacotes_perdidos.all()
            ],
            "adiantamentos": [
                {
                    "id": a.id,
                    "data": a.data.strftime("%d/%m/%Y") if a.data else "",
                    "tipo": a.tipo,
                    "tipo_display": a.get_tipo_display(),
                    "descricao": a.descricao,
                    "valor": str(a.valor),
                    "documento_referencia": a.documento_referencia,
                    "paid_by_source": a.paid_by_source,
                    "paid_by_lender_id": a.paid_by_lender_id,
                    "paid_by_lender_nome": (
                        a.paid_by_lender.nome
                        if a.paid_by_lender_id else ""
                    ),
                }
                for a in pf.adiantamentos.select_related(
                    "paid_by_lender",
                ).all()
            ],
            "total_comissoes_indicacao": str(pf.total_comissoes_indicacao),
            "comissoes_indicacao_detalhe": _get_referral_commission_detail(pf),
        })

    return JsonResponse({
        "pre_invoices": data,
        "driver_nome": driver.nome_completo,
        "driver_id": driver.id,
        "is_fleet_driver": bool(driver.empresa_parceira_id),
        "empresa_parceira_id": driver.empresa_parceira_id,
        "empresa_parceira_nome": (
            driver.empresa_parceira.nome
            if driver.empresa_parceira_id else ""
        ),
        "parceiros": parceiros,
    })


@login_required
@require_http_methods(["POST"])
def driver_pre_invoice_create(request, driver_id):
    """Cria nova pré-fatura para o motorista.

    Regra: APENAS pacotes com task_status='Delivered' são contabilizados
    para pagamento. 'Attempt Failure' aparece nas estatísticas mas
    nunca entra na linha de facturação.

    Auto-popula:
      - 1 linha CAINIAO com total_pacotes = delivered count
      - taxa_por_entrega = driver.price_per_package (override) ou
        partner.price_per_package (default)
      - PreInvoiceBonus para domingos/feriados com >=30 entregas
    """
    import json
    from datetime import timedelta
    from decimal import Decimal, InvalidOperation
    from drivers_app.models import DriverProfile
    from .models import (
        BonusBlackoutDate, CainiaoOperationTask, Holiday, PreInvoiceBonus,
        PreInvoiceLine,
    )
    from core.models import Partner

    driver = get_object_or_404(DriverProfile, id=driver_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    def to_dec(val, default="0.00"):
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError):
            return Decimal(default)

    # Motorista de frota não emite PF individual — a sua factura é
    # gerada via lote (FleetInvoice) da empresa parceira a que pertence.
    if driver.empresa_parceira_id:
        empresa_nome = (
            driver.empresa_parceira.nome
            if driver.empresa_parceira else "?"
        )
        return JsonResponse(
            {"success": False,
             "error": (
                 f"Motorista vinculado à frota '{empresa_nome}'. "
                 "PFs são geradas via lote da frota — não é possível "
                 "emitir PF individual."
             ),
             "reason": "fleet_driver",
             "empresa_parceira_id": driver.empresa_parceira_id,
             "empresa_parceira_nome": empresa_nome},
            status=400,
        )

    periodo_inicio = parse_date(body.get("periodo_inicio") or "")
    periodo_fim = parse_date(body.get("periodo_fim") or "")
    if not periodo_inicio or not periodo_fim:
        return JsonResponse(
            {"success": False, "error":
             "periodo_inicio e periodo_fim são obrigatórios"},
            status=400,
        )
    if periodo_inicio > periodo_fim:
        periodo_inicio, periodo_fim = periodo_fim, periodo_inicio

    # Gerar número sequencial
    ultimo = (
        DriverPreInvoice.objects.filter(numero__startswith="PF-")
        .order_by("-numero")
        .first()
    )
    if ultimo:
        try:
            seq = int(ultimo.numero.split("-")[1]) + 1
        except (IndexError, ValueError):
            seq = 1
    else:
        seq = 1
    numero = f"PF-{seq:04d}"

    pf = DriverPreInvoice.objects.create(
        numero=numero,
        driver=driver,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        ajuste_manual=to_dec(body.get("ajuste_manual", 0)),
        penalizacoes_gerais=to_dec(body.get("penalizacoes_gerais", 0)),
        observacoes=body.get("observacoes", ""),
        created_by=request.user,
    )

    # ── Auto-popular linhas CAINIAO + bónus (regra: só Delivered) ──────
    # Cada login distinto vira UMA PreInvoiceLine própria (com seu
    # courier_id) e cada login tem os seus próprios bónus por dia.
    cainiao_partner = Partner.objects.filter(
        name__iexact="CAINIAO",
    ).first()
    if cainiao_partner:
        # Construir lista de logins únicos (cid, cname)
        logins = []
        seen_pairs = set()

        def _add_login(cid, cname, source):
            cid = (cid or "").strip()
            cname = (cname or "").strip()
            if not cid and not cname:
                return
            key = (cid, cname)
            if key in seen_pairs:
                return
            seen_pairs.add(key)
            logins.append({
                "courier_id": cid, "courier_name": cname,
                "source": source,
            })

        if driver.courier_id_cainiao or driver.apelido:
            _add_login(
                driver.courier_id_cainiao, driver.apelido, "perfil",
            )
        for m in driver.courier_mappings.filter(partner=cainiao_partner):
            _add_login(m.courier_id, m.courier_name, "mapping")

        if logins:
            from core.finance import resolve_driver_price
            price, _src = resolve_driver_price(driver, cainiao_partner)

            # PUDO: se activo, tasks com delivery_type=PUDO são pagas
            # pela fórmula 1ª + (N-1) × adicional por PUDO distinto.
            # São acumuladas em pudo_tasks_global e calculadas no fim
            # como uma única PreInvoiceLine agregada.
            pudo_active = bool(getattr(cainiao_partner, "pudo_enabled", False))
            pudo_tasks_global = []  # lista de tasks PUDO de todos os logins

            base_qs = CainiaoOperationTask.objects.filter(
                task_date__range=(periodo_inicio, periodo_fim),
                task_status="Delivered",  # ← REGRA: só Delivered
            )

            # Overrides de transferência + preço
            from .models import (
                WaybillAttributionOverride, PackagePriceOverride,
            )
            outgoing = set(
                WaybillAttributionOverride.objects.filter(
                    task_date__range=(periodo_inicio, periodo_fim),
                ).exclude(attributed_to_driver=driver)
                .values_list("waybill_number", flat=True),
            )
            incoming = list(
                WaybillAttributionOverride.objects.filter(
                    attributed_to_driver=driver,
                    task_date__range=(periodo_inicio, periodo_fim),
                ).values_list("waybill_number", flat=True),
            )
            price_overrides_map = {
                po.waybill_number: po
                for po in PackagePriceOverride.objects.filter(
                    task_date__range=(periodo_inicio, periodo_fim),
                )
            }

            seen_task_ids = set()  # evita dupla-contagem entre logins

            for ln in logins:
                cid = ln["courier_id"]
                cname = ln["courier_name"]
                login_q = Q()
                if cid:
                    login_q |= Q(courier_id_cainiao=cid)
                if cname:
                    login_q |= Q(courier_name=cname)
                if not login_q:
                    continue
                qs_login = base_qs.filter(login_q)
                if outgoing:
                    qs_login = qs_login.exclude(
                        waybill_number__in=outgoing,
                    )
                if seen_task_ids:
                    qs_login = qs_login.exclude(id__in=seen_task_ids)
                rows_this = list(
                    qs_login.values(
                        "id", "waybill_number", "delivery_type",
                        "receiver_latitude", "receiver_longitude",
                        "actual_latitude", "actual_longitude",
                        "zip_code", "detailed_address",
                    ),
                )
                if not rows_this:
                    continue
                ids_this = [r["id"] for r in rows_this]
                seen_task_ids.update(ids_this)

                login_label = cname or cid

                # Separar PUDO de Door (PUDO calculado depois,
                # agregado por todos os logins do driver).
                rows_door = []
                if pudo_active:
                    from settlements.services_pudo import _PudoTaskLite_dict_to_obj
                    for r in rows_this:
                        dt = (r.get("delivery_type") or "").upper().strip()
                        if dt == "PUDO":
                            pudo_tasks_global.append(
                                _PudoTaskLite_dict_to_obj(r)
                            )
                        else:
                            rows_door.append(r)
                else:
                    rows_door = rows_this

                if not rows_door:
                    # Tudo era PUDO — sem linha door para este login
                    continue

                # Agrupar por preço (base + cada override) — só Door
                price_groups = {}
                for r in rows_door:
                    po = price_overrides_map.get(r["waybill_number"])
                    unit_price = po.price if po else price
                    if unit_price not in price_groups:
                        price_groups[unit_price] = {
                            "n": 0, "reasons": set(),
                        }
                    price_groups[unit_price]["n"] += 1
                    if po and po.reason:
                        price_groups[unit_price]["reasons"].add(po.reason)

                # Criar 1 PreInvoiceLine por grupo de preço
                for unit_price, info in sorted(
                    price_groups.items(), reverse=True,
                ):
                    is_special = unit_price != price
                    obs_extra = ""
                    if is_special:
                        reasons_txt = ", ".join(info["reasons"]) or "—"
                        obs_extra = (
                            f" · Preço especial €{unit_price}: {reasons_txt}"
                        )
                    line = PreInvoiceLine(
                        pre_invoice=pf,
                        parceiro=cainiao_partner,
                        courier_id=cid or cname,
                        total_pacotes=info["n"],
                        taxa_por_entrega=unit_price,
                        dsr_percentual=Decimal("0"),
                        api_source=(
                            "auto:per_login_special"
                            if is_special else "auto:per_login"
                        ),
                        observacoes=(
                            f"Login Cainiao: {login_label} "
                            f"(via {ln['source']}). "
                            f"Só entregas Delivered.{obs_extra}"
                        ),
                    )
                    line.calcular_e_salvar()

                # Bónus por domingo/feriado (≥30 entregas) só desta login
                qs_login_only = CainiaoOperationTask.objects.filter(
                    id__in=ids_this,
                )
                for row in qs_login_only.values("task_date").annotate(
                    n=Count("id"),
                ).order_by("task_date"):
                    d = row["task_date"]
                    n = row["n"]
                    is_sun = d.weekday() == 6
                    h = Holiday.get_holiday(d)
                    if not (is_sun or h):
                        continue
                    if BonusBlackoutDate.is_blocked(d):
                        continue
                    if n < PreInvoiceBonus.LIMIAR_30:
                        continue
                    PreInvoiceBonus.objects.create(
                        pre_invoice=pf,
                        data=d,
                        tipo="FERIADO" if h else "DOMINGO",
                        qtd_entregas_elegiveis=n,
                        observacoes=(
                            f"{h.name if h else 'Domingo'} "
                            f"· Login: {login_label}"
                        ),
                    )

            # Linha de transferências recebidas
            if incoming:
                inc_qs = base_qs.filter(
                    waybill_number__in=incoming,
                )
                if seen_task_ids:
                    inc_qs = inc_qs.exclude(id__in=seen_task_ids)
                rows_inc = list(
                    inc_qs.values(
                        "id", "waybill_number", "delivery_type",
                        "receiver_latitude", "receiver_longitude",
                        "actual_latitude", "actual_longitude",
                        "zip_code", "detailed_address",
                    ),
                )
                if rows_inc:
                    ids_inc = [r["id"] for r in rows_inc]

                    # Separar PUDO de Door
                    if pudo_active:
                        from settlements.services_pudo import (
                            _PudoTaskLite_dict_to_obj,
                        )
                        rows_inc_door = []
                        for r in rows_inc:
                            dt = (r.get("delivery_type") or "").upper().strip()
                            if dt == "PUDO":
                                pudo_tasks_global.append(
                                    _PudoTaskLite_dict_to_obj(r)
                                )
                            else:
                                rows_inc_door.append(r)
                    else:
                        rows_inc_door = rows_inc

                    # Agrupar por preço (base + cada override) — só Door
                    price_groups_inc = {}
                    for r in rows_inc_door:
                        po = price_overrides_map.get(r["waybill_number"])
                        unit_price = po.price if po else price
                        if unit_price not in price_groups_inc:
                            price_groups_inc[unit_price] = {
                                "n": 0, "reasons": set(),
                            }
                        price_groups_inc[unit_price]["n"] += 1
                        if po and po.reason:
                            price_groups_inc[unit_price]["reasons"].add(
                                po.reason,
                            )

                    # Criar 1 PreInvoiceLine por grupo de preço
                    for unit_price, info in sorted(
                        price_groups_inc.items(), reverse=True,
                    ):
                        is_special = unit_price != price
                        obs_extra = ""
                        if is_special:
                            reasons_txt = (
                                ", ".join(info["reasons"]) or "—"
                            )
                            obs_extra = (
                                f" · Preço especial €{unit_price}: "
                                f"{reasons_txt}"
                            )
                        line_inc = PreInvoiceLine(
                            pre_invoice=pf,
                            parceiro=cainiao_partner,
                            courier_id="(transferências)",
                            total_pacotes=info["n"],
                            taxa_por_entrega=unit_price,
                            dsr_percentual=Decimal("0"),
                            api_source=(
                                "auto:transfer_in_special"
                                if is_special else "auto:transfer_in"
                            ),
                            observacoes=(
                                "Login: ↻ Transferências recebidas. "
                                "Pacotes entregues por outros drivers "
                                "atribuídos a este por override."
                                f"{obs_extra}"
                            ),
                        )
                        line_inc.calcular_e_salvar()

                    inc_only = CainiaoOperationTask.objects.filter(
                        id__in=ids_inc,
                    )
                    for row in inc_only.values("task_date").annotate(
                        n=Count("id"),
                    ).order_by("task_date"):
                        d = row["task_date"]
                        n = row["n"]
                        is_sun = d.weekday() == 6
                        h = Holiday.get_holiday(d)
                        if not (is_sun or h):
                            continue
                        if BonusBlackoutDate.is_blocked(d):
                            continue
                        if n < PreInvoiceBonus.LIMIAR_30:
                            continue
                        PreInvoiceBonus.objects.create(
                            pre_invoice=pf,
                            data=d,
                            tipo="FERIADO" if h else "DOMINGO",
                            qtd_entregas_elegiveis=n,
                            observacoes=(
                                f"{h.name if h else 'Domingo'} "
                                f"· Login: ↻ Transferências recebidas"
                            ),
                        )

            # ── PUDO: linha agregada por driver ────────────────────
            # Aplica fórmula 1ª + (N-1) × adicional por PUDO distinto.
            if pudo_active and pudo_tasks_global:
                from settlements.services_pudo import (
                    compute_pudo_total_for_driver,
                )
                pudo_total, n_pudos, _bd = compute_pudo_total_for_driver(
                    pudo_tasks_global, cainiao_partner,
                )
                n_pudo_pkgs = len(pudo_tasks_global)
                # Taxa "média" para o modelo (que multiplica
                # total_pacotes × taxa). Guardamos em 4 decimais para
                # preservar o total exacto.
                avg_rate = (
                    pudo_total / Decimal(n_pudo_pkgs)
                    if n_pudo_pkgs else Decimal("0")
                ).quantize(Decimal("0.0001"))
                # Pequeno ajuste: garantir que pacotes × taxa = total
                # (arredondamento). Se diferir, somamos a diferença.
                line_pudo = PreInvoiceLine(
                    pre_invoice=pf,
                    parceiro=cainiao_partner,
                    courier_id="(PUDO)",
                    total_pacotes=n_pudo_pkgs,
                    taxa_por_entrega=avg_rate,
                    dsr_percentual=Decimal("0"),
                    api_source="auto:pudo",
                    observacoes=(
                        f"PUDO: {n_pudo_pkgs} pacote{'s' if n_pudo_pkgs > 1 else ''} "
                        f"em {n_pudos} PUDO{'s' if n_pudos > 1 else ''} distinto"
                        f"{'s' if n_pudos > 1 else ''}. "
                        f"1ª: €{cainiao_partner.pudo_first_delivery_price} · "
                        f"adicional: €{cainiao_partner.pudo_additional_delivery_price} · "
                        f"total: €{pudo_total:.2f}"
                    )[:300],
                )
                line_pudo.calcular_e_salvar()

    pf.recalcular()

    # Auto-inclui DriverClaim APPROVED do período como pacotes perdidos
    from .services_claims_in_pf import auto_include_approved_claims
    claims_result = auto_include_approved_claims(pf)
    if claims_result["included"]:
        pf.recalcular()  # recalcular para somar os novos pacotes perdidos

    # Auto-inclui TODOS os PreInvoiceAdvance PENDENTE do motorista —
    # limpa a conta-corrente para esta PF. Se a PF for cancelada/apagada
    # antes de ser paga, são automaticamente libertados (ver
    # DriverPreInvoice.save/delete).
    from .cash_entry_services import auto_attach_all_pending
    attached_count = auto_attach_all_pending(pf)
    if attached_count:
        pf.recalcular()

    return JsonResponse({
        "success": True, "numero": pf.numero, "id": pf.id,
        "total_a_receber": str(pf.total_a_receber),
        "claims_auto_included": claims_result["included"],
        "advances_auto_attached": attached_count,
    })


@login_required
@require_http_methods(["POST"])
def driver_pre_invoice_update(request, pre_invoice_id):
    """Atualiza campos da pré-fatura."""
    import json
    from decimal import Decimal, InvalidOperation

    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    def to_dec(val, default="0.00"):
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError):
            return Decimal(default)

    if "ajuste_manual" in body:
        pf.ajuste_manual = to_dec(body["ajuste_manual"])
    if "penalizacoes_gerais" in body:
        pf.penalizacoes_gerais = to_dec(body["penalizacoes_gerais"])
    if "observacoes" in body:
        pf.observacoes = body["observacoes"]
    if "status" in body:
        pf.status = body["status"]
    if "data_pagamento" in body and body["data_pagamento"]:
        pf.data_pagamento = body["data_pagamento"]
    if "referencia_pagamento" in body:
        pf.referencia_pagamento = body["referencia_pagamento"]

    pf.recalcular()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def pre_invoice_add_bonus(request, pre_invoice_id):
    """Adiciona uma bonificação domingo/feriado."""
    import json
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    bonus = PreInvoiceBonus.objects.create(
        pre_invoice=pf,
        data=body.get("data"),
        tipo=body.get("tipo", "DOMINGO"),
        qtd_entregas_elegiveis=int(body.get("qtd_entregas_elegiveis", 0) or 0),
        observacoes=body.get("observacoes", ""),
    )
    pf.recalcular()
    return JsonResponse({
        "success": True,
        "bonus_calculado": str(bonus.bonus_calculado),
        "id": bonus.id,
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_delete_bonus(request, bonus_id):
    """Remove uma bonificação."""
    bonus = get_object_or_404(PreInvoiceBonus, id=bonus_id)
    pf = bonus.pre_invoice
    bonus.delete()
    pf.recalcular()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def pre_invoice_add_lost_package(request, pre_invoice_id):
    """Adiciona um pacote perdido."""
    import json
    from decimal import Decimal, InvalidOperation
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    try:
        valor = Decimal(str(body.get("valor", "50.00")))
    except (InvalidOperation, TypeError):
        valor = Decimal("50.00")

    try:
        iva_percentual = Decimal(str(body.get("iva_percentual", "0")))
    except (InvalidOperation, TypeError):
        iva_percentual = Decimal("0")

    pkg = PreInvoiceLostPackage.objects.create(
        pre_invoice=pf,
        data=body.get("data") or None,
        numero_pacote=body.get("numero_pacote", ""),
        descricao=body.get("descricao", ""),
        valor=valor,
        iva_percentual=iva_percentual,
        observacoes=body.get("observacoes", ""),
    )
    pf.recalcular()
    return JsonResponse({
        "success": True,
        "id": pkg.id,
        "valor": str(pkg.valor),
        "iva_percentual": str(pkg.iva_percentual),
        "valor_com_iva": str(pkg.valor_com_iva),
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_delete_lost_package(request, package_id):
    """Remove um pacote perdido."""
    pkg = get_object_or_404(PreInvoiceLostPackage, id=package_id)
    pf = pkg.pre_invoice
    pkg.delete()
    pf.recalcular()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def pre_invoice_add_advance(request, pre_invoice_id):
    """Adiciona adiantamento/combustível.

    Body extras:
      paid_by_source: "EMPRESA" (default) | "TERCEIRO"
      paid_by_lender_id: int (obrigatório quando paid_by_source=="TERCEIRO")

    Quando TERCEIRO, o save() do model cria automaticamente um
    ThirdPartyReimbursement PENDENTE para o sócio.
    """
    import json
    from decimal import Decimal, InvalidOperation
    from .models import Shareholder
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    try:
        valor = Decimal(str(body.get("valor", 0)))
    except (InvalidOperation, TypeError):
        valor = Decimal("0.00")

    paid_by_source = body.get("paid_by_source", "EMPRESA")
    if paid_by_source not in ("EMPRESA", "TERCEIRO"):
        paid_by_source = "EMPRESA"

    paid_by_lender = None
    if paid_by_source == "TERCEIRO":
        lender_id = body.get("paid_by_lender_id")
        if not lender_id:
            return JsonResponse(
                {"success": False,
                 "error": "paid_by_lender_id é obrigatório "
                          "quando paid_by_source=='TERCEIRO'"},
                status=400,
            )
        paid_by_lender = Shareholder.objects.filter(
            id=lender_id, ativo=True,
        ).first()
        if not paid_by_lender:
            return JsonResponse(
                {"success": False,
                 "error": "Sócio inválido ou inativo"},
                status=400,
            )

    adv = PreInvoiceAdvance.objects.create(
        driver=pf.driver,
        pre_invoice=pf,
        status="INCLUIDO_PF",  # criado já dentro da PF (fluxo manual antigo)
        data=body.get("data") or None,
        tipo=body.get("tipo", "ADIANTAMENTO"),
        descricao=body.get("descricao", ""),
        valor=valor,
        documento_referencia=body.get("documento_referencia", ""),
        paid_by_source=paid_by_source,
        paid_by_lender=paid_by_lender,
    )
    pf.recalcular()
    reembolso = adv.reembolsos_terceiros.filter(status="PENDENTE").first()
    return JsonResponse({
        "success": True,
        "id": adv.id,
        "reembolso_id": reembolso.id if reembolso else None,
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_delete_advance(request, advance_id):
    """Remove um adiantamento."""
    adv = get_object_or_404(PreInvoiceAdvance, id=advance_id)
    pf = adv.pre_invoice
    adv.delete()
    pf.recalcular()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def pre_invoice_line_add(request, pre_invoice_id):
    """Adiciona uma linha de trabalho (parceiro + pacotes) à pré-fatura."""
    import json
    from decimal import Decimal, InvalidOperation
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    def to_dec(val, default="0.00"):
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError):
            return Decimal(default)

    from core.models import Partner
    parceiro = None
    pid = body.get("parceiro_id")
    if pid:
        parceiro = Partner.objects.filter(id=pid).first()

    linha = PreInvoiceLine(
        pre_invoice=pf,
        parceiro=parceiro,
        courier_id=body.get("courier_id", ""),
        total_pacotes=int(body.get("total_pacotes", 0) or 0),
        taxa_por_entrega=to_dec(body.get("taxa_por_entrega", 0)),
        dsr_percentual=to_dec(body.get("dsr_percentual", 0)),
        observacoes=body.get("observacoes", ""),
    )
    linha.calcular_e_salvar()  # calcula base_entregas e chama recalcular() na pré-fatura

    return JsonResponse({
        "success": True,
        "id": linha.id,
        "base_entregas": str(linha.base_entregas),
        "total_a_receber": str(pf.total_a_receber),
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_line_update(request, line_id):
    """Actualiza uma linha de trabalho."""
    import json
    from decimal import Decimal, InvalidOperation
    linha = get_object_or_404(PreInvoiceLine, id=line_id)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    def to_dec(val, default="0.00"):
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError):
            return Decimal(default)

    if "parceiro_id" in body:
        from core.models import Partner
        pid = body["parceiro_id"]
        linha.parceiro = Partner.objects.filter(id=pid).first() if pid else None
    if "courier_id" in body:
        linha.courier_id = body["courier_id"]
    if "total_pacotes" in body:
        linha.total_pacotes = int(body.get("total_pacotes", 0) or 0)
    if "taxa_por_entrega" in body:
        linha.taxa_por_entrega = to_dec(body["taxa_por_entrega"])
    if "dsr_percentual" in body:
        linha.dsr_percentual = to_dec(body["dsr_percentual"])
    if "observacoes" in body:
        linha.observacoes = body["observacoes"]

    linha.calcular_e_salvar()
    pf = linha.pre_invoice

    return JsonResponse({
        "success": True,
        "base_entregas": str(linha.base_entregas),
        "total_a_receber": str(pf.total_a_receber),
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_line_delete(request, line_id):
    """Remove uma linha de trabalho."""
    linha = get_object_or_404(PreInvoiceLine, id=line_id)
    pf = linha.pre_invoice
    linha.delete()
    pf.recalcular()
    return JsonResponse({"success": True, "total_a_receber": str(pf.total_a_receber)})


@login_required
@require_http_methods(["POST"])
def driver_pre_invoice_recalculate(request, pre_invoice_id):
    """Força recálculo completo de uma pré-fatura (incluindo comissões de indicação)."""
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)
    pf.recalcular()
    return JsonResponse({
        "success": True,
        "total_a_receber": str(pf.total_a_receber),
        "base_entregas": str(pf.base_entregas),
        "total_bonus": str(pf.total_bonus),
        "penalizacoes_gerais": str(pf.penalizacoes_gerais),
        "total_pacotes_perdidos": str(pf.total_pacotes_perdidos),
        "total_adiantamentos": str(pf.total_adiantamentos),
        "subtotal_bruto": str(pf.subtotal_bruto),
        "total_comissoes_indicacao": str(pf.total_comissoes_indicacao),
        "comissoes_indicacao_detalhe": _get_referral_commission_detail(pf),
    })


@login_required
@require_http_methods(["POST"])
def driver_pre_invoice_delete(request, pre_invoice_id):
    """Elimina uma pré-fatura e todas as suas linhas."""
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)
    numero = pf.numero
    pf.delete()
    return JsonResponse({"success": True, "numero": numero})


@login_required
@require_http_methods(["POST"])
def pre_invoice_upload_fatura(request, pre_invoice_id):
    """Faz upload da fatura emitida pelo motorista para a pré-fatura."""
    pf = get_object_or_404(DriverPreInvoice, id=pre_invoice_id)
    ficheiro = request.FILES.get("fatura")
    if not ficheiro:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)
    pf.fatura_ficheiro = ficheiro
    pf.save(update_fields=["fatura_ficheiro"])
    return JsonResponse({"success": True, "fatura_url": pf.fatura_ficheiro.url})


@login_required
def driver_pre_invoice_pdf(request, pre_invoice_id):
    """Gera e devolve o PDF de uma pré-fatura."""
    from django.http import HttpResponse
    pf = get_object_or_404(
        DriverPreInvoice.objects.select_related("driver")
        .prefetch_related(
            "linhas__parceiro",
            "bonificacoes",
            "pacotes_perdidos",
            "adiantamentos",
        ),
        id=pre_invoice_id,
    )

    try:
        generator = PDFGenerator()
        pdf_buffer = generator.generate_pre_invoice_pdf(pf)
        filename = f"PreFatura_{pf.numero}_{pf.driver.nome_completo.replace(' ', '_')}.pdf"
        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        return JsonResponse(
            {"error": "ReportLab não está instalado no servidor."},
            status=500,
        )


# ============================================================================
# PRÉ-FATURAS — EMPRESA PARCEIRA LANCAMENTO STATUS CHANGE
# ============================================================================


@login_required
@require_http_methods(["POST"])
def empresa_lancamento_change_status(request, lancamento_id):
    """Muda o status de um lançamento de empresa parceira."""
    import json
    from drivers_app.models import EmpresaParceiraLancamento

    lanc = get_object_or_404(EmpresaParceiraLancamento, id=lancamento_id)

    TRANSICOES = {
        "RASCUNHO":  ["APROVADO", "CANCELADO"],
        "APROVADO":  ["PENDENTE", "CANCELADO", "RASCUNHO"],
        "PENDENTE":  ["PAGO", "APROVADO"],
        "PAGO":      [],
        "CANCELADO": ["RASCUNHO"],
    }

    # Aceita JSON ou multipart (para upload de comprovante)
    if request.content_type and "application/json" in request.content_type:
        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            body = {}
    else:
        body = request.POST.dict()

    novo_status = (body.get("status") or "").strip().upper()
    nota = (body.get("nota") or "").strip()

    permitidos = TRANSICOES.get(lanc.status, [])
    if novo_status not in permitidos:
        return JsonResponse(
            {"success": False, "error": f"Transição {lanc.status} → {novo_status} não permitida."},
            status=400,
        )

    lanc.status = novo_status

    if novo_status == "PAGO":
        from datetime import date
        lanc.data_pagamento = lanc.data_pagamento or date.today()
        ref = body.get("referencia_pagamento", "").strip()
        if ref:
            lanc.referencia_pagamento = ref
        if request.FILES.get("comprovante"):
            lanc.comprovante_pagamento = request.FILES["comprovante"]

    if nota:
        lanc.notas = (lanc.notas + f"\n[{novo_status}] {nota}").strip()

    lanc.save()

    status_labels = dict(EmpresaParceiraLancamento.STATUS_CHOICES)
    return JsonResponse({
        "success": True,
        "status": lanc.status,
        "status_display": status_labels.get(lanc.status, lanc.status),
        "comprovante_url": lanc.comprovante_pagamento.url if lanc.comprovante_pagamento else "",
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_bulk_status(request):
    """Marca várias pré-faturas de motoristas como PAGO em lote."""
    import json
    from datetime import date

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST.dict()

    ids = body.get("ids", [])
    novo_status = (body.get("status") or "PAGO").strip().upper()
    ref = (body.get("referencia_pagamento") or "").strip()
    nota = (body.get("nota") or "").strip()

    if not ids:
        return JsonResponse({"success": False, "error": "Nenhum ID fornecido."}, status=400)

    updated = 0
    for pf_id in ids:
        try:
            pf = DriverPreInvoice.objects.get(id=pf_id)
            permitidos = DriverPreInvoice.TRANSICOES.get(pf.status, [])
            if novo_status in permitidos:
                pf.status = novo_status
                if novo_status == "PAGO":
                    pf.data_pagamento = pf.data_pagamento or date.today()
                    if ref:
                        pf.referencia_pagamento = ref
                if nota:
                    pf.observacoes = (pf.observacoes + f"\n[{novo_status}] {nota}").strip()
                pf.save()
                updated += 1
        except DriverPreInvoice.DoesNotExist:
            pass

    return JsonResponse({"success": True, "updated": updated})


# ============================================================================
# IMPORTAÇÃO DE PLANILHAS — Geração automática de Pré-Faturas
# ============================================================================

def _get_referral_commission_detail(pf):
    """Devolve detalhe das comissões de indicação para uma pré-fatura."""
    from decimal import Decimal
    result = []
    for ref in pf.driver.referrals_given.filter(ativo=True):
        referred_pfs = DriverPreInvoice.objects.filter(
            driver=ref.referred,
            periodo_inicio=pf.periodo_inicio,
            periodo_fim=pf.periodo_fim,
        )
        for rpf in referred_pfs:
            total_pcts = sum(l.total_pacotes for l in rpf.linhas.all())
            valor = Decimal(total_pcts) * ref.comissao_por_pacote
            result.append({
                "referred_nome": ref.referred.nome_completo,
                "total_pacotes": total_pcts,
                "comissao_por_pacote": str(ref.comissao_por_pacote),
                "valor_total": str(valor),
            })
    return result


def _gerar_numero_pre_fatura():
    """Gera o próximo número sequencial único para DriverPreInvoice (ex: PF-0042)."""
    ultimo = (
        DriverPreInvoice.objects.filter(numero__startswith="PF-")
        .order_by("-numero")
        .first()
    )
    if ultimo:
        try:
            seq = int(ultimo.numero.split("-")[1]) + 1
        except (IndexError, ValueError):
            seq = 1
    else:
        seq = 1
    return f"PF-{seq:04d}"


def _gerar_nif_placeholder():
    """Gera um NIF placeholder único (999XXXXXX) para drivers auto-importados."""
    from drivers_app.models import DriverProfile
    import random
    for _ in range(100):
        candidate = f"999{random.randint(100000, 999999)}"
        if not DriverProfile.objects.filter(nif=candidate).exists():
            return candidate
    raise ValueError("Não foi possível gerar NIF placeholder único.")


def _parse_cainiao_sheet(workbook):
    """
    Parseia um ficheiro Cainiao (xlsx).

    Colunas Cainiao:
      Courier ID        — identificador único do courier
      Courier Name      — login/nome no sistema Cainiao
      DSP Name          — nome da DSP (Leguas Franzinas)
      Total Parcels     — pacotes atribuídos e tentados
      Delivery Success Rate — DSR em % (ex: "94%")
      Signed Parcels    — pacotes com entrega confirmada (base de faturação)
      Courier Status    — Enable / Disable
      Dispatch Date     — período (ex: "2026-03-16-2026-03-31")

    Retorna lista de dicts com todos os campos acima.
    """
    import calendar
    from datetime import date, datetime as dt
    from decimal import Decimal as D

    ws = workbook.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Localizar linha de header (contém "Courier ID")
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        row_vals = [str(v).strip().lower() if v else "" for v in row]
        if any("courier id" in v for v in row_vals):
            header_row = row_vals
            header_idx = i
            break

    if header_row is None:
        return [], []

    found_headers = [h for h in header_row if h]

    def col(name):
        for j, h in enumerate(header_row):
            if name.lower() in h:
                return j
        return None

    col_id     = col("courier id")
    col_name   = col("courier name")
    col_total  = col("total parcels")
    col_dsr    = col("delivery success rate")
    col_signed = col("signed parcels")
    col_status = col("courier status")
    col_date   = col("dispatch date")

    if col_id is None:
        return []

    results = []
    for row in rows[header_idx + 1:]:
        if not row or not row[col_id]:
            continue

        courier_id   = str(row[col_id]).strip()
        courier_name = str(row[col_name]).strip() if col_name is not None and row[col_name] else ""

        # Total Parcels — tentados/atribuídos (referência, não base de faturação)
        total_parcels = 0
        if col_total is not None and row[col_total]:
            try:
                total_parcels = int(float(str(row[col_total]).replace(",", ".")))
            except (ValueError, TypeError):
                total_parcels = 0

        # Delivery Success Rate — ex: "94%" → 94.00
        dsr_percentual = D("0")
        if col_dsr is not None and row[col_dsr]:
            raw = str(row[col_dsr]).strip().replace("%", "").replace(",", ".")
            try:
                dsr_percentual = D(raw)
            except Exception:
                dsr_percentual = D("0")

        # Signed Parcels — base de faturação (pacotes entregues com sucesso)
        signed_parcels = 0
        if col_signed is not None and row[col_signed]:
            try:
                signed_parcels = int(float(str(row[col_signed]).replace(",", ".")))
            except (ValueError, TypeError):
                signed_parcels = 0

        courier_status = str(row[col_status]).strip() if col_status is not None and row[col_status] else "Enable"

        # Dispatch Date → período: "2026-03-16-2026-03-31"
        periodo_inicio = periodo_fim = None
        if col_date is not None and row[col_date]:
            raw_val = row[col_date]
            # openpyxl may return a datetime/date object for date cells
            if isinstance(raw_val, dt):
                raw_val = raw_val.date()
            if isinstance(raw_val, date):
                # Single date — use first and last day of that month
                periodo_inicio = raw_val.replace(day=1)
                last_day = calendar.monthrange(raw_val.year, raw_val.month)[1]
                periodo_fim = raw_val.replace(day=last_day)
            else:
                date_str = str(raw_val).strip()
                # Formato Cainiao: "2026-03-16~2026-03-31" (separador tilde)
                if "~" in date_str:
                    tilde_parts = date_str.split("~")
                    if len(tilde_parts) == 2:
                        try:
                            periodo_inicio = date.fromisoformat(
                                tilde_parts[0].strip()
                            )
                            periodo_fim = date.fromisoformat(
                                tilde_parts[1].strip()
                            )
                        except ValueError:
                            pass
                else:
                    # Outros formatos: "2026-03-16-2026-03-31" (6 partes) ou
                    # "2026-03-16 - 2026-03-31" (espaços à volta do traço)
                    parts = date_str.replace(" ", "").split("-")
                    if len(parts) == 6:
                        try:
                            periodo_inicio = date(
                                int(parts[0]), int(parts[1]), int(parts[2])
                            )
                            periodo_fim = date(
                                int(parts[3]), int(parts[4]), int(parts[5])
                            )
                        except (ValueError, IndexError):
                            pass
                    elif len(parts) == 3:
                        try:
                            d = date(int(parts[0]), int(parts[1]), int(parts[2]))
                            periodo_inicio = d.replace(day=1)
                            last_day = calendar.monthrange(d.year, d.month)[1]
                            periodo_fim = d.replace(day=last_day)
                        except (ValueError, IndexError):
                            pass

        # Fallback: se col_date não foi encontrado ou parsing falhou,
        # percorrer todas as células da linha à procura de um intervalo de datas
        if periodo_inicio is None:
            import re
            date_range_re = re.compile(
                r'(\d{4}[/\-]\d{2}[/\-]\d{2})\s*[-–]\s*(\d{4}[/\-]\d{2}[/\-]\d{2})'
            )
            for cell_val in row:
                if not cell_val or not isinstance(cell_val, str):
                    continue
                m = date_range_re.search(cell_val.strip())
                if m:
                    try:
                        s = m.group(1).replace("/", "-")
                        e = m.group(2).replace("/", "-")
                        periodo_inicio = date.fromisoformat(s)
                        periodo_fim = date.fromisoformat(e)
                        break
                    except ValueError:
                        pass

        # Guardar o valor raw da célula de data para depuração
        raw_date_val = (
            str(row[col_date]) if col_date is not None and row[col_date] else None
        )

        results.append({
            "courier_id":     courier_id,
            "courier_name":   courier_name,
            "total_parcels":  total_parcels,    # tentados (para referência)
            "dsr_percentual": dsr_percentual,   # lido directamente da planilha
            "signed_parcels": signed_parcels,   # entregues — base de faturação
            "periodo_inicio": periodo_inicio,
            "periodo_fim":    periodo_fim,
            "courier_status": courier_status,
            "_raw_date":      raw_date_val,
        })
    return results, found_headers


@login_required
@require_http_methods(["POST"])
def import_partner_sheet_preview(request, partner_id):
    """
    Passo 1 da importação: parseia o ficheiro e devolve preview de mapeamento de drivers.
    Não cria nada na BD — apenas lê e compara com DriverCourierMapping existentes.
    """
    from core.models import Partner
    from drivers_app.models import DriverProfile

    partner = get_object_or_404(Partner, id=partner_id)
    formato = request.POST.get("formato", "CAINIAO")

    uploaded_file = request.FILES.get("planilha")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao abrir ficheiro: {e}"}, status=400)

    if formato == "CAINIAO":
        rows, found_headers = _parse_cainiao_sheet(wb)
    else:
        return JsonResponse(
            {"success": False, "error": f"Formato '{formato}' não suportado ainda."},
            status=400,
        )

    if not rows:
        return JsonResponse(
            {"success": False, "error": f"Nenhuma linha encontrada. Colunas: {found_headers}"},
            status=400,
        )

    # Carregar todos os mapeamentos existentes para este parceiro
    existing_mappings = {
        m.courier_id: {"driver_id": m.driver_id, "driver_nome": m.driver.nome_completo}
        for m in DriverCourierMapping.objects.filter(partner=partner).select_related("driver")
    }

    # Lista de todos os motoristas activos para o selector
    all_drivers = list(
        DriverProfile.objects.filter(is_active=True)
        .values("id", "nome_completo")
        .order_by("nome_completo")
    )

    # Construir linhas de preview (agrupa por courier_id para evitar duplicados por período)
    seen = {}
    for row in rows:
        cid = row["courier_id"]
        if cid in seen:
            continue
        mapping = existing_mappings.get(cid)
        seen[cid] = {
            "courier_id": cid,
            "courier_name": row["courier_name"],
            "signed_parcels": row["signed_parcels"],
            "total_parcels": row["total_parcels"],
            "dsr": row["dsr_percentual"],
            "periodo": f"{row['periodo_inicio']} → {row['periodo_fim']}" if row["periodo_inicio"] else "—",
            "matched": mapping is not None,
            "driver_id": mapping["driver_id"] if mapping else None,
            "driver_nome": mapping["driver_nome"] if mapping else None,
        }

    return JsonResponse({
        "success": True,
        "rows": list(seen.values()),
        "all_drivers": all_drivers,
        "total_rows": len(rows),
    })


@login_required
@require_http_methods(["POST"])
def import_partner_sheet(request, partner_id):
    """
    Importa planilha de um parceiro e gera/actualiza pré-faturas automaticamente.
    Suporta: Cainiao (xlsx).
    """
    import json
    from decimal import Decimal
    from core.models import Partner
    from drivers_app.models import DriverProfile

    partner = get_object_or_404(Partner, id=partner_id)
    formato = request.POST.get("formato", "CAINIAO")
    raw_taxa = (request.POST.get("taxa_por_entrega", "0") or "0").replace(",", ".")
    try:
        taxa_por_entrega = Decimal(raw_taxa)
    except Exception:
        taxa_por_entrega = Decimal("0")

    # Overrides manuais: {courier_id: driver_id} escolhidos na etapa de preview
    # Pode vir como JSON no campo "driver_overrides" ou como campos individuais "override_<courier_id>"
    driver_overrides = {}
    raw_overrides = request.POST.get("driver_overrides", "")
    if raw_overrides:
        try:
            driver_overrides = json.loads(raw_overrides)
        except (json.JSONDecodeError, ValueError):
            pass

    uploaded_file = request.FILES.get("planilha")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "Nenhum ficheiro enviado."}, status=400)

    # Parse
    try:
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erro ao abrir ficheiro: {e}"}, status=400)

    if formato == "CAINIAO":
        rows, found_headers = _parse_cainiao_sheet(wb)
    else:
        return JsonResponse(
            {"success": False, "error": f"Formato '{formato}' não suportado ainda."},
            status=400,
        )

    if not rows:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    "Nenhuma linha encontrada no ficheiro. "
                    f"Colunas detectadas: {found_headers}"
                ),
            },
            status=400,
        )

    created_drivers = []
    created_invoices = []
    updated_invoices = []
    errors = []

    for row in rows:
        courier_id     = row["courier_id"]
        courier_name   = row["courier_name"]
        signed_parcels = row["signed_parcels"]   # base de faturação
        total_parcels  = row["total_parcels"]    # tentados (referência)
        dsr_planilha   = row["dsr_percentual"]   # DSR lido da planilha
        periodo_inicio = row["periodo_inicio"]
        periodo_fim    = row["periodo_fim"]

        # Ignorar linhas sem pacotes entregues
        if not signed_parcels:
            continue

        # Ignorar linhas sem período
        if not periodo_inicio or not periodo_fim:
            raw_date = row.get("_raw_date")
            hint = f" (valor na célula: {raw_date!r})" if raw_date else \
                   f" (colunas: {found_headers})"
            errors.append(
                f"{courier_id}: período não encontrado na planilha.{hint}"
            )
            continue

        # 1. Encontrar ou criar driver
        mapping = DriverCourierMapping.objects.filter(
            partner=partner, courier_id=courier_id
        ).select_related("driver").first()

        # Override manual do utilizador tem prioridade
        override_driver_id = driver_overrides.get(str(courier_id))
        if override_driver_id:
            try:
                override_driver = DriverProfile.objects.get(id=int(override_driver_id))
                if mapping:
                    # Redirigir mapping existente para o driver escolhido
                    mapping.driver = override_driver
                    mapping.courier_name = courier_name
                    mapping.save()
                else:
                    mapping = DriverCourierMapping.objects.create(
                        driver=override_driver,
                        partner=partner,
                        courier_id=courier_id,
                        courier_name=courier_name,
                    )
                driver = override_driver
            except DriverProfile.DoesNotExist:
                errors.append(f"{courier_id}: driver override ID {override_driver_id} não encontrado.")
                continue
        elif mapping:
            driver = mapping.driver
            # Actualizar nome se mudou
            if mapping.courier_name != courier_name:
                mapping.courier_name = courier_name
                mapping.save(update_fields=["courier_name"])
        else:
            # Criar driver stub
            try:
                nif_ph = _gerar_nif_placeholder()
                driver = DriverProfile.objects.create(
                    nif=nif_ph,
                    nome_completo=courier_name or f"Driver {courier_id}",
                    telefone="000000000",
                    email=f"driver.{courier_id}@import.local",
                    status="PENDENTE",
                    is_active=False,
                    importado_auto=True,
                )
                DriverCourierMapping.objects.create(
                    driver=driver,
                    partner=partner,
                    courier_id=courier_id,
                    courier_name=courier_name,
                )
                created_drivers.append(courier_name or courier_id)
            except Exception as e:
                errors.append(f"{courier_id}: erro ao criar driver — {e}")
                continue

        # 2. Encontrar ou criar pré-fatura para o período
        pf = DriverPreInvoice.objects.filter(
            driver=driver,
            periodo_inicio=periodo_inicio,
            periodo_fim=periodo_fim,
        ).first()

        pf_created = False
        if not pf:
            pf = DriverPreInvoice.objects.create(
                driver=driver,
                numero=_gerar_numero_pre_fatura(),
                periodo_inicio=periodo_inicio,
                periodo_fim=periodo_fim,
                status="RASCUNHO",
            )
            pf_created = True

        # 3. Encontrar ou criar linha para este parceiro nesta pré-fatura
        #    total_pacotes = Signed Parcels (entregues com sucesso — base de faturação)
        #    dsr_percentual = Delivery Success Rate lido directamente da planilha
        #    observacoes = regista Total Parcels para referência
        obs = f"Total atribuídos: {total_parcels} | Entregues: {signed_parcels} | DSR: {dsr_planilha}%"

        linha = PreInvoiceLine.objects.filter(
            pre_invoice=pf,
            parceiro=partner,
        ).first()

        if linha:
            linha.total_pacotes    = signed_parcels
            linha.taxa_por_entrega = taxa_por_entrega
            linha.dsr_percentual   = dsr_planilha
            linha.courier_id       = courier_id
            linha.api_source       = formato
            linha.observacoes      = obs
            linha.save()
            linha.calcular_e_salvar()
            updated_invoices.append(pf.numero)
        else:
            linha = PreInvoiceLine.objects.create(
                pre_invoice=pf,
                parceiro=partner,
                courier_id=courier_id,
                total_pacotes=signed_parcels,
                taxa_por_entrega=taxa_por_entrega,
                dsr_percentual=dsr_planilha,
                api_source=formato,
                observacoes=obs,
            )
            linha.calcular_e_salvar()
            if pf_created:
                created_invoices.append(pf.numero)
            else:
                updated_invoices.append(pf.numero)

    return JsonResponse({
        "success": True,
        "linhas_processadas": len(rows),
        "drivers_criados": len(created_drivers),
        "drivers_criados_lista": created_drivers,
        "prefaturas_criadas": len(created_invoices),
        "prefaturas_atualizadas": len(updated_invoices),
        "erros": errors,
    })


# ============================================================================
# RELATÓRIO DE PRÉ-FATURAS — MODAL + PDF RESUMO
# ============================================================================

def _build_report_qs(request):
    """Aplica os mesmos filtros da listagem ao queryset."""
    from decimal import Decimal
    from drivers_app.models import DriverReferral

    qs = (
        DriverPreInvoice.objects.select_related("driver")
        .prefetch_related("linhas__parceiro", "bonificacoes",
                          "pacotes_perdidos", "adiantamentos")
        .order_by("driver__nome_completo", "-periodo_fim")
    )
    status_f   = request.GET.get("status", "").strip()
    driver_f   = request.GET.get("driver", "").strip()
    date_from  = request.GET.get("date_from", "").strip()
    date_to    = request.GET.get("date_to", "").strip()
    parceiro_f = request.GET.get("parceiro", "").strip()
    if status_f:
        qs = qs.filter(status=status_f)
    if driver_f:
        qs = qs.filter(driver__nome_completo__icontains=driver_f)
    if date_from:
        qs = qs.filter(periodo_inicio__gte=date_from)
    if date_to:
        qs = qs.filter(periodo_fim__lte=date_to)
    if parceiro_f:
        qs = qs.filter(linhas__parceiro_id=parceiro_f).distinct()
    return qs


@login_required
def pre_invoice_report_summary(request):
    """Devolve JSON com totais e detalhe por motorista + empresas parceiras."""
    from decimal import Decimal
    from drivers_app.models import EmpresaParceiraLancamento

    qs = _build_report_qs(request)

    total_pfs = 0
    total_pacotes = 0
    total_base = Decimal("0.00")
    total_bonus = Decimal("0.00")
    total_comissoes = Decimal("0.00")
    total_descontos = Decimal("0.00")
    total_adiantamentos = Decimal("0.00")
    total_a_receber = Decimal("0.00")

    motoristas = []
    for pf in qs:
        total_pfs += 1
        pcts = sum(l.total_pacotes for l in pf.linhas.all())
        total_pacotes += pcts
        total_base += pf.base_entregas
        total_bonus += pf.total_bonus
        total_comissoes += pf.total_comissoes_indicacao
        total_descontos += pf.penalizacoes_gerais + pf.total_pacotes_perdidos
        total_adiantamentos += pf.total_adiantamentos
        total_a_receber += pf.total_a_receber

        parceiros_str = ", ".join(
            sorted(set(l.parceiro.name for l in pf.linhas.all() if l.parceiro))
        ) or "—"
        motoristas.append({
            "nome": pf.driver.nome_completo,
            "numero": pf.numero,
            "periodo": f"{pf.periodo_inicio.strftime('%d/%m/%Y')} → {pf.periodo_fim.strftime('%d/%m/%Y')}",
            "parceiros": parceiros_str,
            "pacotes": pcts,
            "base_entregas": str(pf.base_entregas),
            "total_bonus": str(pf.total_bonus),
            "total_comissoes": str(pf.total_comissoes_indicacao),
            "descontos": str(pf.penalizacoes_gerais + pf.total_pacotes_perdidos),
            "adiantamentos": str(pf.total_adiantamentos),
            "total_a_receber": str(pf.total_a_receber),
            "status": pf.get_status_display(),
            "status_code": pf.status,
        })

    # ── Lançamentos de Empresas Parceiras ────────────────────────────────
    lqs = (
        EmpresaParceiraLancamento.objects
        .select_related("empresa")
        .exclude(status__in=["CANCELADO", "RASCUNHO"])
        .order_by("empresa__nome", "-periodo_inicio")
    )
    date_from = request.GET.get("date_from", "").strip()
    date_to   = request.GET.get("date_to", "").strip()
    if date_from:
        lqs = lqs.filter(periodo_inicio__gte=date_from)
    if date_to:
        lqs = lqs.filter(periodo_fim__lte=date_to)

    total_empresas = Decimal("0.00")
    lancamentos_data = []
    for l in lqs:
        valor = l.total_a_receber
        total_empresas += valor
        lancamentos_data.append({
            "empresa": l.empresa.nome if l.empresa else "—",
            "descricao": l.descricao or "—",
            "periodo": (
                f"{l.periodo_inicio.strftime('%d/%m/%Y')} → {l.periodo_fim.strftime('%d/%m/%Y')}"
                if l.periodo_inicio and l.periodo_fim else "—"
            ),
            "total": str(valor),
            "status": l.get_status_display(),
            "status_code": l.status,
        })

    grand_total = total_a_receber + total_empresas

    return JsonResponse({
        "totais": {
            "num_prefaturas": total_pfs,
            "total_pacotes": total_pacotes,
            "total_base": str(total_base),
            "total_bonus": str(total_bonus),
            "total_comissoes": str(total_comissoes),
            "total_descontos": str(total_descontos),
            "total_adiantamentos": str(total_adiantamentos),
            "total_a_receber": str(total_a_receber),
            "total_empresas": str(total_empresas),
            "grand_total": str(grand_total),
        },
        "motoristas": motoristas,
        "lancamentos": lancamentos_data,
    })


@login_required
def pre_invoice_report_pdf(request):
    """Gera PDF de resumo de pagamentos (uma página, todos os motoristas filtrados)."""
    from decimal import Decimal
    from datetime import datetime
    from django.http import HttpResponse
    from io import BytesIO

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
    except ImportError:
        return HttpResponse("reportlab não instalado", status=500)

    qs = _build_report_qs(request)
    styles = getSampleStyleSheet()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    purple = colors.HexColor("#7C3AED")
    green  = colors.HexColor("#059669")
    gray   = colors.HexColor("#F3F4F6")
    dark   = colors.HexColor("#111827")
    red    = colors.HexColor("#DC2626")
    teal   = colors.HexColor("#0D9488")

    title_style = ParagraphStyle("T", parent=styles["Normal"],
        fontSize=16, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER)
    sub_style = ParagraphStyle("S", parent=styles["Normal"],
        fontSize=8, textColor=colors.white, alignment=TA_CENTER)
    th_style = ParagraphStyle("TH", parent=styles["Normal"],
        fontSize=7, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER)
    td_style = ParagraphStyle("TD", parent=styles["Normal"],
        fontSize=7, textColor=dark, alignment=TA_CENTER)
    td_left = ParagraphStyle("TDL", parent=styles["Normal"],
        fontSize=7, textColor=dark, alignment=TA_LEFT)

    elements = []

    # ── Cabeçalho ────────────────────────────────────────────────────────
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    periodo_label = ""
    if date_from or date_to:
        periodo_label = f"Período: {date_from or '—'} a {date_to or '—'}"
    else:
        periodo_label = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    header_table = Table([[
        Paragraph("RESUMO DE PAGAMENTOS — PRÉ-FATURAS", title_style),
        Paragraph(periodo_label, sub_style),
    ]], colWidths=[12 * cm, 5.7 * cm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), purple),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (0, -1), 12),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.4 * cm))

    # ── Tabela principal ─────────────────────────────────────────────────
    col_widths = [4.5*cm, 1.8*cm, 2.5*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.5*cm, 2.6*cm]

    header_row = [
        Paragraph("Motorista", th_style),
        Paragraph("Pré-Fatura", th_style),
        Paragraph("Parceiro(s)", th_style),
        Paragraph("Pacotes", th_style),
        Paragraph("Base (€)", th_style),
        Paragraph("Bónus (€)", th_style),
        Paragraph("Desc. (€)", th_style),
        Paragraph("A Receber (€)", th_style),
    ]

    total_pacotes = 0
    total_base = Decimal("0.00")
    total_bonus = Decimal("0.00")
    total_descontos = Decimal("0.00")
    total_a_receber = Decimal("0.00")
    total_comissoes = Decimal("0.00")

    data_rows = []
    for pf in qs:
        pcts = sum(l.total_pacotes for l in pf.linhas.all())
        total_pacotes += pcts
        total_base += pf.base_entregas
        total_bonus += pf.total_bonus
        descontos = pf.penalizacoes_gerais + pf.total_pacotes_perdidos
        total_descontos += descontos
        total_a_receber += pf.total_a_receber
        total_comissoes += pf.total_comissoes_indicacao

        parceiros_str = ", ".join(
            sorted(set(l.parceiro.name for l in pf.linhas.all() if l.parceiro))
        ) or "—"

        desc_val = float(descontos)
        desc_str = f"-€{desc_val:.2f}" if desc_val > 0 else "€0.00"

        com_val = float(pf.total_comissoes_indicacao)
        nome_cell = pf.driver.nome_completo
        if com_val > 0:
            nome_cell += f"\n+€{com_val:.2f} indicação"

        data_rows.append([
            Paragraph(nome_cell, td_left),
            Paragraph(pf.numero, td_style),
            Paragraph(parceiros_str, td_style),
            Paragraph(str(pcts), td_style),
            Paragraph(f"€{float(pf.base_entregas):.2f}", td_style),
            Paragraph(f"€{float(pf.total_bonus):.2f}", td_style),
            Paragraph(desc_str, td_style),
            Paragraph(f"€{float(pf.total_a_receber):.2f}", ParagraphStyle(
                "TAR", parent=styles["Normal"],
                fontSize=8, fontName="Helvetica-Bold",
                textColor=green, alignment=TA_CENTER,
            )),
        ])

    # Linha de totais
    data_rows.append([
        Paragraph("TOTAIS", ParagraphStyle("TOT", parent=styles["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=dark)),
        Paragraph(str(len(data_rows)), ParagraphStyle("TOT2", parent=styles["Normal"],
            fontSize=7, textColor=dark, alignment=TA_CENTER)),
        Paragraph("", td_style),
        Paragraph(str(total_pacotes), ParagraphStyle("TOTP", parent=styles["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=dark, alignment=TA_CENTER)),
        Paragraph(f"€{float(total_base):.2f}", ParagraphStyle("TOTB", parent=styles["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=dark, alignment=TA_CENTER)),
        Paragraph(f"€{float(total_bonus):.2f}", ParagraphStyle("TOTBN", parent=styles["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=dark, alignment=TA_CENTER)),
        Paragraph(f"-€{float(total_descontos):.2f}", ParagraphStyle("TOTD", parent=styles["Normal"],
            fontSize=8, fontName="Helvetica-Bold", textColor=red, alignment=TA_CENTER)),
        Paragraph(f"€{float(total_a_receber):.2f}", ParagraphStyle("TOTR", parent=styles["Normal"],
            fontSize=10, fontName="Helvetica-Bold", textColor=green, alignment=TA_CENTER)),
    ])

    all_rows = [header_row] + data_rows
    main_table = Table(all_rows, colWidths=col_widths, repeatRows=1)

    n = len(data_rows)
    ts = [
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0FDF4")),
        ("ROWBACKGROUNDS", (0, 1), (-1, n - 1), [colors.white, gray]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("LINEABOVE", (0, -1), (-1, -1), 1.5, green),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if total_comissoes > 0:
        ts.append(("BACKGROUND", (7, 1), (7, n - 1), colors.HexColor("#F0FDFA")))

    main_table.setStyle(TableStyle(ts))
    elements.append(main_table)

    # ── Secção Empresas Parceiras ─────────────────────────────────────────
    from decimal import Decimal as _Dec
    from drivers_app.models import EmpresaParceiraLancamento
    lqs = (
        EmpresaParceiraLancamento.objects
        .select_related("empresa")
        .exclude(status__in=["CANCELADO", "RASCUNHO"])
        .order_by("empresa__nome", "-periodo_inicio")
    )
    date_from_raw = request.GET.get("date_from", "").strip()
    date_to_raw   = request.GET.get("date_to", "").strip()
    if date_from_raw:
        lqs = lqs.filter(periodo_inicio__gte=date_from_raw)
    if date_to_raw:
        lqs = lqs.filter(periodo_fim__lte=date_to_raw)

    total_empresas = _Dec("0.00")
    emp_rows = []
    for l in lqs:
        valor = l.total_a_receber
        total_empresas += valor
        periodo_str = (
            f"{l.periodo_inicio.strftime('%d/%m/%Y')} → {l.periodo_fim.strftime('%d/%m/%Y')}"
            if l.periodo_inicio and l.periodo_fim else "—"
        )
        emp_rows.append([
            Paragraph(l.empresa.nome if l.empresa else "—", td_left),
            Paragraph(l.descricao or "—", td_style),
            Paragraph(periodo_str, td_style),
            Paragraph(f"€{float(valor):.2f}", ParagraphStyle(
                "EV", parent=styles["Normal"],
                fontSize=7, fontName="Helvetica-Bold",
                textColor=teal, alignment=TA_CENTER,
            )),
        ])

    if emp_rows:
        elements.append(Spacer(1, 0.5 * cm))
        emp_header_row = [
            Paragraph("Empresa Parceira", th_style),
            Paragraph("Descrição", th_style),
            Paragraph("Período", th_style),
            Paragraph("Total (€)", th_style),
        ]
        emp_col_widths = [5.5*cm, 6*cm, 4*cm, 2.2*cm]
        emp_rows.append([
            Paragraph("TOTAL EMPRESAS PARCEIRAS", ParagraphStyle(
                "ETE", parent=styles["Normal"],
                fontSize=8, fontName="Helvetica-Bold", textColor=dark)),
            Paragraph("", td_style),
            Paragraph("", td_style),
            Paragraph(f"€{float(total_empresas):.2f}", ParagraphStyle(
                "ETET", parent=styles["Normal"],
                fontSize=8, fontName="Helvetica-Bold", textColor=teal, alignment=TA_CENTER)),
        ])
        emp_table = Table([emp_header_row] + emp_rows, colWidths=emp_col_widths, repeatRows=1)
        ne = len(emp_rows)
        emp_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), teal),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0FDFA")),
            ("ROWBACKGROUNDS", (0, 1), (-1, ne - 1), [colors.white, gray]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
            ("LINEABOVE", (0, -1), (-1, -1), 1.5, teal),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(Paragraph("Lançamentos de Empresas Parceiras", ParagraphStyle(
            "EH", parent=styles["Normal"],
            fontSize=9, fontName="Helvetica-Bold", textColor=teal,
        )))
        elements.append(Spacer(1, 0.2 * cm))
        elements.append(emp_table)

    # ── Cards de totais no fundo ──────────────────────────────────────────
    elements.append(Spacer(1, 0.5 * cm))

    cards_data = [
        ["Pré-Faturas", str(len(data_rows) - 1)],
        ["Total Pacotes", str(total_pacotes)],
        ["Base Entregas", f"€{float(total_base):.2f}"],
        ["Bónus", f"€{float(total_bonus):.2f}"],
    ]
    if total_comissoes > 0:
        cards_data.append(["Comissões Indicação", f"+€{float(total_comissoes):.2f}"])
    cards_data.append(["TOTAL A PAGAR", f"€{float(total_a_receber):.2f}"])

    card_cols = len(cards_data)
    available_w = 17.7 * cm
    card_w = available_w / card_cols

    card_header = [[Paragraph(c[0], ParagraphStyle("CH", parent=styles["Normal"],
        fontSize=7, textColor=colors.HexColor("#6B7280"), alignment=TA_CENTER))
        for c in cards_data]]
    card_values = [[Paragraph(c[1], ParagraphStyle("CV", parent=styles["Normal"],
        fontSize=10 if c[0] != "TOTAL A PAGAR" else 13,
        fontName="Helvetica-Bold",
        textColor=green if "TOTAL" in c[0] else (teal if "Comissões" in c[0] else dark),
        alignment=TA_CENTER))
        for c in cards_data]]

    cards_table = Table(
        [card_header[0], card_values[0]],
        colWidths=[card_w] * card_cols,
    )
    cards_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), gray),
        ("BACKGROUND", (-1, 0), (-1, -1), colors.HexColor("#F0FDF4")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEABOVE", (-1, 0), (-1, -1), 2, green),
    ]))
    elements.append(cards_table)

    # ── Rodapé ────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph(
        f"Documento gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        "Léguas Franzinas - Sistema de Gestão Logística",
        ParagraphStyle("F", parent=styles["Normal"],
            fontSize=7, textColor=colors.grey, alignment=TA_CENTER),
    ))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="resumo_prefaturas_{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


# ============================================================================
# EMPRESA PARCEIRA LANÇAMENTO — DETALHE / EDITAR / PDF
# ============================================================================


@login_required
@require_http_methods(["GET"])
def empresa_lancamento_detail_api(request, lancamento_id):
    """Devolve JSON com todos os dados de um lançamento de empresa parceira."""
    from drivers_app.models import EmpresaParceiraLancamento

    lanc = get_object_or_404(
        EmpresaParceiraLancamento.objects.select_related("empresa", "created_by"),
        id=lancamento_id,
    )

    return JsonResponse({
        "id": lanc.id,
        "empresa_nome": lanc.empresa.nome,
        "empresa_nif": lanc.empresa.nif or "",
        "empresa_iban": lanc.empresa.iban or "",
        "descricao": lanc.descricao,
        "qtd_entregas": lanc.qtd_entregas,
        "valor_por_entrega": str(lanc.valor_por_entrega),
        "valor_base": str(lanc.valor_base),
        "valor_bonus": str(lanc.valor_bonus),
        "pacotes_perdidos": str(lanc.pacotes_perdidos),
        "adiantamentos": str(lanc.adiantamentos),
        "taxa_iva": str(lanc.taxa_iva),
        "total_a_receber": str(lanc.total_a_receber),
        "valor_iva": str(lanc.valor_iva),
        "total_com_iva": str(lanc.total_com_iva),
        "periodo_inicio": lanc.periodo_inicio.strftime("%d/%m/%Y"),
        "periodo_fim": lanc.periodo_fim.strftime("%d/%m/%Y"),
        "periodo_inicio_iso": lanc.periodo_inicio.isoformat(),
        "periodo_fim_iso": lanc.periodo_fim.isoformat(),
        "status": lanc.status,
        "status_display": lanc.get_status_display(),
        "data_pagamento": lanc.data_pagamento.strftime("%d/%m/%Y") if lanc.data_pagamento else "",
        "referencia_pagamento": lanc.referencia_pagamento or "",
        "comprovante_url": lanc.comprovante_pagamento.url if lanc.comprovante_pagamento else "",
        "notas": lanc.notas or "",
        "created_at": lanc.created_at.strftime("%d/%m/%Y %H:%M"),
        "created_by": lanc.created_by.get_full_name() or lanc.created_by.username if lanc.created_by else "",
    })


@login_required
@require_http_methods(["POST"])
def empresa_lancamento_update(request, lancamento_id):
    """Actualiza os campos editáveis de um lançamento de empresa parceira."""
    import json
    from decimal import Decimal, InvalidOperation
    from drivers_app.models import EmpresaParceiraLancamento

    lanc = get_object_or_404(EmpresaParceiraLancamento, id=lancamento_id)

    if request.content_type and "application/json" in request.content_type:
        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            body = {}
    else:
        body = request.POST.dict()

    def dec(val, default=Decimal("0.00")):
        try:
            return Decimal(str(val)) if val not in (None, "") else default
        except InvalidOperation:
            return default

    if body.get("descricao") is not None:
        lanc.descricao = body["descricao"].strip()
    if body.get("qtd_entregas") is not None:
        try:
            lanc.qtd_entregas = int(body["qtd_entregas"])
        except (ValueError, TypeError):
            pass
    if body.get("valor_por_entrega") is not None:
        lanc.valor_por_entrega = dec(body["valor_por_entrega"])
    if body.get("valor_base") is not None:
        lanc.valor_base = dec(body["valor_base"])
    if body.get("valor_bonus") is not None:
        lanc.valor_bonus = dec(body["valor_bonus"])
    if body.get("pacotes_perdidos") is not None:
        lanc.pacotes_perdidos = dec(body["pacotes_perdidos"])
    if body.get("adiantamentos") is not None:
        lanc.adiantamentos = dec(body["adiantamentos"])
    if body.get("periodo_inicio"):
        try:
            from datetime import date
            lanc.periodo_inicio = date.fromisoformat(body["periodo_inicio"])
        except ValueError:
            pass
    if body.get("periodo_fim"):
        try:
            from datetime import date
            lanc.periodo_fim = date.fromisoformat(body["periodo_fim"])
        except ValueError:
            pass
    if body.get("notas") is not None:
        lanc.notas = body["notas"].strip()
    if body.get("taxa_iva") is not None:
        lanc.taxa_iva = dec(body["taxa_iva"], Decimal("23.00"))

    # Auto-calcular valor_base se qtd e valor_por_entrega forem enviados
    if body.get("qtd_entregas") is not None and body.get("valor_por_entrega") is not None:
        lanc.valor_base = Decimal(str(lanc.qtd_entregas)) * lanc.valor_por_entrega

    lanc.save()

    return JsonResponse({
        "success": True,
        "total_a_receber": str(lanc.total_a_receber),
        "taxa_iva": str(lanc.taxa_iva),
        "valor_iva": str(lanc.valor_iva),
        "total_com_iva": str(lanc.total_com_iva),
        "descricao": lanc.descricao,
        "periodo_inicio": lanc.periodo_inicio.strftime("%d/%m/%Y"),
        "periodo_fim": lanc.periodo_fim.strftime("%d/%m/%Y"),
        "periodo_inicio_iso": lanc.periodo_inicio.isoformat(),
        "periodo_fim_iso": lanc.periodo_fim.isoformat(),
    })


@login_required
def empresa_lancamento_pdf(request, lancamento_id):
    """Gera e devolve o PDF de um lançamento de empresa parceira."""
    from django.http import HttpResponse
    from drivers_app.models import EmpresaParceiraLancamento

    lanc = get_object_or_404(
        EmpresaParceiraLancamento.objects.select_related("empresa"),
        id=lancamento_id,
    )

    try:
        from .reports.pdf_generator import PDFGenerator
        generator = PDFGenerator()
        pdf_buffer = generator.generate_lancamento_pdf(lanc)
        empresa_slug = lanc.empresa.nome.replace(" ", "_")[:30]
        periodo = f"{lanc.periodo_inicio.strftime('%Y%m')}"
        filename = f"Lancamento_{empresa_slug}_{periodo}.pdf"
        response = HttpResponse(pdf_buffer, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        return JsonResponse(
            {"error": "ReportLab não está instalado no servidor."},
            status=500,
        )


# ──────────────────────────────────────────────────────────────────────────
#  Driver Delivery Report — Fase 2 (modal Pré-faturas)
#
#  Devolve para um driver + período:
#   - Lista de entregas (waybill, data, CP4, status, parceiro, courier_id)
#   - Aggregates por dia e por CP4
#   - Bónus auto-detectados (domingos + feriados)
#   - Pré-faturas existentes que cobrem qualquer parte do período (overlap)
#   - Reclamações ligadas a pacotes do período
# ──────────────────────────────────────────────────────────────────────────

@login_required
def driver_delivery_report(request, driver_id):
    """Relatório de entregas de um motorista por período.

    Query params:
      from=YYYY-MM-DD  (default: 7 dias atrás)
      to=YYYY-MM-DD    (default: hoje)
      cp4=XXXX         (opcional, multi)
      include_packages=1  (default 1) — lista detalhada de waybills
      page=N&page_size=M
    """
    from datetime import timedelta
    from decimal import Decimal
    from django.db.models.functions import Substr
    from drivers_app.models import DriverProfile
    from .models import (
        BonusBlackoutDate, CainiaoOperationTask, DriverPreInvoice, Holiday, DriverClaim,
    )

    driver = get_object_or_404(
        DriverProfile.objects.select_related("empresa_parceira"),
        id=driver_id,
    )

    today = timezone.now().date()
    date_from = (parse_date(request.GET.get("from") or "")
                 or (today - timedelta(days=7)))
    date_to = parse_date(request.GET.get("to") or "") or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    cp4_filter = [c for c in request.GET.getlist("cp4") if c.strip()]
    include_packages = request.GET.get("include_packages", "1") != "0"

    # Driver courier_ids (Cainiao) via mappings + profile
    cainiao_courier_ids = set()
    cainiao_courier_names = set()
    if driver.courier_id_cainiao:
        cainiao_courier_ids.add(driver.courier_id_cainiao)
    if driver.apelido:
        cainiao_courier_names.add(driver.apelido)
    for m in driver.courier_mappings.filter(partner__name__iexact="CAINIAO"):
        if m.courier_id:
            cainiao_courier_ids.add(m.courier_id)
        if m.courier_name:
            cainiao_courier_names.add(m.courier_name)

    base_qs = CainiaoOperationTask.objects.filter(
        task_date__range=(date_from, date_to),
    )
    if cainiao_courier_ids or cainiao_courier_names:
        driver_q = Q()
        if cainiao_courier_ids:
            driver_q |= Q(courier_id_cainiao__in=cainiao_courier_ids)
        if cainiao_courier_names:
            driver_q |= Q(courier_name__in=cainiao_courier_names)
        base_qs = base_qs.filter(driver_q)
    else:
        base_qs = base_qs.none()
    if cp4_filter:
        cp4_q = Q()
        for cp4 in cp4_filter:
            cp4_q |= Q(zip_code__startswith=cp4)
        base_qs = base_qs.filter(cp4_q)

    # ── Aplicar overrides de transferência ────────────────────────────
    # Excluir waybills que foram transferidos PARA outros drivers.
    # Adicionar (UNION) waybills transferidos PARA este driver.
    from .models import WaybillAttributionOverride
    outgoing = set(
        WaybillAttributionOverride.objects.filter(
            task_date__range=(date_from, date_to),
        ).exclude(attributed_to_driver=driver)
        .values_list("waybill_number", flat=True),
    )
    incoming = list(
        WaybillAttributionOverride.objects.filter(
            attributed_to_driver=driver,
            task_date__range=(date_from, date_to),
        ).values_list("waybill_number", flat=True),
    )
    if outgoing:
        base_qs = base_qs.exclude(waybill_number__in=outgoing)
    if incoming:
        incoming_qs = CainiaoOperationTask.objects.filter(
            waybill_number__in=incoming,
            task_date__range=(date_from, date_to),
        )
        # Combinar via UNION (manter rows nativas + transferidas)
        base_qs = (base_qs | incoming_qs).distinct()

    total = base_qs.count()
    delivered_qs = base_qs.filter(task_status="Delivered")
    delivered = delivered_qs.count()
    failures = base_qs.filter(task_status="Attempt Failure").count()
    n_transferred_in = len(incoming)
    n_transferred_out = len(outgoing) if outgoing else 0

    # Cascata: Driver > Frota > Parceiro (driver_default) > 0
    from core.models import Partner
    from core.finance import resolve_driver_price, resolve_partner_price
    cainiao_partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    price, price_source = resolve_driver_price(driver, cainiao_partner)
    partner_price = resolve_partner_price(cainiao_partner)
    margin_per_pkg = partner_price - price
    receita_estimada = partner_price * delivered  # leguas recebe

    # ── Aplicar PackagePriceOverride ao custo (valor_estimado) ────────
    # Cada waybill com override entregue paga o preço especial em vez
    # do preço base. Margem = receita_estimada - valor_estimado.
    from .models import PackagePriceOverride
    delivered_wb_qs = delivered_qs.values_list("waybill_number", flat=True)
    overrides_in_period = list(
        PackagePriceOverride.objects.filter(
            waybill_number__in=delivered_wb_qs,
        ).values_list("waybill_number", "price"),
    )
    n_special = len(overrides_in_period)
    sum_special = sum(
        (p for _wb, p in overrides_in_period), Decimal("0"),
    )
    n_base = max(0, delivered - n_special)
    valor_estimado = price * n_base + sum_special
    margem_estimada = receita_estimada - valor_estimado

    # ── Comissão de indicação ───────────────────────────────────────────
    # Se este motorista foi indicado por alguém, parte da margem da Léguas
    # vai para o indicador. Se este motorista é indicador, recebe comissão
    # pelas entregas dos indicados no período.
    referred_info = None
    referral_per_pkg = Decimal("0.0000")
    referral_commission_total = Decimal("0.00")
    try:
        ref = driver.referral_received  # OneToOne — pode não existir
    except Exception:
        ref = None
    if ref and ref.ativo:
        referral_per_pkg = ref.comissao_por_pacote
        referral_commission_total = referral_per_pkg * delivered
        referred_info = {
            "referrer_id": ref.referrer_id,
            "referrer_name": ref.referrer.nome_completo,
            "comissao_per_pkg": str(referral_per_pkg),
            "total": f"{referral_commission_total:.2f}",
        }
    margem_liquida = margem_estimada - referral_commission_total

    # Indicador (driver indicou outros motoristas)
    referrals_given = list(
        driver.referrals_given.filter(ativo=True).select_related("referred")
    )
    referrer_breakdown = []
    referrer_income_total = Decimal("0.00")
    if referrals_given:
        from .models import CainiaoOperationTask as _COT
        for rg in referrals_given:
            ref_driver = rg.referred
            ref_courier_ids = list(
                DriverCourierMapping.objects
                .filter(driver=ref_driver)
                .values_list("courier_id", flat=True)
            )
            ref_courier_names = list(
                DriverCourierMapping.objects
                .filter(driver=ref_driver)
                .values_list("courier_name", flat=True)
            )
            ref_q = _COT.objects.filter(
                task_date__gte=date_from,
                task_date__lte=date_to,
                task_status="Delivered",
            )
            dq = Q()
            if ref_courier_ids:
                dq |= Q(courier_id_cainiao__in=ref_courier_ids)
            if ref_courier_names:
                dq |= Q(courier_name__in=ref_courier_names)
            if not dq:
                continue
            ref_delivered = ref_q.filter(dq).count()
            if ref_delivered <= 0:
                continue
            commission = rg.comissao_por_pacote * ref_delivered
            referrer_income_total += commission
            referrer_breakdown.append({
                "referred_id": ref_driver.id,
                "referred_name": ref_driver.nome_completo,
                "delivered": ref_delivered,
                "comissao_per_pkg": str(rg.comissao_por_pacote),
                "total": f"{commission:.2f}",
            })

    # By day (bónus + heatmap)
    by_day_raw = list(
        delivered_qs.values("task_date")
        .annotate(n=Count("id"))
        .order_by("task_date")
    )
    by_day = []
    bonus_days = 0
    # Pré-fetch dos blackouts no período para evitar N queries
    if by_day_raw:
        _bb_dates = BonusBlackoutDate.dates_in_range(
            by_day_raw[0]["task_date"], by_day_raw[-1]["task_date"],
        )
    else:
        _bb_dates = set()
    for row in by_day_raw:
        d = row["task_date"]
        is_sun = d.weekday() == 6
        h = Holiday.get_holiday(d)
        is_blocked = d in _bb_dates
        is_bonus = (is_sun or bool(h)) and not is_blocked
        if is_bonus:
            bonus_days += 1
        by_day.append({
            "date": d.strftime("%Y-%m-%d"),
            "weekday": d.weekday(),
            "n": row["n"],
            "is_sunday": is_sun,
            "is_holiday": bool(h),
            "holiday_name": h.name if h else "",
            "is_bonus": is_bonus,
            "is_bonus_blocked": is_blocked,
        })

    # By CP4
    by_cp4_raw = list(
        delivered_qs.exclude(zip_code="")
        .annotate(cp4=Substr("zip_code", 1, 4))
        .values("cp4")
        .annotate(n=Count("id"))
        .order_by("-n")[:25]
    )
    by_cp4 = [{"cp4": r["cp4"], "n": r["n"]} for r in by_cp4_raw]

    # Pré-faturas existentes que sobrepõem este período
    overlapping_pfs = list(
        DriverPreInvoice.objects.filter(
            driver=driver,
            periodo_inicio__lte=date_to,
            periodo_fim__gte=date_from,
        ).order_by("-periodo_fim")
    )
    existing_pre_invoices = [
        {
            "id": pf.id,
            "numero": pf.numero,
            "periodo_inicio": pf.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": pf.periodo_fim.strftime("%Y-%m-%d"),
            "status": pf.status,
            "status_display": pf.get_status_display(),
            "total_a_receber": str(pf.total_a_receber),
        }
        for pf in overlapping_pfs
    ]

    # Mapa task_date → PF para indicar "já facturado"
    pf_lookup = {}
    for pf in overlapping_pfs:
        d = pf.periodo_inicio
        while d <= pf.periodo_fim:
            pf_lookup[d] = pf
            d += timedelta(days=1)

    # Reclamações ligadas a pacotes do período
    claim_by_waybill = {
        c.waybill_number: c for c in DriverClaim.objects.filter(
            driver=driver,
        ).exclude(waybill_number="")
    }

    packages = []
    total_pkg = 0
    if include_packages:
        try:
            page = max(1, int(request.GET.get("page") or 1))
            page_size = min(500, max(10, int(request.GET.get("page_size") or 100)))
        except (TypeError, ValueError):
            page, page_size = 1, 100

        # Tabela só mostra pacotes Delivered (que entram na pré-fatura).
        # Falhas (Attempt Failure) ficam apenas no card "Falhas" para visual.
        qs_pkg = (
            delivered_qs
            .order_by("-task_date", "courier_id_cainiao", "waybill_number")
            .values(
                "id", "waybill_number", "task_date", "task_status",
                "courier_name", "courier_id_cainiao", "zip_code",
                "destination_city", "delivery_time",
            )
        )
        total_pkg = qs_pkg.count()
        start = (page - 1) * page_size
        page_rows = list(qs_pkg[start:start + page_size])

        # Para cada waybill na página, descobrir o status FINAL do pacote
        # (linha mais recente na BD entre todos os drivers/datas). Não
        # afecta cálculo de pagamento, só serve para o operador ver o
        # contexto: ex. uma "Attempt Failure" do driver A foi
        # eventualmente entregue pelo driver B no dia seguinte.
        wb_list = [r["waybill_number"] for r in page_rows]
        final_by_wb = {}
        if wb_list:
            from django.db.models import Max
            latest_dates = (
                CainiaoOperationTask.objects
                .filter(waybill_number__in=wb_list)
                .values("waybill_number")
                .annotate(latest=Max("task_date"))
            )
            latest_q = Q()
            for ld in latest_dates:
                latest_q |= Q(
                    waybill_number=ld["waybill_number"],
                    task_date=ld["latest"],
                )
            if latest_q:
                for r in CainiaoOperationTask.objects.filter(latest_q).values(
                    "waybill_number", "task_date", "task_status",
                    "courier_name",
                ):
                    final_by_wb[r["waybill_number"]] = {
                        "task_date": r["task_date"].strftime("%Y-%m-%d"),
                        "status": r["task_status"],
                        "courier_name": r["courier_name"],
                    }

        # Map waybill -> PackagePriceOverride para a página
        po_by_wb_page = {}
        if wb_list:
            for po in PackagePriceOverride.objects.filter(
                waybill_number__in=wb_list,
            ):
                po_by_wb_page[po.waybill_number] = po

        for r in page_rows:
            pf = pf_lookup.get(r["task_date"])
            cp4 = (r["zip_code"] or "")[:4]
            wb = r["waybill_number"]
            claim = claim_by_waybill.get(wb)
            row_status = r["task_status"]
            final = final_by_wb.get(wb)
            po = po_by_wb_page.get(wb)
            # final_status_differs = True quando o status final do
            # waybill é diferente do desta linha do driver.
            final_differs = bool(
                final and (
                    final["status"] != row_status
                    or final["task_date"] != r["task_date"].strftime("%Y-%m-%d")
                )
            )
            packages.append({
                "id": r["id"],
                "waybill_number": wb,
                "task_date": r["task_date"].strftime("%Y-%m-%d"),
                "task_status": row_status,
                "courier_name": r["courier_name"],
                "courier_id_cainiao": r["courier_id_cainiao"],
                "cp4": cp4,
                "zip_code": r["zip_code"],
                "city": r["destination_city"],
                "invoiced_in_pf_id": pf.id if pf else None,
                "invoiced_in_pf_numero": pf.numero if pf else "",
                "invoiced_pf_status": pf.status if pf else "",
                "claim_id": claim.id if claim else None,
                "claim_status": claim.status if claim else "",
                # Contexto do waybill (não-pago, só visual)
                "final_status": final["status"] if final else row_status,
                "final_status_date": final["task_date"] if final else r["task_date"].strftime("%Y-%m-%d"),
                "final_status_courier": final["courier_name"] if final else r["courier_name"],
                "final_status_differs": final_differs,
                # Override de preço (se existir)
                "price_override": (
                    {
                        "id": po.id,
                        "price": str(po.price),
                        "reason": po.reason,
                    } if po else None
                ),
            })

    # Reclamações geral (lista para tab)
    claims = []
    for c in DriverClaim.objects.filter(driver=driver).order_by("-created_at"):
        claims.append({
            "id": c.id,
            "waybill_number": c.waybill_number,
            "valor_a_descontar": str(c.valor_a_descontar),
            "status": c.status,
            "status_display": c.get_status_display(),
            "auto_detected": c.auto_detected,
            "created_at": c.created_at.strftime("%Y-%m-%d"),
            "descricao": (c.descricao or "")[:200],
        })

    # ── Comparação com período anterior (mesma duração, imediatamente antes)
    period_days = (date_to - date_from).days + 1
    prev_to = date_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=period_days - 1)
    prev_qs = CainiaoOperationTask.objects.filter(
        task_date__range=(prev_from, prev_to),
    )
    if cainiao_courier_ids or cainiao_courier_names:
        prev_q = Q()
        if cainiao_courier_ids:
            prev_q |= Q(courier_id_cainiao__in=cainiao_courier_ids)
        if cainiao_courier_names:
            prev_q |= Q(courier_name__in=cainiao_courier_names)
        prev_qs = prev_qs.filter(prev_q)
    else:
        prev_qs = prev_qs.none()
    if cp4_filter:
        prev_cp4_q = Q()
        for cp4 in cp4_filter:
            prev_cp4_q |= Q(zip_code__startswith=cp4)
        prev_qs = prev_qs.filter(prev_cp4_q)
    prev_total = prev_qs.count()
    prev_delivered = prev_qs.filter(task_status="Delivered").count()
    prev_failures = prev_qs.filter(task_status="Attempt Failure").count()
    prev_valor = price * prev_delivered

    def _delta_pct(current, previous):
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return round((current - previous) * 100.0 / previous, 1)

    comparison = {
        "previous_period": {
            "from": prev_from.strftime("%Y-%m-%d"),
            "to": prev_to.strftime("%Y-%m-%d"),
        },
        "previous": {
            "total": prev_total,
            "delivered": prev_delivered,
            "failures": prev_failures,
            "valor_estimado": f"{prev_valor:.2f}",
        },
        "delta": {
            "total_pct": _delta_pct(total, prev_total),
            "delivered_pct": _delta_pct(delivered, prev_delivered),
            "failures_pct": _delta_pct(failures, prev_failures),
            "valor_pct": _delta_pct(float(valor_estimado), float(prev_valor)),
            "delivered_abs": delivered - prev_delivered,
            "failures_abs": failures - prev_failures,
        },
    }

    return JsonResponse({
        "success": True,
        "driver": {
            "id": driver.id,
            "name": driver.nome_completo,
            "apelido": driver.apelido,
            "fleet": (driver.empresa_parceira.nome
                      if driver.empresa_parceira else ""),
            "courier_ids": sorted(cainiao_courier_ids),
        },
        "period": {
            "from": date_from.strftime("%Y-%m-%d"),
            "to": date_to.strftime("%Y-%m-%d"),
            "days": (date_to - date_from).days + 1,
        },
        "filters": {"cp4": cp4_filter},
        "summary": {
            "total": total,
            "delivered": delivered,
            "failures": failures,
            "success_rate": round(delivered * 100.0 / total, 1) if total else 0,
            "price_per_package": str(price),
            "price_source": price_source,
            "valor_estimado": f"{valor_estimado:.2f}",
            "n_priced_overrides": n_special,
            "sum_priced_overrides": f"{sum_special:.2f}",
            "bonus_days": bonus_days,
            # Margem (Fase 6.5)
            "partner_price_per_package": str(partner_price),
            "margin_per_package": str(margin_per_pkg),
            "receita_estimada": f"{receita_estimada:.2f}",
            "margem_estimada": f"{margem_estimada:.2f}",
            "margin_negative": margin_per_pkg < 0,
            # Indicação (referral)
            "is_referred": referred_info is not None,
            "referred_by": referred_info,
            "referral_commission_per_pkg": str(referral_per_pkg),
            "referral_commission_total": f"{referral_commission_total:.2f}",
            "margem_liquida": f"{margem_liquida:.2f}",
            "is_referrer": bool(referrer_breakdown),
            "referrer_breakdown": referrer_breakdown,
            "referrer_income_total": f"{referrer_income_total:.2f}",
        },
        "comparison": comparison,
        "by_day": by_day,
        "by_cp4": by_cp4,
        "packages": packages,
        "packages_total": total_pkg if include_packages else None,
        "existing_pre_invoices": existing_pre_invoices,
        "claims": claims,
    })


@login_required
def driver_account_statement(request, driver_id):
    """Conta-corrente do motorista — saldo a receber, pago, em aberto.

    Para cada pré-fatura (ordem cronológica), devolve:
      - linha com totais e status
      - running_balance: saldo acumulado em aberto (não pagas)
      - lifetime_received: total já pago

    Query params:
      year=YYYY (default: tudo)
    """
    from drivers_app.models import DriverProfile
    from decimal import Decimal

    driver = get_object_or_404(DriverProfile, id=driver_id)
    qs = DriverPreInvoice.objects.filter(driver=driver).order_by(
        "periodo_inicio", "id",
    )
    year = request.GET.get("year")
    if year:
        try:
            qs = qs.filter(periodo_inicio__year=int(year))
        except ValueError:
            pass

    rows = []
    open_balance = Decimal("0")
    paid_total = Decimal("0")
    cancelled_total = Decimal("0")
    for pf in qs:
        amount = pf.total_a_receber
        is_paid = pf.status == "PAGO"
        is_cancelled = pf.status == "CANCELADO"

        if is_paid:
            paid_total += amount
        elif is_cancelled:
            cancelled_total += amount
        else:
            open_balance += amount

        rows.append({
            "id": pf.id,
            "numero": pf.numero,
            "periodo_inicio": pf.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": pf.periodo_fim.strftime("%Y-%m-%d"),
            "status": pf.status,
            "status_display": pf.get_status_display(),
            "amount": str(amount),
            "is_paid": is_paid,
            "is_cancelled": is_cancelled,
            "data_pagamento": (
                pf.data_pagamento.strftime("%Y-%m-%d")
                if pf.data_pagamento else ""
            ),
            "running_open": str(open_balance),
            "lifetime_received": str(paid_total),
        })

    total_amount = sum(
        (Decimal(r["amount"]) for r in rows if not r["is_cancelled"]),
        Decimal("0"),
    )

    return JsonResponse({
        "success": True,
        "driver": {
            "id": driver.id,
            "name": driver.nome_completo,
            "apelido": driver.apelido,
        },
        "summary": {
            "total_pre_invoices": len(rows),
            "open_balance": str(open_balance),
            "paid_total": str(paid_total),
            "cancelled_total": str(cancelled_total),
            "total_amount": str(total_amount),
        },
        "rows": rows,
    })


@login_required
@require_http_methods(["POST"])
def pre_invoice_send_whatsapp(request, pre_invoice_id):
    """Envia uma pré-fatura ao motorista via WhatsApp (WPPConnect).

    Usa o telefone do DriverProfile, formata mensagem com sumário, e
    inclui link público do PDF (se BASE_URL configurado).
    """
    from django.conf import settings
    from system_config.whatsapp_helper import WhatsAppWPPConnectAPI

    pf = get_object_or_404(
        DriverPreInvoice.objects.select_related("driver"),
        id=pre_invoice_id,
    )
    driver = pf.driver
    telefone = (driver.telefone or "").strip()
    if not telefone:
        return JsonResponse(
            {"success": False, "error":
             "Motorista não tem telefone configurado."},
            status=400,
        )

    # Sanitizar telefone (só dígitos). Se começar por 9 e tiver 9 dígitos,
    # assume PT e prefixa 351.
    digits = "".join(c for c in telefone if c.isdigit())
    if len(digits) == 9 and digits.startswith("9"):
        digits = "351" + digits

    # Construir mensagem
    pdf_url = ""
    base = (getattr(settings, "BASE_URL", "") or "").rstrip("/")
    if base:
        pdf_url = f"{base}/settlements/pre-invoices/{pf.id}/pdf/"

    msg = (
        f"📋 Pré-Fatura *{pf.numero}*\n"
        f"👤 {driver.nome_completo}\n"
        f"🗓️ {pf.periodo_inicio.strftime('%d/%m/%Y')}"
        f" → {pf.periodo_fim.strftime('%d/%m/%Y')}\n\n"
        f"📦 Entregas: €{pf.base_entregas}\n"
        f"🎁 Bónus: €{pf.total_bonus}\n"
        f"📉 Pacotes perdidos: -€{pf.total_pacotes_perdidos}\n"
        f"💰 Total a receber: *€{pf.total_a_receber}*\n"
        f"\nEstado: {pf.get_status_display()}"
    )
    if pdf_url:
        msg += f"\n📎 PDF: {pdf_url}"

    if not settings.WPPCONNECT_URL:
        return JsonResponse(
            {"success": False, "error":
             "WPPCONNECT_URL não configurado em settings."},
            status=500,
        )

    try:
        api = WhatsAppWPPConnectAPI(
            base_url=settings.WPPCONNECT_URL,
            session_name=settings.WPPCONNECT_SESSION,
            auth_token=settings.WPPCONNECT_TOKEN,
            secret_key=settings.WPPCONNECT_SECRET,
        )
        result = api.send_text(number=digits, text=msg)
    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Erro a enviar: {e}"},
            status=502,
        )

    return JsonResponse({
        "success": True,
        "phone": digits,
        "message_preview": msg[:200],
        "wpp_response": result,
    })


@login_required
def driver_yearly_heatmap(request, driver_id):
    """Heatmap-calendário anual (estilo GitHub) para um motorista.

    Devolve para cada dia de um ano:
      - n: número de entregas (Delivered)
      - level: 0-4 (intensidade pa coloração)
      - is_sunday, is_holiday, holiday_name

    Query params:
      year=YYYY (default: ano actual)
    """
    from datetime import date, timedelta
    from drivers_app.models import DriverProfile
    from .models import BonusBlackoutDate, CainiaoOperationTask, Holiday

    driver = get_object_or_404(DriverProfile, id=driver_id)
    try:
        year = int(request.GET.get("year") or timezone.now().year)
    except (TypeError, ValueError):
        year = timezone.now().year

    cainiao_courier_ids = set()
    cainiao_courier_names = set()
    if driver.courier_id_cainiao:
        cainiao_courier_ids.add(driver.courier_id_cainiao)
    if driver.apelido:
        cainiao_courier_names.add(driver.apelido)
    for m in driver.courier_mappings.filter(partner__name__iexact="CAINIAO"):
        if m.courier_id:
            cainiao_courier_ids.add(m.courier_id)
        if m.courier_name:
            cainiao_courier_names.add(m.courier_name)

    year_start = date(year, 1, 1)
    year_end = date(year, 12, 31)
    qs = CainiaoOperationTask.objects.filter(
        task_date__range=(year_start, year_end),
        task_status="Delivered",
    )
    if cainiao_courier_ids or cainiao_courier_names:
        driver_q = Q()
        if cainiao_courier_ids:
            driver_q |= Q(courier_id_cainiao__in=cainiao_courier_ids)
        if cainiao_courier_names:
            driver_q |= Q(courier_name__in=cainiao_courier_names)
        qs = qs.filter(driver_q)
    else:
        qs = qs.none()

    by_day = {
        r["task_date"]: r["n"]
        for r in qs.values("task_date").annotate(n=Count("id"))
    }

    # Calcular threshold de níveis (0-4) com base no max do ano
    max_n = max(by_day.values()) if by_day else 0

    def _level(n):
        if not n:
            return 0
        # Faixas: 1-25%, 25-50%, 50-75%, 75-100% do max
        ratio = n / max_n if max_n else 0
        if ratio <= 0.25:
            return 1
        if ratio <= 0.50:
            return 2
        if ratio <= 0.75:
            return 3
        return 4

    blocked_dates = BonusBlackoutDate.dates_in_range(year_start, year_end)

    days = []
    total_year = 0
    days_active = 0
    d = year_start
    while d <= year_end:
        n = by_day.get(d, 0)
        h = Holiday.get_holiday(d)
        days.append({
            "date": d.strftime("%Y-%m-%d"),
            "weekday": d.weekday(),
            "n": n,
            "level": _level(n),
            "is_sunday": d.weekday() == 6,
            "is_holiday": bool(h),
            "holiday_name": h.name if h else "",
            "is_bonus_blocked": d in blocked_dates,
        })
        total_year += n
        if n > 0:
            days_active += 1
        d += timedelta(days=1)

    return JsonResponse({
        "success": True,
        "year": year,
        "max_n": max_n,
        "total": total_year,
        "days_active": days_active,
        "days": days,
    })


@login_required
def driver_pre_invoice_preview(request, driver_id):
    """Preview da pré-fatura antes de criar (não persiste).

    Query params: from=YYYY-MM-DD&to=YYYY-MM-DD

    Se o motorista tem múltiplos logins/aliases na Cainiao, calcula
    entregas + bónus separadamente por login e devolve um breakdown
    detalhado (para auditoria), mas o subtotal final é unificado.
    """
    from datetime import timedelta
    from decimal import Decimal
    from drivers_app.models import DriverProfile
    from .models import (
        BonusBlackoutDate, CainiaoOperationTask, DriverPreInvoice, Holiday,
        PreInvoiceBonus,
    )

    driver = get_object_or_404(DriverProfile, id=driver_id)
    today = timezone.now().date()
    date_from = parse_date(request.GET.get("from") or "") or today
    date_to = parse_date(request.GET.get("to") or "") or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    # Construir lista de logins únicos do motorista. Cada "login" é
    # uma combinação (courier_id, courier_name) que pode aparecer
    # nos tasks da Cainiao.
    # logins = [{cid, cname, source}], dedup por par (cid, cname).
    logins = []
    seen_pairs = set()

    def _add_login(cid, cname, source):
        cid = (cid or "").strip()
        cname = (cname or "").strip()
        if not cid and not cname:
            return
        key = (cid, cname)
        if key in seen_pairs:
            return
        seen_pairs.add(key)
        logins.append({
            "courier_id": cid, "courier_name": cname, "source": source,
        })

    if driver.courier_id_cainiao or driver.apelido:
        _add_login(
            driver.courier_id_cainiao, driver.apelido,
            "perfil",
        )
    for m in driver.courier_mappings.filter(
        partner__name__iexact="CAINIAO",
    ):
        _add_login(m.courier_id, m.courier_name, "mapping")

    # Pricing
    from core.models import Partner
    from core.finance import resolve_driver_price
    cainiao_partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    price, price_source = resolve_driver_price(driver, cainiao_partner)

    # Para cada login, calcular entregas + bónus
    base_qs = CainiaoOperationTask.objects.filter(
        task_date__range=(date_from, date_to),
        task_status="Delivered",
    )

    def _bonus_for(n):
        if n >= PreInvoiceBonus.LIMIAR_60:
            return PreInvoiceBonus.BONUS_50
        if n >= PreInvoiceBonus.LIMIAR_30:
            return PreInvoiceBonus.BONUS_30
        return Decimal("0")

    blocked_dates = BonusBlackoutDate.dates_in_range(date_from, date_to)

    def _bonus_lines_for_qs(qs):
        out = []
        total = Decimal("0")
        for row in qs.values("task_date").annotate(
            n=Count("id"),
        ).order_by("task_date"):
            d = row["task_date"]
            n = row["n"]
            is_sun = d.weekday() == 6
            h = Holiday.get_holiday(d)
            if not (is_sun or h):
                continue
            if d in blocked_dates:
                continue
            b = _bonus_for(n)
            out.append({
                "date": d.strftime("%Y-%m-%d"),
                "weekday_name":
                    ["Seg","Ter","Qua","Qui","Sex","Sáb","Dom"][d.weekday()],
                "type": "FERIADO" if h else "DOMINGO",
                "type_display": h.name if h else "Domingo",
                "deliveries": n,
                "bonus": str(b),
            })
            total += b
        return out, total

    # ── Overrides de transferência ─────────────────────────────────────
    from .models import (
        WaybillAttributionOverride, PackagePriceOverride,
    )
    outgoing = set(
        WaybillAttributionOverride.objects.filter(
            task_date__range=(date_from, date_to),
        ).exclude(attributed_to_driver=driver)
        .values_list("waybill_number", flat=True),
    )
    incoming = list(
        WaybillAttributionOverride.objects.filter(
            attributed_to_driver=driver,
            task_date__range=(date_from, date_to),
        ).values_list("waybill_number", flat=True),
    )

    # ── Overrides de preço (preço especial por pacote) ─────────────────
    price_overrides_map = {
        po.waybill_number: po
        for po in PackagePriceOverride.objects.filter(
            task_date__range=(date_from, date_to),
        )
    }

    # Breakdown por login
    login_breakdown = []
    seen_task_ids = set()  # para evitar dupla-contagem entre logins
    total_delivered = 0
    total_base = Decimal("0")
    total_bonus = Decimal("0")
    for ln in logins:
        cid = ln["courier_id"]
        cname = ln["courier_name"]
        login_q = Q()
        if cid:
            login_q |= Q(courier_id_cainiao=cid)
        if cname:
            login_q |= Q(courier_name=cname)
        if not login_q:
            continue
        # Tasks deste login (ainda não vistas em logins anteriores)
        qs_login = base_qs.filter(login_q)
        if outgoing:
            qs_login = qs_login.exclude(waybill_number__in=outgoing)
        if seen_task_ids:
            qs_login = qs_login.exclude(id__in=seen_task_ids)
        rows_this = list(qs_login.values("id", "waybill_number"))
        if not rows_this:
            continue
        ids_this = [r["id"] for r in rows_this]
        seen_task_ids.update(ids_this)
        n_login = len(ids_this)

        # Agrupar por preço (base + cada override distinto)
        price_groups = {}  # price → {n, waybills}
        for r in rows_this:
            po = price_overrides_map.get(r["waybill_number"])
            unit_price = po.price if po else price
            key = unit_price
            if key not in price_groups:
                price_groups[key] = {"n": 0, "reasons": set()}
            price_groups[key]["n"] += 1
            if po and po.reason:
                price_groups[key]["reasons"].add(po.reason)

        # Linhas de preço (uma por grupo)
        price_breakdown_lines = []
        base_login = Decimal("0")
        for unit_price, info in sorted(
            price_groups.items(), reverse=True,
        ):
            line_total = unit_price * info["n"]
            base_login += line_total
            price_breakdown_lines.append({
                "price": str(unit_price),
                "n": info["n"],
                "total": f"{line_total:.2f}",
                "is_special": unit_price != price,
                "reasons": sorted(info["reasons"]),
            })

        b_lines, b_total = _bonus_lines_for_qs(
            CainiaoOperationTask.objects.filter(id__in=ids_this),
        )
        login_breakdown.append({
            "label": cname or cid,
            "courier_id": cid,
            "courier_name": cname,
            "source": ln["source"],
            "deliveries": n_login,
            "base_entregas": f"{base_login:.2f}",
            "price_breakdown": price_breakdown_lines,
            "bonus_lines": b_lines,
            "total_bonus": f"{b_total:.2f}",
            "subtotal": f"{(base_login + b_total):.2f}",
        })
        total_delivered += n_login
        total_base += base_login
        total_bonus += b_total

    # Linha extra: transferências recebidas (waybills que outros drivers
    # entregaram mas foram atribuídos a este driver)
    if incoming:
        incoming_qs = CainiaoOperationTask.objects.filter(
            waybill_number__in=incoming,
            task_date__range=(date_from, date_to),
            task_status="Delivered",
        )
        if seen_task_ids:
            incoming_qs = incoming_qs.exclude(id__in=seen_task_ids)
        rows_inc = list(incoming_qs.values("id", "waybill_number"))
        if rows_inc:
            ids_inc = [r["id"] for r in rows_inc]
            n_inc = len(ids_inc)

            # Aplicar overrides de preço também nos transferidos
            price_groups_inc = {}
            for r in rows_inc:
                po = price_overrides_map.get(r["waybill_number"])
                unit_price = po.price if po else price
                if unit_price not in price_groups_inc:
                    price_groups_inc[unit_price] = {
                        "n": 0, "reasons": set(),
                    }
                price_groups_inc[unit_price]["n"] += 1
                if po and po.reason:
                    price_groups_inc[unit_price]["reasons"].add(po.reason)

            price_breakdown_inc = []
            base_inc = Decimal("0")
            for unit_price, info in sorted(
                price_groups_inc.items(), reverse=True,
            ):
                line_total = unit_price * info["n"]
                base_inc += line_total
                price_breakdown_inc.append({
                    "price": str(unit_price),
                    "n": info["n"],
                    "total": f"{line_total:.2f}",
                    "is_special": unit_price != price,
                    "reasons": sorted(info["reasons"]),
                })

            b_lines_inc, b_total_inc = _bonus_lines_for_qs(
                CainiaoOperationTask.objects.filter(id__in=ids_inc),
            )
            login_breakdown.append({
                "label": "↻ Transferências recebidas",
                "courier_id": "",
                "courier_name": "(transferências)",
                "source": "transfer",
                "deliveries": n_inc,
                "base_entregas": f"{base_inc:.2f}",
                "price_breakdown": price_breakdown_inc,
                "bonus_lines": b_lines_inc,
                "total_bonus": f"{b_total_inc:.2f}",
                "subtotal": f"{(base_inc + b_total_inc):.2f}",
                "is_transfer": True,
            })
            total_delivered += n_inc
            total_base += base_inc
            total_bonus += b_total_inc

    # Bónus agregado (lista plana p/ retrocompatibilidade — é a soma de todos)
    bonus_lines_all = []
    for lb in login_breakdown:
        for bl in lb["bonus_lines"]:
            bonus_lines_all.append({**bl, "login": lb["label"]})

    overlapping = list(DriverPreInvoice.objects.filter(
        driver=driver,
        periodo_inicio__lte=date_to,
        periodo_fim__gte=date_from,
    ).values("id", "numero", "status", "periodo_inicio", "periodo_fim"))
    has_overlap = bool(overlapping)
    overlapping_serializable = [
        {
            "id": o["id"], "numero": o["numero"],
            "status": o["status"],
            "periodo_inicio": o["periodo_inicio"].strftime("%Y-%m-%d"),
            "periodo_fim": o["periodo_fim"].strftime("%Y-%m-%d"),
        }
        for o in overlapping
    ]

    return JsonResponse({
        "success": True,
        "period": {
            "from": date_from.strftime("%Y-%m-%d"),
            "to": date_to.strftime("%Y-%m-%d"),
        },
        "deliveries": total_delivered,
        "price_per_package": str(price),
        "price_source": price_source,
        "base_entregas": f"{total_base:.2f}",
        "bonus_lines": bonus_lines_all,
        "total_bonus": f"{total_bonus:.2f}",
        "subtotal": f"{(total_base + total_bonus):.2f}",
        "login_breakdown": login_breakdown,
        "n_logins": len(login_breakdown),
        "has_overlap": has_overlap,
        "overlapping": overlapping_serializable,
        # PF individual proibida quando motorista pertence a frota
        "is_fleet_driver": bool(driver.empresa_parceira_id),
        "empresa_parceira_id": driver.empresa_parceira_id,
        "empresa_parceira_nome": (
            driver.empresa_parceira.nome
            if driver.empresa_parceira_id else ""
        ),
        "can_create": (
            (not has_overlap)
            and total_delivered > 0
            and not driver.empresa_parceira_id
        ),
    })


# ──────────────────────────────────────────────────────────────────────────
#  Fase 6.2 — Lote de PFs por Frota (EmpresaParceira)
# ──────────────────────────────────────────────────────────────────────────

def _empresa_lote_compute(empresa, date_from, date_to):
    """Calcula lote de PFs para todos os drivers activos de uma frota.

    Para cada driver:
      - Identifica entregas Cainiao no período
      - Calcula bónus (domingos + feriados com ≥30 entregas)
      - Resolve preço via cascata
    Detecta overlap com PFs existentes.

    Returns:
        dict com {drivers: [...], totals, has_any_overlap}
    """
    from datetime import timedelta
    from decimal import Decimal
    from drivers_app.models import DriverProfile
    from core.models import Partner
    from core.finance import resolve_driver_price
    from .models import (
        BonusBlackoutDate, CainiaoOperationTask, DriverClaim, DriverPreInvoice,
        Holiday, PreInvoiceBonus,
    )

    blocked_dates = BonusBlackoutDate.dates_in_range(date_from, date_to)

    cainiao_partner = Partner.objects.filter(
        name__iexact="CAINIAO"
    ).first()
    drivers = list(
        DriverProfile.objects
        .filter(empresa_parceira=empresa, is_active=True)
        .select_related("empresa_parceira")
        .order_by("nome_completo")
    )

    rows = []
    total_deliveries = 0
    total_bonus = Decimal("0")
    total_amount = Decimal("0")
    has_any_overlap = False

    for d in drivers:
        # Identidade Cainiao — construir lista de logins distintos
        # (cada combinação courier_id+courier_name é um "login" para
        # efeitos de bónus, consistente com a PF individual).
        logins = []
        seen_pairs = set()

        def _add_login(cid, cname, source):
            cid = (cid or "").strip()
            cname = (cname or "").strip()
            if not cid and not cname:
                return
            key = (cid, cname)
            if key in seen_pairs:
                return
            seen_pairs.add(key)
            logins.append({
                "courier_id": cid, "courier_name": cname,
                "source": source,
            })

        if d.courier_id_cainiao or d.apelido:
            _add_login(d.courier_id_cainiao, d.apelido, "perfil")
        for m in d.courier_mappings.filter(partner=cainiao_partner):
            _add_login(m.courier_id, m.courier_name, "mapping")

        if not logins:
            rows.append({
                "driver_id": d.id,
                "name": d.nome_completo,
                "apelido": d.apelido,
                "deliveries": 0,
                "bonus_days": 0,
                "bonus_amount": "0.00",
                "price_per_package": "0",
                "price_source": "none",
                "base": "0.00",
                "subtotal": "0.00",
                "has_overlap": False,
                "skip_reason": "sem identidade Cainiao",
                "login_breakdown": [],
            })
            continue

        from django.db.models import Q
        base_qs = CainiaoOperationTask.objects.filter(
            task_date__range=(date_from, date_to),
            task_status="Delivered",
        )

        price, price_source = resolve_driver_price(d, cainiao_partner)

        # Iterar por login: cada login conta as entregas que NÃO foram
        # já vistas em logins anteriores (evita dupla contagem quando
        # cid/cname se sobrepõem).
        n_delivered = 0
        bonus_days = 0
        bonus_amount = Decimal("0")
        login_breakdown = []
        seen_task_ids = set()

        for ln in logins:
            cid = ln["courier_id"]
            cname = ln["courier_name"]
            login_q = Q()
            if cid:
                login_q |= Q(courier_id_cainiao=cid)
            if cname:
                login_q |= Q(courier_name=cname)
            if not login_q:
                continue
            qs_login = base_qs.filter(login_q)
            if seen_task_ids:
                qs_login = qs_login.exclude(id__in=seen_task_ids)
            ids_this = list(qs_login.values_list("id", flat=True))
            if not ids_this:
                continue
            seen_task_ids.update(ids_this)
            n_login = len(ids_this)
            n_delivered += n_login

            # Bónus por dia DESTE login (regra: por login, não agregado)
            by_day = list(
                CainiaoOperationTask.objects.filter(id__in=ids_this)
                .values("task_date").annotate(n=Count("id"))
                .order_by("task_date")
            )
            login_bonus = Decimal("0")
            login_bonus_days = 0
            for row in by_day:
                day = row["task_date"]
                n = row["n"]
                is_sun = day.weekday() == 6
                h = Holiday.get_holiday(day)
                if not (is_sun or h):
                    continue
                if day in blocked_dates:
                    continue
                if n >= PreInvoiceBonus.LIMIAR_60:
                    login_bonus += PreInvoiceBonus.BONUS_50
                    login_bonus_days += 1
                elif n >= PreInvoiceBonus.LIMIAR_30:
                    login_bonus += PreInvoiceBonus.BONUS_30
                    login_bonus_days += 1
            bonus_amount += login_bonus
            bonus_days += login_bonus_days

            login_breakdown.append({
                "label": cname or cid,
                "courier_id": cid,
                "courier_name": cname,
                "source": ln["source"],
                "deliveries": n_login,
                "bonus_days": login_bonus_days,
                "bonus_amount": f"{login_bonus:.2f}",
            })

        base = price * n_delivered

        # Claims do driver no período (mesma regra que empresa_lote_emit:
        # APPROVED + PENDING; exclui REJECTED/APPEALED).
        claims_count = 0
        claims_amount = Decimal("0")
        for c in DriverClaim.objects.filter(driver=d).exclude(
            status__in=("REJECTED", "APPEALED")
        ).exclude(waybill_number=""):
            if c.operation_task_date is not None:
                in_period = date_from <= c.operation_task_date <= date_to
            else:
                in_period = CainiaoOperationTask.objects.filter(
                    waybill_number=c.waybill_number,
                    task_date__range=(date_from, date_to),
                ).exists()
            if not in_period:
                continue
            claims_amount += Decimal(str(c.amount or 0))
            claims_count += 1

        # Overlap
        overlap = DriverPreInvoice.objects.filter(
            driver=d,
            periodo_inicio__lte=date_to,
            periodo_fim__gte=date_from,
        ).exclude(status="CANCELADO").exists()
        if overlap:
            has_any_overlap = True

        subtotal = base + bonus_amount - claims_amount
        total_deliveries += n_delivered
        total_bonus += bonus_amount
        total_amount += subtotal

        rows.append({
            "driver_id": d.id,
            "name": d.nome_completo,
            "apelido": d.apelido,
            "deliveries": n_delivered,
            "bonus_days": bonus_days,
            "bonus_amount": str(bonus_amount),
            "claims_count": claims_count,
            "claims_amount": str(claims_amount),
            "price_per_package": str(price),
            "price_source": price_source,
            "base": f"{base:.2f}",
            "subtotal": f"{subtotal:.2f}",
            "has_overlap": overlap,
            "skip_reason": "" if n_delivered > 0 else "sem entregas no período",
            "login_breakdown": login_breakdown,
            "n_logins": len(login_breakdown),
        })

    total_claims = sum(
        Decimal(r.get("claims_amount", "0")) for r in rows
    )
    return {
        "rows": rows,
        "totals": {
            "drivers": len(drivers),
            "drivers_with_deliveries": sum(
                1 for r in rows if r["deliveries"] > 0
            ),
            "deliveries": total_deliveries,
            "bonus": str(total_bonus),
            "claims": str(total_claims),
            "amount": f"{total_amount:.2f}",
        },
        "has_any_overlap": has_any_overlap,
    }


@login_required
def empresa_lote_preview(request, empresa_id):
    """Preview do lote de PFs para uma frota num período."""
    from datetime import timedelta
    from drivers_app.models import EmpresaParceira

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    today = timezone.now().date()
    date_from = parse_date(request.GET.get("from") or "") or today
    date_to = parse_date(request.GET.get("to") or "") or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    result = _empresa_lote_compute(empresa, date_from, date_to)
    return JsonResponse({
        "success": True,
        "empresa": {
            "id": empresa.id,
            "nome": empresa.nome,
            "default_price": (
                str(empresa.driver_default_price_per_package)
                if empresa.driver_default_price_per_package else ""
            ),
        },
        "period": {
            "from": date_from.strftime("%Y-%m-%d"),
            "to": date_to.strftime("%Y-%m-%d"),
        },
        **result,
    })


@login_required
@require_http_methods(["POST"])
def empresa_lote_emit(request, empresa_id):
    """Emite UMA FleetInvoice global da frota com 1 linha por driver.

    Body JSON:
      from=YYYY-MM-DD, to=YYYY-MM-DD
      driver_ids=[id, ...] (opcional) — emitir só para subset
      skip_overlap=bool (default True) — pula drivers já incluídos
        em outra FleetInvoice no mesmo período
    """
    import json
    from datetime import timedelta
    from decimal import Decimal
    from django.db import transaction
    from django.db.models import Q
    from drivers_app.models import EmpresaParceira, DriverProfile
    from .models import (
        BonusBlackoutDate, CainiaoOperationTask, FleetInvoice,
        FleetInvoiceDriverLine, FleetInvoiceBonusDay, FleetInvoiceClaim,
        Holiday, PreInvoiceBonus, DriverClaim,
    )
    from core.models import Partner
    from core.finance import resolve_driver_price, resolve_partner_price

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    date_from = parse_date(body.get("from") or "")
    date_to = parse_date(body.get("to") or "")
    if not date_from or not date_to:
        return JsonResponse(
            {"success": False, "error":
             "from e to obrigatórios"}, status=400,
        )
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    skip_overlap = body.get("skip_overlap", True)
    driver_ids_filter = body.get("driver_ids") or []

    cainiao_partner = Partner.objects.filter(
        name__iexact="CAINIAO"
    ).first()
    if not cainiao_partner:
        return JsonResponse(
            {"success": False, "error":
             "Parceiro CAINIAO não configurado."}, status=400,
        )

    # Detecção de overlap a nível de FleetInvoice (uma única por período)
    if skip_overlap:
        existing_overlap = FleetInvoice.objects.filter(
            empresa=empresa,
            periodo_inicio__lte=date_to,
            periodo_fim__gte=date_from,
        ).exclude(status="CANCELADO").first()
        if existing_overlap:
            return JsonResponse({
                "success": False,
                "error": (
                    f"Já existe FleetInvoice {existing_overlap.numero} "
                    f"({existing_overlap.periodo_inicio}→"
                    f"{existing_overlap.periodo_fim}) que sobrepõe este "
                    f"período. Cancela-a ou usa outro intervalo."
                ),
                "existing_id": existing_overlap.id,
            }, status=409)

    # Próximo número FF-NNNN
    ultimo = (
        FleetInvoice.objects.filter(numero__startswith="FF-")
        .order_by("-numero").first()
    )
    seq = 1
    if ultimo:
        try:
            seq = int(ultimo.numero.split("-")[1]) + 1
        except (IndexError, ValueError):
            seq = 1

    drivers_qs = DriverProfile.objects.filter(
        empresa_parceira=empresa, is_active=True,
    )
    if driver_ids_filter:
        drivers_qs = drivers_qs.filter(id__in=driver_ids_filter)

    blocked_dates = BonusBlackoutDate.dates_in_range(date_from, date_to)

    partner_price = resolve_partner_price(cainiao_partner)
    skipped = []
    lines_data = []  # acumulamos antes do create

    for d in drivers_qs.order_by("nome_completo"):
        # Construir lista de logins distintos (consistente com PF individual)
        logins = []
        seen_pairs = set()

        def _add(cid, cname):
            cid = (cid or "").strip()
            cname = (cname or "").strip()
            if not cid and not cname:
                return
            key = (cid, cname)
            if key in seen_pairs:
                return
            seen_pairs.add(key)
            logins.append({"courier_id": cid, "courier_name": cname})

        if d.courier_id_cainiao or d.apelido:
            _add(d.courier_id_cainiao, d.apelido)
        for m in d.courier_mappings.filter(partner=cainiao_partner):
            _add(m.courier_id, m.courier_name)

        if not logins:
            skipped.append({
                "driver_id": d.id, "name": d.nome_completo,
                "reason": "sem identidade Cainiao",
            })
            continue

        base_qs = CainiaoOperationTask.objects.filter(
            task_date__range=(date_from, date_to),
            task_status="Delivered",
        )

        # Iterar por login: cada login conta entregas próprias e gera os
        # seus próprios bónus (≥30/≥60 entregas em domingo/feriado).
        # Evita dupla-contagem via seen_task_ids.
        n_delivered = 0
        bonus_days_detail = []
        bonus_amount = Decimal("0")
        seen_task_ids = set()
        all_ids = set()
        all_names = set()
        for ln in logins:
            cid = ln["courier_id"]
            cname = ln["courier_name"]
            if cid:
                all_ids.add(cid)
            if cname:
                all_names.add(cname)
            login_q = Q()
            if cid:
                login_q |= Q(courier_id_cainiao=cid)
            if cname:
                login_q |= Q(courier_name=cname)
            if not login_q:
                continue
            qs_login = base_qs.filter(login_q)
            if seen_task_ids:
                qs_login = qs_login.exclude(id__in=seen_task_ids)
            ids_this = list(qs_login.values_list("id", flat=True))
            if not ids_this:
                continue
            seen_task_ids.update(ids_this)
            n_delivered += len(ids_this)

            by_day = list(
                CainiaoOperationTask.objects.filter(id__in=ids_this)
                .values("task_date").annotate(n=Count("id"))
                .order_by("task_date")
            )
            for row in by_day:
                day = row["task_date"]
                n = row["n"]
                is_sun = day.weekday() == 6
                h = Holiday.get_holiday(day)
                if not (is_sun or h):
                    continue
                if day in blocked_dates:
                    continue
                if n >= PreInvoiceBonus.LIMIAR_60:
                    b = PreInvoiceBonus.BONUS_50
                elif n >= PreInvoiceBonus.LIMIAR_30:
                    b = PreInvoiceBonus.BONUS_30
                else:
                    continue
                bonus_amount += b
                bonus_days_detail.append({
                    "data": day,
                    "tipo": "FERIADO" if h else "DOMINGO",
                    "deliveries": n,
                    "bonus": b,
                    "feriado_nome": h.name if h else "",
                    "login": cname or cid,  # contexto: qual login gerou
                })

        if n_delivered == 0:
            skipped.append({
                "driver_id": d.id, "name": d.nome_completo,
                "reason": "sem entregas no período",
            })
            continue

        price, src = resolve_driver_price(d, cainiao_partner)
        base = price * n_delivered
        ids = all_ids
        names = all_names

        # Claims do driver no período. Inclui APPROVED + PENDING
        # (exclui REJECTED/APPEALED — ainda em discussão).
        claims = list(
            DriverClaim.objects.filter(driver=d)
            .exclude(status__in=("REJECTED", "APPEALED"))
            .exclude(waybill_number="")
        )
        claims_detail = []
        claims_amount = Decimal("0")
        for c in claims:
            valor = c.amount or Decimal("0")
            # Período: prefere operation_task_date (mais fiável); cai para
            # presença da waybill no CainiaoOperationTask se vazio.
            in_period = False
            if c.operation_task_date is not None:
                in_period = date_from <= c.operation_task_date <= date_to
            else:
                in_period = CainiaoOperationTask.objects.filter(
                    waybill_number=c.waybill_number,
                    task_date__range=(date_from, date_to),
                ).exists()
            if not in_period:
                continue
            claims_amount += Decimal(str(valor))
            claims_detail.append({
                "claim_id": c.id,
                "waybill_number": c.waybill_number,
                "valor": Decimal(str(valor)),
                "descricao": (c.description or "")[:300],
            })

        first_id = (
            sorted(ids)[0] if ids
            else ""
        )
        first_name = (
            sorted(names)[0] if names else ""
        )
        subtotal = base + bonus_amount - claims_amount

        lines_data.append({
            "driver": d,
            "courier_id_snapshot": first_id,
            "courier_name_snapshot": first_name,
            "deliveries": n_delivered,
            "price": price,
            "src": src,
            "base": base,
            "bonus_days_count": len(bonus_days_detail),
            "bonus_amount": bonus_amount,
            "bonus_days_detail": bonus_days_detail,
            "claims_count": len(claims_detail),
            "claims_amount": claims_amount,
            "claims_detail": claims_detail,
            "subtotal": subtotal,
        })

    if not lines_data:
        return JsonResponse({
            "success": False,
            "error": "Nenhum motorista com entregas no período.",
            "skipped": skipped,
        }, status=400)

    # Criar tudo numa transação
    with transaction.atomic():
        fi = FleetInvoice.objects.create(
            numero=f"FF-{seq:04d}",
            empresa=empresa,
            periodo_inicio=date_from,
            periodo_fim=date_to,
            partner_price_per_package=partner_price,
            status="CALCULADO",
            observacoes=f"Lote auto-emitido para {empresa.nome}",
            created_by=request.user,
        )
        for ld in lines_data:
            line = FleetInvoiceDriverLine.objects.create(
                fleet_invoice=fi,
                driver=ld["driver"],
                driver_name_snapshot=ld["driver"].nome_completo,
                courier_id_snapshot=ld["courier_id_snapshot"],
                courier_name_snapshot=ld["courier_name_snapshot"],
                deliveries=ld["deliveries"],
                price_per_package=ld["price"],
                price_source=ld["src"],
                base_amount=ld["base"],
                bonus_days_count=ld["bonus_days_count"],
                bonus_amount=ld["bonus_amount"],
                claims_count=ld["claims_count"],
                claims_amount=ld["claims_amount"],
                subtotal=ld["subtotal"],
            )
            for bd in ld["bonus_days_detail"]:
                # 'login' é apenas contexto informativo (não persistido)
                bd_clean = {k: v for k, v in bd.items() if k != "login"}
                FleetInvoiceBonusDay.objects.create(
                    line=line, **bd_clean,
                )
            for cd in ld["claims_detail"]:
                FleetInvoiceClaim.objects.create(
                    line=line,
                    driver_claim_id=cd["claim_id"],
                    waybill_number=cd["waybill_number"],
                    valor=cd["valor"],
                    descricao=cd["descricao"],
                )
        fi.recalcular()

    return JsonResponse({
        "success": True,
        "fleet_invoice": {
            "id": fi.id,
            "numero": fi.numero,
            "periodo_inicio": fi.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": fi.periodo_fim.strftime("%Y-%m-%d"),
            "total_deliveries": fi.total_deliveries,
            "total_base": str(fi.total_base),
            "total_bonus": str(fi.total_bonus),
            "total_claims": str(fi.total_claims),
            "total_a_receber": str(fi.total_a_receber),
            "n_drivers": len(lines_data),
        },
        "skipped": skipped,
        "summary": {
            "n_created": 1,  # uma única FleetInvoice
            "n_drivers_in_invoice": len(lines_data),
            "n_skipped": len(skipped),
            "total_amount": str(fi.total_a_receber),
        },
    })


@login_required
def empresa_dashboard_stats(request, empresa_id):
    """Dashboard agregado de uma frota.

    Query params: from=YYYY-MM-DD, to=YYYY-MM-DD (default: mês actual)

    Devolve:
      - Total drivers, drivers activos, drivers c/ entregas no período
      - Total entregas, valor pago, receita, margem
      - Top 5 drivers (ranking)
      - Comparação com período anterior (mesma duração)
    """
    from datetime import timedelta
    from decimal import Decimal
    from django.db.models import Q
    from drivers_app.models import EmpresaParceira, DriverProfile
    from core.models import Partner
    from core.finance import resolve_driver_price, resolve_partner_price
    from .models import CainiaoOperationTask

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    today = timezone.now().date()
    date_from = parse_date(request.GET.get("from") or "")
    date_to = parse_date(request.GET.get("to") or "")
    if not date_from or not date_to:
        # Default: mês actual
        date_from = today.replace(day=1)
        if today.month == 12:
            date_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            date_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    period_days = (date_to - date_from).days + 1
    prev_to = date_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=period_days - 1)

    cainiao_partner = Partner.objects.filter(name__iexact="CAINIAO").first()
    partner_price = resolve_partner_price(cainiao_partner)

    drivers = list(
        DriverProfile.objects
        .filter(empresa_parceira=empresa)
        .select_related("empresa_parceira")
    )
    drivers_active = sum(1 for d in drivers if d.is_active)

    def _agg_for_period(d_from, d_to):
        total_deliveries = 0
        total_amount_driver = Decimal("0")
        ranking = []
        drivers_with = 0
        for d in drivers:
            if not d.is_active:
                continue
            ids = set(); names = set()
            if d.courier_id_cainiao: ids.add(d.courier_id_cainiao)
            if d.apelido: names.add(d.apelido)
            for m in d.courier_mappings.filter(partner=cainiao_partner):
                if m.courier_id: ids.add(m.courier_id)
                if m.courier_name: names.add(m.courier_name)
            if not (ids or names):
                continue
            qs = CainiaoOperationTask.objects.filter(
                task_date__range=(d_from, d_to),
                task_status="Delivered",
            )
            dq = Q()
            if ids: dq |= Q(courier_id_cainiao__in=ids)
            if names: dq |= Q(courier_name__in=names)
            n = qs.filter(dq).count()
            if n == 0:
                continue
            drivers_with += 1
            price, _ = resolve_driver_price(d, cainiao_partner)
            amt = price * n
            total_deliveries += n
            total_amount_driver += amt
            ranking.append({
                "driver_id": d.id,
                "name": d.nome_completo,
                "apelido": d.apelido,
                "deliveries": n,
                "amount": f"{amt:.2f}",
                "price_per_package": str(price),
            })
        ranking.sort(key=lambda r: -r["deliveries"])
        return {
            "deliveries": total_deliveries,
            "drivers_with_deliveries": drivers_with,
            "amount_driver": str(total_amount_driver),
            "ranking": ranking[:5],
        }

    cur = _agg_for_period(date_from, date_to)
    prev = _agg_for_period(prev_from, prev_to)

    receita = partner_price * cur["deliveries"]
    margem = receita - Decimal(cur["amount_driver"])
    avg_per_driver = (
        Decimal(cur["amount_driver"]) / cur["drivers_with_deliveries"]
        if cur["drivers_with_deliveries"] else Decimal("0")
    )

    def _delta_pct(c, p):
        if p == 0:
            return 100.0 if c > 0 else 0.0
        return round((c - p) * 100.0 / p, 1)

    return JsonResponse({
        "success": True,
        "empresa": {
            "id": empresa.id, "nome": empresa.nome,
            "default_price": (
                str(empresa.driver_default_price_per_package)
                if empresa.driver_default_price_per_package else ""
            ),
        },
        "period": {
            "from": date_from.strftime("%Y-%m-%d"),
            "to": date_to.strftime("%Y-%m-%d"),
        },
        "previous_period": {
            "from": prev_from.strftime("%Y-%m-%d"),
            "to": prev_to.strftime("%Y-%m-%d"),
        },
        "summary": {
            "drivers_total": len(drivers),
            "drivers_active": drivers_active,
            "drivers_with_deliveries": cur["drivers_with_deliveries"],
            "deliveries": cur["deliveries"],
            "amount_driver": cur["amount_driver"],
            "receita": f"{receita:.2f}",
            "margem": f"{margem:.2f}",
            "avg_per_driver": f"{avg_per_driver:.2f}",
            "partner_price_per_package": str(partner_price),
        },
        "delta": {
            "deliveries_pct": _delta_pct(cur["deliveries"], prev["deliveries"]),
            "amount_pct": _delta_pct(
                float(cur["amount_driver"]),
                float(prev["amount_driver"]),
            ),
        },
        "ranking": cur["ranking"],
    })


@login_required
def empresa_fleet_invoices_list(request, empresa_id):
    """Lista FleetInvoices de uma frota num ano/mês."""
    from drivers_app.models import EmpresaParceira
    from .models import FleetInvoice

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    qs = FleetInvoice.objects.filter(empresa=empresa).order_by(
        "-periodo_fim", "-id",
    )
    try:
        mes = int(request.GET.get("mes") or 0)
        ano = int(request.GET.get("ano") or 0)
    except (TypeError, ValueError):
        mes, ano = 0, 0
    if mes and ano:
        qs = qs.filter(periodo_inicio__year=ano, periodo_inicio__month=mes)

    rows = []
    for fi in qs:
        rows.append({
            "id": fi.id,
            "numero": fi.numero,
            "periodo_inicio": fi.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": fi.periodo_fim.strftime("%Y-%m-%d"),
            "periodo_display": (
                f"{fi.periodo_inicio.strftime('%d/%m')}–"
                f"{fi.periodo_fim.strftime('%d/%m/%Y')}"
            ),
            "status": fi.status,
            "status_display": fi.get_status_display(),
            "total_deliveries": fi.total_deliveries,
            "total_a_receber": str(fi.total_a_receber),
            "vat_rate": str(fi.vat_rate),
            "vat_amount": str(fi.vat_amount),
            "total_com_iva": str(fi.total_com_iva),
            "n_drivers": fi.lines.count(),
        })
    return JsonResponse({"success": True, "rows": rows})


@login_required
def empresa_fleet_invoice_detail(request, fleet_invoice_id):
    """Detalhe completo de uma FleetInvoice (linhas + bónus + claims)."""
    from .models import FleetInvoice

    fi = get_object_or_404(
        FleetInvoice.objects.select_related("empresa")
        .prefetch_related(
            "lines__driver",
            "lines__bonus_days_detail",
            "lines__claims_detail",
        ),
        id=fleet_invoice_id,
    )

    lines = []
    for line in fi.lines.all().order_by("-deliveries"):
        lines.append({
            "id": line.id,
            "driver_id": line.driver_id,
            "driver_name": (
                line.driver_name_snapshot
                or (line.driver.nome_completo if line.driver else "")
            ),
            "courier_id": line.courier_id_snapshot,
            "courier_name": line.courier_name_snapshot,
            "deliveries": line.deliveries,
            "price_per_package": str(line.price_per_package),
            "price_source": line.price_source,
            "base_amount": str(line.base_amount),
            "bonus_days_count": line.bonus_days_count,
            "bonus_amount": str(line.bonus_amount),
            "claims_count": line.claims_count,
            "claims_amount": str(line.claims_amount),
            "subtotal": str(line.subtotal),
            "bonus_days": [
                {
                    "data": bd.data.strftime("%Y-%m-%d"),
                    "tipo": bd.tipo,
                    "deliveries": bd.deliveries,
                    "bonus": str(bd.bonus),
                    "feriado_nome": bd.feriado_nome,
                }
                for bd in line.bonus_days_detail.all()
            ],
            "claims": [
                {
                    "claim_id": cd.driver_claim_id,
                    "waybill_number": cd.waybill_number,
                    "valor": str(cd.valor),
                    "descricao": cd.descricao,
                }
                for cd in line.claims_detail.all()
            ],
        })

    return JsonResponse({
        "success": True,
        "fleet_invoice": {
            "id": fi.id,
            "numero": fi.numero,
            "empresa_id": fi.empresa_id,
            "empresa_nome": fi.empresa.nome,
            "periodo_inicio": fi.periodo_inicio.strftime("%Y-%m-%d"),
            "periodo_fim": fi.periodo_fim.strftime("%Y-%m-%d"),
            "status": fi.status,
            "status_display": fi.get_status_display(),
            "partner_price_per_package": str(fi.partner_price_per_package),
            "total_deliveries": fi.total_deliveries,
            "total_base": str(fi.total_base),
            "total_bonus": str(fi.total_bonus),
            "total_claims": str(fi.total_claims),
            "total_a_receber": str(fi.total_a_receber),
            "vat_rate": str(fi.vat_rate),
            "vat_amount": str(fi.vat_amount),
            "total_com_iva": str(fi.total_com_iva),
            "data_pagamento": (
                fi.data_pagamento.strftime("%Y-%m-%d")
                if fi.data_pagamento else ""
            ),
            "referencia_pagamento": fi.referencia_pagamento,
            "observacoes": fi.observacoes,
            "lines": lines,
        },
    })


@login_required
@require_http_methods(["POST"])
def empresa_fleet_invoice_change_status(request, fleet_invoice_id):
    """Muda o status de uma FleetInvoice (com validação de transição)."""
    import json
    from datetime import date as _date
    from .models import FleetInvoice

    fi = get_object_or_404(FleetInvoice, id=fleet_invoice_id)
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}
    novo = (body.get("status") or "").strip().upper()
    if novo not in dict(FleetInvoice.STATUS_CHOICES):
        return JsonResponse(
            {"success": False, "error": "Status inválido"}, status=400,
        )
    permitido = FleetInvoice.TRANSICOES.get(fi.status, [])
    if novo not in permitido:
        return JsonResponse({
            "success": False,
            "error": (
                f"Transição {fi.status} → {novo} não permitida. "
                f"Permitidas: {permitido}"
            ),
        }, status=400)
    fi.status = novo
    if novo == "PAGO":
        fi.data_pagamento = fi.data_pagamento or _date.today()
        ref = (body.get("referencia_pagamento") or "").strip()
        if ref:
            fi.referencia_pagamento = ref
    fi.save()
    return JsonResponse({
        "success": True, "status": fi.status,
        "status_display": fi.get_status_display(),
    })


@login_required
@require_http_methods(["POST"])
def empresa_fleet_invoice_delete(request, fleet_invoice_id):
    """Apaga uma FleetInvoice (apenas se em RASCUNHO ou CANCELADO)."""
    from .models import FleetInvoice
    fi = get_object_or_404(FleetInvoice, id=fleet_invoice_id)
    if fi.status not in ("RASCUNHO", "CANCELADO"):
        return JsonResponse({
            "success": False,
            "error": (
                f"Não é possível apagar FleetInvoice em status {fi.status}. "
                f"Cancela primeiro."
            ),
        }, status=400)
    fi.delete()
    return JsonResponse({"success": True})


@login_required
def empresa_fleet_invoice_pdf(request, fleet_invoice_id):
    """Gera PDF de uma FleetInvoice (delegado a PDFGenerator)."""
    from django.http import HttpResponse
    from .models import FleetInvoice

    fi = get_object_or_404(
        FleetInvoice.objects.select_related("empresa")
        .prefetch_related(
            "lines__bonus_days_detail",
            "lines__claims_detail",
        ),
        id=fleet_invoice_id,
    )

    try:
        gen = PDFGenerator()
        buf = gen.generate_fleet_invoice_pdf(fi)
    except Exception as e:
        return JsonResponse(
            {"error": f"Erro ao gerar PDF: {e}"}, status=500,
        )

    safe = fi.empresa.nome.replace(" ", "_")[:30]
    filename = f"{fi.numero}_{safe}.pdf"
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
@require_http_methods(["POST"])
def empresa_whatsapp_lote(request, empresa_id):
    """Envia mensagens WhatsApp em lote a drivers da frota com PFs no período.

    Body JSON:
      from=YYYY-MM-DD, to=YYYY-MM-DD
      pf_ids=[id, ...] (opcional) — só estas PFs

    Loop pelos drivers da frota; se a PF cobrir o período, envia. Reutiliza
    a lógica de pre_invoice_send_whatsapp.
    """
    import json
    from django.conf import settings
    from drivers_app.models import EmpresaParceira
    from system_config.whatsapp_helper import WhatsAppWPPConnectAPI

    empresa = get_object_or_404(EmpresaParceira, id=empresa_id)
    try:
        body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        body = {}

    date_from = parse_date(body.get("from") or "")
    date_to = parse_date(body.get("to") or "")
    pf_ids_filter = body.get("pf_ids") or []

    qs = DriverPreInvoice.objects.filter(
        driver__empresa_parceira=empresa,
    ).select_related("driver").exclude(status="CANCELADO")
    if date_from and date_to:
        qs = qs.filter(
            periodo_inicio__lte=date_to,
            periodo_fim__gte=date_from,
        )
    if pf_ids_filter:
        qs = qs.filter(id__in=pf_ids_filter)

    if not settings.WPPCONNECT_URL:
        return JsonResponse(
            {"success": False, "error":
             "WPPCONNECT_URL não configurado em settings."},
            status=500,
        )
    api = WhatsAppWPPConnectAPI(
        base_url=settings.WPPCONNECT_URL,
        session_name=settings.WPPCONNECT_SESSION,
        auth_token=settings.WPPCONNECT_TOKEN,
        secret_key=settings.WPPCONNECT_SECRET,
    )

    sent = []
    failed = []
    no_phone = []

    base = (getattr(settings, "BASE_URL", "") or "").rstrip("/")

    for pf in qs:
        d = pf.driver
        telefone = (d.telefone or "").strip()
        if not telefone:
            no_phone.append({"pf_id": pf.id, "driver_id": d.id,
                             "name": d.nome_completo})
            continue
        digits = "".join(c for c in telefone if c.isdigit())
        if len(digits) == 9 and digits.startswith("9"):
            digits = "351" + digits

        msg = (
            f"📋 Pré-Fatura *{pf.numero}*\n"
            f"👤 {d.nome_completo}\n"
            f"🗓️ {pf.periodo_inicio.strftime('%d/%m/%Y')}"
            f" → {pf.periodo_fim.strftime('%d/%m/%Y')}\n\n"
            f"📦 Entregas: €{pf.base_entregas}\n"
            f"🎁 Bónus: €{pf.total_bonus}\n"
            f"💰 Total: *€{pf.total_a_receber}*\n"
            f"\nEstado: {pf.get_status_display()}"
        )
        if base:
            msg += f"\n📎 PDF: {base}/settlements/pre-invoices/{pf.id}/pdf/"

        try:
            api.send_text(number=digits, text=msg)
            sent.append({"pf_id": pf.id, "driver_id": d.id,
                         "name": d.nome_completo, "phone": digits})
        except Exception as e:
            failed.append({"pf_id": pf.id, "driver_id": d.id,
                           "name": d.nome_completo,
                           "error": str(e)[:200]})

    return JsonResponse({
        "success": True,
        "sent": sent,
        "failed": failed,
        "no_phone": no_phone,
        "summary": {
            "n_sent": len(sent),
            "n_failed": len(failed),
            "n_no_phone": len(no_phone),
        },
    })


# ============================================================================
# Histórico completo de uma waybill (vida do pacote)
# ============================================================================

@login_required
def waybill_history(request, waybill):
    """Retorna a timeline completa de um pacote.

    Combina:
      - Todos os snapshots em CainiaoOperationTask (1 por task_date)
      - Forecast em CainiaoForecastWaybill (planning original)
      - Detalhes em DriverCainiaoDetail (info do driver/courier)
    """
    from .models import (
        CainiaoOperationTask, CainiaoForecastPackage,
        CainiaoPlanningPackage, CainiaoDriverDetailRecord,
    )

    waybill = (waybill or "").strip()
    if not waybill:
        return JsonResponse(
            {"success": False, "error": "Waybill vazio"}, status=400,
        )

    # Snapshots operacionais (cada task_date é uma versão)
    ops = list(
        CainiaoOperationTask.objects.filter(waybill_number=waybill)
        .order_by("task_date", "id")
    )
    if not ops:
        # Tenta também por lp_number
        ops = list(
            CainiaoOperationTask.objects.filter(lp_number=waybill)
            .order_by("task_date", "id")
        )

    # Forecast Package (data do importe, status simples) — por tracking_number
    forecast = (
        CainiaoForecastPackage.objects
        .filter(tracking_number=waybill)
        .order_by("-operation_date").first()
    )

    # Planning Package (dados ricos: cliente, telefone, email, seller)
    # — por parcel_id
    planning = (
        CainiaoPlanningPackage.objects
        .filter(parcel_id=waybill)
        .order_by("-operation_date").first()
    )

    # Detalhe driver (mais recente)
    detail = (
        CainiaoDriverDetailRecord.objects
        .filter(waybill_number=waybill)
        .order_by("-created_at").first()
    )

    # Constrói timeline: forecast → snapshots ordenados
    timeline = []
    if forecast:
        timeline.append({
            "kind": "forecast",
            "date": forecast.operation_date.strftime("%Y-%m-%d"),
            "label": "Planeado (forecast)",
            "icon": "calendar-clock",
            "color": "blue",
            "details": {
                "city": forecast.receiver_city or "",
                "zip": forecast.receiver_zip or "",
                "lp_number": forecast.lp_number or "",
            },
        })

    # Snapshots por task_date
    last_status = None
    last_courier = None
    for op in ops:
        status = op.task_status or ""
        # Marca mudança de courier
        events = []
        if op.courier_name and op.courier_name != last_courier:
            events.append({
                "type": "courier_change",
                "from": last_courier or "",
                "to": op.courier_name,
            })
            last_courier = op.courier_name
        # Marca mudança de status
        if status != last_status:
            events.append({
                "type": "status_change",
                "from": last_status or "—",
                "to": status,
            })
            last_status = status

        timeline.append({
            "kind": "operation",
            "date": op.task_date.strftime("%Y-%m-%d"),
            "label": status or "Snapshot",
            "icon": (
                "check-circle" if status == "Delivered"
                else ("alert-circle" if "Fail" in status
                      else "package")
            ),
            "color": (
                "emerald" if status == "Delivered"
                else ("rose" if "Fail" in status
                      else ("amber" if status in (
                          "Driver Received", "Assigned",
                      ) else "slate"))
            ),
            "details": {
                "courier_name": op.courier_name or "",
                "courier_id_cainiao": op.courier_id_cainiao or "",
                "dsp_name": op.dsp_name or "",
                "delivery_type": op.delivery_type or "",
                "city": op.destination_city or "",
                "zip": op.zip_code or "",
                "address": op.detailed_address or "",
                "delivery_time": (
                    op.delivery_time.strftime("%Y-%m-%d %H:%M")
                    if op.delivery_time else ""
                ),
                "delivery_failure_time": (
                    op.delivery_failure_time.strftime("%Y-%m-%d %H:%M")
                    if op.delivery_failure_time else ""
                ),
                "exception_type": op.exception_type or "",
                "exception_detail": op.exception_detail or "",
                "weight_g": op.weight_g or 0,
                "lp_number": op.lp_number or "",
                "task_id": op.task_id or "",
                "sign_pod": op.sign_pod or "",
                "creation_time": (
                    op.creation_time.strftime("%Y-%m-%d %H:%M")
                    if op.creation_time else ""
                ),
                "outbound_time": (
                    op.outbound_time.strftime("%Y-%m-%d %H:%M")
                    if op.outbound_time else ""
                ),
                "start_delivery_time": (
                    op.start_delivery_time.strftime("%Y-%m-%d %H:%M")
                    if op.start_delivery_time else ""
                ),
                "lat": op.actual_latitude or "",
                "lng": op.actual_longitude or "",
                "events": events,
            },
        })

    # Estado actual (último snapshot)
    last = ops[-1] if ops else None
    current_status = last.task_status if last else ""
    current_courier = last.courier_name if last else ""

    # Cliente/destinatário (do planning, fallback para forecast)
    customer = {
        "name": "",
        "phone": "",
        "email": "",
    }
    if planning:
        customer["name"] = planning.receiver_name or ""
        customer["phone"] = planning.receiver_phone or ""
        customer["email"] = planning.receiver_email or ""
    elif forecast:
        customer["name"] = forecast.receiver_name or ""
        customer["phone"] = forecast.receiver_phone or ""

    # Telefone normalizado para wa.me
    wa_phone = "".join(c for c in customer["phone"] if c.isdigit())

    # Seller e dimensões (planning > op task > detail)
    seller_name = ""
    if planning and planning.seller_name:
        seller_name = planning.seller_name
    elif last and last.seller_name:
        seller_name = last.seller_name

    weight_g = 0
    if last and last.weight_g:
        weight_g = last.weight_g
    elif detail and detail.weight_g:
        weight_g = detail.weight_g

    dimensions = None
    if last and (last.length or last.width or last.height):
        dimensions = {
            "length": last.length or 0,
            "width": last.width or 0,
            "height": last.height or 0,
        }

    # Endereço/destino (do op task mais recente, fallback planning)
    dest = None
    if last:
        dest = {
            "city": last.destination_city or "",
            "zip": last.zip_code or "",
            "address": last.detailed_address or "",
            "area": last.destination_area or "",
            "country": last.destination_country or "",
            "zone": last.zone or "",
        }
    elif planning:
        dest = {
            "city": planning.receiver_city or "",
            "zip": planning.receiver_zip or "",
            "address": planning.receiver_address or "",
            "area": planning.receiver_province or "",
            "country": "PT",
            "zone": "",
        }

    # Coordenadas (lat/lng) para mapa
    geo = None
    if last and (last.receiver_latitude or last.receiver_longitude):
        try:
            geo = {
                "receiver_lat": (
                    float(last.receiver_latitude)
                    if last.receiver_latitude else None
                ),
                "receiver_lng": (
                    float(last.receiver_longitude)
                    if last.receiver_longitude else None
                ),
                "actual_lat": (
                    float(last.actual_latitude)
                    if last.actual_latitude else None
                ),
                "actual_lng": (
                    float(last.actual_longitude)
                    if last.actual_longitude else None
                ),
                "gap_distance": last.delivery_gap_distance or "",
            }
        except (ValueError, TypeError):
            geo = None

    # Detalhe driver (helper, telefone do driver)
    driver_extra = None
    if detail:
        driver_extra = {
            "courier_helper": detail.courier_helper or "",
            "courier_telephone": detail.courier_telephone or "",
            "courier_status": detail.courier_status or "",
            "pudo_locker_id": detail.pudo_locker_id or "",
        }

    summary = {
        "waybill": waybill,
        "lp_number": (
            (last.lp_number if last else "")
            or (planning.lp_code if planning else "")
            or (forecast.lp_number if forecast else "")
        ),
        "current_status": current_status,
        "current_courier": current_courier,
        "current_courier_id": (
            last.courier_id_cainiao if last else ""
        ),
        "dsp_name": last.dsp_name if last else "",
        "n_snapshots": len(ops),
        "first_seen": (
            ops[0].task_date.strftime("%Y-%m-%d") if ops else None
        ),
        "last_seen": (
            ops[-1].task_date.strftime("%Y-%m-%d") if ops else None
        ),
        "customer": customer,
        "wa_phone": wa_phone,
        "seller": seller_name,
        "weight_g": weight_g,
        "dimensions": dimensions,
        "delivery_type": last.delivery_type if last else "",
        "order_type": last.order_type if last else "",
        "destination": dest,
        "geo": geo,
        "driver_extra": driver_extra,
        "has_forecast": forecast is not None,
        "has_planning": planning is not None,
        "has_detail": detail is not None,
        "sign_pod": last.sign_pod if last else "",
        # Tempos chave do planning (forecast detalhado)
        "planning_times": ({
            "order_creation": (
                planning.order_creation_time.strftime(
                    "%Y-%m-%d %H:%M"
                ) if planning.order_creation_time else ""
            ),
            "sign_time": (
                planning.sign_time.strftime("%Y-%m-%d %H:%M")
                if planning.sign_time else ""
            ),
            "pickup_time": (
                planning.pickup_time.strftime("%Y-%m-%d %H:%M")
                if planning.pickup_time else ""
            ),
            "exception_type": planning.exception_type or "",
            "exception_reason": planning.exception_reason or "",
            "hub": planning.hub or "",
        } if planning else None),
    }

    return JsonResponse({
        "success": True,
        "summary": summary,
        "timeline": timeline,
    })


# ============================================================================
# Transferências de atribuição (waybill ↔ driver)
# ============================================================================


@login_required
def driver_day_waybills(request, driver_id):
    """Lista waybills entregues por um driver num dia específico.
    Inclui flag is_transferred (já tem override existente).
    """
    from drivers_app.models import DriverProfile
    from .models import (
        CainiaoOperationTask, WaybillAttributionOverride,
    )
    driver = get_object_or_404(DriverProfile, id=driver_id)
    date_str = request.GET.get("date", "").strip()
    target_date = parse_date(date_str)
    if not target_date:
        return JsonResponse(
            {"success": False, "error": "data inválida."}, status=400,
        )

    courier_ids = set()
    courier_names = set()
    if driver.courier_id_cainiao:
        courier_ids.add(driver.courier_id_cainiao)
    if driver.apelido:
        courier_names.add(driver.apelido)
    for m in driver.courier_mappings.filter(
        partner__name__iexact="CAINIAO",
    ):
        if m.courier_id:
            courier_ids.add(m.courier_id)
        if m.courier_name:
            courier_names.add(m.courier_name)

    # Native = entregas com login deste driver (sejam ou não transferidas)
    qs = CainiaoOperationTask.objects.filter(
        task_date=target_date, task_status="Delivered",
    )
    drv_q = Q()
    if courier_ids:
        drv_q |= Q(courier_id_cainiao__in=courier_ids)
    if courier_names:
        drv_q |= Q(courier_name__in=courier_names)
    native = qs.filter(drv_q) if drv_q else qs.none()

    # Incoming = pacotes transferidos PARA este driver (qualquer
    # courier original) — incluir mesmo que o courier original
    # não seja login do driver.
    incoming_wbs = list(
        WaybillAttributionOverride.objects.filter(
            attributed_to_driver=driver,
            task_date=target_date,
        ).values_list("waybill_number", flat=True),
    )
    if incoming_wbs:
        incoming_qs = qs.filter(waybill_number__in=incoming_wbs)
        combined = (native | incoming_qs).distinct()
    else:
        combined = native

    tasks = list(combined.order_by("-delivery_time", "waybill_number"))
    waybills = [t.waybill_number for t in tasks]
    overrides = {
        o.waybill_number: o
        for o in WaybillAttributionOverride.objects.filter(
            waybill_number__in=waybills,
        ).select_related("attributed_to_driver")
    }
    from .models import PackagePriceOverride
    price_overrides = {
        po.waybill_number: po
        for po in PackagePriceOverride.objects.filter(
            waybill_number__in=waybills,
        )
    }

    n_outgoing = 0
    n_incoming = 0
    n_self = 0
    n_priced = 0
    rows = []
    for t in tasks:
        ovr = overrides.get(t.waybill_number)
        # direction: 'self' | 'outgoing' | 'incoming'
        if ovr is None:
            direction = "self"
            n_self += 1
        elif ovr.attributed_to_driver_id == driver.id:
            direction = "incoming"
            n_incoming += 1
        else:
            direction = "outgoing"
            n_outgoing += 1
        po = price_overrides.get(t.waybill_number)
        if po:
            n_priced += 1
        rows.append({
            "waybill": t.waybill_number,
            "zip": t.zip_code,
            "city": t.destination_city,
            "address": (t.detailed_address or "")[:80],
            "delivery_time": (
                t.delivery_time.isoformat()
                if t.delivery_time else None
            ),
            "courier_name": t.courier_name,
            "direction": direction,
            "transfer": (
                {
                    "override_id": ovr.id,
                    "reason": ovr.reason,
                    "original_courier_name":
                        ovr.original_courier_name,
                    "original_courier_id":
                        ovr.original_courier_id,
                    "target_id": ovr.attributed_to_driver_id,
                    "target_name":
                        ovr.attributed_to_driver.nome_completo,
                }
                if ovr else None
            ),
            "price_override": (
                {
                    "id": po.id,
                    "price": str(po.price),
                    "reason": po.reason,
                }
                if po else None
            ),
        })

    return JsonResponse({
        "success": True,
        "driver": {
            "id": driver.id, "name": driver.nome_completo,
            "apelido": driver.apelido,
        },
        "date": target_date.isoformat(),
        "total": len(rows),
        "n_self": n_self,
        "n_outgoing": n_outgoing,
        "n_incoming": n_incoming,
        "n_priced": n_priced,
        "rows": rows,
    })


@login_required
@require_http_methods(["POST"])
def waybill_transfer_create(request):
    """Cria override(s) de atribuição para um lote de waybills."""
    import json
    from drivers_app.models import DriverProfile
    from .models import (
        CainiaoOperationTask, WaybillAttributionOverride,
    )
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse(
            {"success": False, "error": "JSON inválido."}, status=400,
        )
    waybills = data.get("waybills") or []
    if not waybills or not isinstance(waybills, list):
        return JsonResponse(
            {"success": False, "error": "waybills obrigatório"},
            status=400,
        )
    target_id = data.get("target_driver_id")
    target = get_object_or_404(DriverProfile, id=target_id)
    reason = (data.get("reason") or "").strip()[:200]

    n_created = 0
    n_updated = 0
    errors = []
    for wb in waybills:
        original = (
            CainiaoOperationTask.objects.filter(
                waybill_number=wb, task_status="Delivered",
            ).order_by("-task_date").first()
        )
        if not original:
            errors.append(f"{wb}: sem row Delivered.")
            continue
        _, created = WaybillAttributionOverride.objects.update_or_create(
            waybill_number=wb,
            defaults={
                "task_date": original.task_date,
                "original_courier_id":
                    original.courier_id_cainiao or "",
                "original_courier_name":
                    original.courier_name or "",
                "attributed_to_driver": target,
                "reason": reason,
                "transferred_by": request.user,
            },
        )
        if created:
            n_created += 1
        else:
            n_updated += 1

    return JsonResponse({
        "success": True,
        "created": n_created,
        "updated": n_updated,
        "errors": errors,
        "target": {"id": target.id, "name": target.nome_completo},
    })


@login_required
@require_http_methods(["POST"])
def waybill_transfer_delete(request, override_id):
    """Remove um override (desfaz transferência)."""
    from .models import WaybillAttributionOverride
    ovr = get_object_or_404(
        WaybillAttributionOverride, id=override_id,
    )
    wb = ovr.waybill_number
    ovr.delete()
    return JsonResponse({"success": True, "waybill": wb})


@login_required
@require_http_methods(["POST"])
def package_price_override_create(request):
    """Cria/actualiza overrides de preço para um lote de waybills.

    POST JSON: { waybills: [...], price: float, reason: str }
    Cada waybill recebe um override individual; existentes são
    sobrescritos (idempotente). Snapshot de cp4 e courier original.
    """
    import json
    from decimal import Decimal, InvalidOperation
    from .models import (
        CainiaoOperationTask, PackagePriceOverride,
    )
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse(
            {"success": False, "error": "JSON inválido"}, status=400,
        )
    waybills = data.get("waybills") or []
    if not waybills or not isinstance(waybills, list):
        return JsonResponse(
            {"success": False, "error": "waybills obrigatório"},
            status=400,
        )
    try:
        price = Decimal(str(data.get("price", "")))
    except (InvalidOperation, TypeError):
        return JsonResponse(
            {"success": False, "error": "preço inválido"}, status=400,
        )
    if price <= 0:
        return JsonResponse(
            {"success": False, "error": "preço deve ser > 0"}, status=400,
        )
    reason = (data.get("reason") or "").strip()[:200]

    n_created = 0
    n_updated = 0
    errors = []
    for wb in waybills:
        original = (
            CainiaoOperationTask.objects.filter(
                waybill_number=wb, task_status="Delivered",
            ).order_by("-task_date").first()
        )
        if not original:
            errors.append(f"{wb}: sem row Delivered")
            continue
        cp4 = (original.zip_code or "")[:4]
        _, created = PackagePriceOverride.objects.update_or_create(
            waybill_number=wb,
            defaults={
                "task_date": original.task_date,
                "cp4": cp4,
                "original_courier_name": original.courier_name or "",
                "price": price,
                "reason": reason,
                "created_by": request.user,
            },
        )
        if created:
            n_created += 1
        else:
            n_updated += 1

    return JsonResponse({
        "success": True,
        "created": n_created,
        "updated": n_updated,
        "errors": errors,
        "price": str(price),
    })


@login_required
@require_http_methods(["POST"])
def package_price_override_delete(request, override_id):
    """Apaga um override de preço."""
    from .models import PackagePriceOverride
    o = get_object_or_404(PackagePriceOverride, id=override_id)
    wb = o.waybill_number
    o.delete()
    return JsonResponse({"success": True, "waybill": wb})


@login_required
def drivers_dropdown(request):
    """Lista de drivers activos (id+name) para popular selects de UI."""
    from drivers_app.models import DriverProfile
    qs = DriverProfile.objects.filter(is_active=True).order_by(
        "nome_completo",
    )
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(nome_completo__icontains=q)
            | Q(apelido__icontains=q),
        )
    drivers = [
        {
            "id": d.id,
            "name": d.nome_completo,
            "apelido": d.apelido or "",
        }
        for d in qs[:100]
    ]
    return JsonResponse({"success": True, "drivers": drivers})
