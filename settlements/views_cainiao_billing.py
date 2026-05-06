"""Views da importação da pré-fatura Cainiao."""
from __future__ import annotations

import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods

log = logging.getLogger(__name__)


@login_required
@require_http_methods(["POST"])
def cainiao_billing_upload(request):
    """Recebe upload XLSX da pré-fatura Cainiao e dispara import.

    Idempotente: se o mesmo ficheiro já foi importado (mesmo SHA256),
    redirecciona para o detalhe existente sem criar nada.
    """
    from .services_cainiao_billing import import_cainiao_billing

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Nenhum ficheiro enviado.")
        return redirect("invoice-list")

    if not file.name.lower().endswith(".xlsx"):
        messages.error(
            request,
            "O ficheiro tem de ser .xlsx (Excel). "
            f"Recebido: {file.name}",
        )
        return redirect("invoice-list")

    try:
        session, outcome = import_cainiao_billing(
            file, file.name, user=request.user,
        )
    except ValueError as e:
        messages.error(request, f"Erro ao processar ficheiro: {e}")
        return redirect("invoice-list")
    except Exception as e:
        log.exception("Falha inesperada a importar pré-fatura Cainiao")
        messages.error(
            request,
            f"Erro inesperado: {e}. Veja os logs do servidor.",
        )
        return redirect("invoice-list")

    if outcome == "already_imported":
        messages.info(
            request,
            f"Este ficheiro já tinha sido importado em "
            f"{session.imported_at.strftime('%d/%m/%Y %H:%M')}. "
            "A abrir o detalhe existente.",
        )
    else:
        messages.success(
            request,
            f"Pré-fatura Cainiao importada: "
            f"{session.total_lines} linhas, "
            f"€{session.total_amount:.2f} líquido "
            f"({session.period_from.strftime('%d/%m')}→"
            f"{session.period_to.strftime('%d/%m/%Y')}).",
        )

    return redirect(
        "cainiao-billing-detail",
        import_id=session.id,
    )


@login_required
def cainiao_billing_detail(request, import_id):
    """Página de detalhe da importação Cainiao com 5 tabs."""
    from collections import defaultdict
    from decimal import Decimal
    from django.core.paginator import Paginator
    from django.db.models import Count, Sum
    from django.db.models.functions import Coalesce
    from .models import CainiaoBillingImport
    from .services_cainiao_billing import reconciliation_for_import

    session = get_object_or_404(
        CainiaoBillingImport.objects.select_related(
            "partner_invoice", "imported_by",
        ),
        id=import_id,
    )

    active_tab = request.GET.get("tab", "summary")
    ctx = {"session": session, "active_tab": active_tab}

    # ── Tab: Resumo ──────────────────────────────────────────────────
    if active_tab == "summary":
        # Top 10 motoristas (envio fee, agregado)
        top_drivers = list(
            session.lines.filter(fee_type="envio fee")
            .values("driver_id", "driver__nome_completo", "staff_id")
            .annotate(
                deliveries=Count("id"),
                total=Coalesce(Sum("amount"), Decimal("0")),
            )
            .order_by("-deliveries")[:10]
        )

        # Breakdown por billing_id (lote/corte)
        billing_breakdown = list(
            session.lines.values("cainiao_billing_id")
            .annotate(
                lines=Count("id"),
                total=Coalesce(Sum("amount"), Decimal("0")),
                envios=Count("id", filter=Q_envio()),
                claims=Count("id", filter=Q_claim()),
            )
            .order_by("-lines")
        )

        # Distribuição de preços (envio fee)
        price_dist = list(
            session.lines.filter(fee_type="envio fee")
            .values("amount")
            .annotate(n=Count("id"))
            .order_by("-n")[:10]
        )

        ctx.update({
            "top_drivers": top_drivers,
            "billing_breakdown": billing_breakdown,
            "price_dist": price_dist,
        })

    # ── Tab: Linhas (paginadas) ──────────────────────────────────────
    elif active_tab == "lines":
        qs = session.lines.select_related("driver", "task")
        # Filtros
        f_fee = request.GET.get("fee_type", "")
        f_billing = request.GET.get("billing_id", "")
        f_search = (request.GET.get("q") or "").strip()
        if f_fee in ("envio fee", "compensacion"):
            qs = qs.filter(fee_type=f_fee)
        if f_billing:
            qs = qs.filter(cainiao_billing_id=f_billing)
        if f_search:
            qs = qs.filter(waybill_number__icontains=f_search)
        paginator = Paginator(qs, 100)
        page = paginator.get_page(request.GET.get("page", 1))
        # Listas para os filtros
        billing_ids = list(
            session.lines.values_list(
                "cainiao_billing_id", flat=True,
            ).distinct().order_by("cainiao_billing_id"),
        )
        ctx.update({
            "lines_page": page,
            "filter_fee_type": f_fee,
            "filter_billing_id": f_billing,
            "filter_search": f_search,
            "billing_ids_filter": billing_ids,
        })

    # ── Tab: Reconciliação ───────────────────────────────────────────
    elif active_tab == "reconciliation":
        from .services_cainiao_billing import (
            diagnose_delivered_no_billing, reconciliation_math,
            delivered_no_billing_by_driver,
        )
        recon = reconciliation_for_import(session)
        diagnosis = diagnose_delivered_no_billing(session)
        recon_math = reconciliation_math(session)
        by_driver = delivered_no_billing_by_driver(session)
        # Paginar as duas listas grandes
        from django.core.paginator import Paginator as P
        ctx.update({
            "recon": recon,
            "diagnosis": diagnosis,
            "recon_math": recon_math,
            "by_driver": by_driver,
            "paid_no_task_page": P(
                recon["paid_no_task"], 50,
            ).get_page(request.GET.get("p1", 1)),
            "delivered_no_billing_page": P(
                recon["delivered_no_billing"], 50,
            ).get_page(request.GET.get("p2", 1)),
        })

    # ── Tab: Preços Especiais ────────────────────────────────────────
    elif active_tab == "special_prices":
        qs = session.lines.filter(
            fee_type="envio fee",
        ).exclude(amount=Decimal("1.60")).select_related(
            "driver", "task", "price_override",
        ).order_by("-amount", "-biz_time")
        paginator = Paginator(qs, 100)
        page = paginator.get_page(request.GET.get("page", 1))
        # Distribuição de preços
        dist = (
            session.lines.filter(fee_type="envio fee")
            .exclude(amount=Decimal("1.60"))
            .values("amount")
            .annotate(n=Count("id"), total=Sum("amount"))
            .order_by("-n")
        )
        already_overridden = qs.filter(price_override__isnull=False).count()
        ctx.update({
            "special_page": page,
            "special_dist": list(dist),
            "special_total": qs.count(),
            "already_overridden": already_overridden,
        })

    # ── Tab: Claims ──────────────────────────────────────────────────
    elif active_tab == "claims":
        from drivers_app.models import DriverProfile
        from .models import DriverClaim
        qs = (
            session.lines.filter(fee_type="compensacion")
            .select_related("driver", "task", "claim")
            .order_by("-biz_time")
        )
        # Drivers activos para o select (compactos)
        drivers_options = list(
            DriverProfile.objects
            .filter(is_active=True)
            .values("id", "nome_completo", "apelido", "courier_id_cainiao")
            .order_by("nome_completo")
        )
        ctx.update({
            "claim_lines": list(qs),
            "claims_total_value": qs.aggregate(
                s=Coalesce(Sum("amount"), Decimal("0")),
            )["s"],
            "claims_unassigned": qs.filter(claim__isnull=True).count(),
            "claims_assigned": qs.filter(claim__isnull=False).count(),
            "claims_with_suggestion": qs.filter(
                claim__isnull=True, driver__isnull=False,
            ).count(),
            "drivers_options": drivers_options,
            "claim_types": DriverClaim.CLAIM_TYPES,
        })

    return render(
        request, "settlements/cainiao_billing_detail.html", ctx,
    )


def Q_envio():
    from django.db.models import Q
    return Q(fee_type="envio fee")


def Q_claim():
    from django.db.models import Q
    return Q(fee_type="compensacion")


@login_required
@require_http_methods(["POST"])
def cainiao_billing_assign_claim(request, import_id, line_id):
    """Cria ou actualiza DriverClaim para uma compensación.

    Body:
      driver_id: ID do DriverProfile (obrigatório)
      claim_type: ORDER_LOSS / OTHER / ... (default ORDER_LOSS)
    """
    from drivers_app.models import DriverProfile
    from .models import (
        CainiaoBillingImport, CainiaoBillingLine, DriverClaim,
    )

    session = get_object_or_404(CainiaoBillingImport, id=import_id)
    line = get_object_or_404(
        CainiaoBillingLine, id=line_id, import_session=session,
    )
    if line.fee_type != "compensacion":
        return JsonResponse({
            "success": False,
            "error": "Esta linha não é uma compensación.",
        }, status=400)
    if line.claim_id:
        return JsonResponse({
            "success": False,
            "error": (
                f"Linha já tem DriverClaim #{line.claim_id} "
                "associado."
            ),
        }, status=400)

    driver_id = request.POST.get("driver_id")
    if not driver_id:
        return JsonResponse({
            "success": False, "error": "driver_id obrigatório.",
        }, status=400)
    driver = get_object_or_404(DriverProfile, id=driver_id)

    claim_type = request.POST.get("claim_type") or "ORDER_LOSS"
    if claim_type not in dict(DriverClaim.CLAIM_TYPES):
        claim_type = "ORDER_LOSS"

    # Valor: o XLSX tem negativo (-30). DriverClaim usa positivo.
    amount = abs(line.amount)

    description = (
        f"Cainiao Billing #{session.id} ({line.cainiao_billing_id}). "
        f"Waybill: {line.waybill_number}. "
        f"Motivo: {line.fb2 or '—'}"
    )[:2000]

    claim = DriverClaim.objects.create(
        driver=driver,
        claim_type=claim_type,
        amount=amount,
        description=description,
        occurred_at=line.biz_time,
        waybill_number=line.waybill_number,
        operation_task_date=(
            line.task.task_date if line.task else line.biz_time.date()
        ),
        auto_detected=False,
        created_by=(
            request.user if request.user.is_authenticated else None
        ),
    )
    line.claim = claim
    line.driver = driver
    line.save(update_fields=["claim", "driver"])

    return JsonResponse({
        "success": True,
        "claim_id": claim.id,
        "driver_name": driver.nome_completo,
        "amount": str(amount),
    })


@login_required
@require_http_methods(["POST"])
def cainiao_billing_assign_claims_bulk(request, import_id):
    """Cria DriverClaim para todas as compensaciones com driver sugerido.

    Driver sugerido = line.driver (resolvido no import via task) ou
    line.task.courier_id_cainiao mapeado para DriverProfile.

    Linhas sem driver sugerido são ignoradas (devolvidas como skipped).
    """
    from .models import (
        CainiaoBillingImport, DriverClaim,
    )
    from django.db import transaction

    session = get_object_or_404(CainiaoBillingImport, id=import_id)
    claim_type = request.POST.get("claim_type") or "ORDER_LOSS"
    if claim_type not in dict(DriverClaim.CLAIM_TYPES):
        claim_type = "ORDER_LOSS"

    qs = session.lines.filter(
        fee_type="compensacion",
        claim__isnull=True,
        driver__isnull=False,
    ).select_related("driver", "task")

    created = 0
    with transaction.atomic():
        for line in qs:
            description = (
                f"Cainiao Billing #{session.id} "
                f"({line.cainiao_billing_id}). "
                f"Waybill: {line.waybill_number}. "
                f"Motivo: {line.fb2 or '—'}"
            )[:2000]
            claim = DriverClaim.objects.create(
                driver=line.driver,
                claim_type=claim_type,
                amount=abs(line.amount),
                description=description,
                occurred_at=line.biz_time,
                waybill_number=line.waybill_number,
                operation_task_date=(
                    line.task.task_date if line.task
                    else line.biz_time.date()
                ),
                auto_detected=False,
                created_by=(
                    request.user if request.user.is_authenticated else None
                ),
            )
            line.claim = claim
            line.save(update_fields=["claim"])
            created += 1

    skipped_no_driver = session.lines.filter(
        fee_type="compensacion", claim__isnull=True, driver__isnull=True,
    ).count()

    if created:
        messages.success(
            request,
            f"{created} DriverClaim criados. "
            f"{skipped_no_driver} ignorados (sem driver sugerido — "
            "atribua manualmente).",
        )
    else:
        messages.warning(
            request,
            f"Nenhum claim criado. "
            f"{skipped_no_driver} linhas precisam de atribuição manual.",
        )
    return redirect(
        f"{reverse('cainiao-billing-detail', args=[session.id])}"
        f"?tab=claims",
    )


@login_required
@require_http_methods(["POST"])
def cainiao_billing_create_overrides(request, import_id):
    """Cria PackagePriceOverride para linhas com preço ≠ €1.60.

    Body:
      mode = "all" | "selected"
      line_ids = "1,2,3" (necessário se mode=selected)
    """
    from decimal import Decimal
    from django.db import transaction
    from .models import (
        CainiaoBillingImport, CainiaoBillingLine, PackagePriceOverride,
    )

    session = get_object_or_404(CainiaoBillingImport, id=import_id)
    mode = request.POST.get("mode", "selected")

    qs = session.lines.filter(
        fee_type="envio fee",
    ).exclude(amount=Decimal("1.60")).filter(price_override__isnull=True)

    if mode == "selected":
        ids = (request.POST.get("line_ids") or "").split(",")
        ids = [int(i) for i in ids if i.strip().isdigit()]
        if not ids:
            messages.error(request, "Nenhuma linha seleccionada.")
            return redirect(
                "cainiao-billing-detail",
                import_id=session.id,
            )
        qs = qs.filter(id__in=ids)

    created = 0
    skipped_existing = 0
    skipped_no_waybill = 0

    with transaction.atomic():
        for line in qs.select_related("task"):
            if not line.waybill_number:
                skipped_no_waybill += 1
                continue
            # Se já existe override (criado fora desta tab), só liga
            existing = PackagePriceOverride.objects.filter(
                waybill_number=line.waybill_number,
            ).first()
            if existing:
                line.price_override = existing
                line.save(update_fields=["price_override"])
                skipped_existing += 1
                continue
            task_date = (
                line.task.task_date if line.task
                else line.biz_time.date()
            )
            cp4 = ""
            original_name = ""
            if line.task:
                cp4 = (line.task.zip_code or "")[:4]
                original_name = line.task.courier_name or ""
            override = PackagePriceOverride.objects.create(
                waybill_number=line.waybill_number,
                task_date=task_date,
                cp4=cp4,
                original_courier_name=original_name,
                price=line.amount,
                reason=(
                    f"Cainiao billing import #{session.id} "
                    f"({session.period_from.strftime('%d/%m')}"
                    f"–{session.period_to.strftime('%d/%m/%Y')})"
                )[:200],
                created_by=(
                    request.user if request.user.is_authenticated else None
                ),
            )
            line.price_override = override
            line.save(update_fields=["price_override"])
            created += 1

    if created:
        messages.success(
            request,
            f"{created} PackagePriceOverride criados. "
            f"{skipped_existing} já existiam (linkados). "
            f"{skipped_no_waybill} ignorados (sem waybill).",
        )
    elif skipped_existing:
        messages.info(
            request,
            f"{skipped_existing} overrides já existiam — todos linkados.",
        )
    else:
        messages.warning(request, "Nenhum override criado.")

    return redirect(
        f"{reverse('cainiao-billing-detail', args=[session.id])}"
        f"?tab=special_prices",
    )


@login_required
def cainiao_billing_export_xlsx(request, import_id, kind):
    """Exporta listas da tab Reconciliação em XLSX.

    kind ∈ {paid_no_task, delivered_no_billing, all_lines, special_prices,
            claims}.
    """
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from .models import CainiaoBillingImport, CainiaoOperationTask
    from .services_cainiao_billing import reconciliation_for_import

    session = get_object_or_404(CainiaoBillingImport, id=import_id)

    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        "solid", fgColor="5B21B6",  # violet
    )

    def _set_headers(headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

    if kind == "paid_no_task":
        ws.title = "Pagas s_ task"
        _set_headers([
            "Data", "Waybill (REFs)", "Staff ID", "Lote (Billing ID)",
            "Cidade", "Valor (€)", "Motivo (FB2)",
        ])
        recon = reconciliation_for_import(session)
        for line in recon["paid_no_task"].order_by("biz_time"):
            ws.append([
                line.biz_time.strftime("%Y-%m-%d %H:%M:%S"),
                line.waybill_number,
                line.staff_id or "",
                line.cainiao_billing_id or "",
                line.ciudad or "",
                float(line.amount),
                (line.fb2 or "")[:500],
            ])
        filename = (
            f"cainiao_pagas_sem_task_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )

    elif kind == "delivered_no_billing":
        ws.title = "Entregues s_ pagamento"
        _set_headers([
            "Task Date", "Waybill", "LP Number", "Courier ID Cainiao",
            "Courier Name", "Status", "Cidade", "CP",
        ])
        # Já filtra só 1ª entrega por waybill (re-entregas excluídas)
        recon = reconciliation_for_import(session)
        for t in recon["delivered_no_billing"].order_by("task_date"):
            ws.append([
                t.task_date.strftime("%Y-%m-%d"),
                t.waybill_number,
                t.lp_number or "",
                t.courier_id_cainiao or "",
                t.courier_name or "",
                t.task_status,
                t.destination_city or "",
                t.zip_code or "",
            ])
        filename = (
            f"cainiao_entregues_sem_pagamento_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )

    elif kind == "all_lines":
        ws.title = "Todas as linhas"
        _set_headers([
            "Tipo", "Data", "Waybill (REFs)", "Staff ID",
            "Driver (sistema)", "Lote (Billing ID)", "Cidade",
            "Valor (€)", "Motivo (FB2)", "Task local?",
            "Task Status", "CP4",
        ])
        for line in session.lines.select_related("driver", "task").iterator(
            chunk_size=1000,
        ):
            ws.append([
                line.fee_type,
                line.biz_time.strftime("%Y-%m-%d %H:%M:%S"),
                line.waybill_number,
                line.staff_id or "",
                line.driver.nome_completo if line.driver else "",
                line.cainiao_billing_id or "",
                line.ciudad or "",
                float(line.amount),
                (line.fb2 or "")[:500],
                "SIM" if line.task_id else "NAO",
                line.task.task_status if line.task else "",
                (line.task.zip_code[:4] if line.task and line.task.zip_code
                 else ""),
            ])
        filename = (
            f"cainiao_todas_linhas_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )

    elif kind == "special_prices":
        from decimal import Decimal as D
        ws.title = "Precos especiais"
        _set_headers([
            "Data", "Waybill", "Staff ID", "Driver", "Lote",
            "CP4", "Valor (€)", "Override Criado?", "Override ID",
        ])
        qs = session.lines.filter(fee_type="envio fee").exclude(
            amount=D("1.60"),
        ).select_related("driver", "task", "price_override")
        for line in qs.order_by("-amount"):
            ws.append([
                line.biz_time.strftime("%Y-%m-%d %H:%M:%S"),
                line.waybill_number,
                line.staff_id or "",
                line.driver.nome_completo if line.driver else "",
                line.cainiao_billing_id or "",
                (line.task.zip_code[:4] if line.task and line.task.zip_code
                 else ""),
                float(line.amount),
                "SIM" if line.price_override_id else "NAO",
                line.price_override_id or "",
            ])
        filename = (
            f"cainiao_precos_especiais_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )

    elif kind == "unpaid_by_driver":
        from .services_cainiao_billing import (
            delivered_no_billing_by_driver,
        )
        from .models import CainiaoOperationTask
        from django.db.models import Q as _Q

        wb.remove(ws)

        # Sheet 1 — Resumo agrupado
        sh = wb.create_sheet(title="Resumo por driver")
        sh.append([
            "Driver (sistema)", "Apelido", "Courier ID Cainiao",
            "Courier Name (XLSX)", "Pacotes não pagos",
            "Perda estimada (€)", "Primeira data", "Última data",
        ])
        for col in range(1, 9):
            cell = sh.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        groups = delivered_no_billing_by_driver(session)
        for g in groups:
            sh.append([
                g["driver_name"] or "(não resolvido)",
                g["apelido"],
                g["courier_id_cainiao"],
                g["courier_name"],
                g["n_packages"],
                float(g["estimated_loss"]),
                (g["first_date"].strftime("%Y-%m-%d")
                 if g["first_date"] else ""),
                (g["last_date"].strftime("%Y-%m-%d")
                 if g["last_date"] else ""),
            ])

        # Sheet 2 — Lista completa de pacotes
        sh2 = wb.create_sheet(title="Pacotes detalhados")
        sh2.append([
            "Task Date", "Waybill", "LP Number", "Driver (sistema)",
            "Courier ID Cainiao", "Courier Name", "Cidade", "CP",
            "Status", "Perda estimada (€)",
        ])
        for col in range(1, 11):
            cell = sh2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        billed_wbs = set(
            session.lines.filter(fee_type="envio fee").values_list(
                "waybill_number", flat=True,
            )
        )
        from .services_cainiao_billing import (
            _build_resolution_caches, _first_delivery_task_ids,
        )
        cid_to_driver, _ = _build_resolution_caches()
        first_ids = _first_delivery_task_ids(
            session.period_from, session.period_to,
        )
        qs = CainiaoOperationTask.objects.filter(
            id__in=first_ids,
        ).exclude(
            _Q(waybill_number__in=billed_wbs)
            | _Q(lp_number__in=billed_wbs),
        ).order_by("courier_name", "task_date")

        for t in qs.iterator(chunk_size=1000):
            d = (
                cid_to_driver.get(t.courier_id_cainiao)
                if t.courier_id_cainiao else None
            )
            sh2.append([
                t.task_date.strftime("%Y-%m-%d"),
                t.waybill_number,
                t.lp_number or "",
                d.nome_completo if d else "",
                t.courier_id_cainiao or "",
                t.courier_name or "",
                t.destination_city or "",
                t.zip_code or "",
                t.task_status,
                1.60,
            ])

        for sheet in (sh, sh2):
            for col_cells in sheet.columns:
                max_len = 0
                for cell in col_cells:
                    v = str(cell.value) if cell.value is not None else ""
                    if len(v) > max_len:
                        max_len = len(v)
                sheet.column_dimensions[
                    col_cells[0].column_letter
                ].width = min(max_len + 2, 60)

        filename = (
            f"cainiao_nao_pagos_por_driver_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        return response

    elif kind == "diagnosis":
        from .services_cainiao_billing import diagnose_delivered_no_billing
        diag = diagnose_delivered_no_billing(session)
        # Remove a sheet default; vamos criar uma por categoria
        wb.remove(ws)
        categories = [
            ("billed_other_import", "Facturado outro import"),
            ("compensation_this_import", "Compensacao este import"),
            ("compensation_other_import", "Compensacao outro import"),
            ("duplicate_locally", "Duplicado localmente"),
            ("genuinely_unpaid", "Genuinamente nao pago"),
        ]
        any_data = False
        for key, sheet_name in categories:
            items = diag.get(key) or []
            if not items:
                continue
            any_data = True
            sh = wb.create_sheet(title=sheet_name[:31])
            sh.append([
                "Task Date", "Waybill", "LP Number", "Courier ID",
                "Courier Name", "Cidade", "Causa identificada",
            ])
            for col in range(1, 8):
                cell = sh.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            for t in items:
                sh.append([
                    t["task_date"].strftime("%Y-%m-%d") if t.get(
                        "task_date") else "",
                    t.get("waybill_number") or "",
                    t.get("lp_number") or "",
                    t.get("courier_id_cainiao") or "",
                    t.get("courier_name") or "",
                    t.get("destination_city") or "",
                    t.get("_reason") or "",
                ])
            for col_cells in sh.columns:
                max_len = 0
                for cell in col_cells:
                    v = str(cell.value) if cell.value is not None else ""
                    if len(v) > max_len:
                        max_len = len(v)
                sh.column_dimensions[col_cells[0].column_letter].width = (
                    min(max_len + 2, 60)
                )
        if not any_data:
            sh = wb.create_sheet(title="Sem dados")
            sh.append(["(Nenhuma divergência detectada)"])
        filename = (
            f"cainiao_diagnostico_reconciliacao_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )
        # ... segue para a parte de stream do response
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'
        )
        return response

    elif kind == "claims":
        ws.title = "Compensacoes (claims)"
        _set_headers([
            "Data", "Waybill", "Lote", "Driver Sugerido",
            "DriverClaim Criado?", "DriverClaim ID", "Valor (€)",
            "Motivo (FB2)",
        ])
        qs = session.lines.filter(fee_type="compensacion").select_related(
            "driver", "task", "claim",
        )
        for line in qs.order_by("biz_time"):
            ws.append([
                line.biz_time.strftime("%Y-%m-%d %H:%M:%S"),
                line.waybill_number,
                line.cainiao_billing_id or "",
                line.driver.nome_completo if line.driver else "",
                "SIM" if line.claim_id else "NAO",
                line.claim_id or "",
                float(abs(line.amount)),
                (line.fb2 or "")[:500],
            ])
        filename = (
            f"cainiao_claims_{session.period_from}_"
            f"{session.period_to}.xlsx"
        )

    else:
        return HttpResponse(
            f"Tipo de export desconhecido: {kind}", status=400,
        )

    # Auto-resize colunas (best effort)
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            v = str(cell.value) if cell.value is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )
    return response


@login_required
@require_http_methods(["POST"])
def cainiao_billing_reresolve(request, import_id):
    """Re-corre a resolução de task/driver para um import existente.

    Útil quando o EPOD Task List é importado DEPOIS da pré-fatura
    Cainiao — as tasks agora existem mas o matching original não
    as encontrou.
    """
    from .models import CainiaoBillingImport
    from .services_cainiao_billing import reresolve_matching

    session = get_object_or_404(CainiaoBillingImport, id=import_id)
    try:
        result = reresolve_matching(session)
    except Exception as e:
        log.exception("reresolve_matching falhou")
        messages.error(request, f"Erro ao reprocessar matching: {e}")
        return redirect(
            "cainiao-billing-detail", import_id=session.id,
        )

    messages.success(
        request,
        f"Matching reprocessado: "
        f"{result['newly_resolved_task']} tasks ligadas, "
        f"{result['newly_resolved_driver']} drivers ligados "
        f"(em {result['total']} linhas).",
    )
    return redirect(
        "cainiao-billing-detail", import_id=session.id,
    )


@login_required
@require_http_methods(["POST"])
def cainiao_billing_delete(request, import_id):
    """Apagar uma sessão de importação (cascade nas linhas).

    Se houver PartnerInvoice associada e estiver PAGA, bloqueia.
    """
    from .models import CainiaoBillingImport

    session = get_object_or_404(CainiaoBillingImport, id=import_id)

    if session.partner_invoice and session.partner_invoice.status == "PAID":
        return JsonResponse({
            "success": False,
            "error": (
                "Não é possível apagar — a PartnerInvoice associada já "
                "está marcada como paga. Cancele primeiro a fatura."
            ),
        }, status=400)

    summary = (
        f"Cainiao {session.period_from}→{session.period_to} "
        f"({session.total_lines} linhas)"
    )
    if session.partner_invoice:
        session.partner_invoice.delete()
    session.delete()

    messages.success(request, f"Importação eliminada: {summary}")
    return redirect("invoice-list")
